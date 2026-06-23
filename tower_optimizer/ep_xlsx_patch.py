"""Patch Effective Paths xlsx without openpyxl save (preserves Excel features)."""
from __future__ import annotations

import re
import xml.etree.ElementTree as ET
import zipfile
from pathlib import Path
from typing import Any, Dict, Mapping, Optional, Tuple

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from .engines.core import (
    ENHANCEMENT_MAX_LEVELS,
    LAB_ALIASES,
    LAB_MAX_LEVELS,
    UW_ATTRIBUTE_META,
    UW_NAMES,
    WORKSHOP_MAX_LEVELS,
)
from .ep_master_sheet import _clean, _detect_columns
from .quality import WORKSHOP_ALIASES

MAIN_NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
NS = {"m": MAIN_NS, "r": "http://schemas.openxmlformats.org/officeDocument/2006/relationships"}


def _cell_ref(row: int, col: int) -> str:
    return f"{get_column_letter(col)}{row}"


def collect_master_sheet_cell_updates(
    source_xlsx: Path,
    profile: Mapping[str, Any],
    *,
    max_row: int = 120,
) -> Dict[str, Any]:
    """Read Master Sheet layout and return {A1-style ref: scalar value}."""
    workbook = load_workbook(source_xlsx, read_only=True, data_only=True)
    try:
        sheet = workbook["Master Sheet"]
        cols = _detect_columns(sheet)
        header_row = cols["header_row"]
        lab_col = cols["lab_col"]
        workshop_col = cols["workshop_col"]
        enhancement_col = cols["enhancement_col"]
        uw_header_col = cols["uw_header_col"]
        uw_name_col = uw_header_col + 1
        uw_attr_col = uw_header_col + 2
        uw_value_col = uw_header_col + 3

        labs = profile.get("labs") or {}
        workshop = profile.get("workshop") or {}
        enhancements = profile.get("enhancements") or {}
        uw_data = profile.get("uw") or {}

        updates: Dict[str, Any] = {}
        stats = {"labs": 0, "workshop": 0, "enhancements": 0, "uw_attributes": 0, "uw_owned": 0}
        current_uw: Optional[str] = None

        for row_idx in range(header_row + 1, max_row + 1):
            lab_name = _clean(sheet.cell(row_idx, lab_col).value)
            if lab_name:
                canonical_lab = LAB_ALIASES.get(lab_name, lab_name)
                if canonical_lab in LAB_MAX_LEVELS and canonical_lab in labs:
                    updates[_cell_ref(row_idx, lab_col + 1)] = int(labs[canonical_lab])
                    stats["labs"] += 1

            workshop_name = _clean(sheet.cell(row_idx, workshop_col).value)
            if workshop_name:
                canonical_ws = WORKSHOP_ALIASES.get(workshop_name, workshop_name)
                if canonical_ws in WORKSHOP_MAX_LEVELS and canonical_ws in workshop:
                    level = int(workshop[canonical_ws])
                    updates[_cell_ref(row_idx, workshop_col + 1)] = level
                    updates[_cell_ref(row_idx, workshop_col + 2)] = level
                    stats["workshop"] += 1

            enhancement_name = _clean(sheet.cell(row_idx, enhancement_col).value)
            if enhancement_name in ENHANCEMENT_MAX_LEVELS and enhancement_name in enhancements:
                updates[_cell_ref(row_idx, enhancement_col + 2)] = int(enhancements[enhancement_name])
                stats["enhancements"] += 1

            name_or_status = _clean(sheet.cell(row_idx, uw_name_col).value)
            attribute = _clean(sheet.cell(row_idx, uw_attr_col).value)

            if name_or_status in UW_NAMES:
                current_uw = name_or_status
            elif name_or_status == "UW Unlocked" and current_uw and current_uw in uw_data:
                owned = bool((uw_data[current_uw] or {}).get("owned"))
                updates[_cell_ref(row_idx, uw_name_col)] = "UW Unlocked" if owned else "UW Locked"
                stats["uw_owned"] += 1
            elif current_uw and current_uw in uw_data and attribute in UW_ATTRIBUTE_META.get(current_uw, {}):
                attrs = (uw_data[current_uw] or {}).get("attributes") or {}
                if attribute in attrs:
                    value = attrs[attribute]
                    meta = UW_ATTRIBUTE_META[current_uw][attribute]
                    updates[_cell_ref(row_idx, uw_value_col)] = (
                        int(round(value)) if isinstance(meta.get("max"), int) else float(value)
                    )
                    stats["uw_attributes"] += 1

        return {"updates": updates, "stats": stats}
    finally:
        workbook.close()


def _master_sheet_path(zf: zipfile.ZipFile) -> str:
    wb_root = ET.fromstring(zf.read("xl/workbook.xml"))
    rel_root = ET.fromstring(zf.read("xl/_rels/workbook.xml.rels"))
    rel_map = {rel.attrib["Id"]: rel.attrib["Target"] for rel in rel_root}
    for sheet in wb_root.find("m:sheets", NS):
        if sheet.attrib.get("name") == "Master Sheet":
            rid = sheet.attrib[f"{{{NS['r']}}}id"]
            target = rel_map[rid]
            return "xl/" + target.lstrip("/")
    raise ValueError("Master Sheet not found in workbook.")


def _load_shared_strings(zf: zipfile.ZipFile) -> Tuple[list[str], Optional[bytes]]:
    try:
        payload = zf.read("xl/sharedStrings.xml")
    except KeyError:
        return [], None
    root = ET.fromstring(payload)
    strings: list[str] = []
    for si in root.findall("m:si", NS):
        parts = [node.text or "" for node in si.iter(f"{{{MAIN_NS}}}t")]
        strings.append("".join(parts))
    return strings, payload


def _append_shared_string(strings: list[str], value: str, shared_xml: Optional[bytes]) -> Tuple[int, bytes]:
    if value in strings:
        return strings.index(value), shared_xml or b""
    strings.append(value)
    if shared_xml is None:
        root = ET.Element(f"{{{MAIN_NS}}}sst", count="1", uniqueCount="1")
    else:
        root = ET.fromstring(shared_xml)
    si = ET.SubElement(root, f"{{{MAIN_NS}}}si")
    t = ET.SubElement(si, f"{{{MAIN_NS}}}t")
    t.text = value
    root.attrib["count"] = str(int(root.attrib.get("count", "0")) + 1)
    root.attrib["uniqueCount"] = str(len(strings))
    return len(strings) - 1, ET.tostring(root, encoding="utf-8", xml_declaration=True)


def _set_cell_value(cell: ET.Element, value: Any, *, string_index: Optional[int] = None) -> None:
    for child in list(cell):
        if child.tag != f"{{{MAIN_NS}}}v":
            cell.remove(child)
    if string_index is not None:
        cell.attrib["t"] = "s"
        value_elem = cell.find("m:v", NS)
        if value_elem is None:
            value_elem = ET.SubElement(cell, f"{{{MAIN_NS}}}v")
        value_elem.text = str(string_index)
        return
    if isinstance(value, str):
        raise ValueError("String cell updates require shared string index.")
    cell.attrib.pop("t", None)
    value_elem = cell.find("m:v", NS)
    if value_elem is None:
        value_elem = ET.SubElement(cell, f"{{{MAIN_NS}}}v")
    value_elem.text = repr(value) if isinstance(value, float) else str(value)


def patch_master_sheet_values(source_xlsx: Path, dest_xlsx: Path, updates: Mapping[str, Any]) -> None:
    """Copy xlsx and patch Master Sheet scalar cells without openpyxl save."""
    with zipfile.ZipFile(source_xlsx, "r") as zin:
        sheet_path = _master_sheet_path(zin)
        shared_strings, shared_xml = _load_shared_strings(zin)
        sheet_root = ET.fromstring(zin.read(sheet_path))
        zip_entries = list(zin.infolist())
    rows_by_index: Dict[int, ET.Element] = {}
    cells_by_ref: Dict[str, ET.Element] = {}
    sheet_data = sheet_root.find("m:sheetData", NS)
    if sheet_data is None:
        raise ValueError("Master Sheet is missing sheetData.")

    for row in sheet_data.findall("m:row", NS):
        row_idx = int(row.attrib["r"])
        rows_by_index[row_idx] = row
        for cell in row.findall("m:c", NS):
            ref = cell.attrib.get("r")
            if ref:
                cells_by_ref[ref] = cell

    shared_changed = False
    for ref, value in updates.items():
        if isinstance(value, str):
            string_index, shared_xml = _append_shared_string(shared_strings, value, shared_xml)
            shared_changed = True
            if ref in cells_by_ref:
                _set_cell_value(cells_by_ref[ref], value, string_index=string_index)
            continue
        if ref not in cells_by_ref:
            match = re.match(r"([A-Z]+)(\d+)", ref)
            if not match:
                continue
            row_num = int(match.group(2))
            row = rows_by_index.get(row_num)
            if row is None:
                row = ET.SubElement(sheet_data, f"{{{MAIN_NS}}}row", r=str(row_num))
                rows_by_index[row_num] = row
            cell = ET.SubElement(row, f"{{{MAIN_NS}}}c", r=ref)
            cells_by_ref[ref] = cell
        _set_cell_value(cells_by_ref[ref], value)

    patched_sheet = ET.tostring(sheet_root, encoding="utf-8", xml_declaration=True)
    patched_shared = shared_xml if shared_changed else None

    with zipfile.ZipFile(source_xlsx, "r") as zin, zipfile.ZipFile(dest_xlsx, "w") as zout:
        for info in zip_entries:
            if info.filename == sheet_path:
                payload = patched_sheet
            elif patched_shared is not None and info.filename == "xl/sharedStrings.xml":
                payload = patched_shared
            else:
                payload = zin.read(info.filename)
            zout.writestr(info, payload)


def fill_effective_paths_workbook(source_xlsx: Path, dest_xlsx: Path, profile: Mapping[str, Any]) -> Dict[str, int]:
    collected = collect_master_sheet_cell_updates(source_xlsx, profile)
    patch_master_sheet_values(source_xlsx, dest_xlsx, collected["updates"])
    return collected["stats"]
