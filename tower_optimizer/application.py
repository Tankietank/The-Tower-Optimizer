import csv
import io
import json
import re
import warnings
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Dict, Optional

import streamlit as st
import pandas as pd
from openpyxl import load_workbook

from .battle_learning import (
    ensure_battle_learning_state, normalize_profile_runs, build_battle_learning_report,
    import_runs, prepare_import_batch,
)
from .battle_parser import parse_battle_report_text, parse_battle_report_batch

from .reliability import (
    SUPPORTED_WORKBOOK_VERSIONS,
    atomic_save_json,
    build_diagnostic_zip,
    compare_cap_maps,
    create_profile_backup,
    extract_effective_paths_caps,
    list_profile_backups,
    restore_profile_backup,
    workbook_compatibility,
)
from .game_data_updater import (
    active_update_health,
    analyze_update_bundle,
    apply_update_candidate,
    export_update_bundle,
    get_active_update,
    list_update_history,
    load_runtime_overlay,
    rollback_active_update,
    stage_update_candidate,
)

warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

st.set_page_config(page_title="Tower Optimizer", layout="wide")

APP_VERSION = "2.0.0-preview.6"
DATA_SOURCE_VERSION = "Effective Paths v5.06.04.00"
DATA_SOURCE_DETAIL = "Bundled standalone game data + v2 visual preview 6 + one-click multi-report importing + expandable navigation + persistent custom icon overrides + reviewed local update overlays"

from .runtime_paths import data_dir, profiles_dir

DATA_DIR = data_dir()
PROFILE_DIR = profiles_dir()
PROFILE_DIR.mkdir(parents=True, exist_ok=True)

# -----------------------------------------------------------------------------
# AUTHORITATIVE GAME METADATA
# -----------------------------------------------------------------------------
# The lists, order, and maximums below were extracted directly from the _IDS
# sheet in Effective Paths v5.06.04.00. Workshop and Lab entries therefore all
# have a known maximum in this build.

WORKSHOP_MAX_LEVELS: Dict[str, int] = {'Damage': 6000,
 'Attack Speed': 99,
 'Critical Chance': 79,
 'Critical Factor': 150,
 'Range': 79,
 'Damage / Meter': 200,
 'Multishot Chance': 99,
 'Multishot Targets': 7,
 'Rapid Fire Chance': 85,
 'Rapid Fire Duration': 99,
 'Bounce Shot Chance': 85,
 'Bounce Shot Targets': 7,
 'Bounce Shot Range': 60,
 'Super Critical Chance': 100,
 'Super Critical Mult': 120,
 'Rend Armor Chance': 299,
 'Rend Armor Mult': 299,
 'Health': 6000,
 'Health Regen': 6000,
 'Defense %': 99,
 'Defense Absolute': 5000,
 'Thorn Damage': 99,
 'Lifesteal': 80,
 'Knockback Chance': 80,
 'Knockback Force': 40,
 'Orb Speed': 38,
 'Orbs': 4,
 'Shockwave Size': 35,
 'Shockwave Frequency': 40,
 'Land Mine Chance': 50,
 'Land Mine Damage': 200,
 'Land Mine Radius': 50,
 'Death Defy': 75,
 'Wall Health': 1800,
 'Wall Rebuild': 300,
 'Cash Bonus': 149,
 'Cash / Wave': 149,
 'Coin / Kill Bonus': 149,
 'Coin / Wave': 149,
 'Free Attack Upgrade': 99,
 'Free Defense Upgrade': 99,
 'Free Utility Upgrade': 99,
 'Interest / Wave': 99,
 'Recovery Amount': 300,
 'Max Amount': 500,
 'Package Chance': 60,
 'Enemy Attack Level Skip': 699,
 'Enemy Health Level Skip': 699}

LAB_MAX_LEVELS: Dict[str, int] = {'Game Speed': 7,
 'Starting Cash': 99,
 'Workshop Attack Discount': 99,
 'Workshop Defense Discount': 99,
 'Workshop Utility Discount': 99,
 'Labs Coin Discount': 99,
 'Labs Speed': 99,
 'Buy Multiplier': 4,
 'More Round Stats': 1,
 'Target Priority': 2,
 'Card Presets': 1,
 'Workshop Respec': 1,
 'Reroll Daily Mission': 1,
 'Workshop Enhancements': 1,
 'Enhancement Attack - Coin Discount': 100,
 'Enhancement Defense - Coin Discount': 100,
 'Enhancement Utility - Coin Discount': 100,
 'Dissonant Echo - Attack': 20,
 'Dissonant Echo - Defense': 20,
 'Dissonant Echo - Utility': 20,
 'Dissonant Echo - Ultimate Weapons': 20,
 'Damage': 100,
 'Attack Speed': 99,
 'Critical Factor': 99,
 'Range': 80,
 'Damage / Meter': 99,
 'Super Crit Chance': 50,
 'Super Crit Multi': 40,
 'Max Rend Armor Multiplier': 30,
 'Light Speed Shots': 1,
 'Health': 100,
 'Health Regen': 100,
 'Defense Absolute': 100,
 'Defense %': 50,
 'Orbs Speed': 20,
 'Land Mine Damage': 20,
 'Land Mine Decay': 35,
 'Shockwave Size': 20,
 'Orb Boss Hit': 10,
 'Wall Health': 50,
 'Wall Rebuild': 20,
 'Wall Regen': 30,
 'Wall Thorns': 20,
 'Wall Invincibility': 10,
 'Wall Fortification': 60,
 'Garlic Thorns': 10,
 'Cash Bonus': 99,
 'Cash / Wave': 99,
 'Coins / Kill Bonus': 99,
 'Coins / Wave': 99,
 'Interest': 99,
 'Max Interest': 15,
 'Package After Boss': 1,
 'Recovery Package Amount': 20,
 'Recovery Package Max': 20,
 'Recovery Package Chance': 20,
 'Enemy Attack Level Skip': 20,
 'Enemy Health Level Skip': 20,
 'Missile Despawn Time': 20,
 'Missiles Explosion': 1,
 'Missile Radius': 20,
 'Chrono Field Duration': 30,
 'Chrono Field Damage Reduction': 1,
 'Chrono Field Reduction %': 30,
 'Swamp Radius': 30,
 'Swamp Stun': 1,
 'Swamp Stun Chance': 30,
 'Swamp Stun Time': 30,
 'Golden Tower Bonus': 25,
 'Golden Tower Duration': 20,
 'Chain Lightning Shock': 1,
 'Shock Chance': 30,
 'Shock Multiplier': 14,
 'Death Wave Health': 30,
 'Death Wave Coin Bonus': 20,
 'Inner Mine Blast Radius': 20,
 'Inner Mine Rotation Speed': 20,
 'Chrono Field Range': 20,
 'Missile Amplifier': 25,
 'Missile Barrage': 1,
 'Missile Barrage Quantity': 6,
 'Inner Mine Stun': 1,
 'Black Hole Damage': 10,
 'Extra Black Hole': 1,
 'Black Hole Coin Bonus': 20,
 'Spotlight Coin Bonus': 20,
 'Spotlight Missiles': 18,
 'Black Hole Disable Ranged Enemies': 1,
 'Recharge Missile Barrage': 7,
 'Swamp Rend': 30,
 'Swamp Rend - Additional Enemies': 6,
 'Chain Thunder': 30,
 'Lightning Amplifier - Scatter': 30,
 'Death Wave Cells Bonus': 20,
 'Death Wave Damage Amplifier': 30,
 'Death Wave Armor Stripping': 10,
 'Inner Land Mine - Chrono Jump': 10,
 'Second Wind Blast': 4,
 'Double Death Ray': 30,
 'Extra Orb Adjuster': 1,
 'Extra Extra Orbs': 2,
 'Energy Shield Extra Hit': 2,
 'Super Tower Bonus': 30,
 'Recharge Second Wind': 7,
 'Recharge Demon Mode': 7,
 'Recharge Nuke': 7,
 'Unlock Perks': 1,
 'Waves Required': 100,
 'Auto Pick Perks': 1,
 'Standard Perks Bonus': 25,
 'Perk Option Quantity': 2,
 'First Perk Choice': 1,
 'Ban Perks': 8,
 'Improve Trade-off Perks': 10,
 'Auto Pick Ranking': 32,
 'Flame Bot - Cooldown': 25,
 'Thunder Bot - Cooldown': 25,
 'Gold Bot - Cooldown': 25,
 'Amp Bot - Cooldown': 25,
 'Bot Bot - Cooldown': 25,
 'Flame Bot - Burn Stack': 5,
 'Thunder Bot - Linger Time': 20,
 'Gold Bot - Duration': 20,
 'Amp Bot - Duration': 20,
 'Bot Bot - Duration': 20,
 'Common Enemy Health': 30,
 'Common Enemy Attack': 30,
 'Fast Enemy Health': 30,
 'Fast Enemy Attack': 30,
 'Fast Enemy Speed': 30,
 'Tank Enemy Health': 30,
 'Tank Enemy Attack': 30,
 'Ranged Enemy Health': 30,
 'Ranged Enemy Attack': 30,
 'Boss Health': 30,
 'Boss Attack': 30,
 'Protector Health': 30,
 'Protector Radius': 30,
 'Protector Damage Reduction': 20,
 'Ray Enemy Attack': 30,
 'Ray Enemy Health': 30,
 'Vampire Enemy Attack': 30,
 'Vampire Enemy Health': 30,
 'Scatter Enemy Attack': 30,
 'Scatter Enemy Health': 30,
 'Ranged Enemy Range': 30,
 'Common Drop Chance': 10,
 'Reroll Shards': 100,
 'Daily Mission Shards': 50,
 'Module Shards Cost': 30,
 'Module Coin Cost': 30,
 'Rare Drop Chance': 10,
 'Unmerge Module': 1,
 'Shatter Shards': 5,
 'Cannon Effect Bans': 4,
 'Armor Effect Bans': 4,
 'Generator Effect Bans': 3,
 'Core Effect Bans': 7,
 'Assist Module Substats - Cannon': 30,
 'Assist Module Substats - Armor': 30,
 'Assist Module Substats - Generator': 30,
 'Assist Module Substats - Core': 30,
 'Assist Module Bonus - Cannon': 30,
 'Assist Module Bonus - Armor': 30,
 'Assist Module Bonus - Generator': 30,
 'Assist Module Bonus - Core': 30,
 'Damage Mastery': 9,
 'Attack Speed Mastery': 9,
 'Health Mastery': 9,
 'Health Regen Mastery': 9,
 'Range Mastery': 9,
 'Cash Mastery': 9,
 'Coins Mastery': 9,
 'Slow Aura Mastery': 9,
 'Critical Chance Mastery': 9,
 'Enemy Balance Mastery': 9,
 'Extra Defense Mastery': 9,
 'Fortress Mastery': 9,
 'Free Upgrades Mastery': 9,
 'Extra Orb Mastery': 9,
 'Plasma Cannon Mastery': 9,
 'Critical Coin Mastery': 9,
 'Wave Skip Mastery': 9,
 'Intro Sprint Mastery': 9,
 'Land Mine Stun Mastery': 9,
 'Recovery Package Chance Mastery': 9,
 'Death Ray Mastery': 9,
 'Energy Net Mastery': 9,
 'Super Tower Mastery': 9,
 'Second Wind Mastery': 9,
 'Demon Mode Mastery': 9,
 'Energy Shield Mastery': 9,
 'Wave Accelerator Mastery': 9,
 'Berserker Mastery': 9,
 'Ultimate Crit Mastery': 9,
 'Nuke Mastery': 9,
 'Area of Effect Mastery': 9,
 'Battle Condition Reduction': 10,
 'Knockback Resistance': 20,
 'Thorns Resistance': 20,
 'Orb Resistance': 20,
 'Plasma Cannon Resistance': 20,
 'Death Ray Resistance': 20,
 'Armored Enemies': 20,
 'Enemy Speed': 20,
 'More Enemies': 20,
 'Enemy Attack Speed': 20,
 "Fast's Ultimate": 10,
 'Ranged Ultimate': 10,
 "Boss's Ultimate": 10,
 "Basic's Ultimate": 10,
 "Tank's Ultimate": 10,
 "Protector's Ultimate": 10,
 'Ultimate Weapon Durations': 10,
 'Death Defy Down': 10,
 'Energy Shields Down': 10,
 'Enemy Level Skip Reduction': 10}

WORKSHOP_GROUPS = {'Attack': ['Damage',
            'Attack Speed',
            'Critical Chance',
            'Critical Factor',
            'Range',
            'Damage / Meter',
            'Multishot Chance',
            'Multishot Targets',
            'Rapid Fire Chance',
            'Rapid Fire Duration',
            'Bounce Shot Chance',
            'Bounce Shot Targets',
            'Bounce Shot Range',
            'Super Critical Chance',
            'Super Critical Mult',
            'Rend Armor Chance',
            'Rend Armor Mult'],
 'Defense': ['Health',
             'Health Regen',
             'Defense %',
             'Defense Absolute',
             'Thorn Damage',
             'Lifesteal',
             'Knockback Chance',
             'Knockback Force',
             'Orb Speed',
             'Orbs',
             'Shockwave Size',
             'Shockwave Frequency',
             'Land Mine Chance',
             'Land Mine Damage',
             'Land Mine Radius',
             'Death Defy',
             'Wall Health',
             'Wall Rebuild'],
 'Utility': ['Cash Bonus',
             'Cash / Wave',
             'Coin / Kill Bonus',
             'Coin / Wave',
             'Free Attack Upgrade',
             'Free Defense Upgrade',
             'Free Utility Upgrade',
             'Interest / Wave',
             'Recovery Amount',
             'Max Amount',
             'Package Chance',
             'Enemy Attack Level Skip',
             'Enemy Health Level Skip']}

LAB_GROUPS = {'General & Unlocks': ['Game Speed',
                       'Starting Cash',
                       'Workshop Attack Discount',
                       'Workshop Defense Discount',
                       'Workshop Utility Discount',
                       'Labs Coin Discount',
                       'Labs Speed',
                       'Buy Multiplier',
                       'More Round Stats',
                       'Target Priority',
                       'Card Presets',
                       'Workshop Respec',
                       'Reroll Daily Mission',
                       'Workshop Enhancements',
                       'Enhancement Attack - Coin Discount',
                       'Enhancement Defense - Coin Discount',
                       'Enhancement Utility - Coin Discount',
                       'Dissonant Echo - Attack',
                       'Dissonant Echo - Defense',
                       'Dissonant Echo - Utility',
                       'Dissonant Echo - Ultimate Weapons'],
 'Attack': ['Damage',
            'Attack Speed',
            'Critical Factor',
            'Range',
            'Damage / Meter',
            'Super Crit Chance',
            'Super Crit Multi',
            'Max Rend Armor Multiplier',
            'Light Speed Shots'],
 'Defense': ['Health',
             'Health Regen',
             'Defense Absolute',
             'Defense %',
             'Orbs Speed',
             'Land Mine Damage',
             'Land Mine Decay',
             'Shockwave Size',
             'Orb Boss Hit',
             'Wall Health',
             'Wall Rebuild',
             'Wall Regen',
             'Wall Thorns',
             'Wall Invincibility',
             'Wall Fortification',
             'Garlic Thorns'],
 'Utility & Economy': ['Cash Bonus',
                       'Cash / Wave',
                       'Coins / Kill Bonus',
                       'Coins / Wave',
                       'Interest',
                       'Max Interest',
                       'Package After Boss',
                       'Recovery Package Amount',
                       'Recovery Package Max',
                       'Recovery Package Chance',
                       'Enemy Attack Level Skip',
                       'Enemy Health Level Skip'],
 'Ultimate Weapons': ['Missile Despawn Time',
                      'Missiles Explosion',
                      'Missile Radius',
                      'Chrono Field Duration',
                      'Chrono Field Damage Reduction',
                      'Chrono Field Reduction %',
                      'Swamp Radius',
                      'Swamp Stun',
                      'Swamp Stun Chance',
                      'Swamp Stun Time',
                      'Golden Tower Bonus',
                      'Golden Tower Duration',
                      'Chain Lightning Shock',
                      'Shock Chance',
                      'Shock Multiplier',
                      'Death Wave Health',
                      'Death Wave Coin Bonus',
                      'Inner Mine Blast Radius',
                      'Inner Mine Rotation Speed',
                      'Chrono Field Range',
                      'Missile Amplifier',
                      'Missile Barrage',
                      'Missile Barrage Quantity',
                      'Inner Mine Stun',
                      'Black Hole Damage',
                      'Extra Black Hole',
                      'Black Hole Coin Bonus',
                      'Spotlight Coin Bonus',
                      'Spotlight Missiles',
                      'Black Hole Disable Ranged Enemies',
                      'Recharge Missile Barrage',
                      'Swamp Rend',
                      'Swamp Rend - Additional Enemies',
                      'Chain Thunder',
                      'Lightning Amplifier - Scatter',
                      'Death Wave Cells Bonus',
                      'Death Wave Damage Amplifier',
                      'Death Wave Armor Stripping',
                      'Inner Land Mine - Chrono Jump',
                      'Second Wind Blast'],
 'Cards': ['Double Death Ray',
           'Extra Orb Adjuster',
           'Extra Extra Orbs',
           'Energy Shield Extra Hit',
           'Super Tower Bonus',
           'Recharge Second Wind',
           'Recharge Demon Mode',
           'Recharge Nuke'],
 'Perks': ['Unlock Perks',
           'Waves Required',
           'Auto Pick Perks',
           'Standard Perks Bonus',
           'Perk Option Quantity',
           'First Perk Choice',
           'Ban Perks',
           'Improve Trade-off Perks',
           'Auto Pick Ranking'],
 'Bots': ['Flame Bot - Cooldown',
          'Thunder Bot - Cooldown',
          'Gold Bot - Cooldown',
          'Amp Bot - Cooldown',
          'Bot Bot - Cooldown',
          'Flame Bot - Burn Stack',
          'Thunder Bot - Linger Time',
          'Gold Bot - Duration',
          'Amp Bot - Duration',
          'Bot Bot - Duration'],
 'Enemies': ['Common Enemy Health',
             'Common Enemy Attack',
             'Fast Enemy Health',
             'Fast Enemy Attack',
             'Fast Enemy Speed',
             'Tank Enemy Health',
             'Tank Enemy Attack',
             'Ranged Enemy Health',
             'Ranged Enemy Attack',
             'Boss Health',
             'Boss Attack',
             'Protector Health',
             'Protector Radius',
             'Protector Damage Reduction',
             'Ray Enemy Attack',
             'Ray Enemy Health',
             'Vampire Enemy Attack',
             'Vampire Enemy Health',
             'Scatter Enemy Attack',
             'Scatter Enemy Health',
             'Ranged Enemy Range'],
 'Modules': ['Common Drop Chance',
             'Reroll Shards',
             'Daily Mission Shards',
             'Module Shards Cost',
             'Module Coin Cost',
             'Rare Drop Chance',
             'Unmerge Module',
             'Shatter Shards',
             'Cannon Effect Bans',
             'Armor Effect Bans',
             'Generator Effect Bans',
             'Core Effect Bans',
             'Assist Module Substats - Cannon',
             'Assist Module Substats - Armor',
             'Assist Module Substats - Generator',
             'Assist Module Substats - Core',
             'Assist Module Bonus - Cannon',
             'Assist Module Bonus - Armor',
             'Assist Module Bonus - Generator',
             'Assist Module Bonus - Core'],
 'Card Masteries': ['Damage Mastery',
                    'Attack Speed Mastery',
                    'Health Mastery',
                    'Health Regen Mastery',
                    'Range Mastery',
                    'Cash Mastery',
                    'Coins Mastery',
                    'Slow Aura Mastery',
                    'Critical Chance Mastery',
                    'Enemy Balance Mastery',
                    'Extra Defense Mastery',
                    'Fortress Mastery',
                    'Free Upgrades Mastery',
                    'Extra Orb Mastery',
                    'Plasma Cannon Mastery',
                    'Critical Coin Mastery',
                    'Wave Skip Mastery',
                    'Intro Sprint Mastery',
                    'Land Mine Stun Mastery',
                    'Recovery Package Chance Mastery',
                    'Death Ray Mastery',
                    'Energy Net Mastery',
                    'Super Tower Mastery',
                    'Second Wind Mastery',
                    'Demon Mode Mastery',
                    'Energy Shield Mastery',
                    'Wave Accelerator Mastery',
                    'Berserker Mastery',
                    'Ultimate Crit Mastery',
                    'Nuke Mastery',
                    'Area of Effect Mastery'],
 'Battle Conditions': ['Battle Condition Reduction',
                       'Knockback Resistance',
                       'Thorns Resistance',
                       'Orb Resistance',
                       'Plasma Cannon Resistance',
                       'Death Ray Resistance',
                       'Armored Enemies',
                       'Enemy Speed',
                       'More Enemies',
                       'Enemy Attack Speed',
                       "Fast's Ultimate",
                       'Ranged Ultimate',
                       "Boss's Ultimate",
                       "Basic's Ultimate",
                       "Tank's Ultimate",
                       "Protector's Ultimate",
                       'Ultimate Weapon Durations',
                       'Death Defy Down',
                       'Energy Shields Down',
                       'Enemy Level Skip Reduction']}

# Names used by older versions of this app that differ from Effective Paths.
WORKSHOP_ALIASES = {
    "Super Crit Chance": "Super Critical Chance",
    "Super Crit Mult": "Super Critical Mult",
    "Thorns": "Thorn Damage",
    "Coins / Kill": "Coin / Kill Bonus",
    "Coins / Wave": "Coin / Wave",
    "Max Recovery": "Max Amount",
    "Recovery Package Chance": "Package Chance",
}

LAB_ALIASES = {
    "Lab Speed": "Labs Speed",
    "Orb Speed": "Orbs Speed",
    "Standard Perk Bonus": "Standard Perks Bonus",
    "Common Module Drop Chance": "Common Drop Chance",
    "Rare Module Drop Chance": "Rare Drop Chance",
    "Module Shard Cost": "Module Shards Cost",
}

# These were incorrectly placed in Workshop by older builds. Move them to Labs
# when the destination is empty.
WORKSHOP_TO_LAB_MIGRATIONS = {
    "Wall Regen": "Wall Regen",
    "Wall Thorns": "Wall Thorns",
    "Wall Fortification": "Wall Fortification",
}

ENHANCEMENT_MAX_LEVELS: Dict[str, int] = {
    "Damage +": 400,
    "Rend Armor Mult +": 400,
    "Critical Factor +": 400,
    "Damage / Meter +": 400,
    "Super Crit Multi +": 400,
    "Attack Speed +": 75,
    "Health +": 400,
    "Health Regen +": 400,
    "Defense Absolute +": 400,
    "Land Mine Damage +": 400,
    "Wall Health +": 400,
    "Orb Size +": 200,
    "Cash Bonus +": 400,
    "Coin Bonus +": 200,
    "Cells / Kill Bonus +": 200,
    "Free Upgrades +": 100,
    "Recovery Package +": 300,
    "Enemy Level Skips +": 60,
}

ENHANCEMENT_GROUPS = {
    "Attack": [
        "Damage +", "Rend Armor Mult +", "Critical Factor +",
        "Damage / Meter +", "Super Crit Multi +", "Attack Speed +",
    ],
    "Defense": [
        "Health +", "Health Regen +", "Defense Absolute +",
        "Land Mine Damage +", "Wall Health +", "Orb Size +",
    ],
    "Utility": [
        "Cash Bonus +", "Coin Bonus +", "Cells / Kill Bonus +",
        "Free Upgrades +", "Recovery Package +", "Enemy Level Skips +",
    ],
}

UW_NAMES = [
    "Chain Lightning",
    "Smart Missiles",
    "Death Wave",
    "Chrono Field",
    "Inner Land Mines",
    "Golden Tower",
    "Poison Swamp",
    "Black Hole",
    "Spotlight",
]

# Attribute values and caps extracted from the All UWs worksheet in
# Effective Paths v5.06.04.00.
UW_ATTRIBUTE_META: Dict[str, Dict[str, Dict[str, Any]]] = {
    "Chain Lightning": {
        "Damage": {"max": 7961.0, "step": 1.0},
        "Quantity": {"max": 5, "step": 1},
        "Chance": {"max": 0.275, "step": 0.005, "display": "percent"},
    },
    "Smart Missiles": {
        "Damage": {"max": 3021.0, "step": 1.0},
        "Quantity": {"max": 20, "step": 1},
        "Cooldown": {"max": 20, "step": 1, "lower_is_better": True, "start": 180},
    },
    "Death Wave": {
        "Damage": {"max": 9119.0, "step": 1.0},
        "Quantity": {"max": 5, "step": 1},
        "Cooldown": {"max": 50, "step": 10, "lower_is_better": True, "start": 300},
    },
    "Chrono Field": {
        "Duration": {"max": 40, "step": 1},
        "Speed Reduction": {"max": 0.75, "step": 0.05, "display": "percent"},
        "Cooldown": {"max": 60, "step": 10, "lower_is_better": True, "start": 180},
    },
    "Inner Land Mines": {
        "Damage": {"max": 3021.0, "step": 1.0},
        "Quantity": {"max": 6, "step": 1},
        "Cooldown": {"max": 50, "step": 10, "lower_is_better": True, "start": 200},
    },
    "Golden Tower": {
        "Multiplier": {"max": 21.0, "step": 0.1},
        "Duration": {"max": 53, "step": 1},
        "Cooldown": {"max": 100, "step": 10, "lower_is_better": True, "start": 300},
    },
    "Poison Swamp": {
        "Damage": {"max": 3021.0, "step": 1.0},
        "Duration": {"max": 100, "step": 1},
        "Cooldown": {"max": 50, "step": 5, "lower_is_better": True, "start": 125},
    },
    "Black Hole": {
        "Size": {"max": 70, "step": 2},
        "Duration": {"max": 38, "step": 1},
        "Cooldown": {"max": 50, "step": 10, "lower_is_better": True, "start": 200},
    },
    "Spotlight": {
        "Multiplier": {"max": 43.0, "step": 0.1},
        "Angle": {"max": 90, "step": 1},
        "Quantity": {"max": 4, "step": 1},
    },
}

MODULE_OPTIONS_BY_SLOT = {
    "Cannon": [
        "", "Astral Deliverance", "Being Annihilator", "Death Penalty",
        "Havoc Bringer", "Shrink Ray", "Amplifying Strike",
    ],
    "Armor": [
        "", "Anti-Cube Portal", "Negative Mass Projector",
        "Wormhole Redirector", "Space Displacer", "Sharp Fortitude",
        "Orbital Augment",
    ],
    "Generator": [
        "", "Singularity Harness", "Galaxy Compressor", "Pulsar Harvester",
        "Black Hole Digestor", "Project Funding", "Restorative Bonus",
    ],
    "Core": [
        "", "Om Chip", "Harmony Conductor", "Dimension Core",
        "Multiverse Nexus", "Magnetic Hook", "Primordial Collapse",
    ],
}

MODULE_RARITY_MAX_LEVELS = {
    "Common": 20,
    "Rare": 30,
    "Rare+": 40,
    "Epic": 60,
    "Epic+": 80,
    "Legendary": 100,
    "Legendary+": 120,
    "Mythic": 140,
    "Mythic+": 160,
    "Ancestral": 200,
    "Ancestral 1*": 220,
    "Ancestral 2*": 240,
    "Ancestral 3*": 260,
    "Ancestral 4*": 280,
    "Ancestral 5*": 300,
}
RARITY_OPTIONS = list(MODULE_RARITY_MAX_LEVELS.keys())


CARD_NAMES = [
    "Damage", "Attack Speed", "Health", "Health Regen", "Range", "Cash",
    "Coins", "Slow Aura", "Critical Chance", "Enemy Balance", "Extra Defense",
    "Fortress", "Free Upgrades", "Extra Orb", "Plasma Cannon", "Critical Coin",
    "Wave Skip", "Intro Sprint", "Land Mine Stun", "Recovery Package Chance",
    "Death Ray", "Energy Net", "Super Tower", "Second Wind", "Demon Mode",
    "Energy Shield", "Wave Accelerator", "Berserker", "Ultimate Crit", "Nuke",
    "Area of Effect",
]

BOT_NAMES = ["Flame Bot", "Thunder Bot", "Golden Bot", "Amplify Bot", "Bot Bot"]
BOT_ATTRIBUTES = {
    "Flame Bot": ["Damage R.", "Cooldown", "Damage", "Range"],
    "Thunder Bot": ["Duration", "Cooldown", "Linger", "Range"],
    "Golden Bot": ["Duration", "Cooldown", "Bonus", "Range"],
    "Amplify Bot": ["Duration", "Cooldown", "Bonus", "Range"],
    "Bot Bot": ["Duration", "Cooldown", "Bonus", "Range"],
}

GUARDIAN_NAMES = ["Attack", "Ally", "Bounty", "Fetch", "Summon", "Scout"]
GUARDIAN_ATTRIBUTES = {
    "Attack": ["Percentage", "Cooldown", "Targets"],
    "Ally": ["Recovery Amount", "Max Recovery", "Cooldown"],
    "Bounty": ["Multiplier", "Cooldown", "Targets"],
    "Fetch": ["Cooldown", "Find Chance", "Double Find Chance"],
    "Summon": ["Cooldown", "Duration", "Cash Bonus"],
    "Scout": ["Cooldown", "Range Bonus", "Duration"],
}

COMPANION_SIGNATURES = {
    "Wv": "workshop",
    "Lv": "laboratory",
    "Uv": "ultimate_weapons",
    "Mv": "modules",
    "Cv": "cards",
    "Rv": "relics",
    "T&Sv": "themes",
    "Bv": "bots",
    "Gv": "guardians",
    "Vv": "vault",
    "Pv": "player",
}

SECTION_LABELS = {
    "player": "Player & Progression",
    "workshop": "Workshop",
    "labs": "Laboratory",
    "enhancements": "Workshop Enhancements",
    "uw": "Ultimate Weapons",
    "module_inventory": "Module Inventory",
    "module_presets": "Module Presets",
    "cards": "Cards",
    "relics": "Relics",
    "themes": "Themes & Songs",
    "bots": "Bots",
    "guardians": "Guardians",
    "vault": "Vault",
}

# -----------------------------------------------------------------------------
# PROFILE HELPERS
# -----------------------------------------------------------------------------

# -----------------------------------------------------------------------------
# PROFILE HELPERS
# -----------------------------------------------------------------------------

# Apply an approved local metadata overlay without modifying installed source files.
# A full Streamlit restart is required after applying or rolling back an update so
# the standalone engine modules load the same metadata.
_ACTIVE_GAME_DATA_UPDATE = load_runtime_overlay()
if _ACTIVE_GAME_DATA_UPDATE:
    incoming_workshop = _ACTIVE_GAME_DATA_UPDATE.get("workshop_max_levels", {})
    incoming_labs = _ACTIVE_GAME_DATA_UPDATE.get("lab_max_levels", {})
    if incoming_workshop:
        WORKSHOP_MAX_LEVELS = {str(name): int(value) for name, value in incoming_workshop.items()}
    if incoming_labs:
        LAB_MAX_LEVELS = {str(name): int(value) for name, value in incoming_labs.items()}
    SUPPORTED_WORKBOOK_VERSIONS.update(_ACTIVE_GAME_DATA_UPDATE.get("workbook_versions", {}))
    if _ACTIVE_GAME_DATA_UPDATE.get("effective_paths_version"):
        DATA_SOURCE_VERSION = f"Effective Paths v{_ACTIVE_GAME_DATA_UPDATE['effective_paths_version']} (local overlay)"


def default_profile() -> Dict[str, Any]:
    return {
        "name": "default",
        "resources": {"coins": 0, "stones": 0, "gems": 0, "medals": 0, "keys": 0, "bits": 0, "reroll_shards": 0, "module_shards": 0},
        "player": {
            "player_id": "", "farming_tier": "", "tourney_league": "",
            "lifetime_coins": 0, "coin_multiplier": "", "packs": {}, "tiers": {},
        },
        "workshop": {},
        "labs": {},
        "enhancements": {},
        "uw": {},
        "modules": {},
        "module_inventory": {},
        "module_presets": {},
        "module_forge": {"fodder": {}},
        "cards": {"slots": 0, "slot_target": 1, "items": {}, "presets": {}},
        "relics": {"summary": {}, "bonuses": {}, "items": {}},
        "themes": {"summary": {}, "items": {}},
        "bots": {},
        "guardians": {},
        "vault": {"keys_spent": 0, "bonuses": {}, "unlocks": {}},
        "runs": [],
        "roi_reference": {"source": {}, "paths": {}, "imported_at": None, "warnings": []},
        "native_econ": {
            "settings": {
                "bh_perk_duration_bonus": 12.0,
                "bh_coverage_divisor": 70.0,
                "dw_tag_share_per_quantity": 0.0382747832266563,
                "lab_speed_multiplier_override": 0.0,
                "allow_desync_cooldowns": False,
            },
            "last_run": None,
        },
        "analysis": {"last_updated": None, "notes": []},
        "calibration": {"history": [], "last_report": None},
        "planner": {},
        "battle_learning": {},
        "setup_wizard": {"completed": False, "last_step": "Profile"},
        "maxed": {
            "workshop": {}, "labs": {}, "enhancements": {}, "uw": {}, "modules": {},
        },
        "custom_max": {"workshop": {}, "labs": {}, "enhancements": {}},
        "settings": {"auto_gold_at_max": True, "show_only_incomplete": False, "visual_theme": "Void Cyan", "visual_density": "Comfortable", "visual_motion": True},
        "sources": {},
        "import_audit": [],
        "metadata": {
            "effective_paths_version": DATA_SOURCE_VERSION,
            "app_version": APP_VERSION,
            "last_import": None,
        },
    }


def move_profile_item(profile: Dict[str, Any], section: str, old: str, new: str) -> None:
    values = profile.setdefault(section, {})
    maxed = profile.setdefault("maxed", {}).setdefault(section, {})
    custom = profile.setdefault("custom_max", {}).setdefault(section, {})

    if old in values:
        old_value = values.get(old, 0)
        if new not in values or int(values.get(new, 0) or 0) == 0:
            values[new] = old_value
        values.pop(old, None)
    if old in maxed:
        maxed.setdefault(new, maxed.get(old, False))
        maxed.pop(old, None)
    if old in custom:
        custom.setdefault(new, custom.get(old, 0))
        custom.pop(old, None)


def migrate_profile(profile: Dict[str, Any]) -> Dict[str, Any]:
    for old, new in WORKSHOP_ALIASES.items():
        move_profile_item(profile, "workshop", old, new)
    for old, new in LAB_ALIASES.items():
        move_profile_item(profile, "labs", old, new)

    for old, new in WORKSHOP_TO_LAB_MIGRATIONS.items():
        old_value = profile.setdefault("workshop", {}).get(old)
        if old_value is not None:
            labs = profile.setdefault("labs", {})
            if int(labs.get(new, 0) or 0) == 0:
                labs[new] = old_value
            profile.setdefault("maxed", {}).setdefault("labs", {}).setdefault(
                new,
                bool(profile.setdefault("maxed", {}).setdefault("workshop", {}).get(old, False)),
            )
            profile["workshop"].pop(old, None)
            profile["maxed"]["workshop"].pop(old, None)

    # Convert the old generic UW shape into attribute dictionaries.
    old_field_map = {
        "bonus_or_damage": None,
        "duration": "Duration",
        "cooldown": "Cooldown",
        "quantity": "Quantity",
    }
    for uw_name in UW_NAMES:
        uw = profile.setdefault("uw", {}).setdefault(uw_name, {})
        attrs = uw.setdefault("attributes", {})
        first_attr = next(iter(UW_ATTRIBUTE_META[uw_name]))
        for old_key, attribute in old_field_map.items():
            if old_key in uw and uw.get(old_key) not in (None, 0, 0.0):
                attrs.setdefault(attribute or first_attr, uw.get(old_key))
        if uw_name == "Golden Tower" and uw.get("gt_bonus") not in (None, 0, 0.0):
            attrs.setdefault("Multiplier", uw.get("gt_bonus"))
        # Old BH coin bonus was a Lab value, not a stone attribute.
        if uw_name == "Black Hole" and uw.get("coin_bonus") not in (None, 0, 0.0):
            profile.setdefault("labs", {}).setdefault("Black Hole Coin Bonus", int(uw.get("coin_bonus")))

        old_max = profile.setdefault("maxed", {}).setdefault("uw", {}).get(uw_name, {})
        if isinstance(old_max, bool):
            profile["maxed"]["uw"][uw_name] = {
                attr: old_max for attr in UW_ATTRIBUTE_META[uw_name]
            }
        else:
            profile["maxed"]["uw"].setdefault(uw_name, {})

    return profile


def ensure_profile_shape(profile: Dict[str, Any]) -> Dict[str, Any]:
    base = default_profile()
    for key, value in base.items():
        profile.setdefault(key, value)

    for section in ["workshop", "labs", "enhancements", "uw", "modules"]:
        profile.setdefault(section, {})
        profile.setdefault("maxed", {}).setdefault(section, {})
    for section in ["workshop", "labs", "enhancements"]:
        profile.setdefault("custom_max", {}).setdefault(section, {})
    profile.setdefault("resources", {})
    for currency in ["coins", "stones", "gems", "medals", "keys", "bits", "reroll_shards", "module_shards"]:
        profile["resources"].setdefault(currency, 0)
    profile.setdefault("player", {})
    for key, value in base["player"].items():
        profile["player"].setdefault(key, value)
    profile.setdefault("module_inventory", {})
    profile.setdefault("module_presets", {})
    profile.setdefault("module_forge", {"fodder": {}})
    if not isinstance(profile.get("module_forge"), dict):
        profile["module_forge"] = {"fodder": {}}
    profile["module_forge"].setdefault("fodder", {})
    profile.setdefault("cards", {"slots": 0, "slot_target": 1, "items": {}, "presets": {}})
    profile["cards"].setdefault("slots", 0)
    profile["cards"].setdefault("slot_target", max(1, int(profile["cards"].get("slots", 0) or 0)))
    profile["cards"].setdefault("presets", {})
    profile["cards"].setdefault("items", {})
    for card in CARD_NAMES:
        profile["cards"]["items"].setdefault(card, {"level": 0, "mastery": 0})
    profile.setdefault("relics", {"summary": {}, "bonuses": {}, "items": {}})
    profile["relics"].setdefault("summary", {})
    profile["relics"].setdefault("bonuses", {})
    profile["relics"].setdefault("items", {})
    profile.setdefault("themes", {"summary": {}, "items": {}})
    profile["themes"].setdefault("summary", {})
    profile["themes"].setdefault("items", {})
    profile.setdefault("bots", {})
    profile.setdefault("guardians", {})
    profile.setdefault("vault", {"keys_spent": 0, "bonuses": {}, "unlocks": {}})
    profile["vault"].setdefault("keys_spent", 0)
    profile["vault"].setdefault("bonuses", {})
    profile["vault"].setdefault("unlocks", {})
    profile.setdefault("runs", [])
    if not isinstance(profile.get("runs"), list):
        profile["runs"] = []
    profile.setdefault("roi_reference", {"source": {}, "paths": {}, "imported_at": None, "warnings": []})
    if not isinstance(profile.get("roi_reference"), dict):
        profile["roi_reference"] = {"source": {}, "paths": {}, "imported_at": None, "warnings": []}
    profile["roi_reference"].setdefault("source", {})
    profile["roi_reference"].setdefault("paths", {})
    profile["roi_reference"].setdefault("imported_at", None)
    profile["roi_reference"].setdefault("warnings", [])
    profile.setdefault("native_econ", {})
    profile["native_econ"].setdefault("settings", {})
    native_defaults = base["native_econ"]["settings"]
    for native_key, native_value in native_defaults.items():
        profile["native_econ"]["settings"].setdefault(native_key, native_value)
    profile["native_econ"].setdefault("last_run", None)
    profile.setdefault("analysis", {"last_updated": None, "notes": []})
    profile["analysis"].setdefault("last_updated", None)
    profile["analysis"].setdefault("notes", [])
    profile.setdefault("calibration", {"history": [], "last_report": None})
    if not isinstance(profile.get("calibration"), dict):
        profile["calibration"] = {"history": [], "last_report": None}
    profile["calibration"].setdefault("history", [])
    profile["calibration"].setdefault("last_report", None)
    profile.setdefault("planner", {})
    ensure_battle_learning_state(profile)
    normalize_profile_runs(profile)
    profile.setdefault("setup_wizard", {"completed": False, "last_step": "Profile"})
    if not isinstance(profile.get("setup_wizard"), dict):
        profile["setup_wizard"] = {"completed": False, "last_step": "Profile"}
    profile["setup_wizard"].setdefault("completed", False)
    profile["setup_wizard"].setdefault("last_step", "Profile")
    profile.setdefault("sources", {})
    profile.setdefault("import_audit", [])

    for key, value in base["settings"].items():
        profile.setdefault("settings", {}).setdefault(key, value)
    profile.setdefault("metadata", {})
    profile["metadata"]["effective_paths_version"] = DATA_SOURCE_VERSION
    profile["metadata"]["app_version"] = APP_VERSION
    profile["metadata"].setdefault("last_import", None)
    return migrate_profile(profile)


def safe_profile_filename(name: str) -> str:
    cleaned = "".join(c for c in name.strip() if c.isalnum() or c in "-_ ")
    return cleaned.strip() or "default"


def save_profile(profile_name: str, data: Dict[str, Any]) -> None:
    clean_name = safe_profile_filename(profile_name)
    data["name"] = clean_name
    data.setdefault("metadata", {})["app_version"] = APP_VERSION
    path = PROFILE_DIR / f"{clean_name}.json"
    atomic_save_json(path, data, PROFILE_DIR)


def load_profile(profile_name: str) -> Dict[str, Any]:
    path = PROFILE_DIR / f"{profile_name}.json"
    with path.open("r", encoding="utf-8") as file:
        return ensure_profile_shape(json.load(file))


def list_profiles() -> list[str]:
    return sorted(path.stem for path in PROFILE_DIR.glob("*.json"))


def bump_revision() -> None:
    st.session_state.profile_revision = st.session_state.get("profile_revision", 0) + 1


def metadata_for(section: str, item_name: str) -> Optional[int]:
    if section == "workshop":
        return WORKSHOP_MAX_LEVELS.get(item_name)
    if section == "labs":
        return LAB_MAX_LEVELS.get(item_name)
    if section == "enhancements":
        return ENHANCEMENT_MAX_LEVELS.get(item_name)
    return None


def effective_max(section: str, item_name: str) -> Optional[int]:
    built_in = metadata_for(section, item_name)
    if built_in is not None:
        return int(built_in)
    custom = st.session_state.profile.get("custom_max", {}).get(section, {}).get(item_name)
    try:
        custom_value = int(custom or 0)
    except (TypeError, ValueError):
        return None
    return custom_value if custom_value > 0 else None


def is_maxed(section: str, item_name: str) -> bool:
    return bool(st.session_state.profile.get("maxed", {}).get(section, {}).get(item_name, False))


def format_cap_text(section: str, item_name: str) -> str:
    cap = effective_max(section, item_name)
    return f"Max {cap:,}" if cap is not None else "Maximum unknown"




# -----------------------------------------------------------------------------
# NATIVE ECONOMY ENGINE DATA (v1.0)
# -----------------------------------------------------------------------------
# Base laboratory durations/costs and UW stone curves were extracted from the
# companion workbooks supplied with Effective Paths. These tables let the app
# calculate economy paths without opening a spreadsheet at runtime.

NATIVE_ECON_LAB_TABLES = {'Coins / Kill Bonus': {1: {'seconds': 14, 'cost': 30.0},
                        2: {'seconds': 384, 'cost': 71.0},
                        3: {'seconds': 984, 'cost': 178.0},
                        4: {'seconds': 1892, 'cost': 398.0},
                        5: {'seconds': 3170, 'cost': 782.0},
                        6: {'seconds': 4800, 'cost': 1350.0},
                        7: {'seconds': 6960, 'cost': 2130.0},
                        8: {'seconds': 9540, 'cost': 3180.0},
                        9: {'seconds': 12660, 'cost': 4520.0},
                        10: {'seconds': 16320, 'cost': 6180.0},
                        11: {'seconds': 20580, 'cost': 8180.0},
                        12: {'seconds': 25380, 'cost': 10570.0},
                        13: {'seconds': 30840, 'cost': 13360.0},
                        14: {'seconds': 36900, 'cost': 16590.0},
                        15: {'seconds': 43620, 'cost': 20280.0},
                        16: {'seconds': 51000, 'cost': 24450.0},
                        17: {'seconds': 59100, 'cost': 29140.0},
                        18: {'seconds': 67920, 'cost': 34370.0},
                        19: {'seconds': 77460, 'cost': 40170.0},
                        20: {'seconds': 87720, 'cost': 46550.0},
                        21: {'seconds': 98820, 'cost': 53540.0},
                        22: {'seconds': 110640, 'cost': 61170.0},
                        23: {'seconds': 123240, 'cost': 69470.0},
                        24: {'seconds': 136680, 'cost': 78440.0},
                        25: {'seconds': 150960, 'cost': 88130.0},
                        26: {'seconds': 166020, 'cost': 98540.0},
                        27: {'seconds': 181980, 'cost': 109710.0},
                        28: {'seconds': 198780, 'cost': 121660.0},
                        29: {'seconds': 216480, 'cost': 134400.0},
                        30: {'seconds': 235080, 'cost': 147960.0},
                        31: {'seconds': 254580, 'cost': 162360.0},
                        32: {'seconds': 274980, 'cost': 177630.0},
                        33: {'seconds': 296340, 'cost': 193790.0},
                        34: {'seconds': 318600, 'cost': 210840.0},
                        35: {'seconds': 341820, 'cost': 228830.0},
                        36: {'seconds': 366000, 'cost': 247770.0},
                        37: {'seconds': 391200, 'cost': 267670.0},
                        38: {'seconds': 417300, 'cost': 288570.0},
                        39: {'seconds': 444480, 'cost': 310480.0},
                        40: {'seconds': 472620, 'cost': 333410.0},
                        41: {'seconds': 501840, 'cost': 357400.0},
                        42: {'seconds': 532020, 'cost': 382460.0},
                        43: {'seconds': 563280, 'cost': 408610.0},
                        44: {'seconds': 595560, 'cost': 435880.0},
                        45: {'seconds': 628920, 'cost': 464270.0},
                        46: {'seconds': 663360, 'cost': 493820.0},
                        47: {'seconds': 698820, 'cost': 524540.0},
                        48: {'seconds': 735420, 'cost': 556440.0},
                        49: {'seconds': 773100, 'cost': 589560.0},
                        50: {'seconds': 811920, 'cost': 623890.0},
                        51: {'seconds': 851820, 'cost': 659500.0},
                        52: {'seconds': 892800, 'cost': 696350.0},
                        53: {'seconds': 934980, 'cost': 734500.0},
                        54: {'seconds': 978300, 'cost': 773950.0},
                        55: {'seconds': 1022760, 'cost': 814720.0},
                        56: {'seconds': 1068360, 'cost': 856840.0},
                        57: {'seconds': 1115160, 'cost': 900310.0},
                        58: {'seconds': 1163100, 'cost': 945160.0},
                        59: {'seconds': 1212300, 'cost': 991420.0},
                        60: {'seconds': 1262640, 'cost': 1040000.0},
                        61: {'seconds': 1314180, 'cost': 1090000.0},
                        62: {'seconds': 1366920, 'cost': 1140000.0},
                        63: {'seconds': 1420860, 'cost': 1190000.0},
                        64: {'seconds': 1476060, 'cost': 1240000.0},
                        65: {'seconds': 1532520, 'cost': 1300000.0},
                        66: {'seconds': 1590180, 'cost': 1360000.0},
                        67: {'seconds': 1649100, 'cost': 1410000.0},
                        68: {'seconds': 1709280, 'cost': 1470000.0},
                        69: {'seconds': 1770720, 'cost': 1530000.0},
                        70: {'seconds': 1833420, 'cost': 1600000.0},
                        71: {'seconds': 1897440, 'cost': 1660000.0},
                        72: {'seconds': 1962720, 'cost': 1730000.0},
                        73: {'seconds': 2029260, 'cost': 1800000.0},
                        74: {'seconds': 2097180, 'cost': 1870000.0},
                        75: {'seconds': 2166360, 'cost': 1940000.0},
                        76: {'seconds': 2236860, 'cost': 2010000.0},
                        77: {'seconds': 2308680, 'cost': 2080000.0},
                        78: {'seconds': 2381820, 'cost': 2160000.0},
                        79: {'seconds': 2456280, 'cost': 2240000.0},
                        80: {'seconds': 2532120, 'cost': 2320000.0},
                        81: {'seconds': 2609280, 'cost': 2400000.0},
                        82: {'seconds': 2687820, 'cost': 2480000.0},
                        83: {'seconds': 2767740, 'cost': 2570000.0},
                        84: {'seconds': 2849040, 'cost': 2650000.0},
                        85: {'seconds': 2931660, 'cost': 2740000.0},
                        86: {'seconds': 3015720, 'cost': 2830000.0},
                        87: {'seconds': 3101160, 'cost': 2930000.0},
                        88: {'seconds': 3187980, 'cost': 3020000.0},
                        89: {'seconds': 3276180, 'cost': 3120000.0},
                        90: {'seconds': 3365820, 'cost': 3220000.0},
                        91: {'seconds': 3456900, 'cost': 3320000.0},
                        92: {'seconds': 3549360, 'cost': 3420000.0},
                        93: {'seconds': 3643260, 'cost': 3520000.0},
                        94: {'seconds': 3738600, 'cost': 3630000.0},
                        95: {'seconds': 3835380, 'cost': 3740000.0},
                        96: {'seconds': 3933600, 'cost': 3850000.0},
                        97: {'seconds': 4033320, 'cost': 3960000.0},
                        98: {'seconds': 4134420, 'cost': 4070000.0},
                        99: {'seconds': 4237080, 'cost': 4190000.0}},
 'Golden Tower Bonus': {1: {'seconds': 144000, 'cost': 1000000.0},
                        2: {'seconds': 174060, 'cost': 1310000.0},
                        3: {'seconds': 204780, 'cost': 1850000.0},
                        4: {'seconds': 237240, 'cost': 3550000.0},
                        5: {'seconds': 272940, 'cost': 8500000.0},
                        6: {'seconds': 313560, 'cost': 20290000.0},
                        7: {'seconds': 361020, 'cost': 44330000.0},
                        8: {'seconds': 417480, 'cost': 88160000.0},
                        9: {'seconds': 485340, 'cost': 161660000.0},
                        10: {'seconds': 567060, 'cost': 277370000.0},
                        11: {'seconds': 665340, 'cost': 450680000.0},
                        12: {'seconds': 783000, 'cost': 700090000.0},
                        13: {'seconds': 922980, 'cost': 1050000000.0},
                        14: {'seconds': 1088460, 'cost': 1520000000.0},
                        15: {'seconds': 1282680, 'cost': 2140000000.0},
                        16: {'seconds': 1508940, 'cost': 2950000000.0},
                        17: {'seconds': 1770840, 'cost': 3980000000.0},
                        18: {'seconds': 2071920, 'cost': 5270000000.0},
                        19: {'seconds': 2415960, 'cost': 6880000000.0},
                        20: {'seconds': 2806800, 'cost': 8840000000.0},
                        21: {'seconds': 3248340, 'cost': 11220000000.0},
                        22: {'seconds': 3744720, 'cost': 14080000000.0},
                        23: {'seconds': 4300020, 'cost': 17480000000.0},
                        24: {'seconds': 4918560, 'cost': 21490000000.0},
                        25: {'seconds': 5604600, 'cost': 26190000000.0}},
 'Golden Tower Duration': {1: {'seconds': 144000, 'cost': 1000000.0},
                           2: {'seconds': 174060, 'cost': 1310000.0},
                           3: {'seconds': 204780, 'cost': 1850000.0},
                           4: {'seconds': 237240, 'cost': 3550000.0},
                           5: {'seconds': 272940, 'cost': 8500000.0},
                           6: {'seconds': 313560, 'cost': 20290000.0},
                           7: {'seconds': 361020, 'cost': 44330000.0},
                           8: {'seconds': 417480, 'cost': 88160000.0},
                           9: {'seconds': 485340, 'cost': 161660000.0},
                           10: {'seconds': 567060, 'cost': 277370000.0},
                           11: {'seconds': 665340, 'cost': 450680000.0},
                           12: {'seconds': 783000, 'cost': 700090000.0},
                           13: {'seconds': 922980, 'cost': 1050000000.0},
                           14: {'seconds': 1088460, 'cost': 1520000000.0},
                           15: {'seconds': 1282680, 'cost': 2140000000.0},
                           16: {'seconds': 1508940, 'cost': 2950000000.0},
                           17: {'seconds': 1770840, 'cost': 3980000000.0},
                           18: {'seconds': 2071920, 'cost': 5270000000.0},
                           19: {'seconds': 2415960, 'cost': 6880000000.0},
                           20: {'seconds': 2806800, 'cost': 8840000000.0}},
 'Black Hole Coin Bonus': {1: {'seconds': 143940, 'cost': 20000000.0},
                           2: {'seconds': 174060, 'cost': 21405405.0},
                           3: {'seconds': 204780, 'cost': 23124324.0},
                           4: {'seconds': 237240, 'cost': 26627027.0},
                           5: {'seconds': 272940, 'cost': 35837838.0},
                           6: {'seconds': 313560, 'cost': 58248649.0},
                           7: {'seconds': 361020, 'cost': 106162162.0},
                           8: {'seconds': 417480, 'cost': 197870270.0},
                           9: {'seconds': 485340, 'cost': 358875676.0},
                           10: {'seconds': 567060, 'cost': 623091892.0},
                           11: {'seconds': 665340, 'cost': 1030000000.0},
                           12: {'seconds': 783000, 'cost': 1640000000.0},
                           13: {'seconds': 922980, 'cost': 2520000000.0},
                           14: {'seconds': 1088460, 'cost': 3740000000.0},
                           15: {'seconds': 1282680, 'cost': 5410000000.0},
                           16: {'seconds': 1508940, 'cost': 7630000000.0},
                           17: {'seconds': 1770840, 'cost': 10520000000.0},
                           18: {'seconds': 2071920, 'cost': 14230000000.0},
                           19: {'seconds': 2415931, 'cost': 18930000000.0},
                           20: {'seconds': 2806800, 'cost': 24800000000.0}},
 'Spotlight Coin Bonus': {1: {'seconds': 143940, 'cost': 20000000.0},
                          2: {'seconds': 174060, 'cost': 21405405.0},
                          3: {'seconds': 204780, 'cost': 23124324.0},
                          4: {'seconds': 237240, 'cost': 26627027.0},
                          5: {'seconds': 272940, 'cost': 35837838.0},
                          6: {'seconds': 313560, 'cost': 58248649.0},
                          7: {'seconds': 361020, 'cost': 106162162.0},
                          8: {'seconds': 417480, 'cost': 197870270.0},
                          9: {'seconds': 485340, 'cost': 358875676.0},
                          10: {'seconds': 567060, 'cost': 623091892.0},
                          11: {'seconds': 665340, 'cost': 1030000000.0},
                          12: {'seconds': 783000, 'cost': 1640000000.0},
                          13: {'seconds': 922980, 'cost': 2520000000.0},
                          14: {'seconds': 1088460, 'cost': 3740000000.0},
                          15: {'seconds': 1282680, 'cost': 5410000000.0},
                          16: {'seconds': 1508940, 'cost': 7630000000.0},
                          17: {'seconds': 1770840, 'cost': 10520000000.0},
                          18: {'seconds': 2071920, 'cost': 14230000000.0},
                          19: {'seconds': 2415931, 'cost': 18930000000.0},
                          20: {'seconds': 2806800, 'cost': 24800000000.0}},
 'Death Wave Coin Bonus': {1: {'seconds': 72000, 'cost': 250000.0},
                           2: {'seconds': 102069, 'cost': 560000.0},
                           3: {'seconds': 132791, 'cost': 1100000.0},
                           4: {'seconds': 165273, 'cost': 2800000.0},
                           5: {'seconds': 200959, 'cost': 7750000.0},
                           6: {'seconds': 241565, 'cost': 19540000.0},
                           7: {'seconds': 289036, 'cost': 43580000.0},
                           8: {'seconds': 345524, 'cost': 87410000.0},
                           9: {'seconds': 413370, 'cost': 160910000.0},
                           10: {'seconds': 495089, 'cost': 276620000.0},
                           11: {'seconds': 593359, 'cost': 449930000.0},
                           12: {'seconds': 711009, 'cost': 699340000.0},
                           13: {'seconds': 851017, 'cost': 1050000000.0},
                           14: {'seconds': 1016497, 'cost': 1520000000.0},
                           15: {'seconds': 1217897, 'cost': 2140000000.0},
                           16: {'seconds': 1436991, 'cost': 2950000000.0},
                           17: {'seconds': 1698879, 'cost': 3980000000.0},
                           18: {'seconds': 1985577, 'cost': 5270000000.0},
                           19: {'seconds': 2344015, 'cost': 6880000000.0},
                           20: {'seconds': 2734837, 'cost': 8840000000.0}},
 'Extra Black Hole': {1: {'seconds': 1499940, 'cost': 15000000000.0}},
 'Labs Coin Discount': {1: {'seconds': 19, 'cost': 40.0},
                        2: {'seconds': 589, 'cost': 83.0},
                        3: {'seconds': 1429, 'cost': 210.0},
                        4: {'seconds': 2672, 'cost': 517.0},
                        5: {'seconds': 4380, 'cost': 1100.0},
                        6: {'seconds': 6720, 'cost': 2070.0},
                        7: {'seconds': 9720, 'cost': 3510.0},
                        8: {'seconds': 13500, 'cost': 5550.0},
                        9: {'seconds': 18060, 'cost': 8270.0},
                        10: {'seconds': 23460, 'cost': 11790.0},
                        11: {'seconds': 29820, 'cost': 16200.0},
                        12: {'seconds': 37140, 'cost': 21620.0},
                        13: {'seconds': 45540, 'cost': 28150.0},
                        14: {'seconds': 54960, 'cost': 35890.0},
                        15: {'seconds': 65580, 'cost': 44940.0},
                        16: {'seconds': 77340, 'cost': 55400.0},
                        17: {'seconds': 90300, 'cost': 67380.0},
                        18: {'seconds': 104580, 'cost': 80990.0},
                        19: {'seconds': 120780, 'cost': 96310.0},
                        20: {'seconds': 137100, 'cost': 113450.0},
                        21: {'seconds': 155460, 'cost': 132510.0},
                        22: {'seconds': 175200, 'cost': 153580.0},
                        23: {'seconds': 196440, 'cost': 176770.0},
                        24: {'seconds': 219240, 'cost': 202170.0},
                        25: {'seconds': 243540, 'cost': 229870.0},
                        26: {'seconds': 269460, 'cost': 259970.0},
                        27: {'seconds': 296940, 'cost': 292560.0},
                        28: {'seconds': 326160, 'cost': 327730.0},
                        29: {'seconds': 357060, 'cost': 365580.0},
                        30: {'seconds': 389640, 'cost': 406200.0},
                        31: {'seconds': 424020, 'cost': 449660.0},
                        32: {'seconds': 460200, 'cost': 496070.0},
                        33: {'seconds': 498180, 'cost': 545500.0},
                        34: {'seconds': 538080, 'cost': 598050.0},
                        35: {'seconds': 579780, 'cost': 653790.0},
                        36: {'seconds': 623460, 'cost': 712810.0},
                        37: {'seconds': 669120, 'cost': 775190.0},
                        38: {'seconds': 716760, 'cost': 841010.0},
                        39: {'seconds': 766380, 'cost': 910340.0},
                        40: {'seconds': 818040, 'cost': 983280.0},
                        41: {'seconds': 871800, 'cost': 1060000.0},
                        42: {'seconds': 927660, 'cost': 1140000.0},
                        43: {'seconds': 985680, 'cost': 1220000.0},
                        44: {'seconds': 1045800, 'cost': 1310000.0},
                        45: {'seconds': 1108200, 'cost': 1400000.0},
                        46: {'seconds': 1172760, 'cost': 1500000.0},
                        47: {'seconds': 1239540, 'cost': 1600000.0},
                        48: {'seconds': 1308600, 'cost': 1710000.0},
                        49: {'seconds': 1380000, 'cost': 1810000.0},
                        50: {'seconds': 1453680, 'cost': 1930000.0},
                        51: {'seconds': 1529940, 'cost': 2040000.0},
                        52: {'seconds': 1608180, 'cost': 2170000.0},
                        53: {'seconds': 1689060, 'cost': 2290000.0},
                        54: {'seconds': 1772340, 'cost': 2420000.0},
                        55: {'seconds': 1858020, 'cost': 2560000.0},
                        56: {'seconds': 1946280, 'cost': 2700000.0},
                        57: {'seconds': 2037000, 'cost': 2840000.0},
                        58: {'seconds': 2130240, 'cost': 2990000.0},
                        59: {'seconds': 2226060, 'cost': 3150000.0},
                        60: {'seconds': 2324400, 'cost': 3310000.0},
                        61: {'seconds': 2425440, 'cost': 3470000.0},
                        62: {'seconds': 2529060, 'cost': 3640000.0},
                        63: {'seconds': 2635320, 'cost': 3810000.0},
                        64: {'seconds': 2744280, 'cost': 3990000.0},
                        65: {'seconds': 2855940, 'cost': 4180000.0},
                        66: {'seconds': 2970300, 'cost': 4360000.0},
                        67: {'seconds': 3087480, 'cost': 4560000.0},
                        68: {'seconds': 3207360, 'cost': 4760000.0},
                        69: {'seconds': 3330060, 'cost': 4960000.0},
                        70: {'seconds': 3455580, 'cost': 5170000.0},
                        71: {'seconds': 3583920, 'cost': 5390000.0},
                        72: {'seconds': 3715140, 'cost': 5610000.0},
                        73: {'seconds': 3849240, 'cost': 5840000.0},
                        74: {'seconds': 3986280, 'cost': 6070000.0},
                        75: {'seconds': 4126200, 'cost': 6300000.0},
                        76: {'seconds': 4269060, 'cost': 6550000.0},
                        77: {'seconds': 4414980, 'cost': 6800000.0},
                        78: {'seconds': 4563780, 'cost': 7050000.0},
                        79: {'seconds': 4715700, 'cost': 7310000.0},
                        80: {'seconds': 4870620, 'cost': 7570000.0},
                        81: {'seconds': 5028600, 'cost': 7840000.0},
                        82: {'seconds': 5189640, 'cost': 8120000.0},
                        83: {'seconds': 5353800, 'cost': 8400000.0},
                        84: {'seconds': 5521080, 'cost': 8690000.0},
                        85: {'seconds': 5691480, 'cost': 8980000.0},
                        86: {'seconds': 5865060, 'cost': 9280000.0},
                        87: {'seconds': 6041880, 'cost': 9580000.0},
                        88: {'seconds': 6221820, 'cost': 9890000.0},
                        89: {'seconds': 6405060, 'cost': 10200000.0},
                        90: {'seconds': 6591480, 'cost': 10530000.0},
                        91: {'seconds': 6781200, 'cost': 10850000.0},
                        92: {'seconds': 6974220, 'cost': 11180000.0},
                        93: {'seconds': 7170540, 'cost': 11520000.0},
                        94: {'seconds': 7370160, 'cost': 11860000.0},
                        95: {'seconds': 7573140, 'cost': 12210000.0},
                        96: {'seconds': 7779540, 'cost': 12570000.0},
                        97: {'seconds': 7989240, 'cost': 12930000.0},
                        98: {'seconds': 8202420, 'cost': 13290000.0},
                        99: {'seconds': 8418960, 'cost': 13670000.0}},
 'Labs Speed': {1: {'seconds': 24, 'cost': 40.0},
                2: {'seconds': 583, 'cost': 83.0},
                3: {'seconds': 1379, 'cost': 211.0},
                4: {'seconds': 2526, 'cost': 522.0},
                5: {'seconds': 4080, 'cost': 1120.0},
                6: {'seconds': 6120, 'cost': 2100.0},
                7: {'seconds': 8700, 'cost': 3580.0},
                8: {'seconds': 11820, 'cost': 5670.0},
                9: {'seconds': 15540, 'cost': 8470.0},
                10: {'seconds': 19920, 'cost': 12120.0},
                11: {'seconds': 24840, 'cost': 16710.0},
                12: {'seconds': 30480, 'cost': 22360.0},
                13: {'seconds': 36720, 'cost': 29200.0},
                14: {'seconds': 43620, 'cost': 37340.0},
                15: {'seconds': 51240, 'cost': 46910.0},
                16: {'seconds': 59460, 'cost': 58010.0},
                17: {'seconds': 68400, 'cost': 70780.0},
                18: {'seconds': 78060, 'cost': 85340.0},
                19: {'seconds': 88380, 'cost': 101810.0},
                20: {'seconds': 99360, 'cost': 120310.0},
                21: {'seconds': 111000, 'cost': 140970.0},
                22: {'seconds': 123360, 'cost': 163910.0},
                23: {'seconds': 136440, 'cost': 189260.0},
                24: {'seconds': 150120, 'cost': 217150.0},
                25: {'seconds': 164520, 'cost': 247700.0},
                26: {'seconds': 179640, 'cost': 281050.0},
                27: {'seconds': 195360, 'cost': 317310.0},
                28: {'seconds': 211800, 'cost': 356620.0},
                29: {'seconds': 228840, 'cost': 399110.0},
                30: {'seconds': 246600, 'cost': 444900.0},
                31: {'seconds': 265020, 'cost': 494140.0},
                32: {'seconds': 284040, 'cost': 546930.0},
                33: {'seconds': 303780, 'cost': 603430.0},
                34: {'seconds': 324120, 'cost': 663760.0},
                35: {'seconds': 345120, 'cost': 728050.0},
                36: {'seconds': 366720, 'cost': 796430.0},
                37: {'seconds': 389040, 'cost': 869040.0},
                38: {'seconds': 411900, 'cost': 946010.0},
                39: {'seconds': 435420, 'cost': 1030000.0},
                40: {'seconds': 459600, 'cost': 1110000.0},
                41: {'seconds': 484320, 'cost': 1200000.0},
                42: {'seconds': 509700, 'cost': 1300000.0},
                43: {'seconds': 535680, 'cost': 1400000.0},
                44: {'seconds': 562260, 'cost': 1510000.0},
                45: {'seconds': 589440, 'cost': 1620000.0},
                46: {'seconds': 617220, 'cost': 1730000.0},
                47: {'seconds': 645600, 'cost': 1860000.0},
                48: {'seconds': 674520, 'cost': 1990000.0},
                49: {'seconds': 704040, 'cost': 2120000.0},
                50: {'seconds': 734160, 'cost': 2260000.0},
                51: {'seconds': 764880, 'cost': 2400000.0},
                52: {'seconds': 796140, 'cost': 2560000.0},
                53: {'seconds': 827940, 'cost': 2720000.0},
                54: {'seconds': 860340, 'cost': 2880000.0},
                55: {'seconds': 893280, 'cost': 3050000.0},
                56: {'seconds': 926760, 'cost': 3230000.0},
                57: {'seconds': 960840, 'cost': 3420000.0},
                58: {'seconds': 995400, 'cost': 3610000.0},
                59: {'seconds': 1030560, 'cost': 3810000.0},
                60: {'seconds': 1066260, 'cost': 4020000.0},
                61: {'seconds': 1102440, 'cost': 4230000.0},
                62: {'seconds': 1139220, 'cost': 4450000.0},
                63: {'seconds': 1176480, 'cost': 4680000.0},
                64: {'seconds': 1214280, 'cost': 4920000.0},
                65: {'seconds': 1252620, 'cost': 5170000.0},
                66: {'seconds': 1291440, 'cost': 5420000.0},
                67: {'seconds': 1330800, 'cost': 5680000.0},
                68: {'seconds': 1370640, 'cost': 5960000.0},
                69: {'seconds': 1411020, 'cost': 6240000.0},
                70: {'seconds': 1451880, 'cost': 6520000.0},
                71: {'seconds': 1493280, 'cost': 6820000.0},
                72: {'seconds': 1535160, 'cost': 7130000.0},
                73: {'seconds': 1577520, 'cost': 7440000.0},
                74: {'seconds': 1620420, 'cost': 7770000.0},
                75: {'seconds': 1663800, 'cost': 8100000.0},
                76: {'seconds': 1707600, 'cost': 8450000.0},
                77: {'seconds': 1751940, 'cost': 8800000.0},
                78: {'seconds': 1796760, 'cost': 9170000.0},
                79: {'seconds': 1842060, 'cost': 9540000.0},
                80: {'seconds': 1887840, 'cost': 9920000.0},
                81: {'seconds': 1934040, 'cost': 10320000.0},
                82: {'seconds': 1980780, 'cost': 10720000.0},
                83: {'seconds': 2027940, 'cost': 11140000.0},
                84: {'seconds': 2075580, 'cost': 11570000.0},
                85: {'seconds': 2123700, 'cost': 12000000.0},
                86: {'seconds': 2172240, 'cost': 12450000.0},
                87: {'seconds': 2221260, 'cost': 12910000.0},
                88: {'seconds': 2270700, 'cost': 13380000.0},
                89: {'seconds': 2320620, 'cost': 13870000.0},
                90: {'seconds': 2371020, 'cost': 14360000.0},
                91: {'seconds': 2421840, 'cost': 14870000.0},
                92: {'seconds': 2473080, 'cost': 15380000.0},
                93: {'seconds': 2524800, 'cost': 15910000.0},
                94: {'seconds': 2576940, 'cost': 16460000.0},
                95: {'seconds': 2629560, 'cost': 17010000.0},
                96: {'seconds': 2682600, 'cost': 17580000.0},
                97: {'seconds': 2736000, 'cost': 18160000.0},
                98: {'seconds': 2789940, 'cost': 18750000.0},
                99: {'seconds': 2844240, 'cost': 19360000.0}}}

NATIVE_ECON_UW_TABLES = {'GT | Bonus': [{'value': 5.0, 'cost': 0.0},
                {'value': 5.8, 'cost': 5.0},
                {'value': 6.6, 'cost': 13.0},
                {'value': 7.4, 'cost': 22.0},
                {'value': 8.2, 'cost': 32.0},
                {'value': 9.0, 'cost': 43.0},
                {'value': 9.8, 'cost': 55.0},
                {'value': 10.6, 'cost': 68.0},
                {'value': 11.4, 'cost': 82.0},
                {'value': 12.2, 'cost': 98.0},
                {'value': 13.0, 'cost': 116.0},
                {'value': 13.8, 'cost': 138.0},
                {'value': 14.6, 'cost': 162.0},
                {'value': 15.4, 'cost': 250.0},
                {'value': 16.2, 'cost': 350.0},
                {'value': 17.0, 'cost': 500.0},
                {'value': 17.8, 'cost': 700.0},
                {'value': 18.6, 'cost': 950.0},
                {'value': 19.4, 'cost': 1250.0},
                {'value': 20.2, 'cost': 1600.0},
                {'value': 21.0, 'cost': 2000.0}],
 'GT | Duration': [{'value': 15.0, 'cost': 0.0},
                   {'value': 16.0, 'cost': 5.0},
                   {'value': 17.0, 'cost': 14.0},
                   {'value': 18.0, 'cost': 23.0},
                   {'value': 19.0, 'cost': 32.0},
                   {'value': 20.0, 'cost': 41.0},
                   {'value': 21.0, 'cost': 50.0},
                   {'value': 22.0, 'cost': 59.0},
                   {'value': 23.0, 'cost': 68.0},
                   {'value': 24.0, 'cost': 77.0},
                   {'value': 25.0, 'cost': 87.0},
                   {'value': 26.0, 'cost': 98.0},
                   {'value': 27.0, 'cost': 110.0},
                   {'value': 28.0, 'cost': 123.0},
                   {'value': 29.0, 'cost': 137.0},
                   {'value': 30.0, 'cost': 152.0},
                   {'value': 31.0, 'cost': 168.0},
                   {'value': 32.0, 'cost': 185.0},
                   {'value': 33.0, 'cost': 203.0},
                   {'value': 34.0, 'cost': 222.0},
                   {'value': 35.0, 'cost': 242.0},
                   {'value': 36.0, 'cost': 263.0},
                   {'value': 37.0, 'cost': 285.0},
                   {'value': 38.0, 'cost': 308.0},
                   {'value': 39.0, 'cost': 332.0},
                   {'value': 40.0, 'cost': 356.0},
                   {'value': 41.0, 'cost': 380.0},
                   {'value': 42.0, 'cost': 404.0},
                   {'value': 43.0, 'cost': 428.0},
                   {'value': 44.0, 'cost': 452.0},
                   {'value': 45.0, 'cost': 476.0},
                   {'value': 46.0, 'cost': 530.0},
                   {'value': 47.0, 'cost': 614.0},
                   {'value': 48.0, 'cost': 728.0},
                   {'value': 49.0, 'cost': 872.0},
                   {'value': 50.0, 'cost': 1046.0},
                   {'value': 51.0, 'cost': 1250.0},
                   {'value': 52.0, 'cost': 1484.0},
                   {'value': 53.0, 'cost': 1748.0}],
 'GT | Cooldown': [{'value': 300.0, 'cost': 0.0},
                   {'value': 290.0, 'cost': 10.0},
                   {'value': 280.0, 'cost': 28.0},
                   {'value': 270.0, 'cost': 46.0},
                   {'value': 260.0, 'cost': 64.0},
                   {'value': 250.0, 'cost': 82.0},
                   {'value': 240.0, 'cost': 100.0},
                   {'value': 230.0, 'cost': 118.0},
                   {'value': 220.0, 'cost': 136.0},
                   {'value': 210.0, 'cost': 154.0},
                   {'value': 200.0, 'cost': 172.0},
                   {'value': 190.0, 'cost': 190.0},
                   {'value': 180.0, 'cost': 208.0},
                   {'value': 170.0, 'cost': 226.0},
                   {'value': 160.0, 'cost': 244.0},
                   {'value': 150.0, 'cost': 262.0},
                   {'value': 140.0, 'cost': 300.0},
                   {'value': 130.0, 'cost': 368.0},
                   {'value': 120.0, 'cost': 476.0},
                   {'value': 110.0, 'cost': 644.0},
                   {'value': 100.0, 'cost': 872.0}],
 'BH | Size': [{'value': 30.0, 'cost': 0.0},
               {'value': 32.0, 'cost': 5.0},
               {'value': 34.0, 'cost': 12.0},
               {'value': 36.0, 'cost': 19.0},
               {'value': 38.0, 'cost': 26.0},
               {'value': 40.0, 'cost': 34.0},
               {'value': 42.0, 'cost': 43.0},
               {'value': 44.0, 'cost': 53.0},
               {'value': 46.0, 'cost': 64.0},
               {'value': 48.0, 'cost': 76.0},
               {'value': 50.0, 'cost': 89.0},
               {'value': 52.0, 'cost': 103.0},
               {'value': 54.0, 'cost': 118.0},
               {'value': 56.0, 'cost': 134.0},
               {'value': 58.0, 'cost': 151.0},
               {'value': 60.0, 'cost': 169.0},
               {'value': 62.0, 'cost': 189.0},
               {'value': 64.0, 'cost': 211.0},
               {'value': 66.0, 'cost': 236.0},
               {'value': 68.0, 'cost': 264.0},
               {'value': 70.0, 'cost': 295.0}],
 'BH | Duration': [{'value': 15.0, 'cost': 0.0},
                   {'value': 16.0, 'cost': 5.0},
                   {'value': 17.0, 'cost': 14.0},
                   {'value': 18.0, 'cost': 23.0},
                   {'value': 19.0, 'cost': 32.0},
                   {'value': 20.0, 'cost': 41.0},
                   {'value': 21.0, 'cost': 50.0},
                   {'value': 22.0, 'cost': 59.0},
                   {'value': 23.0, 'cost': 68.0},
                   {'value': 24.0, 'cost': 77.0},
                   {'value': 25.0, 'cost': 86.0},
                   {'value': 26.0, 'cost': 95.0},
                   {'value': 27.0, 'cost': 104.0},
                   {'value': 28.0, 'cost': 113.0},
                   {'value': 29.0, 'cost': 122.0},
                   {'value': 30.0, 'cost': 131.0},
                   {'value': 31.0, 'cost': 165.0},
                   {'value': 32.0, 'cost': 224.0},
                   {'value': 33.0, 'cost': 308.0},
                   {'value': 34.0, 'cost': 417.0},
                   {'value': 35.0, 'cost': 551.0},
                   {'value': 36.0, 'cost': 710.0},
                   {'value': 37.0, 'cost': 894.0},
                   {'value': 38.0, 'cost': 1103.0}],
 'BH | Cooldown': [{'value': 200.0, 'cost': 0.0},
                   {'value': 190.0, 'cost': 10.0},
                   {'value': 180.0, 'cost': 28.0},
                   {'value': 170.0, 'cost': 46.0},
                   {'value': 160.0, 'cost': 64.0},
                   {'value': 150.0, 'cost': 82.0},
                   {'value': 140.0, 'cost': 100.0},
                   {'value': 130.0, 'cost': 118.0},
                   {'value': 120.0, 'cost': 136.0},
                   {'value': 110.0, 'cost': 154.0},
                   {'value': 100.0, 'cost': 172.0},
                   {'value': 90.0, 'cost': 190.0},
                   {'value': 80.0, 'cost': 208.0},
                   {'value': 70.0, 'cost': 226.0},
                   {'value': 60.0, 'cost': 244.0},
                   {'value': 50.0, 'cost': 262.0}],
 'SL | Angle': [{'value': 30.0, 'cost': 0.0},
                {'value': 31.0, 'cost': 5.0},
                {'value': 32.0, 'cost': 16.0},
                {'value': 33.0, 'cost': 27.0},
                {'value': 34.0, 'cost': 38.0},
                {'value': 35.0, 'cost': 49.0},
                {'value': 36.0, 'cost': 60.0},
                {'value': 37.0, 'cost': 71.0},
                {'value': 38.0, 'cost': 82.0},
                {'value': 39.0, 'cost': 93.0},
                {'value': 40.0, 'cost': 104.0},
                {'value': 41.0, 'cost': 115.0},
                {'value': 42.0, 'cost': 126.0},
                {'value': 43.0, 'cost': 137.0},
                {'value': 44.0, 'cost': 148.0},
                {'value': 45.0, 'cost': 159.0},
                {'value': 46.0, 'cost': 170.0},
                {'value': 47.0, 'cost': 181.0},
                {'value': 48.0, 'cost': 192.0},
                {'value': 49.0, 'cost': 203.0},
                {'value': 50.0, 'cost': 214.0},
                {'value': 51.0, 'cost': 225.0},
                {'value': 52.0, 'cost': 236.0},
                {'value': 53.0, 'cost': 247.0},
                {'value': 54.0, 'cost': 258.0},
                {'value': 55.0, 'cost': 269.0},
                {'value': 56.0, 'cost': 280.0},
                {'value': 57.0, 'cost': 291.0},
                {'value': 58.0, 'cost': 302.0},
                {'value': 59.0, 'cost': 313.0},
                {'value': 60.0, 'cost': 324.0},
                {'value': 61.0, 'cost': 337.0},
                {'value': 62.0, 'cost': 352.0},
                {'value': 63.0, 'cost': 369.0},
                {'value': 64.0, 'cost': 388.0},
                {'value': 65.0, 'cost': 409.0},
                {'value': 66.0, 'cost': 432.0},
                {'value': 67.0, 'cost': 457.0},
                {'value': 68.0, 'cost': 484.0},
                {'value': 69.0, 'cost': 513.0},
                {'value': 70.0, 'cost': 544.0},
                {'value': 71.0, 'cost': 577.0},
                {'value': 72.0, 'cost': 612.0},
                {'value': 73.0, 'cost': 649.0},
                {'value': 74.0, 'cost': 688.0},
                {'value': 75.0, 'cost': 729.0},
                {'value': 76.0, 'cost': 772.0},
                {'value': 77.0, 'cost': 817.0},
                {'value': 78.0, 'cost': 864.0},
                {'value': 79.0, 'cost': 913.0},
                {'value': 80.0, 'cost': 964.0},
                {'value': 81.0, 'cost': 1017.0},
                {'value': 82.0, 'cost': 1072.0},
                {'value': 83.0, 'cost': 1129.0},
                {'value': 84.0, 'cost': 1188.0},
                {'value': 85.0, 'cost': 1249.0},
                {'value': 86.0, 'cost': 1312.0},
                {'value': 87.0, 'cost': 1377.0},
                {'value': 88.0, 'cost': 1444.0},
                {'value': 89.0, 'cost': 1513.0},
                {'value': 90.0, 'cost': 1584.0}],
 'SL | Quantity': [{'value': 1.0, 'cost': 0.0},
                   {'value': 2.0, 'cost': 375.0},
                   {'value': 3.0, 'cost': 850.0},
                   {'value': 4.0, 'cost': 2500.0}],
 'DW | Quantity': [{'value': 1.0, 'cost': 0.0},
                   {'value': 2.0, 'cost': 200.0},
                   {'value': 3.0, 'cost': 500.0},
                   {'value': 4.0, 'cost': 850.0},
                   {'value': 5.0, 'cost': 1400.0}],
 'DW | Cooldown': [{'value': 300.0, 'cost': 0.0},
                   {'value': 290.0, 'cost': 8.0},
                   {'value': 280.0, 'cost': 24.0},
                   {'value': 270.0, 'cost': 40.0},
                   {'value': 260.0, 'cost': 56.0},
                   {'value': 250.0, 'cost': 72.0},
                   {'value': 240.0, 'cost': 88.0},
                   {'value': 230.0, 'cost': 104.0},
                   {'value': 220.0, 'cost': 120.0},
                   {'value': 210.0, 'cost': 136.0},
                   {'value': 200.0, 'cost': 152.0},
                   {'value': 190.0, 'cost': 168.0},
                   {'value': 180.0, 'cost': 184.0},
                   {'value': 170.0, 'cost': 200.0},
                   {'value': 160.0, 'cost': 216.0},
                   {'value': 150.0, 'cost': 232.0},
                   {'value': 140.0, 'cost': 248.0},
                   {'value': 130.0, 'cost': 264.0},
                   {'value': 120.0, 'cost': 280.0},
                   {'value': 110.0, 'cost': 346.0},
                   {'value': 100.0, 'cost': 512.0},
                   {'value': 90.0, 'cost': 688.0},
                   {'value': 80.0, 'cost': 874.0},
                   {'value': 70.0, 'cost': 1070.0},
                   {'value': 60.0, 'cost': 1276.0},
                   {'value': 50.0, 'cost': 1492.0}]}

NATIVE_ECON_ENHANCEMENT_COSTS = {'Coin Bonus +': [0.0,
                  5000000000.0,
                  6250000000.0,
                  12460000000.0,
                  27420000000.0,
                  54500000000.0,
                  96850000000.0,
                  157450000000.0,
                  239170000000.0,
                  344790000000.0,
                  477010000000.0,
                  638460000000.0,
                  831700000000.0,
                  1060000000000.0,
                  1320000000000.0,
                  1630000000000.0,
                  1970000000000.0,
                  2360000000000.0,
                  5590000000000.0,
                  9840000000000.0,
                  15260000000000.0,
                  22020000000000.0,
                  30290000000000.0,
                  40240000000000.0,
                  52080000000000.0,
                  65990000000000.0,
                  82190000000000.0,
                  100890000000000.0,
                  122320000000000.0,
                  146710000000000.0,
                  174290000000000.0,
                  205320000000000.0,
                  288060000000000.0,
                  390250000000000.0,
                  514700000000000.0,
                  664460000000000.0,
                  842820000000000.0,
                  1050000000000000.0,
                  1300000000000000.0,
                  1590000000000000.0,
                  1920000000000000.0,
                  2300000000000000.0,
                  2730000000000000.0,
                  3220000000000000.0,
                  3780000000000000.0,
                  4410000000000000.0,
                  5110000000000000.0,
                  5900000000000000.0,
                  6770000000000000.0,
                  7740000000000000.0,
                  8820000000000000.0,
                  1.001e+16,
                  1.301e+16,
                  1.658e+16,
                  2.077e+16,
                  2.567e+16,
                  3.136e+16,
                  3.793e+16,
                  4.547e+16,
                  5.409e+16,
                  6.39e+16,
                  7.502e+16,
                  8.757e+16,
                  1.0169e+17,
                  1.1752e+17,
                  1.3522e+17,
                  1.5495e+17,
                  1.7687e+17,
                  2.0118e+17,
                  2.2807e+17,
                  2.5774e+17,
                  2.9041e+17,
                  3.263e+17,
                  3.6565e+17,
                  4.0872e+17,
                  4.5577e+17,
                  5.0708e+17,
                  5.6294e+17,
                  6.2366e+17,
                  6.8955e+17,
                  7.6097e+17,
                  8.3825e+17,
                  9.2176e+17,
                  1.01e+18,
                  1.11e+18,
                  1.21e+18,
                  1.33e+18,
                  1.45e+18,
                  1.58e+18,
                  1.72e+18,
                  1.86e+18,
                  2.02e+18,
                  2.19e+18,
                  2.37e+18,
                  2.57e+18,
                  2.77e+18,
                  2.99e+18,
                  3.23e+18,
                  3.48e+18,
                  3.74e+18,
                  4.02e+18,
                  4.31e+18,
                  4.63e+18,
                  4.96e+18,
                  5.31e+18,
                  5.69e+18,
                  6.08e+18,
                  6.49e+18,
                  6.93e+18,
                  7.4e+18,
                  7.88e+18,
                  8.4e+18,
                  8.94e+18,
                  9.51e+18,
                  1.01e+19,
                  1.073e+19,
                  1.139e+19,
                  1.209e+19,
                  1.282e+19,
                  1.358e+19,
                  1.438e+19,
                  1.522e+19,
                  1.61e+19,
                  1.702e+19,
                  1.798e+19,
                  1.899e+19,
                  2.004e+19,
                  2.114e+19,
                  2.229e+19,
                  2.349e+19,
                  2.475e+19,
                  2.606e+19,
                  2.742e+19,
                  2.885e+19,
                  3.033e+19,
                  3.188e+19,
                  3.349e+19,
                  3.516e+19,
                  3.691e+19,
                  3.872e+19,
                  4.061e+19,
                  4.258e+19,
                  4.462e+19,
                  4.674e+19,
                  4.895e+19,
                  5.124e+19,
                  5.361e+19,
                  5.608e+19,
                  5.864e+19,
                  6.13e+19,
                  6.405e+19,
                  6.691e+19,
                  6.987e+19,
                  7.293e+19,
                  7.611e+19,
                  7.94e+19,
                  8.281e+19,
                  8.633e+19,
                  8.998e+19,
                  9.376e+19,
                  9.766e+19,
                  1.017e+20,
                  1.0587e+20,
                  1.1019e+20,
                  1.1465e+20,
                  1.1926e+20,
                  1.2402e+20,
                  1.2893e+20,
                  1.3401e+20,
                  1.3924e+20,
                  1.4465e+20,
                  1.5023e+20,
                  1.5598e+20,
                  1.6192e+20,
                  1.6804e+20,
                  1.7435e+20,
                  1.8085e+20,
                  1.8756e+20,
                  1.9446e+20,
                  2.0158e+20,
                  2.0891e+20,
                  2.1646e+20,
                  2.2423e+20,
                  2.3223e+20,
                  2.4047e+20,
                  2.4894e+20,
                  2.5766e+20,
                  2.6663e+20,
                  2.7586e+20,
                  2.8535e+20,
                  2.9511e+20,
                  3.0514e+20,
                  3.1545e+20,
                  3.2604e+20,
                  3.3693e+20,
                  3.4812e+20,
                  3.5962e+20,
                  3.7142e+20,
                  3.8354e+20,
                  3.9599e+20,
                  4.0877e+20],
 'Cells / Kill Bonus +': [0.0,
                          5000000000.0,
                          6250000000.0,
                          12460000000.0,
                          27420000000.0,
                          54500000000.0,
                          96850000000.0,
                          157450000000.0,
                          239170000000.0,
                          344790000000.0,
                          477010000000.0,
                          638460000000.0,
                          831700000000.0,
                          1060000000000.0,
                          1320000000000.0,
                          1630000000000.0,
                          1970000000000.0,
                          2360000000000.0,
                          5590000000000.0,
                          9840000000000.0,
                          15260000000000.0,
                          22020000000000.0,
                          30290000000000.0,
                          40240000000000.0,
                          52080000000000.0,
                          65990000000000.0,
                          82190000000000.0,
                          100890000000000.0,
                          122320000000000.0,
                          146710000000000.0,
                          174290000000000.0,
                          205320000000000.0,
                          288060000000000.0,
                          390250000000000.0,
                          514700000000000.0,
                          664460000000000.0,
                          842820000000000.0,
                          1050000000000000.0,
                          1300000000000000.0,
                          1590000000000000.0,
                          1920000000000000.0,
                          2300000000000000.0,
                          2730000000000000.0,
                          3220000000000000.0,
                          3780000000000000.0,
                          4410000000000000.0,
                          5110000000000000.0,
                          5900000000000000.0,
                          6770000000000000.0,
                          7740000000000000.0,
                          8820000000000000.0,
                          1.001e+16,
                          1.301e+16,
                          1.658e+16,
                          2.077e+16,
                          2.567e+16,
                          3.136e+16,
                          3.793e+16,
                          4.547e+16,
                          5.409e+16,
                          6.39e+16,
                          7.502e+16,
                          8.757e+16,
                          1.0169e+17,
                          1.1752e+17,
                          1.3522e+17,
                          1.5495e+17,
                          1.7687e+17,
                          2.0118e+17,
                          2.2807e+17,
                          2.5774e+17,
                          2.9041e+17,
                          3.263e+17,
                          3.6565e+17,
                          4.0872e+17,
                          4.5577e+17,
                          5.0708e+17,
                          5.6294e+17,
                          6.2366e+17,
                          6.8955e+17,
                          7.6097e+17,
                          8.3825e+17,
                          9.2176e+17,
                          1.01e+18,
                          1.11e+18,
                          1.21e+18,
                          1.33e+18,
                          1.45e+18,
                          1.58e+18,
                          1.72e+18,
                          1.86e+18,
                          2.02e+18,
                          2.19e+18,
                          2.37e+18,
                          2.57e+18,
                          2.77e+18,
                          2.99e+18,
                          3.23e+18,
                          3.48e+18,
                          3.74e+18,
                          4.02e+18,
                          4.31e+18,
                          4.63e+18,
                          4.96e+18,
                          5.31e+18,
                          5.69e+18,
                          6.08e+18,
                          6.49e+18,
                          6.93e+18,
                          7.4e+18,
                          7.88e+18,
                          8.4e+18,
                          8.94e+18,
                          9.51e+18,
                          1.01e+19,
                          1.073e+19,
                          1.139e+19,
                          1.209e+19,
                          1.282e+19,
                          1.358e+19,
                          1.438e+19,
                          1.522e+19,
                          1.61e+19,
                          1.702e+19,
                          1.798e+19,
                          1.899e+19,
                          2.004e+19,
                          2.114e+19,
                          2.229e+19,
                          2.349e+19,
                          2.475e+19,
                          2.606e+19,
                          2.742e+19,
                          2.885e+19,
                          3.033e+19,
                          3.188e+19,
                          3.349e+19,
                          3.516e+19,
                          3.691e+19,
                          3.872e+19,
                          4.061e+19,
                          4.258e+19,
                          4.462e+19,
                          4.674e+19,
                          4.895e+19,
                          5.124e+19,
                          5.361e+19,
                          5.608e+19,
                          5.864e+19,
                          6.13e+19,
                          6.405e+19,
                          6.691e+19,
                          6.987e+19,
                          7.293e+19,
                          7.611e+19,
                          7.94e+19,
                          8.281e+19,
                          8.633e+19,
                          8.998e+19,
                          9.376e+19,
                          9.766e+19,
                          1.017e+20,
                          1.0587e+20,
                          1.1019e+20,
                          1.1465e+20,
                          1.1926e+20,
                          1.2402e+20,
                          1.2893e+20,
                          1.3401e+20,
                          1.3924e+20,
                          1.4465e+20,
                          1.5023e+20,
                          1.5598e+20,
                          1.6192e+20,
                          1.6804e+20,
                          1.7435e+20,
                          1.8085e+20,
                          1.8756e+20,
                          1.9446e+20,
                          2.0158e+20,
                          2.0891e+20,
                          2.1646e+20,
                          2.2423e+20,
                          2.3223e+20,
                          2.4047e+20,
                          2.4894e+20,
                          2.5766e+20,
                          2.6663e+20,
                          2.7586e+20,
                          2.8535e+20,
                          2.9511e+20,
                          3.0514e+20,
                          3.1545e+20,
                          3.2604e+20,
                          3.3693e+20,
                          3.4812e+20,
                          3.5962e+20,
                          3.7142e+20,
                          3.8354e+20,
                          3.9599e+20,
                          4.0877e+20],
 'Free Upgrades +': [0.0,
                     5000000000.0,
                     6050000000.0,
                     22250000000.0,
                     95560000000.0,
                     299270000000.0,
                     739390000000.0,
                     1560000000000.0,
                     2920000000000.0,
                     5050000000000.0,
                     8180000000000.0,
                     12590000000000.0,
                     18610000000000.0,
                     26590000000000.0,
                     36920000000000.0,
                     50020000000000.0,
                     66380000000000.0,
                     86480000000000.0,
                     221770000000000.0,
                     420490000000000.0,
                     699780000000000.0,
                     1080000000000000.0,
                     1580000000000000.0,
                     2230000000000000.0,
                     3060000000000000.0,
                     4100000000000000.0,
                     5390000000000000.0,
                     6960000000000000.0,
                     8870000000000000.0,
                     1.115e+16,
                     1.387e+16,
                     1.707e+16,
                     2.5e+16,
                     3.529e+16,
                     4.845e+16,
                     6.503e+16,
                     8.565e+16,
                     1.1104e+17,
                     1.4199e+17,
                     1.794e+17,
                     2.2425e+17,
                     2.7766e+17,
                     3.4083e+17,
                     4.1511e+17,
                     5.0197e+17,
                     6.0303e+17,
                     7.2003e+17,
                     8.5491e+17,
                     1.01e+18,
                     1.19e+18,
                     1.39e+18,
                     1.62e+18,
                     2.16e+18,
                     2.82e+18,
                     3.62e+18,
                     4.59e+18,
                     5.74e+18,
                     7.1e+18,
                     8.72e+18,
                     1.06e+19,
                     1.281e+19,
                     1.537e+19,
                     1.833e+19,
                     2.174e+19,
                     2.566e+19,
                     3.013e+19,
                     3.523e+19,
                     4.102e+19,
                     4.758e+19,
                     5.499e+19,
                     6.333e+19,
                     7.271e+19,
                     8.321e+19,
                     9.496e+19,
                     1.0806e+20,
                     1.2266e+20,
                     1.3887e+20,
                     1.5684e+20,
                     1.7674e+20,
                     1.9872e+20,
                     2.2296e+20,
                     2.4965e+20,
                     2.79e+20,
                     3.112e+20,
                     3.465e+20,
                     3.8513e+20,
                     4.2734e+20,
                     4.7341e+20,
                     5.2363e+20,
                     5.7829e+20,
                     6.3772e+20,
                     7.0226e+20,
                     7.7226e+20,
                     8.481e+20,
                     9.3019e+20,
                     1.02e+21,
                     1.11e+21,
                     1.22e+21,
                     1.33e+21,
                     1.45e+21,
                     1.58e+21],
 'Recovery Package +': [0.0,
                        5000000000.0,
                        5040000000.0,
                        5110000000.0,
                        5200000000.0,
                        5330000000.0,
                        5490000000.0,
                        5700000000.0,
                        5930000000.0,
                        6210000000.0,
                        6530000000.0,
                        6880000000.0,
                        7280000000.0,
                        7730000000.0,
                        8210000000.0,
                        8740000000.0,
                        9320000000.0,
                        9940000000.0,
                        21210000000.0,
                        33950000000.0,
                        48300000000.0,
                        64410000000.0,
                        82420000000.0,
                        102490000000.0,
                        124750000000.0,
                        149360000000.0,
                        176480000000.0,
                        206250000000.0,
                        238830000000.0,
                        274390000000.0,
                        313070000000.0,
                        355040000000.0,
                        480540000000.0,
                        629270000000.0,
                        803650000000.0,
                        1010000000000.0,
                        1240000000000.0,
                        1510000000000.0,
                        1810000000000.0,
                        2150000000000.0,
                        2540000000000.0,
                        2970000000000.0,
                        3460000000000.0,
                        3990000000000.0,
                        4590000000000.0,
                        5240000000000.0,
                        5960000000000.0,
                        6760000000000.0,
                        7620000000000.0,
                        8560000000000.0,
                        9590000000000.0,
                        10700000000000.0,
                        13700000000000.0,
                        17180000000000.0,
                        21210000000000.0,
                        25830000000000.0,
                        31110000000000.0,
                        37110000000000.0,
                        43890000000000.0,
                        51520000000000.0,
                        60080000000000.0,
                        69650000000000.0,
                        80310000000000.0,
                        92140000000000.0,
                        105230000000000.0,
                        119690000000000.0,
                        135610000000000.0,
                        153100000000000.0,
                        172260000000000.0,
                        193210000000000.0,
                        216080000000000.0,
                        240980000000000.0,
                        268050000000000.0,
                        297430000000000.0,
                        329250000000000.0,
                        363670000000000.0,
                        400830000000000.0,
                        440910000000000.0,
                        484060000000000.0,
                        530450000000000.0,
                        580270000000000.0,
                        633700000000000.0,
                        690930000000000.0,
                        752160000000000.0,
                        817600000000000.0,
                        887460000000000.0,
                        961960000000000.0,
                        1040000000000000.0,
                        1130000000000000.0,
                        1220000000000000.0,
                        1310000000000000.0,
                        1410000000000000.0,
                        1520000000000000.0,
                        1630000000000000.0,
                        1750000000000000.0,
                        1880000000000000.0,
                        2020000000000000.0,
                        2160000000000000.0,
                        2310000000000000.0,
                        2470000000000000.0,
                        2630000000000000.0,
                        2810000000000000.0,
                        2990000000000000.0,
                        3190000000000000.0,
                        3390000000000000.0,
                        3610000000000000.0,
                        3830000000000000.0,
                        4070000000000000.0,
                        4320000000000000.0,
                        4580000000000000.0,
                        4850000000000000.0,
                        5140000000000000.0,
                        5440000000000000.0,
                        5750000000000000.0,
                        6080000000000000.0,
                        6420000000000000.0,
                        6770000000000000.0,
                        7150000000000000.0,
                        7530000000000000.0,
                        7940000000000000.0,
                        8360000000000000.0,
                        8800000000000000.0,
                        9260000000000000.0,
                        9740000000000000.0,
                        1.024e+16,
                        1.076e+16,
                        1.13e+16,
                        1.186e+16,
                        1.244e+16,
                        1.304e+16,
                        1.367e+16,
                        1.432e+16,
                        1.5e+16,
                        1.571e+16,
                        1.643e+16,
                        1.719e+16,
                        1.797e+16,
                        1.879e+16,
                        1.963e+16,
                        2.05e+16,
                        2.14e+16,
                        2.234e+16,
                        2.33e+16,
                        2.43e+16,
                        2.534e+16,
                        2.641e+16,
                        2.751e+16,
                        2.865e+16,
                        2.983e+16,
                        3.105e+16,
                        3.231e+16,
                        3.361e+16,
                        3.495e+16,
                        3.634e+16,
                        3.776e+16,
                        3.924e+16,
                        4.075e+16,
                        4.232e+16,
                        4.393e+16,
                        4.559e+16,
                        4.731e+16,
                        4.907e+16,
                        5.089e+16,
                        5.276e+16,
                        5.468e+16,
                        5.666e+16,
                        5.87e+16,
                        6.08e+16,
                        6.296e+16,
                        6.518e+16,
                        6.746e+16,
                        6.98e+16,
                        7.221e+16,
                        7.469e+16,
                        7.723e+16,
                        7.985e+16,
                        8.253e+16,
                        8.529e+16,
                        8.812e+16,
                        9.103e+16,
                        9.401e+16,
                        9.707e+16,
                        1.0022e+17,
                        1.0344e+17,
                        1.0674e+17,
                        1.1013e+17,
                        1.1361e+17,
                        1.1718e+17,
                        1.2083e+17,
                        1.2458e+17,
                        1.2841e+17,
                        1.3235e+17,
                        1.3638e+17,
                        1.4051e+17,
                        1.4473e+17,
                        1.4906e+17,
                        1.535e+17,
                        1.5804e+17,
                        1.6269e+17,
                        1.6744e+17,
                        1.7231e+17,
                        1.7729e+17,
                        2.1887e+17,
                        2.6265e+17,
                        3.0871e+17,
                        3.5712e+17,
                        4.0795e+17,
                        4.613e+17,
                        5.1724e+17,
                        5.7585e+17,
                        6.3722e+17,
                        7.0143e+17,
                        7.6857e+17,
                        8.3873e+17,
                        9.12e+17,
                        9.8848e+17,
                        1.07e+18,
                        1.15e+18,
                        1.24e+18,
                        1.33e+18,
                        1.42e+18,
                        1.52e+18,
                        1.62e+18,
                        1.73e+18,
                        1.84e+18,
                        1.95e+18,
                        2.07e+18,
                        2.19e+18,
                        2.32e+18,
                        2.45e+18,
                        2.59e+18,
                        2.73e+18,
                        2.88e+18,
                        3.03e+18,
                        3.19e+18,
                        3.35e+18,
                        3.52e+18,
                        3.7e+18,
                        3.88e+18,
                        4.07e+18,
                        4.26e+18,
                        4.46e+18,
                        4.67e+18,
                        4.88e+18,
                        5.1e+18,
                        5.33e+18,
                        5.56e+18,
                        5.8e+18,
                        6.05e+18,
                        6.31e+18,
                        6.57e+18,
                        6.84e+18,
                        7.13e+18,
                        7.42e+18,
                        7.71e+18,
                        8.02e+18,
                        8.34e+18,
                        8.66e+18,
                        9e+18,
                        9.34e+18,
                        9.7e+18,
                        1.006e+19,
                        1.044e+19,
                        1.082e+19,
                        1.122e+19,
                        1.163e+19,
                        1.205e+19,
                        1.248e+19,
                        1.292e+19,
                        1.337e+19,
                        1.384e+19,
                        1.432e+19,
                        1.481e+19,
                        1.531e+19,
                        1.583e+19,
                        1.636e+19,
                        1.691e+19,
                        1.746e+19,
                        1.804e+19,
                        1.862e+19,
                        1.923e+19,
                        1.984e+19,
                        2.048e+19,
                        2.113e+19,
                        2.179e+19,
                        2.247e+19,
                        2.317e+19,
                        2.389e+19,
                        2.462e+19,
                        2.537e+19,
                        2.614e+19,
                        2.692e+19,
                        2.773e+19,
                        2.855e+19,
                        2.939e+19,
                        3.026e+19,
                        3.114e+19,
                        3.204e+19,
                        3.297e+19,
                        3.391e+19,
                        3.488e+19],
 'Enemy Level Skip +': [0.0,
                        5000000000.0,
                        15500000000.0,
                        202980000000.0,
                        1130000000000.0,
                        3890000000000.0,
                        10140000000000.0,
                        22190000000000.0,
                        43050000000000.0,
                        76440000000000.0,
                        126850000000000.0,
                        199540000000000.0,
                        300610000000000.0,
                        437010000000000.0,
                        616540000000000.0,
                        847920000000000.0,
                        1140000000000000.0,
                        1510000000000000.0,
                        3910000000000000.0,
                        7500000000000000.0,
                        1.261e+16,
                        1.965e+16,
                        2.91e+16,
                        4.145e+16,
                        5.735e+16,
                        7.747e+16,
                        1.026e+17,
                        1.3359e+17,
                        1.7141e+17,
                        2.1713e+17,
                        2.7192e+17,
                        3.37e+17,
                        4.9677e+17,
                        7.0587e+17,
                        9.75e+17,
                        1.32e+18,
                        1.74e+18,
                        2.27e+18,
                        2.92e+18,
                        3.71e+18,
                        4.67e+18,
                        5.81e+18,
                        7.16e+18,
                        8.77e+18,
                        1.065e+19,
                        1.285e+19,
                        1.542e+19,
                        1.839e+19,
                        2.181e+19,
                        2.574e+19,
                        3.024e+19,
                        3.537e+19,
                        4.738e+19,
                        6.214e+19,
                        8.012e+19,
                        1.0184e+20,
                        1.2788e+20,
                        1.5891e+20,
                        1.9563e+20,
                        2.3888e+20,
                        2.8953e+20]}

NATIVE_ECON_LAB_MAX = {name: max(levels) if levels else 0 for name, levels in NATIVE_ECON_LAB_TABLES.items()}

# -----------------------------------------------------------------------------
# IMPORT HELPERS
# -----------------------------------------------------------------------------

def clean_cell(value: Any) -> str:
    return "" if value is None else str(value).strip()


def to_number(value: Any) -> Optional[float]:
    if value is None or isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace(",", "")
    if not text or text in {"-", "Locked", "Max"}:
        return None
    text = text.replace("%", "").replace("x", "")
    try:
        return float(text)
    except ValueError:
        return None


def normalize_export_rows(rows: list[list[Any]]) -> list[list[Any]]:
    rows = [list(row) for row in rows if any(clean_cell(v) for v in row)]
    if not rows:
        return []

    # Dataframe CSV exports often add a header row 0,1,2... and an index column.
    first = rows[0]
    if len(first) >= 4:
        numeric_headers = [clean_cell(v) for v in first[1:5]]
        if numeric_headers == ["0", "1", "2", "3"]:
            rows = rows[1:]

    if len(rows) >= 3:
        seq = True
        for idx, row in enumerate(rows[: min(8, len(rows))]):
            if not row or clean_cell(row[0]) != str(idx):
                seq = False
                break
        if seq:
            rows = [row[1:] for row in rows]

    width = max(len(row) for row in rows)
    return [row + [None] * (width - len(row)) for row in rows]


def find_cells(rows: list[list[Any]], needle: str) -> list[tuple[int, int]]:
    found = []
    for r, row in enumerate(rows[:12]):
        for c, value in enumerate(row):
            if clean_cell(value).casefold() == needle.casefold():
                found.append((r, c))
    return found


def parse_master_rows(raw_rows: list[list[Any]]) -> Dict[str, Any]:
    rows = normalize_export_rows(raw_rows)
    result: Dict[str, Any] = {
        "workshop": {}, "labs": {}, "enhancements": {}, "uw": {},
        "unmapped": {"workshop": [], "labs": [], "enhancements": [], "uw": []},
    }
    if not rows:
        raise ValueError("The uploaded file contains no readable rows.")

    lab_cells = find_cells(rows, "Go to my Laboratory Sheet")
    ws_cells = find_cells(rows, "Go to my Workshop Sheet")
    uw_cells = find_cells(rows, "Go to my Ultimate Weapon Sheet")
    if not lab_cells or len(ws_cells) < 2 or not uw_cells:
        raise ValueError(
            "This does not look like an Effective Paths Master Sheet export. "
            "Export the Master Sheet tab as CSV, or upload the filled XLSX workbook."
        )

    header_row = min(lab_cells[0][0], ws_cells[0][0], uw_cells[0][0])
    lab_col = lab_cells[0][1]
    workshop_col = ws_cells[0][1]
    enhancement_col = ws_cells[1][1]
    uw_header_col = uw_cells[0][1]

    for row in rows[header_row + 1:]:
        # Labs: name, level, target, max
        lab_name = clean_cell(row[lab_col])
        lab_level = to_number(row[lab_col + 1])
        canonical_lab = LAB_ALIASES.get(lab_name, lab_name)
        if canonical_lab in LAB_MAX_LEVELS and lab_level is not None:
            result["labs"][canonical_lab] = int(round(lab_level))
        elif lab_name and lab_name not in {"LABS"} and lab_level is not None:
            result["unmapped"]["labs"].append(lab_name)

        # Workshop: first of the coin/dollar level columns that contains a number.
        workshop_name = clean_cell(row[workshop_col])
        workshop_level = to_number(row[workshop_col + 1])
        if workshop_level is None:
            workshop_level = to_number(row[workshop_col + 2])
        canonical_ws = WORKSHOP_ALIASES.get(workshop_name, workshop_name)
        if canonical_ws in WORKSHOP_MAX_LEVELS and workshop_level is not None:
            result["workshop"][canonical_ws] = int(round(workshop_level))
        elif workshop_name and workshop_name not in {"WORKSHOP"} and workshop_level is not None:
            result["unmapped"]["workshop"].append(workshop_name)

        # Enhancements: name, unlock flag, level, max
        enhancement_name = clean_cell(row[enhancement_col])
        enhancement_level = to_number(row[enhancement_col + 2])
        if enhancement_name in ENHANCEMENT_MAX_LEVELS and enhancement_level is not None:
            result["enhancements"][enhancement_name] = int(round(enhancement_level))
        elif enhancement_name and enhancement_name not in {"WORKSHOP ENHANCEMENTS"} and enhancement_level is not None:
            result["unmapped"]["enhancements"].append(enhancement_name)

    # UW section uses one column after the hyperlink header for the UW/status,
    # then Attribute and Value in the following two columns.
    uw_name_col = uw_header_col + 1
    uw_attr_col = uw_header_col + 2
    uw_value_col = uw_header_col + 3
    current_uw: Optional[str] = None
    for row in rows[header_row + 1:]:
        name_or_status = clean_cell(row[uw_name_col])
        attribute = clean_cell(row[uw_attr_col])
        value = row[uw_value_col]

        if name_or_status in UW_NAMES:
            current_uw = name_or_status
            result["uw"].setdefault(current_uw, {"owned": None, "attributes": {}})
        elif name_or_status == "UW Unlocked" and current_uw:
            result["uw"].setdefault(current_uw, {"owned": None, "attributes": {}})["owned"] = True
        elif name_or_status == "UW Locked" and current_uw:
            result["uw"].setdefault(current_uw, {"owned": None, "attributes": {}})["owned"] = False

        if current_uw and attribute in UW_ATTRIBUTE_META[current_uw]:
            number = to_number(value)
            if number is not None:
                meta = UW_ATTRIBUTE_META[current_uw][attribute]
                if isinstance(meta["max"], int) and not isinstance(meta["max"], bool):
                    number = int(round(number))
                result["uw"].setdefault(current_uw, {"owned": None, "attributes": {}})["attributes"][attribute] = number
        elif current_uw and attribute and attribute not in {"", "UW+"} and to_number(value) is not None:
            result["unmapped"]["uw"].append(f"{current_uw}: {attribute}")

    for section in result["unmapped"]:
        result["unmapped"][section] = sorted(set(result["unmapped"][section]))
    return result


def parse_uploaded_effective_paths(uploaded_file: Any) -> Dict[str, Any]:
    suffix = Path(uploaded_file.name).suffix.lower()
    payload = uploaded_file.getvalue()
    if suffix == ".csv":
        text = payload.decode("utf-8-sig")
        rows = list(csv.reader(io.StringIO(text)))
        return parse_master_rows(rows)
    if suffix == ".xlsx":
        workbook = load_workbook(io.BytesIO(payload), read_only=True, data_only=True)
        if "Master Sheet" not in workbook.sheetnames:
            raise ValueError("The workbook does not contain a 'Master Sheet' tab.")
        sheet = workbook["Master Sheet"]
        rows = [list(row) for row in sheet.iter_rows(values_only=True)]
        return parse_master_rows(rows)
    raise ValueError("Supported file types are .csv and .xlsx.")





# -----------------------------------------------------------------------------
# NATIVE DAMAGE ENGINE DATA
# -----------------------------------------------------------------------------

NATIVE_DAMAGE_LAB_TABLES = {'Damage': {1: {'seconds': 14.0, 'cost': 30.0},
            2: {'seconds': 384.0, 'cost': 71.0},
            3: {'seconds': 984.0, 'cost': 178.0},
            4: {'seconds': 1892.0, 'cost': 398.0},
            5: {'seconds': 3165.0, 'cost': 772.0},
            6: {'seconds': 4800.0, 'cost': 1340.0},
            7: {'seconds': 6960.0, 'cost': 2120.0},
            8: {'seconds': 9540.0, 'cost': 3170.0},
            9: {'seconds': 12660.0, 'cost': 4510.0},
            10: {'seconds': 16320.0, 'cost': 6170.0},
            11: {'seconds': 20580.0, 'cost': 8170.0},
            12: {'seconds': 25380.0, 'cost': 10560.0},
            13: {'seconds': 30840.0, 'cost': 13350.0},
            14: {'seconds': 36900.0, 'cost': 16580.0},
            15: {'seconds': 43620.0, 'cost': 20270.0},
            16: {'seconds': 51000.0, 'cost': 24440.0},
            17: {'seconds': 59100.0, 'cost': 29130.0},
            18: {'seconds': 67920.0, 'cost': 34360.0},
            19: {'seconds': 77460.0, 'cost': 40160.0},
            20: {'seconds': 87720.0, 'cost': 46540.0},
            21: {'seconds': 98760.0, 'cost': 53530.0},
            22: {'seconds': 110640.0, 'cost': 61160.0},
            23: {'seconds': 123240.0, 'cost': 69460.0},
            24: {'seconds': 136680.0, 'cost': 78430.0},
            25: {'seconds': 150960.0, 'cost': 88120.0},
            26: {'seconds': 166020.0, 'cost': 98530.0},
            27: {'seconds': 181980.0, 'cost': 109700.0},
            28: {'seconds': 198780.0, 'cost': 121650.0},
            29: {'seconds': 216480.0, 'cost': 134390.0},
            30: {'seconds': 235080.0, 'cost': 147950.0},
            31: {'seconds': 254580.0, 'cost': 162350.0},
            32: {'seconds': 274980.0, 'cost': 177620.0},
            33: {'seconds': 296280.0, 'cost': 193780.0},
            34: {'seconds': 318600.0, 'cost': 210830.0},
            35: {'seconds': 341820.0, 'cost': 228820.0},
            36: {'seconds': 366000.0, 'cost': 247760.0},
            37: {'seconds': 391200.0, 'cost': 267660.0},
            38: {'seconds': 417300.0, 'cost': 288560.0},
            39: {'seconds': 444480.0, 'cost': 310470.0},
            40: {'seconds': 472620.0, 'cost': 333400.0},
            41: {'seconds': 501840.0, 'cost': 357390.0},
            42: {'seconds': 532020.0, 'cost': 382450.0},
            43: {'seconds': 563280.0, 'cost': 408600.0},
            44: {'seconds': 595560.0, 'cost': 435870.0},
            45: {'seconds': 628920.0, 'cost': 464260.0},
            46: {'seconds': 663300.0, 'cost': 493810.0},
            47: {'seconds': 698820.0, 'cost': 524530.0},
            48: {'seconds': 735420.0, 'cost': 556430.0},
            49: {'seconds': 773100.0, 'cost': 589550.0},
            50: {'seconds': 811860.0, 'cost': 623890.0},
            51: {'seconds': 851820.0, 'cost': 659490.0},
            52: {'seconds': 892800.0, 'cost': 696340.0},
            53: {'seconds': 934980.0, 'cost': 734490.0},
            54: {'seconds': 978300.0, 'cost': 773940.0},
            55: {'seconds': 1022760.0, 'cost': 814710.0},
            56: {'seconds': 1068360.0, 'cost': 856830.0},
            57: {'seconds': 1115160.0, 'cost': 900300.0},
            58: {'seconds': 1163100.0, 'cost': 945160.0},
            59: {'seconds': 1212240.0, 'cost': 991410.0},
            60: {'seconds': 1262580.0, 'cost': 1040000.0},
            61: {'seconds': 1314180.0, 'cost': 1090000.0},
            62: {'seconds': 1366920.0, 'cost': 1140000.0},
            63: {'seconds': 1420860.0, 'cost': 1190000.0},
            64: {'seconds': 1476060.0, 'cost': 1240000.0},
            65: {'seconds': 1532520.0, 'cost': 1300000.0},
            66: {'seconds': 1590180.0, 'cost': 1360000.0},
            67: {'seconds': 1649100.0, 'cost': 1410000.0},
            68: {'seconds': 1709280.0, 'cost': 1470000.0},
            69: {'seconds': 1770720.0, 'cost': 1530000.0},
            70: {'seconds': 1833420.0, 'cost': 1600000.0},
            71: {'seconds': 1897440.0, 'cost': 1660000.0},
            72: {'seconds': 1962720.0, 'cost': 1730000.0},
            73: {'seconds': 2029260.0, 'cost': 1800000.0},
            74: {'seconds': 2097120.0, 'cost': 1870000.0},
            75: {'seconds': 2166360.0, 'cost': 1940000.0},
            76: {'seconds': 2236800.0, 'cost': 2010000.0},
            77: {'seconds': 2308680.0, 'cost': 2080000.0},
            78: {'seconds': 2381820.0, 'cost': 2160000.0},
            79: {'seconds': 2456280.0, 'cost': 2240000.0},
            80: {'seconds': 2532120.0, 'cost': 2320000.0},
            81: {'seconds': 2609280.0, 'cost': 2400000.0},
            82: {'seconds': 2687820.0, 'cost': 2480000.0},
            83: {'seconds': 2767740.0, 'cost': 2570000.0},
            84: {'seconds': 2849040.0, 'cost': 2650000.0},
            85: {'seconds': 2931660.0, 'cost': 2740000.0},
            86: {'seconds': 3015720.0, 'cost': 2830000.0},
            87: {'seconds': 3101160.0, 'cost': 2930000.0},
            88: {'seconds': 3187980.0, 'cost': 3020000.0},
            89: {'seconds': 3276180.0, 'cost': 3120000.0},
            90: {'seconds': 3365820.0, 'cost': 3220000.0},
            91: {'seconds': 3456900.0, 'cost': 3320000.0},
            92: {'seconds': 3549360.0, 'cost': 3420000.0},
            93: {'seconds': 3643260.0, 'cost': 3520000.0},
            94: {'seconds': 3738600.0, 'cost': 3630000.0},
            95: {'seconds': 3835380.0, 'cost': 3740000.0},
            96: {'seconds': 3933600.0, 'cost': 3850000.0},
            97: {'seconds': 4033320.0, 'cost': 3960000.0},
            98: {'seconds': 4134420.0, 'cost': 4070000.0},
            99: {'seconds': 4237020.0, 'cost': 4190000.0},
            100: {'seconds': 4341120.0, 'cost': 4310000.0}},
 'Attack Speed': {1: {'seconds': 14.0, 'cost': 30.0},
                  2: {'seconds': 384.0, 'cost': 71.0},
                  3: {'seconds': 984.0, 'cost': 178.0},
                  4: {'seconds': 1892.0, 'cost': 398.0},
                  5: {'seconds': 3165.0, 'cost': 772.0},
                  6: {'seconds': 4800.0, 'cost': 1340.0},
                  7: {'seconds': 6960.0, 'cost': 2120.0},
                  8: {'seconds': 9540.0, 'cost': 3170.0},
                  9: {'seconds': 12660.0, 'cost': 4510.0},
                  10: {'seconds': 16320.0, 'cost': 6170.0},
                  11: {'seconds': 20580.0, 'cost': 8170.0},
                  12: {'seconds': 25380.0, 'cost': 10560.0},
                  13: {'seconds': 30840.0, 'cost': 13350.0},
                  14: {'seconds': 36900.0, 'cost': 16580.0},
                  15: {'seconds': 43620.0, 'cost': 20270.0},
                  16: {'seconds': 51000.0, 'cost': 24440.0},
                  17: {'seconds': 59100.0, 'cost': 29130.0},
                  18: {'seconds': 67920.0, 'cost': 34360.0},
                  19: {'seconds': 77460.0, 'cost': 40160.0},
                  20: {'seconds': 87720.0, 'cost': 46540.0},
                  21: {'seconds': 98760.0, 'cost': 53530.0},
                  22: {'seconds': 110640.0, 'cost': 61160.0},
                  23: {'seconds': 123240.0, 'cost': 69460.0},
                  24: {'seconds': 136680.0, 'cost': 78430.0},
                  25: {'seconds': 150960.0, 'cost': 88120.0},
                  26: {'seconds': 166020.0, 'cost': 98530.0},
                  27: {'seconds': 181980.0, 'cost': 109700.0},
                  28: {'seconds': 198780.0, 'cost': 121650.0},
                  29: {'seconds': 216480.0, 'cost': 134390.0},
                  30: {'seconds': 235080.0, 'cost': 147950.0},
                  31: {'seconds': 254580.0, 'cost': 162350.0},
                  32: {'seconds': 274980.0, 'cost': 177620.0},
                  33: {'seconds': 296280.0, 'cost': 193780.0},
                  34: {'seconds': 318600.0, 'cost': 210830.0},
                  35: {'seconds': 341820.0, 'cost': 228820.0},
                  36: {'seconds': 366000.0, 'cost': 247760.0},
                  37: {'seconds': 391200.0, 'cost': 267660.0},
                  38: {'seconds': 417300.0, 'cost': 288560.0},
                  39: {'seconds': 444480.0, 'cost': 310470.0},
                  40: {'seconds': 472620.0, 'cost': 333400.0},
                  41: {'seconds': 501840.0, 'cost': 357390.0},
                  42: {'seconds': 532020.0, 'cost': 382450.0},
                  43: {'seconds': 563280.0, 'cost': 408600.0},
                  44: {'seconds': 595560.0, 'cost': 435870.0},
                  45: {'seconds': 628920.0, 'cost': 464260.0},
                  46: {'seconds': 663300.0, 'cost': 493810.0},
                  47: {'seconds': 698820.0, 'cost': 524530.0},
                  48: {'seconds': 735420.0, 'cost': 556430.0},
                  49: {'seconds': 773100.0, 'cost': 589550.0},
                  50: {'seconds': 811860.0, 'cost': 623890.0},
                  51: {'seconds': 851820.0, 'cost': 659490.0},
                  52: {'seconds': 892800.0, 'cost': 696340.0},
                  53: {'seconds': 934980.0, 'cost': 734490.0},
                  54: {'seconds': 978300.0, 'cost': 773940.0},
                  55: {'seconds': 1022760.0, 'cost': 814710.0},
                  56: {'seconds': 1068360.0, 'cost': 856830.0},
                  57: {'seconds': 1115160.0, 'cost': 900300.0},
                  58: {'seconds': 1163100.0, 'cost': 945160.0},
                  59: {'seconds': 1212240.0, 'cost': 991410.0},
                  60: {'seconds': 1262580.0, 'cost': 1040000.0},
                  61: {'seconds': 1314180.0, 'cost': 1090000.0},
                  62: {'seconds': 1366920.0, 'cost': 1140000.0},
                  63: {'seconds': 1420860.0, 'cost': 1190000.0},
                  64: {'seconds': 1476060.0, 'cost': 1240000.0},
                  65: {'seconds': 1532520.0, 'cost': 1300000.0},
                  66: {'seconds': 1590180.0, 'cost': 1360000.0},
                  67: {'seconds': 1649100.0, 'cost': 1410000.0},
                  68: {'seconds': 1709280.0, 'cost': 1470000.0},
                  69: {'seconds': 1770720.0, 'cost': 1530000.0},
                  70: {'seconds': 1833420.0, 'cost': 1600000.0},
                  71: {'seconds': 1897440.0, 'cost': 1660000.0},
                  72: {'seconds': 1962720.0, 'cost': 1730000.0},
                  73: {'seconds': 2029260.0, 'cost': 1800000.0},
                  74: {'seconds': 2097120.0, 'cost': 1870000.0},
                  75: {'seconds': 2166360.0, 'cost': 1940000.0},
                  76: {'seconds': 2236800.0, 'cost': 2010000.0},
                  77: {'seconds': 2308680.0, 'cost': 2080000.0},
                  78: {'seconds': 2381820.0, 'cost': 2160000.0},
                  79: {'seconds': 2456280.0, 'cost': 2240000.0},
                  80: {'seconds': 2532120.0, 'cost': 2320000.0},
                  81: {'seconds': 2609280.0, 'cost': 2400000.0},
                  82: {'seconds': 2687820.0, 'cost': 2480000.0},
                  83: {'seconds': 2767740.0, 'cost': 2570000.0},
                  84: {'seconds': 2849040.0, 'cost': 2650000.0},
                  85: {'seconds': 2931660.0, 'cost': 2740000.0},
                  86: {'seconds': 3015720.0, 'cost': 2830000.0},
                  87: {'seconds': 3101160.0, 'cost': 2930000.0},
                  88: {'seconds': 3187980.0, 'cost': 3020000.0},
                  89: {'seconds': 3276180.0, 'cost': 3120000.0},
                  90: {'seconds': 3365820.0, 'cost': 3220000.0},
                  91: {'seconds': 3456900.0, 'cost': 3320000.0},
                  92: {'seconds': 3549360.0, 'cost': 3420000.0},
                  93: {'seconds': 3643260.0, 'cost': 3520000.0},
                  94: {'seconds': 3738600.0, 'cost': 3630000.0},
                  95: {'seconds': 3835380.0, 'cost': 3740000.0},
                  96: {'seconds': 3933600.0, 'cost': 3850000.0},
                  97: {'seconds': 4033320.0, 'cost': 3960000.0},
                  98: {'seconds': 4134420.0, 'cost': 4070000.0},
                  99: {'seconds': 4237020.0, 'cost': 4190000.0}},
 'Critical Factor': {1: {'seconds': 14.0, 'cost': 30.0},
                     2: {'seconds': 384.0, 'cost': 71.0},
                     3: {'seconds': 984.0, 'cost': 178.0},
                     4: {'seconds': 1892.0, 'cost': 398.0},
                     5: {'seconds': 3165.0, 'cost': 772.0},
                     6: {'seconds': 4800.0, 'cost': 1340.0},
                     7: {'seconds': 6960.0, 'cost': 2120.0},
                     8: {'seconds': 9540.0, 'cost': 3170.0},
                     9: {'seconds': 12660.0, 'cost': 4510.0},
                     10: {'seconds': 16320.0, 'cost': 6170.0},
                     11: {'seconds': 20580.0, 'cost': 8170.0},
                     12: {'seconds': 25380.0, 'cost': 10560.0},
                     13: {'seconds': 30840.0, 'cost': 13350.0},
                     14: {'seconds': 36900.0, 'cost': 16580.0},
                     15: {'seconds': 43620.0, 'cost': 20270.0},
                     16: {'seconds': 51000.0, 'cost': 24440.0},
                     17: {'seconds': 59100.0, 'cost': 29130.0},
                     18: {'seconds': 67920.0, 'cost': 34360.0},
                     19: {'seconds': 77460.0, 'cost': 40160.0},
                     20: {'seconds': 87720.0, 'cost': 46540.0},
                     21: {'seconds': 98760.0, 'cost': 53530.0},
                     22: {'seconds': 110640.0, 'cost': 61160.0},
                     23: {'seconds': 123240.0, 'cost': 69460.0},
                     24: {'seconds': 136680.0, 'cost': 78430.0},
                     25: {'seconds': 150960.0, 'cost': 88120.0},
                     26: {'seconds': 166020.0, 'cost': 98530.0},
                     27: {'seconds': 181980.0, 'cost': 109700.0},
                     28: {'seconds': 198780.0, 'cost': 121650.0},
                     29: {'seconds': 216480.0, 'cost': 134390.0},
                     30: {'seconds': 235080.0, 'cost': 147950.0},
                     31: {'seconds': 254580.0, 'cost': 162350.0},
                     32: {'seconds': 274980.0, 'cost': 177620.0},
                     33: {'seconds': 296280.0, 'cost': 193780.0},
                     34: {'seconds': 318600.0, 'cost': 210830.0},
                     35: {'seconds': 341820.0, 'cost': 228820.0},
                     36: {'seconds': 366000.0, 'cost': 247760.0},
                     37: {'seconds': 391200.0, 'cost': 267660.0},
                     38: {'seconds': 417300.0, 'cost': 288560.0},
                     39: {'seconds': 444480.0, 'cost': 310470.0},
                     40: {'seconds': 472620.0, 'cost': 333400.0},
                     41: {'seconds': 501840.0, 'cost': 357390.0},
                     42: {'seconds': 532020.0, 'cost': 382450.0},
                     43: {'seconds': 563280.0, 'cost': 408600.0},
                     44: {'seconds': 595560.0, 'cost': 435870.0},
                     45: {'seconds': 628920.0, 'cost': 464260.0},
                     46: {'seconds': 663300.0, 'cost': 493810.0},
                     47: {'seconds': 698820.0, 'cost': 524530.0},
                     48: {'seconds': 735420.0, 'cost': 556430.0},
                     49: {'seconds': 773100.0, 'cost': 589550.0},
                     50: {'seconds': 811860.0, 'cost': 623890.0},
                     51: {'seconds': 851820.0, 'cost': 659490.0},
                     52: {'seconds': 892800.0, 'cost': 696340.0},
                     53: {'seconds': 934980.0, 'cost': 734490.0},
                     54: {'seconds': 978300.0, 'cost': 773940.0},
                     55: {'seconds': 1022760.0, 'cost': 814710.0},
                     56: {'seconds': 1068360.0, 'cost': 856830.0},
                     57: {'seconds': 1115160.0, 'cost': 900300.0},
                     58: {'seconds': 1163100.0, 'cost': 945160.0},
                     59: {'seconds': 1212240.0, 'cost': 991410.0},
                     60: {'seconds': 1262580.0, 'cost': 1040000.0},
                     61: {'seconds': 1314180.0, 'cost': 1090000.0},
                     62: {'seconds': 1366920.0, 'cost': 1140000.0},
                     63: {'seconds': 1420860.0, 'cost': 1190000.0},
                     64: {'seconds': 1476060.0, 'cost': 1240000.0},
                     65: {'seconds': 1532520.0, 'cost': 1300000.0},
                     66: {'seconds': 1590180.0, 'cost': 1360000.0},
                     67: {'seconds': 1649100.0, 'cost': 1410000.0},
                     68: {'seconds': 1709280.0, 'cost': 1470000.0},
                     69: {'seconds': 1770720.0, 'cost': 1530000.0},
                     70: {'seconds': 1833420.0, 'cost': 1600000.0},
                     71: {'seconds': 1897440.0, 'cost': 1660000.0},
                     72: {'seconds': 1962720.0, 'cost': 1730000.0},
                     73: {'seconds': 2029260.0, 'cost': 1800000.0},
                     74: {'seconds': 2097120.0, 'cost': 1870000.0},
                     75: {'seconds': 2166360.0, 'cost': 1940000.0},
                     76: {'seconds': 2236800.0, 'cost': 2010000.0},
                     77: {'seconds': 2308680.0, 'cost': 2080000.0},
                     78: {'seconds': 2381820.0, 'cost': 2160000.0},
                     79: {'seconds': 2456280.0, 'cost': 2240000.0},
                     80: {'seconds': 2532120.0, 'cost': 2320000.0},
                     81: {'seconds': 2609280.0, 'cost': 2400000.0},
                     82: {'seconds': 2687820.0, 'cost': 2480000.0},
                     83: {'seconds': 2767740.0, 'cost': 2570000.0},
                     84: {'seconds': 2849040.0, 'cost': 2650000.0},
                     85: {'seconds': 2931660.0, 'cost': 2740000.0},
                     86: {'seconds': 3015720.0, 'cost': 2830000.0},
                     87: {'seconds': 3101160.0, 'cost': 2930000.0},
                     88: {'seconds': 3187980.0, 'cost': 3020000.0},
                     89: {'seconds': 3276180.0, 'cost': 3120000.0},
                     90: {'seconds': 3365820.0, 'cost': 3220000.0},
                     91: {'seconds': 3456900.0, 'cost': 3320000.0},
                     92: {'seconds': 3549360.0, 'cost': 3420000.0},
                     93: {'seconds': 3643260.0, 'cost': 3520000.0},
                     94: {'seconds': 3738600.0, 'cost': 3630000.0},
                     95: {'seconds': 3835380.0, 'cost': 3740000.0},
                     96: {'seconds': 3933600.0, 'cost': 3850000.0},
                     97: {'seconds': 4033320.0, 'cost': 3960000.0},
                     98: {'seconds': 4134420.0, 'cost': 4070000.0},
                     99: {'seconds': 4237020.0, 'cost': 4190000.0}},
 'Range': {1: {'seconds': 14.0, 'cost': 30.0},
           2: {'seconds': 384.0, 'cost': 71.0},
           3: {'seconds': 984.0, 'cost': 178.0},
           4: {'seconds': 1892.0, 'cost': 398.0},
           5: {'seconds': 3165.0, 'cost': 772.0},
           6: {'seconds': 4800.0, 'cost': 1340.0},
           7: {'seconds': 6960.0, 'cost': 2120.0},
           8: {'seconds': 9540.0, 'cost': 3170.0},
           9: {'seconds': 12660.0, 'cost': 4510.0},
           10: {'seconds': 16320.0, 'cost': 6170.0},
           11: {'seconds': 20580.0, 'cost': 8170.0},
           12: {'seconds': 25380.0, 'cost': 10560.0},
           13: {'seconds': 30840.0, 'cost': 13350.0},
           14: {'seconds': 36900.0, 'cost': 16580.0},
           15: {'seconds': 43620.0, 'cost': 20270.0},
           16: {'seconds': 51000.0, 'cost': 24440.0},
           17: {'seconds': 59100.0, 'cost': 29130.0},
           18: {'seconds': 67920.0, 'cost': 34360.0},
           19: {'seconds': 77460.0, 'cost': 40160.0},
           20: {'seconds': 87720.0, 'cost': 46540.0},
           21: {'seconds': 98760.0, 'cost': 53530.0},
           22: {'seconds': 110640.0, 'cost': 61160.0},
           23: {'seconds': 123240.0, 'cost': 69460.0},
           24: {'seconds': 136680.0, 'cost': 78430.0},
           25: {'seconds': 150960.0, 'cost': 88120.0},
           26: {'seconds': 166020.0, 'cost': 98530.0},
           27: {'seconds': 181980.0, 'cost': 109700.0},
           28: {'seconds': 198780.0, 'cost': 121650.0},
           29: {'seconds': 216480.0, 'cost': 134390.0},
           30: {'seconds': 235080.0, 'cost': 147950.0},
           31: {'seconds': 254580.0, 'cost': 162350.0},
           32: {'seconds': 274980.0, 'cost': 177620.0},
           33: {'seconds': 296280.0, 'cost': 193780.0},
           34: {'seconds': 318600.0, 'cost': 210830.0},
           35: {'seconds': 341820.0, 'cost': 228820.0},
           36: {'seconds': 366000.0, 'cost': 247760.0},
           37: {'seconds': 391200.0, 'cost': 267660.0},
           38: {'seconds': 417300.0, 'cost': 288560.0},
           39: {'seconds': 444480.0, 'cost': 310470.0},
           40: {'seconds': 472620.0, 'cost': 333400.0},
           41: {'seconds': 501840.0, 'cost': 357390.0},
           42: {'seconds': 532020.0, 'cost': 382450.0},
           43: {'seconds': 563280.0, 'cost': 408600.0},
           44: {'seconds': 595560.0, 'cost': 435870.0},
           45: {'seconds': 628920.0, 'cost': 464260.0},
           46: {'seconds': 663300.0, 'cost': 493810.0},
           47: {'seconds': 698820.0, 'cost': 524530.0},
           48: {'seconds': 735420.0, 'cost': 556430.0},
           49: {'seconds': 773100.0, 'cost': 589550.0},
           50: {'seconds': 811860.0, 'cost': 623890.0},
           51: {'seconds': 851820.0, 'cost': 659490.0},
           52: {'seconds': 892800.0, 'cost': 696340.0},
           53: {'seconds': 934980.0, 'cost': 734490.0},
           54: {'seconds': 978300.0, 'cost': 773940.0},
           55: {'seconds': 1022760.0, 'cost': 814710.0},
           56: {'seconds': 1068360.0, 'cost': 856830.0},
           57: {'seconds': 1115160.0, 'cost': 900300.0},
           58: {'seconds': 1163100.0, 'cost': 945160.0},
           59: {'seconds': 1212240.0, 'cost': 991410.0},
           60: {'seconds': 1262580.0, 'cost': 1040000.0},
           61: {'seconds': 1314180.0, 'cost': 1090000.0},
           62: {'seconds': 1366920.0, 'cost': 1140000.0},
           63: {'seconds': 1420860.0, 'cost': 1190000.0},
           64: {'seconds': 1476060.0, 'cost': 1240000.0},
           65: {'seconds': 1532520.0, 'cost': 1300000.0},
           66: {'seconds': 1590180.0, 'cost': 1360000.0},
           67: {'seconds': 1649100.0, 'cost': 1410000.0},
           68: {'seconds': 1709280.0, 'cost': 1470000.0},
           69: {'seconds': 1770720.0, 'cost': 1530000.0},
           70: {'seconds': 1833420.0, 'cost': 1600000.0},
           71: {'seconds': 1897440.0, 'cost': 1660000.0},
           72: {'seconds': 1962720.0, 'cost': 1730000.0},
           73: {'seconds': 2029260.0, 'cost': 1800000.0},
           74: {'seconds': 2097120.0, 'cost': 1870000.0},
           75: {'seconds': 2166360.0, 'cost': 1940000.0},
           76: {'seconds': 2236800.0, 'cost': 2010000.0},
           77: {'seconds': 2308680.0, 'cost': 2080000.0},
           78: {'seconds': 2381820.0, 'cost': 2160000.0},
           79: {'seconds': 2456280.0, 'cost': 2240000.0},
           80: {'seconds': 2532120.0, 'cost': 2320000.0}},
 'Damage / Meter': {1: {'seconds': 14.0, 'cost': 30.0},
                    2: {'seconds': 384.0, 'cost': 71.0},
                    3: {'seconds': 984.0, 'cost': 178.0},
                    4: {'seconds': 1892.0, 'cost': 398.0},
                    5: {'seconds': 3170.0, 'cost': 782.0},
                    6: {'seconds': 4800.0, 'cost': 1350.0},
                    7: {'seconds': 6960.0, 'cost': 2130.0},
                    8: {'seconds': 9540.0, 'cost': 3180.0},
                    9: {'seconds': 12660.0, 'cost': 4520.0},
                    10: {'seconds': 16320.0, 'cost': 6180.0},
                    11: {'seconds': 20580.0, 'cost': 8180.0},
                    12: {'seconds': 25380.0, 'cost': 10570.0},
                    13: {'seconds': 30840.0, 'cost': 13360.0},
                    14: {'seconds': 36900.0, 'cost': 16590.0},
                    15: {'seconds': 43620.0, 'cost': 20280.0},
                    16: {'seconds': 51000.0, 'cost': 24450.0},
                    17: {'seconds': 59100.0, 'cost': 29140.0},
                    18: {'seconds': 67920.0, 'cost': 34370.0},
                    19: {'seconds': 77460.0, 'cost': 40170.0},
                    20: {'seconds': 87720.0, 'cost': 46550.0},
                    21: {'seconds': 98820.0, 'cost': 53540.0},
                    22: {'seconds': 110640.0, 'cost': 61170.0},
                    23: {'seconds': 123240.0, 'cost': 69470.0},
                    24: {'seconds': 136680.0, 'cost': 78440.0},
                    25: {'seconds': 150960.0, 'cost': 88130.0},
                    26: {'seconds': 166020.0, 'cost': 98540.0},
                    27: {'seconds': 181980.0, 'cost': 109710.0},
                    28: {'seconds': 198780.0, 'cost': 121660.0},
                    29: {'seconds': 216480.0, 'cost': 134400.0},
                    30: {'seconds': 235080.0, 'cost': 147960.0},
                    31: {'seconds': 254580.0, 'cost': 162360.0},
                    32: {'seconds': 274980.0, 'cost': 177630.0},
                    33: {'seconds': 296340.0, 'cost': 193790.0},
                    34: {'seconds': 318600.0, 'cost': 210840.0},
                    35: {'seconds': 341820.0, 'cost': 228830.0},
                    36: {'seconds': 366000.0, 'cost': 247770.0},
                    37: {'seconds': 391200.0, 'cost': 267670.0},
                    38: {'seconds': 417300.0, 'cost': 288570.0},
                    39: {'seconds': 444480.0, 'cost': 310480.0},
                    40: {'seconds': 472620.0, 'cost': 333410.0},
                    41: {'seconds': 501840.0, 'cost': 357400.0},
                    42: {'seconds': 532020.0, 'cost': 382460.0},
                    43: {'seconds': 563280.0, 'cost': 408610.0},
                    44: {'seconds': 595560.0, 'cost': 435880.0},
                    45: {'seconds': 628920.0, 'cost': 464270.0},
                    46: {'seconds': 663360.0, 'cost': 493820.0},
                    47: {'seconds': 698820.0, 'cost': 524540.0},
                    48: {'seconds': 735420.0, 'cost': 556440.0},
                    49: {'seconds': 773100.0, 'cost': 589560.0},
                    50: {'seconds': 811920.0, 'cost': 623890.0},
                    51: {'seconds': 851820.0, 'cost': 659500.0},
                    52: {'seconds': 892800.0, 'cost': 696350.0},
                    53: {'seconds': 934980.0, 'cost': 734500.0},
                    54: {'seconds': 978300.0, 'cost': 773950.0},
                    55: {'seconds': 1022760.0, 'cost': 814720.0},
                    56: {'seconds': 1068360.0, 'cost': 856840.0},
                    57: {'seconds': 1115160.0, 'cost': 900310.0},
                    58: {'seconds': 1163100.0, 'cost': 945160.0},
                    59: {'seconds': 1212300.0, 'cost': 991420.0},
                    60: {'seconds': 1262640.0, 'cost': 1040000.0},
                    61: {'seconds': 1314180.0, 'cost': 1090000.0},
                    62: {'seconds': 1366920.0, 'cost': 1140000.0},
                    63: {'seconds': 1420860.0, 'cost': 1190000.0},
                    64: {'seconds': 1476060.0, 'cost': 1240000.0},
                    65: {'seconds': 1532520.0, 'cost': 1300000.0},
                    66: {'seconds': 1590180.0, 'cost': 1360000.0},
                    67: {'seconds': 1649100.0, 'cost': 1410000.0},
                    68: {'seconds': 1709280.0, 'cost': 1470000.0},
                    69: {'seconds': 1770720.0, 'cost': 1530000.0},
                    70: {'seconds': 1833420.0, 'cost': 1600000.0},
                    71: {'seconds': 1897440.0, 'cost': 1660000.0},
                    72: {'seconds': 1962720.0, 'cost': 1730000.0},
                    73: {'seconds': 2029260.0, 'cost': 1800000.0},
                    74: {'seconds': 2097180.0, 'cost': 1870000.0},
                    75: {'seconds': 2166360.0, 'cost': 1940000.0},
                    76: {'seconds': 2236860.0, 'cost': 2010000.0},
                    77: {'seconds': 2308680.0, 'cost': 2080000.0},
                    78: {'seconds': 2381820.0, 'cost': 2160000.0},
                    79: {'seconds': 2456280.0, 'cost': 2240000.0},
                    80: {'seconds': 2532120.0, 'cost': 2320000.0},
                    81: {'seconds': 2609280.0, 'cost': 2400000.0},
                    82: {'seconds': 2687820.0, 'cost': 2480000.0},
                    83: {'seconds': 2767740.0, 'cost': 2570000.0},
                    84: {'seconds': 2849040.0, 'cost': 2650000.0},
                    85: {'seconds': 2931660.0, 'cost': 2740000.0},
                    86: {'seconds': 3015720.0, 'cost': 2830000.0},
                    87: {'seconds': 3101160.0, 'cost': 2930000.0},
                    88: {'seconds': 3187980.0, 'cost': 3020000.0},
                    89: {'seconds': 3276180.0, 'cost': 3120000.0},
                    90: {'seconds': 3365820.0, 'cost': 3220000.0},
                    91: {'seconds': 3456900.0, 'cost': 3320000.0},
                    92: {'seconds': 3549360.0, 'cost': 3420000.0},
                    93: {'seconds': 3643260.0, 'cost': 3520000.0},
                    94: {'seconds': 3738600.0, 'cost': 3630000.0},
                    95: {'seconds': 3835380.0, 'cost': 3740000.0},
                    96: {'seconds': 3933600.0, 'cost': 3850000.0},
                    97: {'seconds': 4033320.0, 'cost': 3960000.0},
                    98: {'seconds': 4134420.0, 'cost': 4070000.0},
                    99: {'seconds': 4237080.0, 'cost': 4190000.0}},
 'Super Crit Chance': {1: {'seconds': 99960.0, 'cost': 200000.0},
                       2: {'seconds': 150060.0, 'cost': 401000.0},
                       3: {'seconds': 201240.0, 'cost': 625990.0},
                       4: {'seconds': 257040.0, 'cost': 974770.0},
                       5: {'seconds': 323640.0, 'cost': 1680000.0},
                       6: {'seconds': 410340.0, 'cost': 3130000.0},
                       7: {'seconds': 529800.0, 'cost': 5940000.0},
                       8: {'seconds': 697980.0, 'cost': 10970000.0},
                       9: {'seconds': 934560.0, 'cost': 19360000.0},
                       10: {'seconds': 1262700.0, 'cost': 32540000.0},
                       11: {'seconds': 1709400.0, 'cost': 52320000.0},
                       12: {'seconds': 2305560.0, 'cost': 80840000.0},
                       13: {'seconds': 3085920.0, 'cost': 120670000.0},
                       14: {'seconds': 4089300.0, 'cost': 174800000.0},
                       15: {'seconds': 5358600.0, 'cost': 246670000.0},
                       16: {'seconds': 6940860.0, 'cost': 340200000.0},
                       17: {'seconds': 8887320.0, 'cost': 459820000.0},
                       18: {'seconds': 11253420.0, 'cost': 610490000.0},
                       19: {'seconds': 14099100.0, 'cost': 797730000.0},
                       20: {'seconds': 17488500.0, 'cost': 1030000000.0},
                       21: {'seconds': 21490260.0, 'cost': 1310000000.0},
                       22: {'seconds': 26177520.0, 'cost': 1640000000.0},
                       23: {'seconds': 31627920.0, 'cost': 2040000000.0},
                       24: {'seconds': 37923660.0, 'cost': 2520000000.0},
                       25: {'seconds': 45151500.0, 'cost': 3070000000.0},
                       26: {'seconds': 53403000.0, 'cost': 3720000000.0},
                       27: {'seconds': 62774160.0, 'cost': 4480000000.0},
                       28: {'seconds': 73365960.0, 'cost': 5340000000.0},
                       29: {'seconds': 85284060.0, 'cost': 6340000000.0},
                       30: {'seconds': 98638800.0, 'cost': 7480000000.0},
                       31: {'seconds': 113545560.0, 'cost': 8700000000.0},
                       32: {'seconds': 130124460.0, 'cost': 10230000000.0},
                       33: {'seconds': 148500540.0, 'cost': 11870000000.0},
                       34: {'seconds': 168803760.0, 'cost': 13720000000.0},
                       35: {'seconds': 191169120.0, 'cost': 15780000000.0},
                       36: {'seconds': 215736540.0, 'cost': 18080000000.0},
                       37: {'seconds': 242651040.0, 'cost': 20640000000.0},
                       38: {'seconds': 272062680.0, 'cost': 23480000000.0},
                       39: {'seconds': 304126440.0, 'cost': 26610000000.0},
                       40: {'seconds': 339002760.0, 'cost': 30070000000.0},
                       41: {'seconds': 376856940.0, 'cost': 33870000000.0},
                       42: {'seconds': 417859440.0, 'cost': 38030000000.0},
                       43: {'seconds': 462186120.0, 'cost': 42600000000.0},
                       44: {'seconds': 510017880.0, 'cost': 47580000000.0},
                       45: {'seconds': 561540960.0, 'cost': 53000000000.0},
                       46: {'seconds': 616946760.0, 'cost': 58910000000.0},
                       47: {'seconds': 676432080.0, 'cost': 65320000000.0},
                       48: {'seconds': 740199120.0, 'cost': 72260000000.0},
                       49: {'seconds': 808455120.0, 'cost': 79780000000.0},
                       50: {'seconds': 881413080.0, 'cost': 87900000000.0}},
 'Super Crit Multi': {1: {'seconds': 99960.0, 'cost': 200000.0},
                      2: {'seconds': 150060.0, 'cost': 401000.0},
                      3: {'seconds': 201060.0, 'cost': 625990.0},
                      4: {'seconds': 255660.0, 'cost': 974770.0},
                      5: {'seconds': 317880.0, 'cost': 1680000.0},
                      6: {'seconds': 393720.0, 'cost': 3130000.0},
                      7: {'seconds': 490680.0, 'cost': 5940000.0},
                      8: {'seconds': 618060.0, 'cost': 10970000.0},
                      9: {'seconds': 786660.0, 'cost': 19360000.0},
                      10: {'seconds': 1009260.0, 'cost': 32540000.0},
                      11: {'seconds': 1299960.0, 'cost': 52320000.0},
                      12: {'seconds': 1674840.0, 'cost': 80840000.0},
                      13: {'seconds': 2151480.0, 'cost': 120670000.0},
                      14: {'seconds': 2749260.0, 'cost': 174800000.0},
                      15: {'seconds': 3489060.0, 'cost': 246670000.0},
                      16: {'seconds': 4393740.0, 'cost': 340200000.0},
                      17: {'seconds': 5487480.0, 'cost': 459820000.0},
                      18: {'seconds': 6796440.0, 'cost': 610490000.0},
                      19: {'seconds': 8348280.0, 'cost': 797730000.0},
                      20: {'seconds': 10172460.0, 'cost': 1030000000.0},
                      21: {'seconds': 12299940.0, 'cost': 1310000000.0},
                      22: {'seconds': 14763660.0, 'cost': 1640000000.0},
                      23: {'seconds': 17597880.0, 'cost': 2040000000.0},
                      24: {'seconds': 20838840.0, 'cost': 2520000000.0},
                      25: {'seconds': 24524280.0, 'cost': 3070000000.0},
                      26: {'seconds': 28693740.0, 'cost': 3720000000.0},
                      27: {'seconds': 33388260.0, 'cost': 4480000000.0},
                      28: {'seconds': 38650860.0, 'cost': 5340000000.0},
                      29: {'seconds': 44525880.0, 'cost': 6340000000.0},
                      30: {'seconds': 51059640.0, 'cost': 7480000000.0},
                      31: {'seconds': 58299960.0, 'cost': 8770000000.0},
                      32: {'seconds': 66296460.0, 'cost': 10230000000.0},
                      33: {'seconds': 75100260.0, 'cost': 11870000000.0},
                      34: {'seconds': 84764460.0, 'cost': 13720000000.0},
                      35: {'seconds': 95343480.0, 'cost': 15780000000.0},
                      36: {'seconds': 106893720.0, 'cost': 18080000000.0},
                      37: {'seconds': 119473080.0, 'cost': 20640000000.0},
                      38: {'seconds': 133141260.0, 'cost': 23480000000.0},
                      39: {'seconds': 147959460.0, 'cost': 26610000000.0},
                      40: {'seconds': 163990860.0, 'cost': 30070000000.0}},
 'Max Rend Armor Multiplier': {1: {'seconds': 299940.0, 'cost': 200000000000.0},
                               2: {'seconds': 350040.0, 'cost': 240000000000.0},
                               3: {'seconds': 400320.0, 'cost': 280010000000.0},
                               4: {'seconds': 450780.0, 'cost': 320030000000.0},
                               5: {'seconds': 501420.0, 'cost': 360100000000.0},
                               6: {'seconds': 552360.0, 'cost': 400230000000.0},
                               7: {'seconds': 603600.0, 'cost': 440450000000.0},
                               8: {'seconds': 655020.0, 'cost': 480810000000.0},
                               9: {'seconds': 706740.0, 'cost': 521350000000.0},
                               10: {'seconds': 758760.0, 'cost': 562110000000.0},
                               11: {'seconds': 811080.0, 'cost': 603150000000.0},
                               12: {'seconds': 863640.0, 'cost': 644530000000.0},
                               13: {'seconds': 916560.0, 'cost': 686310000000.0},
                               14: {'seconds': 969720.0, 'cost': 728550000000.0},
                               15: {'seconds': 1023240.0, 'cost': 771330000000.0},
                               16: {'seconds': 1077060.0, 'cost': 814730000000.0},
                               17: {'seconds': 1131180.0, 'cost': 858820000000.0},
                               18: {'seconds': 1185600.0, 'cost': 903700000000.0},
                               19: {'seconds': 1240380.0, 'cost': 949440000000.0},
                               20: {'seconds': 1295520.0, 'cost': 996160000000.0},
                               21: {'seconds': 1350960.0, 'cost': 1040000000000.0},
                               22: {'seconds': 1406700.0, 'cost': 1090000000000.0},
                               23: {'seconds': 1462860.0, 'cost': 1140000000000.0},
                               24: {'seconds': 1519320.0, 'cost': 1190000000000.0},
                               25: {'seconds': 1576080.0, 'cost': 1250000000000.0},
                               26: {'seconds': 1633260.0, 'cost': 1300000000000.0},
                               27: {'seconds': 1690740.0, 'cost': 1360000000000.0},
                               28: {'seconds': 1748640.0, 'cost': 1420000000000.0},
                               29: {'seconds': 1806840.0, 'cost': 1480000000000.0},
                               30: {'seconds': 1865400.0, 'cost': 1540000000000.0}},
 'Shock Chance': {1: {'seconds': 72000.0, 'cost': 250000.0},
                  2: {'seconds': 102060.0, 'cost': 560000.0},
                  3: {'seconds': 132780.0, 'cost': 1100000.0},
                  4: {'seconds': 165240.0, 'cost': 2800000.0},
                  5: {'seconds': 200940.0, 'cost': 7750000.0},
                  6: {'seconds': 241560.0, 'cost': 19540000.0},
                  7: {'seconds': 289020.0, 'cost': 43580000.0},
                  8: {'seconds': 345480.0, 'cost': 87410000.0},
                  9: {'seconds': 413340.0, 'cost': 160910000.0},
                  10: {'seconds': 495060.0, 'cost': 276620000.0},
                  11: {'seconds': 593340.0, 'cost': 449930000.0},
                  12: {'seconds': 711000.0, 'cost': 699340000.0},
                  13: {'seconds': 850980.0, 'cost': 1050000000.0},
                  14: {'seconds': 1016460.0, 'cost': 1520000000.0},
                  15: {'seconds': 1210680.0, 'cost': 2140000000.0},
                  16: {'seconds': 1436940.0, 'cost': 2950000000.0},
                  17: {'seconds': 1698840.0, 'cost': 3980000000.0},
                  18: {'seconds': 1999920.0, 'cost': 5270000000.0},
                  19: {'seconds': 2343960.0, 'cost': 6880000000.0},
                  20: {'seconds': 2734800.0, 'cost': 8840000000.0},
                  21: {'seconds': 3176340.0, 'cost': 11220000000.0},
                  22: {'seconds': 3672720.0, 'cost': 14080000000.0},
                  23: {'seconds': 4228020.0, 'cost': 17480000000.0},
                  24: {'seconds': 4846560.0, 'cost': 21490000000.0},
                  25: {'seconds': 5532600.0, 'cost': 26190000000.0},
                  26: {'seconds': 6290700.0, 'cost': 31660000000.0},
                  27: {'seconds': 7125420.0, 'cost': 37990000000.0},
                  28: {'seconds': 8041260.0, 'cost': 45280000000.0},
                  29: {'seconds': 9043080.0, 'cost': 53620000000.0},
                  30: {'seconds': 10135680.0, 'cost': 63130000000.0}},
 'Shock Multiplier': {1: {'seconds': 72000.0, 'cost': 250000.0},
                      2: {'seconds': 102060.0, 'cost': 560000.0},
                      3: {'seconds': 132780.0, 'cost': 1100000.0},
                      4: {'seconds': 165240.0, 'cost': 2800000.0},
                      5: {'seconds': 200940.0, 'cost': 7750000.0},
                      6: {'seconds': 241560.0, 'cost': 19540000.0},
                      7: {'seconds': 289020.0, 'cost': 43580000.0},
                      8: {'seconds': 345480.0, 'cost': 87410000.0},
                      9: {'seconds': 413340.0, 'cost': 160910000.0},
                      10: {'seconds': 495060.0, 'cost': 276620000.0},
                      11: {'seconds': 593340.0, 'cost': 449930000.0},
                      12: {'seconds': 711000.0, 'cost': 699340000.0},
                      13: {'seconds': 850980.0, 'cost': 1050000000.0},
                      14: {'seconds': 1016460.0, 'cost': 1520000000.0}},
 'Super Tower Bonus': {1: {'seconds': 17940.0, 'cost': 2000000000.0},
                       2: {'seconds': 48060.0, 'cost': 3010000000.0},
                       3: {'seconds': 78780.0, 'cost': 4070000000.0},
                       4: {'seconds': 111240.0, 'cost': 5330000000.0},
                       5: {'seconds': 146940.0, 'cost': 6970000000.0},
                       6: {'seconds': 187560.0, 'cost': 9260000000.0},
                       7: {'seconds': 235020.0, 'cost': 8230000000.0},
                       8: {'seconds': 291480.0, 'cost': 12530000000.0},
                       9: {'seconds': 359340.0, 'cost': 17130000000.0},
                       10: {'seconds': 441060.0, 'cost': 32140000000.0},
                       11: {'seconds': 539340.0, 'cost': 43550000000.0},
                       12: {'seconds': 657000.0, 'cost': 58320000000.0},
                       13: {'seconds': 796980.0, 'cost': 77080000000.0},
                       14: {'seconds': 962460.0, 'cost': 100500000000.0},
                       15: {'seconds': 1156680.0, 'cost': 129310000000.0},
                       16: {'seconds': 1382940.0, 'cost': 164270000000.0},
                       17: {'seconds': 1644840.0, 'cost': 206200000000.0},
                       18: {'seconds': 1945920.0, 'cost': 255960000000.0},
                       19: {'seconds': 2289960.0, 'cost': 314450000000.0},
                       20: {'seconds': 2680800.0, 'cost': 382600000000.0},
                       21: {'seconds': 3122340.0, 'cost': 461420000000.0},
                       22: {'seconds': 3618720.0, 'cost': 551940000000.0},
                       23: {'seconds': 4174020.0, 'cost': 655210000000.0},
                       24: {'seconds': 4792560.0, 'cost': 750900000000.0},
                       25: {'seconds': 5478600.0, 'cost': 772370000000.0},
                       26: {'seconds': 6236700.0, 'cost': 904560000000.0},
                       27: {'seconds': 7071420.0, 'cost': 1220000000000.0},
                       28: {'seconds': 7987260.0, 'cost': 1400000000000.0},
                       29: {'seconds': 8989080.0, 'cost': 1610000000000.0},
                       30: {'seconds': 10081680.0, 'cost': 1830000000000.0}},
 'Missile Amplifier': {1: {'seconds': 71940.0, 'cost': 500000.0},
                       2: {'seconds': 102060.0, 'cost': 860000.0},
                       3: {'seconds': 132780.0, 'cost': 1450000.0},
                       4: {'seconds': 165240.0, 'cost': 3200000.0},
                       5: {'seconds': 200940.0, 'cost': 8200000.0},
                       6: {'seconds': 241560.0, 'cost': 20040000.0},
                       7: {'seconds': 289020.0, 'cost': 44130000.0},
                       8: {'seconds': 345480.0, 'cost': 88010000.0},
                       9: {'seconds': 413340.0, 'cost': 161560000.0},
                       10: {'seconds': 495060.0, 'cost': 277320000.0},
                       11: {'seconds': 593340.0, 'cost': 450680000.0},
                       12: {'seconds': 711000.0, 'cost': 700140000.0},
                       13: {'seconds': 850980.0, 'cost': 1050000000.0},
                       14: {'seconds': 1016460.0, 'cost': 1520000000.0},
                       15: {'seconds': 1210680.0, 'cost': 2140000000.0},
                       16: {'seconds': 1436940.0, 'cost': 2950000000.0},
                       17: {'seconds': 1698840.0, 'cost': 3980000000.0},
                       18: {'seconds': 1999920.0, 'cost': 5270000000.0},
                       19: {'seconds': 2343960.0, 'cost': 6880000000.0},
                       20: {'seconds': 2734800.0, 'cost': 8840000000.0},
                       21: {'seconds': 3176340.0, 'cost': 11220000000.0},
                       22: {'seconds': 3672720.0, 'cost': 14080000000.0},
                       23: {'seconds': 4228020.0, 'cost': 17480000000.0},
                       24: {'seconds': 4846560.0, 'cost': 21490000000.0},
                       25: {'seconds': 5532600.0, 'cost': 26190000000.0}},
 'Missile Radius': {1: {'seconds': 139980.0, 'cost': 800000.0},
                    2: {'seconds': 152040.0, 'cost': 835000.0},
                    3: {'seconds': 164760.0, 'cost': 1050000.0},
                    4: {'seconds': 179220.0, 'cost': 2000000.0},
                    5: {'seconds': 196920.0, 'cost': 4780000.0},
                    6: {'seconds': 219540.0, 'cost': 11050000.0},
                    7: {'seconds': 249000.0, 'cost': 23130000.0},
                    8: {'seconds': 287520.0, 'cost': 44020000.0},
                    9: {'seconds': 337320.0, 'cost': 77430000.0},
                    10: {'seconds': 401040.0, 'cost': 127860000.0},
                    11: {'seconds': 481320.0, 'cost': 200580000.0},
                    12: {'seconds': 580980.0, 'cost': 301670000.0},
                    13: {'seconds': 702960.0, 'cost': 438100000.0},
                    14: {'seconds': 850440.0, 'cost': 617660000.0},
                    15: {'seconds': 1026660.0, 'cost': 849060000.0},
                    16: {'seconds': 1234980.0, 'cost': 1140000000.0},
                    17: {'seconds': 1478880.0, 'cost': 1510000000.0},
                    18: {'seconds': 1761960.0, 'cost': 1960000000.0},
                    19: {'seconds': 2088000.0, 'cost': 2500000000.0},
                    20: {'seconds': 2460780.0, 'cost': 3150000000.0}},
 'Death Wave Damage Amplifier': {1: {'seconds': 720000.0, 'cost': 100000000000.0},
                                 2: {'seconds': 806400.0, 'cost': 150000000000.0},
                                 3: {'seconds': 903180.0, 'cost': 225000000000.0},
                                 4: {'seconds': 1011548.0, 'cost': 337500000000.0},
                                 5: {'seconds': 1132933.0, 'cost': 506250000000.0},
                                 6: {'seconds': 1268887.0, 'cost': 759375000000.0},
                                 7: {'seconds': 1421152.0, 'cost': 1140000000000.0},
                                 8: {'seconds': 1591690.0, 'cost': 1710000000000.0},
                                 9: {'seconds': 1782693.0, 'cost': 2560000000000.0},
                                 10: {'seconds': 1996616.0, 'cost': 3840000000000.0},
                                 11: {'seconds': 2236210.0, 'cost': 5770000000000.0},
                                 12: {'seconds': 2504555.0, 'cost': 8650000000000.0},
                                 13: {'seconds': 2805102.0, 'cost': 12975000000000.0},
                                 14: {'seconds': 3141714.0, 'cost': 19462500000000.0},
                                 15: {'seconds': 3518720.0, 'cost': 29187500000000.0},
                                 16: {'seconds': 3940967.0, 'cost': 43792500000000.0},
                                 17: {'seconds': 4413883.0, 'cost': 65684000000000.0},
                                 18: {'seconds': 4943549.0, 'cost': 98526000000000.0},
                                 19: {'seconds': 5540375.0, 'cost': 147790000000000.0},
                                 20: {'seconds': 6201188.0, 'cost': 221680000000000.0},
                                 21: {'seconds': 6945331.0, 'cost': 332530000000000.0},
                                 22: {'seconds': 7778771.0, 'cost': 498790000000000.0},
                                 23: {'seconds': 8712223.0, 'cost': 748180000000000.0},
                                 24: {'seconds': 9757690.0, 'cost': 1120000000000000.0},
                                 25: {'seconds': 10928613.0, 'cost': 1680000000000000.0},
                                 26: {'seconds': 12240047.0, 'cost': 2530000000000000.0},
                                 27: {'seconds': 13708852.0, 'cost': 3790000000000000.0},
                                 28: {'seconds': 15382714.0, 'cost': 5680000000000000.0},
                                 29: {'seconds': 17196385.0, 'cost': 8520000000000000.0},
                                 30: {'seconds': 19259951.0, 'cost': 1.278e+16}},
 'Black Hole Damage': {1: {'seconds': 143940.0, 'cost': 20000000.0},
                       2: {'seconds': 174060.0, 'cost': 20810000.0},
                       3: {'seconds': 204780.0, 'cost': 21850000.0},
                       4: {'seconds': 237240.0, 'cost': 24050000.0},
                       5: {'seconds': 272940.0, 'cost': 29500000.0},
                       6: {'seconds': 313560.0, 'cost': 41790000.0},
                       7: {'seconds': 361020.0, 'cost': 66330000.0},
                       8: {'seconds': 417480.0, 'cost': 110660000.0},
                       9: {'seconds': 485340.0, 'cost': 184660000.0},
                       10: {'seconds': 567060.0, 'cost': 300870000.0}}}

NATIVE_DAMAGE_UW_TABLES = {'CL | Damage': [{'value': 2.0, 'cost': 0.0},
                 {'value': 3.0, 'cost': 5.0},
                 {'value': 5.0, 'cost': 11.0},
                 {'value': 9.0, 'cost': 17.0},
                 {'value': 14.0, 'cost': 23.0},
                 {'value': 22.0, 'cost': 29.0},
                 {'value': 32.0, 'cost': 35.0},
                 {'value': 46.0, 'cost': 41.0},
                 {'value': 63.0, 'cost': 47.0},
                 {'value': 85.0, 'cost': 53.0},
                 {'value': 113.0, 'cost': 61.0},
                 {'value': 148.0, 'cost': 71.0},
                 {'value': 191.0, 'cost': 84.0},
                 {'value': 244.0, 'cost': 100.0},
                 {'value': 309.0, 'cost': 120.0},
                 {'value': 387.0, 'cost': 144.0},
                 {'value': 482.0, 'cost': 174.0},
                 {'value': 596.0, 'cost': 210.0},
                 {'value': 733.0, 'cost': 252.0},
                 {'value': 898.0, 'cost': 302.0},
                 {'value': 1094.0, 'cost': 362.0},
                 {'value': 1328.0, 'cost': 434.0},
                 {'value': 1607.0, 'cost': 525.0},
                 {'value': 1937.0, 'cost': 636.0},
                 {'value': 2329.0, 'cost': 767.0},
                 {'value': 2794.0, 'cost': 923.0},
                 {'value': 3342.0, 'cost': 1109.0},
                 {'value': 3990.0, 'cost': 1295.0},
                 {'value': 4755.0, 'cost': 1521.0},
                 {'value': 5655.0, 'cost': 1787.0},
                 {'value': 6715.0, 'cost': 2103.0},
                 {'value': 7961.0, 'cost': 2469.0}],
 'CL | Quantity': [{'value': 1.0, 'cost': 0.0},
                   {'value': 2.0, 'cost': 30.0},
                   {'value': 3.0, 'cost': 75.0},
                   {'value': 4.0, 'cost': 150.0},
                   {'value': 5.0, 'cost': 400.0}],
 'CL | Chance': [{'value': 0.05, 'cost': 0.0},
                 {'value': 0.065, 'cost': 8.0},
                 {'value': 0.08, 'cost': 26.0},
                 {'value': 0.095, 'cost': 44.0},
                 {'value': 0.11, 'cost': 62.0},
                 {'value': 0.125, 'cost': 80.0},
                 {'value': 0.14, 'cost': 98.0},
                 {'value': 0.155, 'cost': 116.0},
                 {'value': 0.17, 'cost': 134.0},
                 {'value': 0.185, 'cost': 152.0},
                 {'value': 0.2, 'cost': 170.0},
                 {'value': 0.215, 'cost': 188.0},
                 {'value': 0.23, 'cost': 206.0},
                 {'value': 0.245, 'cost': 224.0},
                 {'value': 0.26, 'cost': 242.0},
                 {'value': 0.275, 'cost': 260.0}],
 'SM | Damage': [{'value': 10.0, 'cost': 0.0},
                 {'value': 11.0, 'cost': 5.0},
                 {'value': 13.0, 'cost': 11.0},
                 {'value': 16.0, 'cost': 17.0},
                 {'value': 20.0, 'cost': 23.0},
                 {'value': 26.0, 'cost': 29.0},
                 {'value': 34.0, 'cost': 35.0},
                 {'value': 43.0, 'cost': 41.0},
                 {'value': 55.0, 'cost': 47.0},
                 {'value': 69.0, 'cost': 53.0},
                 {'value': 87.0, 'cost': 61.0},
                 {'value': 108.0, 'cost': 71.0},
                 {'value': 134.0, 'cost': 84.0},
                 {'value': 164.0, 'cost': 100.0},
                 {'value': 200.0, 'cost': 120.0},
                 {'value': 243.0, 'cost': 144.0},
                 {'value': 293.0, 'cost': 174.0},
                 {'value': 352.0, 'cost': 210.0},
                 {'value': 421.0, 'cost': 252.0},
                 {'value': 502.0, 'cost': 302.0},
                 {'value': 597.0, 'cost': 362.0},
                 {'value': 708.0, 'cost': 432.0},
                 {'value': 838.0, 'cost': 528.0},
                 {'value': 989.0, 'cost': 654.0},
                 {'value': 1165.0, 'cost': 810.0},
                 {'value': 1370.0, 'cost': 996.0},
                 {'value': 1608.0, 'cost': 1222.0},
                 {'value': 1886.0, 'cost': 1488.0},
                 {'value': 2209.0, 'cost': 1804.0},
                 {'value': 2585.0, 'cost': 2180.0},
                 {'value': 3021.0, 'cost': 2636.0}],
 'SM | Quantity': [{'value': 5.0, 'cost': 0.0},
                   {'value': 6.0, 'cost': 4.0},
                   {'value': 7.0, 'cost': 12.0},
                   {'value': 8.0, 'cost': 35.0},
                   {'value': 9.0, 'cost': 70.0},
                   {'value': 10.0, 'cost': 120.0},
                   {'value': 11.0, 'cost': 180.0},
                   {'value': 12.0, 'cost': 275.0},
                   {'value': 13.0, 'cost': 350.0},
                   {'value': 14.0, 'cost': 420.0},
                   {'value': 15.0, 'cost': 500.0},
                   {'value': 16.0, 'cost': 600.0},
                   {'value': 17.0, 'cost': 750.0},
                   {'value': 18.0, 'cost': 950.0},
                   {'value': 19.0, 'cost': 1200.0},
                   {'value': 20.0, 'cost': 1500.0}],
 'SM | Cooldown': [{'value': 180.0, 'cost': 0.0},
                   {'value': 170.0, 'cost': 8.0},
                   {'value': 160.0, 'cost': 24.0},
                   {'value': 150.0, 'cost': 40.0},
                   {'value': 140.0, 'cost': 56.0},
                   {'value': 130.0, 'cost': 72.0},
                   {'value': 120.0, 'cost': 88.0},
                   {'value': 110.0, 'cost': 104.0},
                   {'value': 100.0, 'cost': 120.0},
                   {'value': 90.0, 'cost': 136.0},
                   {'value': 80.0, 'cost': 152.0},
                   {'value': 70.0, 'cost': 168.0},
                   {'value': 60.0, 'cost': 184.0},
                   {'value': 50.0, 'cost': 200.0},
                   {'value': 40.0, 'cost': 216.0},
                   {'value': 30.0, 'cost': 232.0},
                   {'value': 20.0, 'cost': 750.0}],
 'DW | Damage': [{'value': 2.0, 'cost': 0.0},
                 {'value': 3.0, 'cost': 5.0},
                 {'value': 5.0, 'cost': 11.0},
                 {'value': 9.0, 'cost': 17.0},
                 {'value': 14.0, 'cost': 23.0},
                 {'value': 22.0, 'cost': 29.0},
                 {'value': 32.0, 'cost': 35.0},
                 {'value': 46.0, 'cost': 41.0},
                 {'value': 63.0, 'cost': 47.0},
                 {'value': 85.0, 'cost': 53.0},
                 {'value': 113.0, 'cost': 61.0},
                 {'value': 148.0, 'cost': 71.0},
                 {'value': 191.0, 'cost': 84.0},
                 {'value': 244.0, 'cost': 100.0},
                 {'value': 309.0, 'cost': 120.0},
                 {'value': 387.0, 'cost': 144.0},
                 {'value': 482.0, 'cost': 174.0},
                 {'value': 596.0, 'cost': 210.0},
                 {'value': 723.0, 'cost': 254.0},
                 {'value': 877.0, 'cost': 308.0},
                 {'value': 1064.0, 'cost': 374.0},
                 {'value': 1290.0, 'cost': 452.0},
                 {'value': 1569.0, 'cost': 558.0},
                 {'value': 1916.0, 'cost': 694.0},
                 {'value': 2356.0, 'cost': 880.0},
                 {'value': 2919.0, 'cost': 1126.0},
                 {'value': 3637.0, 'cost': 1432.0},
                 {'value': 4544.0, 'cost': 1813.0},
                 {'value': 5678.0, 'cost': 2269.0},
                 {'value': 7078.0, 'cost': 2800.0},
                 {'value': 9119.0, 'cost': 3406.0}],
 'DW | Quantity': [{'value': 1.0, 'cost': 0.0},
                   {'value': 2.0, 'cost': 200.0},
                   {'value': 3.0, 'cost': 500.0},
                   {'value': 4.0, 'cost': 850.0},
                   {'value': 5.0, 'cost': 1400.0}],
 'DW | Cooldown': [{'value': 300.0, 'cost': 0.0},
                   {'value': 290.0, 'cost': 8.0},
                   {'value': 280.0, 'cost': 24.0},
                   {'value': 270.0, 'cost': 40.0},
                   {'value': 260.0, 'cost': 56.0},
                   {'value': 250.0, 'cost': 72.0},
                   {'value': 240.0, 'cost': 88.0},
                   {'value': 230.0, 'cost': 104.0},
                   {'value': 220.0, 'cost': 120.0},
                   {'value': 210.0, 'cost': 136.0},
                   {'value': 200.0, 'cost': 152.0},
                   {'value': 190.0, 'cost': 168.0},
                   {'value': 180.0, 'cost': 184.0},
                   {'value': 170.0, 'cost': 200.0},
                   {'value': 160.0, 'cost': 216.0},
                   {'value': 150.0, 'cost': 232.0},
                   {'value': 140.0, 'cost': 248.0},
                   {'value': 130.0, 'cost': 264.0},
                   {'value': 120.0, 'cost': 280.0},
                   {'value': 110.0, 'cost': 346.0},
                   {'value': 100.0, 'cost': 512.0},
                   {'value': 90.0, 'cost': 688.0},
                   {'value': 80.0, 'cost': 874.0},
                   {'value': 70.0, 'cost': 1070.0},
                   {'value': 60.0, 'cost': 1276.0},
                   {'value': 50.0, 'cost': 1492.0}],
 'BH | Size': [{'value': 30.0, 'cost': 0.0},
               {'value': 32.0, 'cost': 5.0},
               {'value': 34.0, 'cost': 12.0},
               {'value': 36.0, 'cost': 19.0},
               {'value': 38.0, 'cost': 26.0},
               {'value': 40.0, 'cost': 34.0},
               {'value': 42.0, 'cost': 43.0},
               {'value': 44.0, 'cost': 53.0},
               {'value': 46.0, 'cost': 64.0},
               {'value': 48.0, 'cost': 76.0},
               {'value': 50.0, 'cost': 89.0},
               {'value': 52.0, 'cost': 103.0},
               {'value': 54.0, 'cost': 118.0},
               {'value': 56.0, 'cost': 134.0},
               {'value': 58.0, 'cost': 151.0},
               {'value': 60.0, 'cost': 169.0},
               {'value': 62.0, 'cost': 189.0},
               {'value': 64.0, 'cost': 211.0},
               {'value': 66.0, 'cost': 236.0},
               {'value': 68.0, 'cost': 264.0},
               {'value': 70.0, 'cost': 295.0}],
 'BH | Duration': [{'value': 15.0, 'cost': 0.0},
                   {'value': 16.0, 'cost': 5.0},
                   {'value': 17.0, 'cost': 14.0},
                   {'value': 18.0, 'cost': 23.0},
                   {'value': 19.0, 'cost': 32.0},
                   {'value': 20.0, 'cost': 41.0},
                   {'value': 21.0, 'cost': 50.0},
                   {'value': 22.0, 'cost': 59.0},
                   {'value': 23.0, 'cost': 68.0},
                   {'value': 24.0, 'cost': 77.0},
                   {'value': 25.0, 'cost': 86.0},
                   {'value': 26.0, 'cost': 95.0},
                   {'value': 27.0, 'cost': 104.0},
                   {'value': 28.0, 'cost': 113.0},
                   {'value': 29.0, 'cost': 122.0},
                   {'value': 30.0, 'cost': 131.0},
                   {'value': 31.0, 'cost': 165.0},
                   {'value': 32.0, 'cost': 224.0},
                   {'value': 33.0, 'cost': 308.0},
                   {'value': 34.0, 'cost': 417.0},
                   {'value': 35.0, 'cost': 551.0},
                   {'value': 36.0, 'cost': 710.0},
                   {'value': 37.0, 'cost': 894.0},
                   {'value': 38.0, 'cost': 1103.0}],
 'BH | Cooldown': [{'value': 200.0, 'cost': 0.0},
                   {'value': 190.0, 'cost': 10.0},
                   {'value': 180.0, 'cost': 28.0},
                   {'value': 170.0, 'cost': 46.0},
                   {'value': 160.0, 'cost': 64.0},
                   {'value': 150.0, 'cost': 82.0},
                   {'value': 140.0, 'cost': 100.0},
                   {'value': 130.0, 'cost': 118.0},
                   {'value': 120.0, 'cost': 136.0},
                   {'value': 110.0, 'cost': 154.0},
                   {'value': 100.0, 'cost': 172.0},
                   {'value': 90.0, 'cost': 190.0},
                   {'value': 80.0, 'cost': 208.0},
                   {'value': 70.0, 'cost': 226.0},
                   {'value': 60.0, 'cost': 244.0},
                   {'value': 50.0, 'cost': 262.0}],
 'SL | Multiplier': [{'value': 8.0, 'cost': 0.0},
                     {'value': 9.4, 'cost': 5.0},
                     {'value': 10.8, 'cost': 13.0},
                     {'value': 12.200000000000001, 'cost': 21.0},
                     {'value': 13.600000000000001, 'cost': 30.0},
                     {'value': 15.000000000000002, 'cost': 40.0},
                     {'value': 16.400000000000002, 'cost': 52.0},
                     {'value': 17.8, 'cost': 65.0},
                     {'value': 19.2, 'cost': 80.0},
                     {'value': 20.599999999999998, 'cost': 95.0},
                     {'value': 21.999999999999996, 'cost': 112.0},
                     {'value': 23.399999999999995, 'cost': 133.0},
                     {'value': 24.799999999999994, 'cost': 150.0},
                     {'value': 26.199999999999992, 'cost': 180.0},
                     {'value': 27.59999999999999, 'cost': 220.0},
                     {'value': 28.99999999999999, 'cost': 280.0},
                     {'value': 30.399999999999988, 'cost': 320.0},
                     {'value': 31.799999999999986, 'cost': 360.0},
                     {'value': 33.19999999999999, 'cost': 420.0},
                     {'value': 34.59999999999999, 'cost': 500.0},
                     {'value': 35.999999999999986, 'cost': 600.0},
                     {'value': 37.399999999999984, 'cost': 720.0},
                     {'value': 38.79999999999998, 'cost': 850.0},
                     {'value': 40.19999999999998, 'cost': 1000.0},
                     {'value': 41.59999999999998, 'cost': 1175.0},
                     {'value': 42.99999999999998, 'cost': 1400.0}],
 'SL | Angle': [{'value': 30.0, 'cost': 0.0},
                {'value': 31.0, 'cost': 5.0},
                {'value': 32.0, 'cost': 16.0},
                {'value': 33.0, 'cost': 27.0},
                {'value': 34.0, 'cost': 38.0},
                {'value': 35.0, 'cost': 49.0},
                {'value': 36.0, 'cost': 60.0},
                {'value': 37.0, 'cost': 71.0},
                {'value': 38.0, 'cost': 82.0},
                {'value': 39.0, 'cost': 93.0},
                {'value': 40.0, 'cost': 104.0},
                {'value': 41.0, 'cost': 115.0},
                {'value': 42.0, 'cost': 126.0},
                {'value': 43.0, 'cost': 137.0},
                {'value': 44.0, 'cost': 148.0},
                {'value': 45.0, 'cost': 159.0},
                {'value': 46.0, 'cost': 170.0},
                {'value': 47.0, 'cost': 181.0},
                {'value': 48.0, 'cost': 192.0},
                {'value': 49.0, 'cost': 203.0},
                {'value': 50.0, 'cost': 214.0},
                {'value': 51.0, 'cost': 225.0},
                {'value': 52.0, 'cost': 236.0},
                {'value': 53.0, 'cost': 247.0},
                {'value': 54.0, 'cost': 258.0},
                {'value': 55.0, 'cost': 269.0},
                {'value': 56.0, 'cost': 280.0},
                {'value': 57.0, 'cost': 291.0},
                {'value': 58.0, 'cost': 302.0},
                {'value': 59.0, 'cost': 313.0},
                {'value': 60.0, 'cost': 324.0},
                {'value': 61.0, 'cost': 337.0},
                {'value': 62.0, 'cost': 352.0},
                {'value': 63.0, 'cost': 369.0},
                {'value': 64.0, 'cost': 388.0},
                {'value': 65.0, 'cost': 409.0},
                {'value': 66.0, 'cost': 432.0},
                {'value': 67.0, 'cost': 457.0},
                {'value': 68.0, 'cost': 484.0},
                {'value': 69.0, 'cost': 513.0},
                {'value': 70.0, 'cost': 544.0},
                {'value': 71.0, 'cost': 577.0},
                {'value': 72.0, 'cost': 612.0},
                {'value': 73.0, 'cost': 649.0},
                {'value': 74.0, 'cost': 688.0},
                {'value': 75.0, 'cost': 729.0},
                {'value': 76.0, 'cost': 772.0},
                {'value': 77.0, 'cost': 817.0},
                {'value': 78.0, 'cost': 864.0},
                {'value': 79.0, 'cost': 913.0},
                {'value': 80.0, 'cost': 964.0},
                {'value': 81.0, 'cost': 1017.0},
                {'value': 82.0, 'cost': 1072.0},
                {'value': 83.0, 'cost': 1129.0},
                {'value': 84.0, 'cost': 1188.0},
                {'value': 85.0, 'cost': 1249.0},
                {'value': 86.0, 'cost': 1312.0},
                {'value': 87.0, 'cost': 1377.0},
                {'value': 88.0, 'cost': 1444.0},
                {'value': 89.0, 'cost': 1513.0},
                {'value': 90.0, 'cost': 1584.0}],
 'SL | Quantity': [{'value': 1.0, 'cost': 0.0},
                   {'value': 2.0, 'cost': 375.0},
                   {'value': 3.0, 'cost': 850.0},
                   {'value': 4.0, 'cost': 2500.0}]}

NATIVE_DAMAGE_LAB_EFFECTS = {
    "Damage": {"kind": "mult", "per_level": 0.02, "confidence": "High"},
    "Attack Speed": {"kind": "mult", "per_level": 0.02, "confidence": "High"},
    "Critical Factor": {"kind": "crit_factor", "per_level": 0.03, "confidence": "Medium"},
    "Range": {"kind": "mult", "per_level": 0.0035, "confidence": "Regression-calibrated"},
    "Damage / Meter": {"kind": "mult", "per_level": 0.002, "confidence": "Medium"},
    "Super Crit Chance": {"kind": "super_chance", "per_level": 0.003, "confidence": "Medium"},
    "Super Crit Multi": {"kind": "super_mult", "per_level": 0.10, "confidence": "Medium"},
    "Max Rend Armor Multiplier": {"kind": "rend", "per_level": 0.10, "confidence": "Low-Medium"},
    "Shock Chance": {"kind": "shock_chance", "per_level": 0.02, "confidence": "Regression-calibrated"},
    "Shock Multiplier": {"kind": "shock_mult", "per_level": 0.10, "confidence": "Regression-calibrated"},
    "Super Tower Bonus": {"kind": "mult", "per_level": 0.02, "confidence": "Low-Medium"},
    "Missile Amplifier": {"kind": "sm_amp", "per_level": 0.25, "confidence": "Medium"},
    "Missile Radius": {"kind": "sm_radius", "per_level": 0.02, "confidence": "Low"},
    "Death Wave Damage Amplifier": {"kind": "dw_amp", "per_level": 0.05, "confidence": "Medium"},
    "Black Hole Damage": {"kind": "bh_damage", "per_level": 0.002, "confidence": "High"},
}

NATIVE_DAMAGE_KEY_UPGRADES = [
    {"Upgrade": "Bounce Shot Chance", "Cost": 20.0, "Gain": 0.08, "Mode": "mult", "Notes": "Power vault first-tier node."},
    {"Upgrade": "Damage", "Cost": 15.0, "Gain": 0.05, "Mode": "mult", "Notes": "Power vault first-tier node."},
    {"Upgrade": "Critical Chance", "Cost": 15.0, "Gain": 0.01, "Mode": "crit_chance", "Notes": "Power vault first-tier node."},
    {"Upgrade": "Super Crit Chance", "Cost": 25.0, "Gain": 0.02, "Mode": "super_chance", "Notes": "Power vault first-tier node."},
    {"Upgrade": "Ultimate Weapon Damage", "Cost": 15.0, "Gain": 0.05, "Mode": "uw_mult", "Notes": "Power vault first-tier node."},
    {"Upgrade": "Critical Factor", "Cost": 25.0, "Gain": 0.05, "Mode": "crit_factor_mult", "Notes": "Power vault first-tier node."},
    {"Upgrade": "Damage / Meter", "Cost": 10.0, "Gain": 0.05, "Mode": "mult", "Notes": "Power vault first-tier node."},
    {"Upgrade": "Attack Speed", "Cost": 25.0, "Gain": 0.05, "Mode": "mult", "Notes": "Power vault first-tier node."},
    {"Upgrade": "Super Crit Mult", "Cost": 25.0, "Gain": 0.05, "Mode": "super_mult_mult", "Notes": "Power vault first-tier node."},
]

# -----------------------------------------------------------------------------
# EFFECTIVE PATHS ROI REFERENCE IMPORT
# -----------------------------------------------------------------------------

ROI_PATH_SPECS: Dict[str, Dict[str, Any]] = {
    "econ_lab": {
        "sheet": "eEcon", "start_row": 6, "end_row": 100,
        "columns": {"Upgrade": "F", "Level": "G", "Cost": "H", "Duration": "I", "ROI": "J", "Result": "K", "Cumulative": "L"},
        "title": "Economy · Lab time path", "resource": "Time", "metric": "CPK",
    },
    "econ_stone": {
        "sheet": "eEcon", "start_row": 6, "end_row": 100,
        "columns": {"Upgrade": "N", "Level": "O", "Cost": "P", "ROI": "Q", "Result": "R", "Cumulative": "S"},
        "title": "Economy · Stone path", "resource": "Stones", "metric": "CPK",
    },
    "econ_coin": {
        "sheet": "eEcon", "start_row": 6, "end_row": 100,
        "columns": {"Upgrade": "U", "Level": "V", "Cost": "W", "ROI": "X", "Result": "Y", "Cumulative": "Z"},
        "title": "Economy · Coin path", "resource": "Coins", "metric": "CPK",
    },
    "econ_discount": {
        "sheet": "eEcon", "start_row": 6, "end_row": 100,
        "columns": {"Upgrade": "AB", "Level": "AC", "Duration": "AD", "ROI": "AE", "Result": "AF", "Cumulative": "AG"},
        "title": "Economy · Lab discount path", "resource": "Time", "metric": "Effective coin value",
    },
    "damage_lab": {
        "sheet": "eDamage", "start_row": 6, "end_row": 150,
        "columns": {"Upgrade": "F", "Level": "G", "Cost": "H", "Duration": "I", "ROI": "J", "Result": "K", "Cumulative": "L"},
        "title": "Damage · Lab time path", "resource": "Time", "metric": "eDMG",
    },
    "damage_stone": {
        "sheet": "eDamage", "start_row": 6, "end_row": 150,
        "columns": {"Upgrade": "N", "Level": "O", "Cost": "P", "ROI": "Q", "Result": "R", "Cumulative": "S"},
        "title": "Damage · Stone path", "resource": "Stones", "metric": "eDMG",
    },
    "damage_coin": {
        "sheet": "eDamage", "start_row": 6, "end_row": 150,
        "columns": {"Upgrade": "U", "Level": "V", "Cost": "W", "ROI": "X", "Result": "Y", "Cumulative": "Z"},
        "title": "Damage · Coin path", "resource": "Coins", "metric": "eDMG",
    },
    "damage_key": {
        "sheet": "eDamage", "start_row": 6, "end_row": 150,
        "columns": {"Upgrade": "AB", "Level": "AC", "Cost": "AD", "ROI": "AE", "Result": "AF", "Cumulative": "AG"},
        "title": "Damage · Key path", "resource": "Keys", "metric": "eDMG",
    },
    "health_lab": {
        "sheet": "eHP", "start_row": 6, "end_row": 150,
        "columns": {"Upgrade": "F", "Level": "G", "Cost": "H", "Duration": "I", "ROI": "J", "Result": "K", "Cumulative": "L"},
        "title": "Health · Lab time path", "resource": "Time", "metric": "eHP",
    },
    "health_stone": {
        "sheet": "eHP", "start_row": 6, "end_row": 150,
        "columns": {"Upgrade": "N", "Level": "O", "Cost": "P", "ROI": "Q", "Result": "R", "Cumulative": "S"},
        "title": "Health · Stone path", "resource": "Stones", "metric": "eHP",
    },
    "health_coin": {
        "sheet": "eHP", "start_row": 6, "end_row": 150,
        "columns": {"Upgrade": "U", "Level": "V", "Cost": "W", "ROI": "X", "Result": "Y", "Cumulative": "Z"},
        "title": "Health · Coin path", "resource": "Coins", "metric": "eHP",
    },
    "regen_lab": {
        "sheet": "eHP", "start_row": 6, "end_row": 150,
        "columns": {"Upgrade": "AB", "Level": "AC", "Cost": "AD", "Duration": "AE", "ROI": "AF", "Result": "AG", "Cumulative": "AH"},
        "title": "Wall regen · Lab time path", "resource": "Time", "metric": "Wall regen",
    },
}


def serialize_sheet_value(value: Any) -> Any:
    """Convert openpyxl values to JSON-safe display values."""
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.isoformat()
    # datetime.timedelta is not imported directly, so detect by attributes.
    if hasattr(value, "total_seconds") and not isinstance(value, (str, bytes)):
        try:
            seconds = int(value.total_seconds())
            days, seconds = divmod(seconds, 86400)
            hours, seconds = divmod(seconds, 3600)
            minutes, seconds = divmod(seconds, 60)
            parts = []
            if days:
                parts.append(f"{days}d")
            if hours or days:
                parts.append(f"{hours}h")
            if minutes or hours or days:
                parts.append(f"{minutes}m")
            parts.append(f"{seconds}s")
            return " ".join(parts)
        except Exception:
            return str(value)
    if isinstance(value, (int, float, bool, str)):
        return value
    return str(value)


def parse_roi_path_table(workbook: Any, path_key: str, spec: Dict[str, Any]) -> list[Dict[str, Any]]:
    sheet_name = spec["sheet"]
    if sheet_name not in workbook.sheetnames:
        return []
    sheet = workbook[sheet_name]
    rows: list[Dict[str, Any]] = []
    blank_streak = 0
    for row_number in range(int(spec["start_row"]), int(spec["end_row"]) + 1):
        row: Dict[str, Any] = {}
        for field, column in spec["columns"].items():
            row[field] = serialize_sheet_value(sheet[f"{column}{row_number}"].value)
        upgrade = row.get("Upgrade")
        roi = row.get("ROI")
        if upgrade in (None, ""):
            blank_streak += 1
            if blank_streak >= 12 and rows:
                break
            continue
        blank_streak = 0
        upgrade_text = str(upgrade).strip()
        if not upgrade_text or upgrade_text.lower().startswith("not seeing path"):
            continue
        # A valid recommendation needs an upgrade and at least one useful output.
        if all(row.get(field) in (None, "") for field in ["Level", "Cost", "Duration", "ROI", "Result"]):
            continue
        row["Rank"] = len(rows) + 1
        row["Path"] = path_key
        row["Resource"] = spec.get("resource")
        row["Metric"] = spec.get("metric")
        try:
            row["ROI Numeric"] = float(roi) if roi not in (None, "") else None
        except (TypeError, ValueError):
            row["ROI Numeric"] = None
        row["Cost Numeric"] = parse_tower_number(row.get("Cost"))
        rows.append(row)
    return rows


def detect_effective_paths_version(workbook: Any, filename: str) -> str:
    # Prefer a version from a visible path title, then fall back to filename.
    for sheet_name, cell in [("eEcon", "F3"), ("eDamage", "F3"), ("eHP", "F3")]:
        if sheet_name in workbook.sheetnames:
            text = clean_cell(workbook[sheet_name][cell].value)
            match = re.search(r"v(\d+(?:\.\d+)*)", text, re.IGNORECASE)
            if match:
                return f"v{match.group(1)}"
    match = re.search(r"v\s*(\d+(?:\.\d+)+)", filename, re.IGNORECASE)
    return f"v{match.group(1)}" if match else "Unknown"


def parse_effective_paths_roi_reference(uploaded_file: Any) -> Dict[str, Any]:
    if Path(uploaded_file.name).suffix.lower() != ".xlsx":
        raise ValueError("ROI reference import requires the filled Effective Paths .xlsx workbook.")
    payload = uploaded_file.getvalue()
    workbook = load_workbook(io.BytesIO(payload), read_only=False, data_only=True)
    missing = sorted({spec["sheet"] for spec in ROI_PATH_SPECS.values()} - set(workbook.sheetnames))
    if missing:
        raise ValueError("Missing Effective Paths result sheets: " + ", ".join(missing))

    paths: Dict[str, Any] = {}
    warnings_list: list[str] = []
    for path_key, spec in ROI_PATH_SPECS.items():
        rows = parse_roi_path_table(workbook, path_key, spec)
        paths[path_key] = {
            "title": spec["title"],
            "resource": spec["resource"],
            "metric": spec["metric"],
            "rows": rows,
        }
        if not rows:
            warnings_list.append(
                f"{spec['title']} had no cached recommendation rows. Open/recalculate the workbook in Google Sheets or Excel, save it, and import it again."
            )

    nonempty = sum(1 for path in paths.values() if path.get("rows"))
    if nonempty == 0:
        raise ValueError(
            "The workbook contains the result sheets, but no cached ROI paths were readable. "
            "Recalculate and save the workbook before importing it."
        )

    return {
        "source": {
            "filename": uploaded_file.name,
            "effective_paths_version": detect_effective_paths_version(workbook, uploaded_file.name),
            "imported_at": datetime.now(timezone.utc).isoformat(),
            "mode": "cached spreadsheet outputs",
        },
        "paths": paths,
        "warnings": warnings_list,
    }


def apply_roi_reference(reference: Dict[str, Any]) -> None:
    st.session_state.profile["roi_reference"] = reference
    st.session_state.profile["sources"]["roi_reference"] = reference.get("source", {})
    st.session_state.profile["metadata"]["last_import"] = {
        "at": reference.get("source", {}).get("imported_at"),
        "source": f"Effective Paths ROI reference ({reference.get('source', {}).get('filename', 'workbook')})",
        "app_version": APP_VERSION,
    }


def roi_path_rows(profile: Dict[str, Any], path_key: str) -> list[Dict[str, Any]]:
    path = profile.get("roi_reference", {}).get("paths", {}).get(path_key, {})
    rows = path.get("rows", []) if isinstance(path, dict) else []
    return rows if isinstance(rows, list) else []


def recommendation_is_gold_boxed(profile: Dict[str, Any], path_key: str, upgrade: str) -> bool:
    """Best-effort filtering of exact Lab/Enhancement recommendations already at cap."""
    text = str(upgrade or "").strip()
    if path_key.endswith("_lab") or path_key == "econ_discount":
        canonical = LAB_ALIASES.get(text, text)
        return bool(profile.get("maxed", {}).get("labs", {}).get(canonical, False))
    if path_key.endswith("_coin"):
        canonical = text
        if canonical.endswith(" +"):
            canonical = canonical[:-2].strip() + " +"
        return bool(profile.get("maxed", {}).get("enhancements", {}).get(canonical, False))
    return False


def roi_summary_rows(profile: Dict[str, Any], include_gold: bool = False, top_per_path: int = 3) -> list[Dict[str, Any]]:
    output: list[Dict[str, Any]] = []
    paths = profile.get("roi_reference", {}).get("paths", {})
    for path_key, path in paths.items():
        rows = path.get("rows", []) if isinstance(path, dict) else []
        added = 0
        for row in rows:
            if not include_gold and recommendation_is_gold_boxed(profile, path_key, row.get("Upgrade", "")):
                continue
            output.append({
                "Category": path.get("title", path_key),
                "Upgrade": row.get("Upgrade"),
                "Level": row.get("Level"),
                "Cost": row.get("Cost"),
                "Duration": row.get("Duration"),
                "ROI": row.get("ROI"),
                "Result": row.get("Result"),
                "Rank in path": row.get("Rank"),
                "Path key": path_key,
            })
            added += 1
            if added >= top_per_path:
                break
    return output


def auto_mark_gold(section: str, name: str, value: int) -> None:
    cap = metadata_for(section, name)
    st.session_state.profile["maxed"][section][name] = bool(cap is not None and value >= cap)


def apply_import(imported: Dict[str, Any], replace_values: bool) -> None:
    profile = st.session_state.profile
    for section in ["workshop", "labs", "enhancements"]:
        if replace_values:
            profile[section] = {}
            profile["maxed"][section] = {}
        for name, value in imported.get(section, {}).items():
            profile[section][name] = int(value)
            auto_mark_gold(section, name, int(value))

    if replace_values:
        profile["uw"] = {}
        profile["maxed"]["uw"] = {}
    for uw_name, data in imported.get("uw", {}).items():
        uw = profile["uw"].setdefault(uw_name, {"owned": False, "attributes": {}})
        if data.get("owned") is not None:
            uw["owned"] = bool(data["owned"])
        attrs = uw.setdefault("attributes", {})
        gold = profile["maxed"]["uw"].setdefault(uw_name, {})
        for attr, value in data.get("attributes", {}).items():
            attrs[attr] = value
            meta = UW_ATTRIBUTE_META[uw_name][attr]
            gold[attr] = value >= meta["max"] if not meta.get("lower_is_better") else value <= meta["max"]

    profile["metadata"]["last_import"] = {
        "at": datetime.now(timezone.utc).isoformat(),
        "source": st.session_state.get("import_filename", "uploaded file"),
        "app_version": APP_VERSION,
    }



# -----------------------------------------------------------------------------
# IDS COMPANION WORKBOOK IMPORTERS
# -----------------------------------------------------------------------------

def sheet_rows(sheet: Any) -> list[list[Any]]:
    return [list(row) for row in sheet.iter_rows(values_only=True)]


def row_value(row: list[Any], index: int) -> Any:
    return row[index] if 0 <= index < len(row) else None


def to_bool(value: Any) -> Optional[bool]:
    if isinstance(value, bool):
        return value
    number = to_number(value)
    if number is not None:
        return bool(int(round(number)))
    text = clean_cell(value).lower()
    if text in {"true", "yes", "owned", "unlocked"}: return True
    if text in {"false", "no", "locked"}: return False
    return None


def detect_companion_kind(workbook: Any) -> tuple[str, str]:
    if "EXPORT" not in workbook.sheetnames:
        raise ValueError("Workbook has no EXPORT sheet and is not a supported IDS companion file.")
    export = workbook["EXPORT"]
    signature = clean_cell(export["A1"].value)
    second = clean_cell(export["A2"].value)
    if second == "MASTER":
        return "ids_master", signature
    for prefix, kind in COMPANION_SIGNATURES.items():
        if signature.startswith(prefix):
            return kind, signature
    raise ValueError(f"Unknown IDS workbook signature: {signature or '(blank)'}")


def parse_workshop_companion(workbook: Any) -> Dict[str, Any]:
    rows = sheet_rows(workbook["EXPORT"])
    patch = {"workshop": {}, "enhancements": {}}
    for row in rows[3:]:
        name = clean_cell(row_value(row, 2))
        level = to_number(row_value(row, 3))
        if level is None:
            level = to_number(row_value(row, 4))
        canonical = WORKSHOP_ALIASES.get(name, name)
        if canonical in WORKSHOP_MAX_LEVELS and level is not None:
            patch["workshop"][canonical] = int(round(level))
        enhancement = clean_cell(row_value(row, 15))
        enhancement_level = to_number(row_value(row, 17))
        if enhancement in ENHANCEMENT_MAX_LEVELS and enhancement_level is not None:
            patch["enhancements"][enhancement] = int(round(enhancement_level))
    return patch


def parse_laboratory_companion(workbook: Any) -> Dict[str, Any]:
    rows = sheet_rows(workbook["EXPORT"])
    labs: Dict[str, int] = {}
    for row in rows[4:]:
        name = clean_cell(row_value(row, 1))
        level = to_number(row_value(row, 2))
        canonical = LAB_ALIASES.get(name, name)
        if canonical in LAB_MAX_LEVELS and level is not None:
            labs[canonical] = int(round(level))
    return {"labs": labs}


def parse_uw_companion(workbook: Any) -> Dict[str, Any]:
    rows = sheet_rows(workbook["EXPORT"])
    result: Dict[str, Any] = {"uw": {}, "resources": {}}
    stones = to_number(row_value(rows[1], 4)) if len(rows) > 1 else None
    if stones is not None:
        result["resources"]["stones_spent_uw"] = int(round(stones))
    current: Optional[str] = None
    for row in rows[4:]:
        possible_name = clean_cell(row_value(row, 2))
        status = clean_cell(row_value(row, 3))
        attribute = clean_cell(row_value(row, 4))
        value = row_value(row, 5)
        if possible_name in UW_NAMES:
            current = possible_name
            result["uw"].setdefault(current, {"owned": None, "attributes": {}, "plus": {}})
        if current and status in {"UW Unlocked", "UW Locked"}:
            result["uw"][current]["owned"] = status == "UW Unlocked"
        if current and attribute in UW_ATTRIBUTE_META[current]:
            number = to_number(value)
            if number is not None:
                meta = UW_ATTRIBUTE_META[current][attribute]
                result["uw"][current]["attributes"][attribute] = int(round(number)) if isinstance(meta["max"], int) else float(number)
        if current and possible_name == "UW+" and attribute:
            result["uw"][current]["plus"] = {"name": attribute, "value": value}
    return result


def parse_cards_companion(workbook: Any) -> Dict[str, Any]:
    rows = sheet_rows(workbook["EXPORT"])
    cards = {"slots": 0, "items": {}}
    slots = to_number(row_value(rows[1], 2)) if len(rows) > 1 else None
    if slots is not None:
        cards["slots"] = int(round(slots))
    for row in rows[4:]:
        name = clean_cell(row_value(row, 1))
        if name not in CARD_NAMES:
            continue
        level = to_number(row_value(row, 2)) or 0
        mastery = to_number(row_value(row, 3)) or 0
        cards["items"][name] = {"level": int(round(level)), "mastery": int(round(mastery))}
    return {"cards": cards}


def parse_player_companion(workbook: Any) -> Dict[str, Any]:
    rows = sheet_rows(workbook["EXPORT"])
    player: Dict[str, Any] = {"packs": {}, "tiers": {}}
    resources: Dict[str, Any] = {}
    for row in rows[2:]:
        tier = clean_cell(row_value(row, 1))
        if tier.startswith("Tier "):
            player["tiers"][tier] = {
                "wave": int(round(to_number(row_value(row, 2)) or 0)),
                "attack": int(round(to_number(row_value(row, 3)) or 0)),
                "defense": int(round(to_number(row_value(row, 4)) or 0)),
                "utility": int(round(to_number(row_value(row, 5)) or 0)),
                "ultimate_weapon": int(round(to_number(row_value(row, 6)) or 0)),
                "premium": int(round(to_number(row_value(row, 7)) or 0)),
            }
        stat = clean_cell(row_value(row, 9))
        value = row_value(row, 10)
        if stat == "Player ID": player["player_id"] = clean_cell(value)
        elif stat == "Farming Tier": player["farming_tier"] = clean_cell(value)
        elif stat == "Tourney League": player["tourney_league"] = clean_cell(value)
        elif stat == "Lifetime Coins":
            n = to_number(value)
            if n is not None: player["lifetime_coins"] = int(round(n))
        elif stat == "Stones":
            n = to_number(value)
            if n is not None: resources["stones"] = int(round(n))
        elif stat == "Gems":
            n = to_number(value)
            if n is not None: resources["gems"] = int(round(n))
        elif stat == "Coin Multiplier": player["coin_multiplier"] = clean_cell(value)
        elif stat in {"Premium Packs", "Disable Ads", "Starter Pack", "Epic Pack"}:
            player["packs"][stat] = value
    return {"player": player, "resources": resources}


def parse_relics_companion(workbook: Any) -> Dict[str, Any]:
    result = {"relics": {"summary": {}, "bonuses": {}, "items": {}}}
    export_rows = sheet_rows(workbook["EXPORT"])
    mode = "summary"
    for row in export_rows[1:]:
        name = clean_cell(row_value(row, 1))
        active = row_value(row, 2)
        total = row_value(row, 4)
        if name == "Total Bonuses":
            mode = "bonuses"; continue
        if not name or name in {"Misc.", "Damage‎", "Defense", "Utility"}:
            continue
        if mode == "summary":
            result["relics"]["summary"][name] = {"owned": active, "total": total}
        else:
            result["relics"]["bonuses"][name] = {"active": active, "total": total}
    if "Relics" in workbook.sheetnames:
        rows = sheet_rows(workbook["Relics"])
        for row in rows[2:]:
            name = clean_cell(row_value(row, 3))
            unlocked = to_bool(row_value(row, 6))
            if not name or unlocked is None:
                continue
            result["relics"]["items"][name] = {
                "owned": unlocked,
                "rarity": clean_cell(row_value(row, 1)),
                "number": to_number(row_value(row, 2)),
                "bonus_type": clean_cell(row_value(row, 4)),
                "value": to_number(row_value(row, 5)),
                "event": clean_cell(row_value(row, 7)),
                "unlocked_by": clean_cell(row_value(row, 8)),
                "type": clean_cell(row_value(row, 9)),
                "last_seen": row_value(row, 10),
            }
    return result


def parse_themes_companion(workbook: Any) -> Dict[str, Any]:
    result = {"themes": {"summary": {}, "items": {}}}
    export_rows = sheet_rows(workbook["EXPORT"])
    for row in export_rows[2:]:
        name = clean_cell(row_value(row, 1))
        if name:
            result["themes"]["summary"][name] = {
                "active": row_value(row, 2), "bonus": row_value(row, 3),
                "total": row_value(row, 5), "total_bonus": row_value(row, 6),
            }
    if "Themes & Songs" not in workbook.sheetnames:
        return result
    rows = sheet_rows(workbook["Themes & Songs"])
    items: Dict[str, Any] = {}
    for idx, row in enumerate(rows[2:], start=3):
        tower_name = clean_cell(row_value(row, 2))
        tower_owned = to_bool(row_value(row, 1))
        if tower_name and tower_owned is not None:
            items[f"Event Tower::{tower_name}"] = {"type": "Event Tower", "name": tower_name, "owned": tower_owned, "event": clean_cell(row_value(row, 7))}
        bg_name = clean_cell(row_value(row, 5))
        bg_owned = to_bool(row_value(row, 4))
        if bg_name and bg_owned is not None:
            items[f"Event Background::{bg_name}"] = {"type": "Event Background", "name": bg_name, "owned": bg_owned, "event": clean_cell(row_value(row, 7))}
        if idx <= 23:
            skin_name = clean_cell(row_value(row, 12))
            skin_owned = to_bool(row_value(row, 11))
            if skin_name and skin_owned is not None:
                items[f"Tier Skin::{skin_name}"] = {"type": "Tier Skin", "name": skin_name, "owned": skin_owned, "tier": clean_cell(row_value(row, 13)), "source": clean_cell(row_value(row, 14))}
    # Additional tables below the event/tier sections.
    current_left = None
    current_right = None
    for row in rows[23:]:
        left_marker = clean_cell(row_value(row, 11))
        right_marker = clean_cell(row_value(row, 16))
        if left_marker in {"Songs", "Guardians"}: current_left = left_marker
        if right_marker in {"Menu", "Profile Banner"}: current_right = right_marker
        left_owned = to_bool(row_value(row, 11))
        left_name = clean_cell(row_value(row, 12))
        if current_left and left_owned is not None and left_name:
            items[f"{current_left}::{left_name}"] = {"type": current_left, "name": left_name, "owned": left_owned}
        right_owned = to_bool(row_value(row, 16))
        right_name = clean_cell(row_value(row, 17))
        if current_right and right_owned is not None and right_name:
            items[f"{current_right}::{right_name}"] = {"type": current_right, "name": right_name, "owned": right_owned}
    result["themes"]["items"] = items
    return result


def parse_bots_companion(workbook: Any) -> Dict[str, Any]:
    rows = sheet_rows(workbook["EXPORT"])
    result: Dict[str, Any] = {"bots": {}, "resources": {}}
    medals = to_number(row_value(rows[1], 4)) if len(rows) > 1 else None
    if medals is not None: result["resources"]["medals_spent_bots"] = int(round(medals))
    current: Optional[str] = None
    for row in rows[4:]:
        bot_cell = clean_cell(row_value(row, 2))
        status = clean_cell(row_value(row, 3))
        attr = clean_cell(row_value(row, 4))
        value = row_value(row, 5)
        if bot_cell in BOT_NAMES:
            current = bot_cell
            result["bots"].setdefault(current, {"unlocked": None, "attributes": {}, "plus": {}})
        if current and status in {"Unlocked", "Locked"}:
            result["bots"][current]["unlocked"] = status == "Unlocked"
        if current and attr in BOT_ATTRIBUTES.get(current, []):
            number = to_number(value)
            result["bots"][current]["attributes"][attr] = number if number is not None else value
        if current and bot_cell == "Bot +" and attr:
            result["bots"][current]["plus"] = {"name": attr, "value": value}
    return result


def parse_guardians_companion(workbook: Any) -> Dict[str, Any]:
    rows = sheet_rows(workbook["EXPORT"])
    result: Dict[str, Any] = {"guardians": {}, "resources": {}}
    bits = to_number(row_value(rows[1], 3)) if len(rows) > 1 else None
    if bits is not None: result["resources"]["bits_spent_guardians"] = int(round(bits))
    current: Optional[str] = None
    for row in rows[4:]:
        name_or_flag = clean_cell(row_value(row, 1))
        status = clean_cell(row_value(row, 2))
        attr = clean_cell(row_value(row, 3))
        value = row_value(row, 4)
        spent = to_number(row_value(row, 6))
        if name_or_flag in GUARDIAN_NAMES:
            current = name_or_flag
            result["guardians"].setdefault(current, {"unlocked": None, "attributes": {}, "bits_spent": 0})
        if current and status in {"Unlocked", "Locked"}:
            result["guardians"][current]["unlocked"] = status == "Unlocked"
        if current and attr in GUARDIAN_ATTRIBUTES.get(current, []):
            number = to_number(value)
            result["guardians"][current]["attributes"][attr] = number if number is not None else value
            if spent is not None: result["guardians"][current]["bits_spent"] += int(round(spent))
    return result


def parse_vault_companion(workbook: Any) -> Dict[str, Any]:
    rows = sheet_rows(workbook["EXPORT"])
    vault = {"keys_spent": 0, "bonuses": {}, "unlocks": {}}
    keys = to_number(row_value(rows[1], 2)) if len(rows) > 1 else None
    if keys is not None: vault["keys_spent"] = int(round(keys))
    in_unlocks = False
    for row in rows[4:]:
        name = clean_cell(row_value(row, 1))
        active = row_value(row, 2)
        total = row_value(row, 4)
        if name == "Unlocks": in_unlocks = True; continue
        if not name or name in {"Misc.", "Attack", "Defense", "Utility"}: continue
        if in_unlocks:
            vault["unlocks"][name] = bool(int(round(to_number(active) or 0)))
        else:
            vault["bonuses"][name] = {"active": active, "total": total}
    return {"vault": vault, "resources": {"keys": vault["keys_spent"]}}


def parse_modules_companion(workbook: Any) -> Dict[str, Any]:
    result: Dict[str, Any] = {"module_inventory": {}, "module_presets": {}}
    if "Inventory" in workbook.sheetnames:
        rows = sheet_rows(workbook["Inventory"])
        blocks = [(2, "Cannon", 15), (15, "Armor", 28), (28, "Generator", 41), (41, "Core", len(rows))]
        for start, slot, end in blocks:
            header = rows[start]
            for base in range(5, min(len(header), 40), 5):
                name = clean_cell(row_value(header, base))
                if not name or name == "Any Other": continue
                data_row = rows[start + 2] if start + 2 < len(rows) else []
                rarity = clean_cell(row_value(data_row, base))
                level = to_number(row_value(data_row, base + 1)) or 0
                stat = to_number(row_value(data_row, base + 2)) or 1
                substats = []
                for r in rows[start + 4:end]:
                    subname = clean_cell(row_value(r, base))
                    subrarity = clean_cell(row_value(r, base + 1))
                    display = clean_cell(row_value(r, base + 2))
                    numeric = to_number(row_value(r, base + 3))
                    if subname and subrarity:
                        substats.append({"name": subname, "rarity": subrarity, "display": display, "value": numeric})
                result["module_inventory"][f"{slot}::{name}"] = {
                    "slot": slot, "name": name, "rarity": rarity or "None",
                    "level": int(round(level)), "stat": stat, "substats": substats,
                }
    if "EXPORT" in workbook.sheetnames:
        rows = sheet_rows(workbook["EXPORT"])
        slot_cols = [(1, "Cannon"), (6, "Armor"), (11, "Generator"), (16, "Core")]
        for col, slot in slot_cols:
            current_preset = None
            for row in rows[6:25]:
                marker = clean_cell(row_value(row, col))
                value = clean_cell(row_value(row, col + 1))
                if marker in {"Farming", "Tourney", "Preset 3", "Preset 4", "Preset 5"}:
                    current_preset = marker
                    result["module_presets"].setdefault(current_preset, {})
                elif current_preset and marker in {"Primary Slot", "Assist Slot"}:
                    result["module_presets"][current_preset].setdefault(slot, {})["primary" if marker == "Primary Slot" else "assist"] = value
    return result


COMPANION_PARSERS = {
    "workshop": parse_workshop_companion,
    "laboratory": parse_laboratory_companion,
    "ultimate_weapons": parse_uw_companion,
    "modules": parse_modules_companion,
    "cards": parse_cards_companion,
    "relics": parse_relics_companion,
    "themes": parse_themes_companion,
    "bots": parse_bots_companion,
    "guardians": parse_guardians_companion,
    "vault": parse_vault_companion,
    "player": parse_player_companion,
    "ids_master": lambda workbook: {},
}


def count_leaf_values(value: Any) -> int:
    if isinstance(value, dict):
        return sum(count_leaf_values(v) for v in value.values())
    if isinstance(value, list):
        return sum(count_leaf_values(v) for v in value)
    return 1


def parse_companion_upload(uploaded_file: Any) -> Dict[str, Any]:
    payload = uploaded_file.getvalue()
    workbook = load_workbook(io.BytesIO(payload), read_only=False, data_only=True)
    kind, version = detect_companion_kind(workbook)
    patch = COMPANION_PARSERS[kind](workbook)
    recognized = count_leaf_values(patch)
    return {"filename": uploaded_file.name, "kind": kind, "version": version, "patch": patch, "recognized": recognized}


def merge_dict_recursive(target: Dict[str, Any], patch: Dict[str, Any], replace: bool = False) -> None:
    for key, value in patch.items():
        if isinstance(value, dict):
            if replace or key not in target or not isinstance(target.get(key), dict):
                target[key] = {} if isinstance(value, dict) else value
            merge_dict_recursive(target[key], value, replace=replace)
        else:
            target[key] = value


def apply_companion_bundle(results: list[Dict[str, Any]], replace: bool = False) -> None:
    profile = st.session_state.profile
    now = datetime.now(timezone.utc).isoformat()
    audit = []
    for result in results:
        patch = result.get("patch", {})
        for section, values in patch.items():
            if section == "resources":
                for key, value in values.items():
                    profile["resources"][key] = value
                continue
            if section in {"workshop", "labs", "enhancements"}:
                if replace: profile[section] = {}
                for name, value in values.items():
                    profile[section][name] = int(value)
                    auto_mark_gold(section, name, int(value))
            elif section == "uw":
                if replace: profile["uw"] = {}
                merge_dict_recursive(profile["uw"], values, replace=False)
            else:
                if replace and section in profile: profile[section] = {}
                if section not in profile or not isinstance(profile.get(section), dict): profile[section] = {}
                merge_dict_recursive(profile[section], values, replace=False)
            profile["sources"][section] = {"filename": result["filename"], "version": result["version"], "imported_at": now}
        audit.append({"filename": result["filename"], "kind": result["kind"], "version": result["version"], "recognized": result["recognized"], "status": "Applied"})
    profile["import_audit"] = audit
    profile["metadata"]["last_import"] = {"at": now, "source": f"IDS bundle ({len(results)} files)", "app_version": APP_VERSION}


def completeness_rows(profile: Dict[str, Any]) -> list[Dict[str, Any]]:
    rows = []
    def add(section: str, filled: int, total: int, note: str = "") -> None:
        percent = 100.0 if total == 0 and filled > 0 else (100.0 * filled / total if total else 0.0)
        rows.append({"Section": SECTION_LABELS.get(section, section.title()), "Filled": filled, "Expected": total, "Coverage %": round(percent, 1), "Source": profile.get("sources", {}).get(section, {}).get("filename", "Manual / none"), "Notes": note})
    add("player", sum(bool(profile["player"].get(k)) for k in ["player_id", "farming_tier", "tourney_league", "lifetime_coins"]), 4)
    add("workshop", len(profile.get("workshop", {})), len(WORKSHOP_MAX_LEVELS))
    add("labs", len(profile.get("labs", {})), len(LAB_MAX_LEVELS))
    add("enhancements", len(profile.get("enhancements", {})), len(ENHANCEMENT_MAX_LEVELS))
    uw_filled = sum(1 + len(v.get("attributes", {})) for v in profile.get("uw", {}).values())
    uw_total = len(UW_NAMES) + sum(len(v) for v in UW_ATTRIBUTE_META.values())
    add("uw", uw_filled, uw_total)
    add("module_inventory", len(profile.get("module_inventory", {})), sum(len(v) - 1 for v in MODULE_OPTIONS_BY_SLOT.values()), "Any Other excluded")
    add("module_presets", len(profile.get("module_presets", {})), 5)
    card_items = profile.get("cards", {}).get("items", {})
    cards_imported = "cards" in profile.get("sources", {})
    card_filled = len(CARD_NAMES) if cards_imported else sum(1 for c in CARD_NAMES if int(card_items.get(c, {}).get("level", 0) or 0) > 0 or int(card_items.get(c, {}).get("mastery", 0) or 0) > 0)
    add("cards", card_filled, len(CARD_NAMES))
    relic_items = profile.get("relics", {}).get("items", {})
    add("relics", sum(1 for v in relic_items.values() if "owned" in v), len(relic_items), "Expected is the imported relic catalog")
    theme_items = profile.get("themes", {}).get("items", {})
    add("themes", sum(1 for v in theme_items.values() if "owned" in v), len(theme_items), "Expected is the imported theme catalog")
    add("bots", len(profile.get("bots", {})), len(BOT_NAMES))
    add("guardians", len(profile.get("guardians", {})), len(GUARDIAN_NAMES))
    vault = profile.get("vault", {})
    add("vault", len(vault.get("bonuses", {})) + len(vault.get("unlocks", {})), max(1, len(vault.get("bonuses", {})) + len(vault.get("unlocks", {}))))
    return rows





# -----------------------------------------------------------------------------
# NATIVE DAMAGE ENGINE
# -----------------------------------------------------------------------------

def native_damage_settings(profile_data: Dict[str, Any]) -> Dict[str, Any]:
    damage = profile_data.setdefault("native_damage", {})
    settings = damage.setdefault("settings", {})
    settings.setdefault("cl_weight", 0.12)
    settings.setdefault("dw_weight", 0.0025)
    settings.setdefault("sm_weight", 0.0040)
    settings.setdefault("bh_weight", 0.15)
    settings.setdefault("shock_events_per_cycle", 50.0)
    settings.setdefault("base_shock_chance", 0.20)
    settings.setdefault("base_shock_multiplier", 1.50)
    settings.setdefault("show_unowned_uw", False)
    return settings


def native_vault_active(profile_data: Dict[str, Any], name: str) -> float:
    record = profile_data.get("vault", {}).get("bonuses", {}).get(name, {})
    if isinstance(record, dict):
        return native_number(record.get("active"), 0.0)
    return native_number(record, 0.0)


def native_damage_state(profile_data: Dict[str, Any]) -> Dict[str, Any]:
    return {
        "labs": {name: int(value or 0) for name, value in profile_data.get("labs", {}).items()},
        "enhancements": {name: int(value or 0) for name, value in profile_data.get("enhancements", {}).items()},
        "uw": {
            name: {
                "owned": bool(record.get("owned", False)),
                "attributes": {key: native_number(value) for key, value in record.get("attributes", {}).items()},
            }
            for name, record in profile_data.get("uw", {}).items()
            if isinstance(record, dict)
        },
        "key_bonuses": {},
    }


def native_damage_components(profile_data: Dict[str, Any], state: Optional[Dict[str, Any]] = None) -> Dict[str, float]:
    state = state or native_damage_state(profile_data)
    labs = state.get("labs", {})
    enhancements = state.get("enhancements", {})
    workshop = profile_data.get("workshop", {})
    key_bonuses = state.get("key_bonuses", {})
    settings = native_damage_settings(profile_data)

    damage_lab = int(labs.get("Damage", 0) or 0)
    attack_speed_lab = int(labs.get("Attack Speed", 0) or 0)
    crit_factor_lab = int(labs.get("Critical Factor", 0) or 0)
    range_lab = int(labs.get("Range", 0) or 0)
    dpm_lab = int(labs.get("Damage / Meter", 0) or 0)
    super_chance_lab = int(labs.get("Super Crit Chance", 0) or 0)
    super_mult_lab = int(labs.get("Super Crit Multi", 0) or 0)
    rend_lab = int(labs.get("Max Rend Armor Multiplier", 0) or 0)

    damage_mult = (1.0 + 0.02 * damage_lab)
    damage_mult *= 1.0 + 0.01 * int(enhancements.get("Damage +", 0) or 0)
    damage_mult *= 1.0 + native_vault_active(profile_data, "Damage")
    damage_mult *= 1.0 + native_number(key_bonuses.get("Damage"), 0.0)

    attack_speed_mult = 1.0 + 0.02 * attack_speed_lab
    attack_speed_mult *= 1.0 + 0.01 * int(enhancements.get("Attack Speed +", 0) or 0)
    attack_speed_mult *= 1.0 + native_vault_active(profile_data, "Attack Speed")
    attack_speed_mult *= 1.0 + native_number(key_bonuses.get("Attack Speed"), 0.0)

    crit_chance = min(0.99, 0.05 + 0.01 * int(workshop.get("Critical Chance", 0) or 0))
    crit_chance += native_vault_active(profile_data, "Critical Chance")
    crit_chance += native_number(key_bonuses.get("Critical Chance"), 0.0)
    crit_chance = min(0.99, max(0.0, crit_chance))

    crit_factor = 1.0 + 0.08 * int(workshop.get("Critical Factor", 0) or 0) + 0.03 * crit_factor_lab
    crit_factor *= 1.0 + 0.01 * int(enhancements.get("Critical Factor +", 0) or 0)
    crit_factor *= 1.0 + native_vault_active(profile_data, "Critical Factor")
    crit_factor *= 1.0 + native_number(key_bonuses.get("Critical Factor", 0.0))
    expected_crit = 1.0 + crit_chance * max(0.0, crit_factor - 1.0)

    super_chance = 0.0005 * int(workshop.get("Super Critical Chance", 0) or 0) + 0.003 * super_chance_lab
    super_chance += native_vault_active(profile_data, "Super Crit Chance")
    super_chance += native_number(key_bonuses.get("Super Crit Chance"), 0.0)
    super_chance = min(0.75, max(0.0, super_chance))

    super_mult = 1.0 + 0.10 * int(workshop.get("Super Critical Mult", 0) or 0) + 0.10 * super_mult_lab
    super_mult *= 1.0 + 0.01 * int(enhancements.get("Super Crit Multi +", 0) or 0)
    super_mult *= 1.0 + native_vault_active(profile_data, "Super Crit Mult")
    super_mult *= 1.0 + native_number(key_bonuses.get("Super Crit Mult", 0.0))
    expected_super = 1.0 + super_chance * max(0.0, super_mult - 1.0)

    range_dpm_mult = (1.0 + 0.0035 * range_lab) * (1.0 + 0.002 * dpm_lab)
    range_dpm_mult *= 1.0 + 0.01 * int(enhancements.get("Damage / Meter +", 0) or 0)
    range_dpm_mult *= 1.0 + native_vault_active(profile_data, "Damage / Meter")
    range_dpm_mult *= 1.0 + native_number(key_bonuses.get("Damage / Meter"), 0.0)

    rend_chance = min(1.0, 0.003 * int(workshop.get("Rend Armor Chance", 0) or 0))
    rend_mult = 1.0 + 0.01 * int(workshop.get("Rend Armor Mult", 0) or 0) + 0.10 * rend_lab
    rend_mult *= 1.0 + 0.01 * int(enhancements.get("Rend Armor Mult +", 0) or 0)
    rend_expected = 1.0 + rend_chance * max(0.0, rend_mult - 1.0)

    core_scalar = damage_mult * expected_crit * expected_super * range_dpm_mult
    bullet = core_scalar * attack_speed_mult * rend_expected

    uw = state.get("uw", {})
    cl = uw.get("Chain Lightning", {})
    cl_attrs = cl.get("attributes", {})
    cl_owned = bool(cl.get("owned", False))
    cl_damage = native_number(cl_attrs.get("Damage"), 0.0)
    cl_quantity = native_number(cl_attrs.get("Quantity"), 0.0)
    cl_chance = native_number(cl_attrs.get("Chance"), 0.0)
    uw_damage_mult = 1.0 + native_vault_active(profile_data, "Ultimate Weapon Damage")
    uw_damage_mult *= 1.0 + native_number(key_bonuses.get("Ultimate Weapon Damage"), 0.0)
    cl_component = (core_scalar * cl_damage * cl_quantity * cl_chance * settings["cl_weight"] * uw_damage_mult) if cl_owned else 0.0

    dw = uw.get("Death Wave", {})
    dw_attrs = dw.get("attributes", {})
    dw_owned = bool(dw.get("owned", False))
    dw_damage = native_number(dw_attrs.get("Damage"), 0.0)
    dw_quantity = native_number(dw_attrs.get("Quantity"), 0.0)
    dw_cooldown = max(1.0, native_number(dw_attrs.get("Cooldown"), 300.0))
    dw_amp = 1.0 + 0.05 * int(labs.get("Death Wave Damage Amplifier", 0) or 0)
    dw_component = (core_scalar * dw_damage * dw_quantity * (300.0 / dw_cooldown) * settings["dw_weight"] * uw_damage_mult * dw_amp) if dw_owned else 0.0

    sm = uw.get("Smart Missiles", {})
    sm_attrs = sm.get("attributes", {})
    sm_owned = bool(sm.get("owned", False))
    sm_damage = native_number(sm_attrs.get("Damage"), 0.0)
    sm_quantity = native_number(sm_attrs.get("Quantity"), 0.0)
    sm_cooldown = max(1.0, native_number(sm_attrs.get("Cooldown"), 180.0))
    sm_amp = 1.0 + 0.25 * int(labs.get("Missile Amplifier", 0) or 0)
    sm_radius = 1.0 + 0.02 * int(labs.get("Missile Radius", 0) or 0)
    sm_component = (core_scalar * sm_damage * sm_quantity * (180.0 / sm_cooldown) * settings["sm_weight"] * uw_damage_mult * sm_amp * sm_radius) if sm_owned else 0.0

    bh = uw.get("Black Hole", {})
    bh_owned = bool(bh.get("owned", False))
    bh_damage = 0.002 * int(labs.get("Black Hole Damage", 0) or 0)
    bh_component = settings["bh_weight"] * bh_damage if bh_owned else 0.0

    pre_global = bullet + cl_component + dw_component + sm_component + bh_component

    sl = uw.get("Spotlight", {})
    sl_attrs = sl.get("attributes", {})
    sl_owned = bool(sl.get("owned", False))
    sl_multiplier = native_number(sl_attrs.get("Multiplier"), 1.0)
    sl_angle = native_number(sl_attrs.get("Angle"), 0.0)
    sl_quantity = native_number(sl_attrs.get("Quantity"), 0.0)
    sl_coverage = min(1.0, max(0.0, sl_angle * sl_quantity / 360.0)) if sl_owned else 0.0
    spotlight_factor = 1.0 + sl_coverage * max(0.0, sl_multiplier - 1.0)

    shock_unlocked = int(labs.get("Chain Lightning Shock", 0) or 0) > 0 and cl_owned
    shock_chance = settings["base_shock_chance"] + 0.02 * int(labs.get("Shock Chance", 0) or 0)
    shock_chance = min(1.0, max(0.0, shock_chance))
    shock_multiplier = settings["base_shock_multiplier"] + 0.06 * int(labs.get("Shock Multiplier", 0) or 0)
    if shock_unlocked:
        event_rate = max(0.0, cl_chance * cl_quantity * settings["shock_events_per_cycle"] * shock_chance)
        shock_uptime = 1.0 - pow(2.718281828459045, -event_rate)
    else:
        shock_uptime = 0.0
    shock_factor = 1.0 + shock_uptime * max(0.0, shock_multiplier - 1.0)

    total = max(1e-12, pre_global * spotlight_factor * shock_factor)
    return {
        "Total": total,
        "Bullet": bullet,
        "Chain Lightning": cl_component,
        "Death Wave": dw_component,
        "Smart Missiles": sm_component,
        "Black Hole": bh_component,
        "Spotlight Factor": spotlight_factor,
        "Spotlight Coverage": sl_coverage,
        "Shock Factor": shock_factor,
        "Shock Uptime": shock_uptime,
        "Crit Expected": expected_crit,
        "Super Crit Expected": expected_super,
        "Rend Expected": rend_expected,
    }


def native_damage_score(profile_data: Dict[str, Any], state: Optional[Dict[str, Any]] = None) -> float:
    return native_damage_components(profile_data, state)["Total"]


def native_damage_lab_allowed(profile_data: Dict[str, Any], name: str) -> tuple[bool, str]:
    uw = profile_data.get("uw", {})
    if name in {"Shock Chance", "Shock Multiplier"}:
        if not bool(uw.get("Chain Lightning", {}).get("owned", False)):
            return False, "Requires Chain Lightning."
        if int(profile_data.get("labs", {}).get("Chain Lightning Shock", 0) or 0) < 1:
            return False, "Requires Chain Lightning Shock."
    if name in {"Missile Amplifier", "Missile Radius"} and not bool(uw.get("Smart Missiles", {}).get("owned", False)):
        return False, "Requires Smart Missiles."
    if name == "Death Wave Damage Amplifier" and not bool(uw.get("Death Wave", {}).get("owned", False)):
        return False, "Requires Death Wave."
    if name == "Black Hole Damage" and not bool(uw.get("Black Hole", {}).get("owned", False)):
        return False, "Requires Black Hole."
    return True, ""


def native_damage_lab_candidates(profile_data: Dict[str, Any], state: Dict[str, Any]) -> list[Dict[str, Any]]:
    before = native_damage_score(profile_data, state)
    labs = state.get("labs", {})
    speed = native_lab_speed_multiplier(profile_data, labs)
    coin_mult = native_lab_coin_multiplier(labs)
    rows = []
    for name, table in NATIVE_DAMAGE_LAB_TABLES.items():
        allowed, reason = native_damage_lab_allowed(profile_data, name)
        if not allowed:
            continue
        current = int(labs.get(name, 0) or 0)
        next_level = current + 1
        record = table.get(next_level)
        if not record or next_level > LAB_MAX_LEVELS.get(name, next_level):
            continue
        if profile_data.get("maxed", {}).get("labs", {}).get(name, False):
            continue
        next_state = json.loads(json.dumps(state))
        next_state["labs"][name] = next_level
        after = native_damage_score(profile_data, next_state)
        relative_gain = after / before - 1.0
        duration = native_number(record.get("seconds"), 0.0) / max(speed, 1e-12)
        roi = relative_gain / max(duration / 86400.0, 1e-12)
        rows.append({
            "Upgrade": name,
            "Level": next_level,
            "Cost Numeric": native_number(record.get("cost"), 0.0) * coin_mult,
            "Cost": format_large_number(native_number(record.get("cost"), 0.0) * coin_mult),
            "Duration Seconds": duration,
            "Duration": native_format_duration(duration),
            "Relative Gain": relative_gain,
            "Gain %": relative_gain * 100.0,
            "ROI Numeric": roi,
            "ROI": roi,
            "Result Numeric": after,
            "Result": f"x{after:.6f}",
            "Confidence": NATIVE_DAMAGE_LAB_EFFECTS.get(name, {}).get("confidence", "Medium"),
            "Why": f"Estimated {relative_gain * 100:.3f}% effective-damage gain in {native_format_duration(duration)}.",
        })
    return sorted(rows, key=lambda row: row["ROI Numeric"], reverse=True)


def build_native_damage_lab_path(profile_data: Dict[str, Any], steps: int = 50) -> list[Dict[str, Any]]:
    state = native_damage_state(profile_data)
    cumulative_seconds = 0.0
    rows = []
    for rank in range(1, steps + 1):
        candidates = native_damage_lab_candidates(profile_data, state)
        if not candidates:
            break
        chosen = dict(candidates[0])
        state["labs"][chosen["Upgrade"]] = int(chosen["Level"])
        cumulative_seconds += chosen["Duration Seconds"]
        chosen.update({"Rank": rank, "Path": "native_damage_lab", "Resource": "Time", "Cumulative": native_format_duration(cumulative_seconds)})
        rows.append(chosen)
    return rows


def native_damage_stone_candidates(profile_data: Dict[str, Any], state: Dict[str, Any]) -> list[Dict[str, Any]]:
    before = native_damage_score(profile_data, state)
    rows = []
    mapping = {
        "CL | Damage": ("Chain Lightning", "Damage"), "CL | Quantity": ("Chain Lightning", "Quantity"), "CL | Chance": ("Chain Lightning", "Chance"),
        "SM | Damage": ("Smart Missiles", "Damage"), "SM | Quantity": ("Smart Missiles", "Quantity"), "SM | Cooldown": ("Smart Missiles", "Cooldown"),
        "DW | Damage": ("Death Wave", "Damage"), "DW | Quantity": ("Death Wave", "Quantity"), "DW | Cooldown": ("Death Wave", "Cooldown"),
        "BH | Size": ("Black Hole", "Size"), "BH | Duration": ("Black Hole", "Duration"), "BH | Cooldown": ("Black Hole", "Cooldown"),
        "SL | Multiplier": ("Spotlight", "Multiplier"), "SL | Angle": ("Spotlight", "Angle"), "SL | Quantity": ("Spotlight", "Quantity"),
    }
    damage_relevant = {"CL | Damage", "CL | Quantity", "CL | Chance", "SM | Damage", "SM | Quantity", "SM | Cooldown", "DW | Damage", "DW | Quantity", "DW | Cooldown", "SL | Multiplier", "SL | Angle", "SL | Quantity"}
    if int(state.get("labs", {}).get("Black Hole Damage", 0) or 0) > 0:
        damage_relevant.update({"BH | Size", "BH | Duration", "BH | Cooldown"})
    for upgrade in damage_relevant:
        uw_name, attribute = mapping[upgrade]
        uw_record = state.get("uw", {}).get(uw_name, {})
        if not bool(uw_record.get("owned", False)):
            continue
        table = NATIVE_DAMAGE_UW_TABLES.get(upgrade, [])
        current = native_number(uw_record.get("attributes", {}).get(attribute), 0.0)
        index = native_table_index(table, current)
        next_index = index + 1
        if index < 0 or next_index >= len(table):
            continue
        record = table[next_index]
        cost = native_number(record.get("cost"), 0.0)
        if cost <= 0:
            continue
        next_state = json.loads(json.dumps(state))
        next_state["uw"].setdefault(uw_name, {"owned": True, "attributes": {}})["attributes"][attribute] = record["value"]
        after = native_damage_score(profile_data, next_state)
        relative_gain = after / before - 1.0
        if relative_gain <= 0:
            continue
        rows.append({
            "Upgrade": upgrade,
            "Level": f"lvl {next_index}",
            "Value": record["value"],
            "Cost Numeric": cost,
            "Cost": cost,
            "Relative Gain": relative_gain,
            "Gain %": relative_gain * 100.0,
            "ROI Numeric": relative_gain / cost,
            "ROI": relative_gain / cost,
            "Result Numeric": after,
            "Result": f"x{after:.6f}",
            "Confidence": "High tables / Medium model",
            "Why": f"Exact UW value/cost step; estimated {relative_gain * 100:.3f}% eDamage gain.",
        })
    return sorted(rows, key=lambda row: row["ROI Numeric"], reverse=True)


def build_native_damage_stone_path(profile_data: Dict[str, Any], steps: int = 50) -> list[Dict[str, Any]]:
    state = native_damage_state(profile_data)
    cumulative = 0.0
    rows = []
    for rank in range(1, steps + 1):
        candidates = native_damage_stone_candidates(profile_data, state)
        if not candidates:
            break
        chosen = dict(candidates[0])
        uw_name, attribute = {
            "CL | Damage": ("Chain Lightning", "Damage"), "CL | Quantity": ("Chain Lightning", "Quantity"), "CL | Chance": ("Chain Lightning", "Chance"),
            "SM | Damage": ("Smart Missiles", "Damage"), "SM | Quantity": ("Smart Missiles", "Quantity"), "SM | Cooldown": ("Smart Missiles", "Cooldown"),
            "DW | Damage": ("Death Wave", "Damage"), "DW | Quantity": ("Death Wave", "Quantity"), "DW | Cooldown": ("Death Wave", "Cooldown"),
            "BH | Size": ("Black Hole", "Size"), "BH | Duration": ("Black Hole", "Duration"), "BH | Cooldown": ("Black Hole", "Cooldown"),
            "SL | Multiplier": ("Spotlight", "Multiplier"), "SL | Angle": ("Spotlight", "Angle"), "SL | Quantity": ("Spotlight", "Quantity"),
        }[chosen["Upgrade"]]
        state["uw"][uw_name]["attributes"][attribute] = chosen["Value"]
        cumulative += chosen["Cost Numeric"]
        chosen.update({"Rank": rank, "Path": "native_damage_stone", "Resource": "Stones", "Cumulative": f"{cumulative:,.0f}"})
        rows.append(chosen)
    return rows


def native_damage_enhancement_apply(state: Dict[str, Any], name: str, level: int) -> None:
    state.setdefault("enhancements", {})[name] = level


def build_native_damage_coin_path(profile_data: Dict[str, Any], steps: int = 50) -> list[Dict[str, Any]]:
    state = native_damage_state(profile_data)
    candidates_names = ["Damage +", "Critical Factor +", "Damage / Meter +", "Super Crit Multi +", "Attack Speed +", "Rend Armor Mult +"]
    costs = NATIVE_ECON_ENHANCEMENT_COSTS.get("Coin Bonus +", [])
    discount_level = int(profile_data.get("labs", {}).get("Enhancement Attack - Coin Discount", 0) or 0)
    discount = max(0.05, 1.0 - 0.01 * discount_level)
    cumulative = 0.0
    rows = []
    for rank in range(1, steps + 1):
        before = native_damage_score(profile_data, state)
        candidates = []
        for name in candidates_names:
            current = int(state.get("enhancements", {}).get(name, 0) or 0)
            next_level = current + 1
            if next_level >= len(costs) or next_level > ENHANCEMENT_MAX_LEVELS.get(name, 200):
                continue
            cost = native_number(costs[next_level], 0.0) * discount
            if cost <= 0:
                continue
            next_state = json.loads(json.dumps(state))
            native_damage_enhancement_apply(next_state, name, next_level)
            after = native_damage_score(profile_data, next_state)
            relative_gain = after / before - 1.0
            candidates.append({
                "Upgrade": name, "Level": next_level, "Cost Numeric": cost, "Cost": format_large_number(cost),
                "Relative Gain": relative_gain, "Gain %": relative_gain * 100.0,
                "ROI Numeric": relative_gain / max(cost / 1e9, 1e-12), "ROI": relative_gain / max(cost / 1e9, 1e-12),
                "Result Numeric": after, "Result": f"x{after:.6f}", "Confidence": "Medium",
                "Why": f"Estimated {relative_gain * 100:.3f}% damage gain per enhancement step.",
            })
        if not candidates:
            break
        chosen = max(candidates, key=lambda row: row["ROI Numeric"])
        state["enhancements"][chosen["Upgrade"]] = int(chosen["Level"])
        cumulative += chosen["Cost Numeric"]
        chosen.update({"Rank": rank, "Path": "native_damage_coin", "Resource": "Coins", "Cumulative": format_large_number(cumulative)})
        rows.append(chosen)
    return rows


def build_native_damage_key_path(profile_data: Dict[str, Any], steps: int = 25) -> list[Dict[str, Any]]:
    state = native_damage_state(profile_data)
    cumulative = 0.0
    rows = []
    used = set()
    for rank in range(1, min(steps, len(NATIVE_DAMAGE_KEY_UPGRADES)) + 1):
        before = native_damage_score(profile_data, state)
        candidates = []
        for item in NATIVE_DAMAGE_KEY_UPGRADES:
            name = item["Upgrade"]
            if name in used:
                continue
            next_state = json.loads(json.dumps(state))
            key = next_state.setdefault("key_bonuses", {})
            gain = native_number(item["Gain"])
            mode = item["Mode"]
            if mode in {"mult", "uw_mult", "crit_factor_mult", "super_mult_mult"}:
                key[name] = native_number(key.get(name), 0.0) + gain
            elif mode in {"crit_chance", "super_chance"}:
                key[name] = native_number(key.get(name), 0.0) + gain
            after = native_damage_score(profile_data, next_state)
            relative_gain = after / before - 1.0
            cost = native_number(item["Cost"])
            candidates.append({
                "Upgrade": name, "Level": 1, "Cost Numeric": cost, "Cost": cost,
                "Relative Gain": relative_gain, "Gain %": relative_gain * 100.0,
                "ROI Numeric": relative_gain / cost, "ROI": relative_gain / cost,
                "Result Numeric": after, "Result": f"x{after:.6f}", "Confidence": "Medium-Low",
                "Why": f"{item['Notes']} Estimated {relative_gain * 100:.3f}% eDamage gain.", "Next State": next_state,
            })
        if not candidates:
            break
        chosen = max(candidates, key=lambda row: row["ROI Numeric"])
        state = chosen.pop("Next State")
        used.add(chosen["Upgrade"])
        cumulative += chosen["Cost Numeric"]
        chosen.update({"Rank": rank, "Path": "native_damage_key", "Resource": "Keys", "Cumulative": cumulative})
        rows.append(chosen)
    return rows


def build_native_damage_paths(profile_data: Dict[str, Any], steps: int = 50) -> Dict[str, list[Dict[str, Any]]]:
    result = {
        "damage_lab": build_native_damage_lab_path(profile_data, steps),
        "damage_stone": build_native_damage_stone_path(profile_data, steps),
        "damage_coin": build_native_damage_coin_path(profile_data, steps),
        "damage_key": build_native_damage_key_path(profile_data, min(steps, 25)),
    }
    profile_data.setdefault("native_damage", {})["last_run"] = {
        "at": datetime.now(timezone.utc).isoformat(), "steps": steps,
        "coverage": {
            "damage_lab": "Exact lab cost/time tables; relative damage model is regression-calibrated.",
            "damage_stone": "Exact UW value/cost tables; CL calibration high confidence for the supplied profile.",
            "damage_coin": "Exact enhancement cost curve; effect model medium confidence.",
            "damage_key": "First-tier Power vault effects only; prerequisite graph not yet enforced.",
        },
    }
    return result



# -----------------------------------------------------------------------------
# NATIVE HEALTH / REGEN ENGINE (v1.2)
# -----------------------------------------------------------------------------
# Cost and duration tables are embedded from Laboratory v3.0.5.  The score is a
# relative comparison model: it is designed to rank the next upgrade, not claim
# an exact displayed HP value or a guaranteed wave result.

NATIVE_HEALTH_LAB_TABLES = {'Health': {1: {'seconds': 14, 'cost': 30.0},
            2: {'seconds': 384, 'cost': 71.0},
            3: {'seconds': 984, 'cost': 178.0},
            4: {'seconds': 1892, 'cost': 398.0},
            5: {'seconds': 3165, 'cost': 772.0},
            6: {'seconds': 4800, 'cost': 1340.0},
            7: {'seconds': 6960, 'cost': 2120.0},
            8: {'seconds': 9540, 'cost': 3170.0},
            9: {'seconds': 12660, 'cost': 4510.0},
            10: {'seconds': 16320, 'cost': 6170.0},
            11: {'seconds': 20580, 'cost': 8170.0},
            12: {'seconds': 25380, 'cost': 10560.0},
            13: {'seconds': 30840, 'cost': 13350.0},
            14: {'seconds': 36900, 'cost': 16580.0},
            15: {'seconds': 43620, 'cost': 20270.0},
            16: {'seconds': 51000, 'cost': 24440.0},
            17: {'seconds': 59100, 'cost': 29130.0},
            18: {'seconds': 67920, 'cost': 34360.0},
            19: {'seconds': 77460, 'cost': 40160.0},
            20: {'seconds': 87720, 'cost': 46540.0},
            21: {'seconds': 98760, 'cost': 53530.0},
            22: {'seconds': 110640, 'cost': 61160.0},
            23: {'seconds': 123240, 'cost': 69460.0},
            24: {'seconds': 136680, 'cost': 78430.0},
            25: {'seconds': 150960, 'cost': 88120.0},
            26: {'seconds': 166020, 'cost': 98530.0},
            27: {'seconds': 181980, 'cost': 109700.0},
            28: {'seconds': 198780, 'cost': 121650.0},
            29: {'seconds': 216480, 'cost': 134390.0},
            30: {'seconds': 235080, 'cost': 147950.0},
            31: {'seconds': 254580, 'cost': 162350.0},
            32: {'seconds': 274980, 'cost': 177620.0},
            33: {'seconds': 296280, 'cost': 193780.0},
            34: {'seconds': 318600, 'cost': 210830.0},
            35: {'seconds': 341820, 'cost': 228820.0},
            36: {'seconds': 366000, 'cost': 247760.0},
            37: {'seconds': 391200, 'cost': 267660.0},
            38: {'seconds': 417300, 'cost': 288560.0},
            39: {'seconds': 444480, 'cost': 310470.0},
            40: {'seconds': 472620, 'cost': 333400.0},
            41: {'seconds': 501840, 'cost': 357390.0},
            42: {'seconds': 532020, 'cost': 382450.0},
            43: {'seconds': 563280, 'cost': 408600.0},
            44: {'seconds': 595560, 'cost': 435870.0},
            45: {'seconds': 628920, 'cost': 464260.0},
            46: {'seconds': 663300, 'cost': 493810.0},
            47: {'seconds': 698820, 'cost': 524530.0},
            48: {'seconds': 735420, 'cost': 556430.0},
            49: {'seconds': 773100, 'cost': 589550.0},
            50: {'seconds': 811860, 'cost': 623890.0},
            51: {'seconds': 851820, 'cost': 659490.0},
            52: {'seconds': 892800, 'cost': 696340.0},
            53: {'seconds': 934980, 'cost': 734490.0},
            54: {'seconds': 978300, 'cost': 773940.0},
            55: {'seconds': 1022760, 'cost': 814710.0},
            56: {'seconds': 1068360, 'cost': 856830.0},
            57: {'seconds': 1115160, 'cost': 900300.0},
            58: {'seconds': 1163100, 'cost': 945160.0},
            59: {'seconds': 1212240, 'cost': 991410.0},
            60: {'seconds': 1262580, 'cost': 1040000.0},
            61: {'seconds': 1314180, 'cost': 1090000.0},
            62: {'seconds': 1366920, 'cost': 1140000.0},
            63: {'seconds': 1420860, 'cost': 1190000.0},
            64: {'seconds': 1476060, 'cost': 1240000.0},
            65: {'seconds': 1532520, 'cost': 1300000.0},
            66: {'seconds': 1590180, 'cost': 1360000.0},
            67: {'seconds': 1649100, 'cost': 1410000.0},
            68: {'seconds': 1709280, 'cost': 1470000.0},
            69: {'seconds': 1770720, 'cost': 1530000.0},
            70: {'seconds': 1833420, 'cost': 1600000.0},
            71: {'seconds': 1897440, 'cost': 1660000.0},
            72: {'seconds': 1962720, 'cost': 1730000.0},
            73: {'seconds': 2029260, 'cost': 1800000.0},
            74: {'seconds': 2097120, 'cost': 1870000.0},
            75: {'seconds': 2166360, 'cost': 1940000.0},
            76: {'seconds': 2236800, 'cost': 2010000.0},
            77: {'seconds': 2308680, 'cost': 2080000.0},
            78: {'seconds': 2381820, 'cost': 2160000.0},
            79: {'seconds': 2456280, 'cost': 2240000.0},
            80: {'seconds': 2532120, 'cost': 2320000.0},
            81: {'seconds': 2609280, 'cost': 2400000.0},
            82: {'seconds': 2687820, 'cost': 2480000.0},
            83: {'seconds': 2767740, 'cost': 2570000.0},
            84: {'seconds': 2849040, 'cost': 2650000.0},
            85: {'seconds': 2931660, 'cost': 2740000.0},
            86: {'seconds': 3015720, 'cost': 2830000.0},
            87: {'seconds': 3101160, 'cost': 2930000.0},
            88: {'seconds': 3187980, 'cost': 3020000.0},
            89: {'seconds': 3276180, 'cost': 3120000.0},
            90: {'seconds': 3365820, 'cost': 3220000.0},
            91: {'seconds': 3456900, 'cost': 3320000.0},
            92: {'seconds': 3549360, 'cost': 3420000.0},
            93: {'seconds': 3643260, 'cost': 3520000.0},
            94: {'seconds': 3738600, 'cost': 3630000.0},
            95: {'seconds': 3835380, 'cost': 3740000.0},
            96: {'seconds': 3933600, 'cost': 3850000.0},
            97: {'seconds': 4033320, 'cost': 3960000.0},
            98: {'seconds': 4134420, 'cost': 4070000.0},
            99: {'seconds': 4237020, 'cost': 4190000.0},
            100: {'seconds': 4341120, 'cost': 4310000.0}},
 'Health Regen': {1: {'seconds': 14, 'cost': 30.0},
                  2: {'seconds': 384, 'cost': 71.0},
                  3: {'seconds': 984, 'cost': 178.0},
                  4: {'seconds': 1892, 'cost': 398.0},
                  5: {'seconds': 3165, 'cost': 772.0},
                  6: {'seconds': 4800, 'cost': 1340.0},
                  7: {'seconds': 6960, 'cost': 2120.0},
                  8: {'seconds': 9540, 'cost': 3170.0},
                  9: {'seconds': 12660, 'cost': 4510.0},
                  10: {'seconds': 16320, 'cost': 6170.0},
                  11: {'seconds': 20580, 'cost': 8170.0},
                  12: {'seconds': 25380, 'cost': 10560.0},
                  13: {'seconds': 30840, 'cost': 13350.0},
                  14: {'seconds': 36900, 'cost': 16580.0},
                  15: {'seconds': 43620, 'cost': 20270.0},
                  16: {'seconds': 51000, 'cost': 24440.0},
                  17: {'seconds': 59100, 'cost': 29130.0},
                  18: {'seconds': 67920, 'cost': 34360.0},
                  19: {'seconds': 77460, 'cost': 40160.0},
                  20: {'seconds': 87720, 'cost': 46540.0},
                  21: {'seconds': 98760, 'cost': 53530.0},
                  22: {'seconds': 110640, 'cost': 61160.0},
                  23: {'seconds': 123240, 'cost': 69460.0},
                  24: {'seconds': 136680, 'cost': 78430.0},
                  25: {'seconds': 150960, 'cost': 88120.0},
                  26: {'seconds': 166020, 'cost': 98530.0},
                  27: {'seconds': 181980, 'cost': 109700.0},
                  28: {'seconds': 198780, 'cost': 121650.0},
                  29: {'seconds': 216480, 'cost': 134390.0},
                  30: {'seconds': 235080, 'cost': 147950.0},
                  31: {'seconds': 254580, 'cost': 162350.0},
                  32: {'seconds': 274980, 'cost': 177620.0},
                  33: {'seconds': 296280, 'cost': 193780.0},
                  34: {'seconds': 318600, 'cost': 210830.0},
                  35: {'seconds': 341820, 'cost': 228820.0},
                  36: {'seconds': 366000, 'cost': 247760.0},
                  37: {'seconds': 391200, 'cost': 267660.0},
                  38: {'seconds': 417300, 'cost': 288560.0},
                  39: {'seconds': 444480, 'cost': 310470.0},
                  40: {'seconds': 472620, 'cost': 333400.0},
                  41: {'seconds': 501840, 'cost': 357390.0},
                  42: {'seconds': 532020, 'cost': 382450.0},
                  43: {'seconds': 563280, 'cost': 408600.0},
                  44: {'seconds': 595560, 'cost': 435870.0},
                  45: {'seconds': 628920, 'cost': 464260.0},
                  46: {'seconds': 663300, 'cost': 493810.0},
                  47: {'seconds': 698820, 'cost': 524530.0},
                  48: {'seconds': 735420, 'cost': 556430.0},
                  49: {'seconds': 773100, 'cost': 589550.0},
                  50: {'seconds': 811860, 'cost': 623890.0},
                  51: {'seconds': 851820, 'cost': 659490.0},
                  52: {'seconds': 892800, 'cost': 696340.0},
                  53: {'seconds': 934980, 'cost': 734490.0},
                  54: {'seconds': 978300, 'cost': 773940.0},
                  55: {'seconds': 1022760, 'cost': 814710.0},
                  56: {'seconds': 1068360, 'cost': 856830.0},
                  57: {'seconds': 1115160, 'cost': 900300.0},
                  58: {'seconds': 1163100, 'cost': 945160.0},
                  59: {'seconds': 1212240, 'cost': 991410.0},
                  60: {'seconds': 1262580, 'cost': 1040000.0},
                  61: {'seconds': 1314180, 'cost': 1090000.0},
                  62: {'seconds': 1366920, 'cost': 1140000.0},
                  63: {'seconds': 1420860, 'cost': 1190000.0},
                  64: {'seconds': 1476060, 'cost': 1240000.0},
                  65: {'seconds': 1532520, 'cost': 1300000.0},
                  66: {'seconds': 1590180, 'cost': 1360000.0},
                  67: {'seconds': 1649100, 'cost': 1410000.0},
                  68: {'seconds': 1709280, 'cost': 1470000.0},
                  69: {'seconds': 1770720, 'cost': 1530000.0},
                  70: {'seconds': 1833420, 'cost': 1600000.0},
                  71: {'seconds': 1897440, 'cost': 1660000.0},
                  72: {'seconds': 1962720, 'cost': 1730000.0},
                  73: {'seconds': 2029260, 'cost': 1800000.0},
                  74: {'seconds': 2097120, 'cost': 1870000.0},
                  75: {'seconds': 2166360, 'cost': 1940000.0},
                  76: {'seconds': 2236800, 'cost': 2010000.0},
                  77: {'seconds': 2308680, 'cost': 2080000.0},
                  78: {'seconds': 2381820, 'cost': 2160000.0},
                  79: {'seconds': 2456280, 'cost': 2240000.0},
                  80: {'seconds': 2532120, 'cost': 2320000.0},
                  81: {'seconds': 2609280, 'cost': 2400000.0},
                  82: {'seconds': 2687820, 'cost': 2480000.0},
                  83: {'seconds': 2767740, 'cost': 2570000.0},
                  84: {'seconds': 2849040, 'cost': 2650000.0},
                  85: {'seconds': 2931660, 'cost': 2740000.0},
                  86: {'seconds': 3015720, 'cost': 2830000.0},
                  87: {'seconds': 3101160, 'cost': 2930000.0},
                  88: {'seconds': 3187980, 'cost': 3020000.0},
                  89: {'seconds': 3276180, 'cost': 3120000.0},
                  90: {'seconds': 3365820, 'cost': 3220000.0},
                  91: {'seconds': 3456900, 'cost': 3320000.0},
                  92: {'seconds': 3549360, 'cost': 3420000.0},
                  93: {'seconds': 3643260, 'cost': 3520000.0},
                  94: {'seconds': 3738600, 'cost': 3630000.0},
                  95: {'seconds': 3835380, 'cost': 3740000.0},
                  96: {'seconds': 3933600, 'cost': 3850000.0},
                  97: {'seconds': 4033320, 'cost': 3960000.0},
                  98: {'seconds': 4134420, 'cost': 4070000.0},
                  99: {'seconds': 4237020, 'cost': 4190000.0},
                  100: {'seconds': 4341120, 'cost': 4310000.0}},
 'Defense Absolute': {1: {'seconds': 14, 'cost': 30.0},
                      2: {'seconds': 384, 'cost': 71.0},
                      3: {'seconds': 984, 'cost': 178.0},
                      4: {'seconds': 1892, 'cost': 398.0},
                      5: {'seconds': 3165, 'cost': 772.0},
                      6: {'seconds': 4800, 'cost': 1340.0},
                      7: {'seconds': 6960, 'cost': 2120.0},
                      8: {'seconds': 9540, 'cost': 3170.0},
                      9: {'seconds': 12660, 'cost': 4510.0},
                      10: {'seconds': 16320, 'cost': 6170.0},
                      11: {'seconds': 20580, 'cost': 8170.0},
                      12: {'seconds': 25380, 'cost': 10560.0},
                      13: {'seconds': 30840, 'cost': 13350.0},
                      14: {'seconds': 36900, 'cost': 16580.0},
                      15: {'seconds': 43620, 'cost': 20270.0},
                      16: {'seconds': 51000, 'cost': 24440.0},
                      17: {'seconds': 59100, 'cost': 29130.0},
                      18: {'seconds': 67920, 'cost': 34360.0},
                      19: {'seconds': 77460, 'cost': 40160.0},
                      20: {'seconds': 87720, 'cost': 46540.0},
                      21: {'seconds': 98760, 'cost': 53530.0},
                      22: {'seconds': 110640, 'cost': 61160.0},
                      23: {'seconds': 123240, 'cost': 69460.0},
                      24: {'seconds': 136680, 'cost': 78430.0},
                      25: {'seconds': 150960, 'cost': 88120.0},
                      26: {'seconds': 166020, 'cost': 98530.0},
                      27: {'seconds': 181980, 'cost': 109700.0},
                      28: {'seconds': 198780, 'cost': 121650.0},
                      29: {'seconds': 216480, 'cost': 134390.0},
                      30: {'seconds': 235080, 'cost': 147950.0},
                      31: {'seconds': 254580, 'cost': 162350.0},
                      32: {'seconds': 274980, 'cost': 177620.0},
                      33: {'seconds': 296280, 'cost': 193780.0},
                      34: {'seconds': 318600, 'cost': 210830.0},
                      35: {'seconds': 341820, 'cost': 228820.0},
                      36: {'seconds': 366000, 'cost': 247760.0},
                      37: {'seconds': 391200, 'cost': 267660.0},
                      38: {'seconds': 417300, 'cost': 288560.0},
                      39: {'seconds': 444480, 'cost': 310470.0},
                      40: {'seconds': 472620, 'cost': 333400.0},
                      41: {'seconds': 501840, 'cost': 357390.0},
                      42: {'seconds': 532020, 'cost': 382450.0},
                      43: {'seconds': 563280, 'cost': 408600.0},
                      44: {'seconds': 595560, 'cost': 435870.0},
                      45: {'seconds': 628920, 'cost': 464260.0},
                      46: {'seconds': 663300, 'cost': 493810.0},
                      47: {'seconds': 698820, 'cost': 524530.0},
                      48: {'seconds': 735420, 'cost': 556430.0},
                      49: {'seconds': 773100, 'cost': 589550.0},
                      50: {'seconds': 811860, 'cost': 623890.0},
                      51: {'seconds': 851820, 'cost': 659490.0},
                      52: {'seconds': 892800, 'cost': 696340.0},
                      53: {'seconds': 934980, 'cost': 734490.0},
                      54: {'seconds': 978300, 'cost': 773940.0},
                      55: {'seconds': 1022760, 'cost': 814710.0},
                      56: {'seconds': 1068360, 'cost': 856830.0},
                      57: {'seconds': 1115160, 'cost': 900300.0},
                      58: {'seconds': 1163100, 'cost': 945160.0},
                      59: {'seconds': 1212240, 'cost': 991410.0},
                      60: {'seconds': 1262580, 'cost': 1040000.0},
                      61: {'seconds': 1314180, 'cost': 1090000.0},
                      62: {'seconds': 1366920, 'cost': 1140000.0},
                      63: {'seconds': 1420860, 'cost': 1190000.0},
                      64: {'seconds': 1476060, 'cost': 1240000.0},
                      65: {'seconds': 1532520, 'cost': 1300000.0},
                      66: {'seconds': 1590180, 'cost': 1360000.0},
                      67: {'seconds': 1649100, 'cost': 1410000.0},
                      68: {'seconds': 1709280, 'cost': 1470000.0},
                      69: {'seconds': 1770720, 'cost': 1530000.0},
                      70: {'seconds': 1833420, 'cost': 1600000.0},
                      71: {'seconds': 1897440, 'cost': 1660000.0},
                      72: {'seconds': 1962720, 'cost': 1730000.0},
                      73: {'seconds': 2029260, 'cost': 1800000.0},
                      74: {'seconds': 2097120, 'cost': 1870000.0},
                      75: {'seconds': 2166360, 'cost': 1940000.0},
                      76: {'seconds': 2236800, 'cost': 2010000.0},
                      77: {'seconds': 2308680, 'cost': 2080000.0},
                      78: {'seconds': 2381820, 'cost': 2160000.0},
                      79: {'seconds': 2456280, 'cost': 2240000.0},
                      80: {'seconds': 2532120, 'cost': 2320000.0},
                      81: {'seconds': 2609280, 'cost': 2400000.0},
                      82: {'seconds': 2687820, 'cost': 2480000.0},
                      83: {'seconds': 2767740, 'cost': 2570000.0},
                      84: {'seconds': 2849040, 'cost': 2650000.0},
                      85: {'seconds': 2931660, 'cost': 2740000.0},
                      86: {'seconds': 3015720, 'cost': 2830000.0},
                      87: {'seconds': 3101160, 'cost': 2930000.0},
                      88: {'seconds': 3187980, 'cost': 3020000.0},
                      89: {'seconds': 3276180, 'cost': 3120000.0},
                      90: {'seconds': 3365820, 'cost': 3220000.0},
                      91: {'seconds': 3456900, 'cost': 3320000.0},
                      92: {'seconds': 3549360, 'cost': 3420000.0},
                      93: {'seconds': 3643260, 'cost': 3520000.0},
                      94: {'seconds': 3738600, 'cost': 3630000.0},
                      95: {'seconds': 3835380, 'cost': 3740000.0},
                      96: {'seconds': 3933600, 'cost': 3850000.0},
                      97: {'seconds': 4033320, 'cost': 3960000.0},
                      98: {'seconds': 4134420, 'cost': 4070000.0},
                      99: {'seconds': 4237020, 'cost': 4190000.0},
                      100: {'seconds': 4341120, 'cost': 4310000.0}},
 'Defense %': {1: {'seconds': 3599, 'cost': 5000.0},
               2: {'seconds': 6660, 'cost': 7500.0},
               3: {'seconds': 10260, 'cost': 24000.0},
               4: {'seconds': 15180, 'cost': 90500.0},
               5: {'seconds': 22380, 'cost': 267000.0},
               6: {'seconds': 32760, 'cost': 637500.0},
               7: {'seconds': 47460, 'cost': 1310000.0},
               8: {'seconds': 67620, 'cost': 2420000.0},
               9: {'seconds': 94440, 'cost': 4110000.0},
               10: {'seconds': 129240, 'cost': 6580000.0},
               11: {'seconds': 173220, 'cost': 10020000.0},
               12: {'seconds': 227880, 'cost': 14660000.0},
               13: {'seconds': 294480, 'cost': 20760000.0},
               14: {'seconds': 374520, 'cost': 28590000.0},
               15: {'seconds': 469500, 'cost': 38440000.0},
               16: {'seconds': 580920, 'cost': 50650000.0},
               17: {'seconds': 710280, 'cost': 65570000.0},
               18: {'seconds': 859140, 'cost': 83550000.0},
               19: {'seconds': 1029180, 'cost': 105010000.0},
               20: {'seconds': 1221960, 'cost': 130350000.0},
               21: {'seconds': 1439160, 'cost': 160040000.0},
               22: {'seconds': 1682520, 'cost': 194520000.0},
               23: {'seconds': 1953600, 'cost': 234290000.0},
               24: {'seconds': 2254320, 'cost': 279880000.0},
               25: {'seconds': 2586300, 'cost': 331820000.0},
               26: {'seconds': 2951340, 'cost': 390670000.0},
               27: {'seconds': 3351300, 'cost': 457020000.0},
               28: {'seconds': 3787980, 'cost': 531490000.0},
               29: {'seconds': 4263180, 'cost': 614700000.0},
               30: {'seconds': 4778820, 'cost': 707330000.0},
               31: {'seconds': 5336820, 'cost': 810050000.0},
               32: {'seconds': 5938980, 'cost': 923570000.0},
               33: {'seconds': 6587280, 'cost': 1050000000.0},
               34: {'seconds': 7283700, 'cost': 1190000000.0},
               35: {'seconds': 8030160, 'cost': 1340000000.0},
               36: {'seconds': 8828700, 'cost': 1500000000.0},
               37: {'seconds': 9681240, 'cost': 1680000000.0},
               38: {'seconds': 10589820, 'cost': 1870000000.0},
               39: {'seconds': 11556480, 'cost': 2090000000.0},
               40: {'seconds': 12529260, 'cost': 2310000000.0},
               41: {'seconds': 13672200, 'cost': 2560000000.0},
               42: {'seconds': 14825460, 'cost': 2830000000.0},
               43: {'seconds': 16045080, 'cost': 3110000000.0},
               44: {'seconds': 17333160, 'cost': 3420000000.0},
               45: {'seconds': 18691860, 'cost': 3750000000.0},
               46: {'seconds': 20087340, 'cost': 4100000000.0},
               47: {'seconds': 21629700, 'cost': 4480000000.0},
               48: {'seconds': 23213160, 'cost': 4880000000.0},
               49: {'seconds': 24875880, 'cost': 5310000000.0},
               50: {'seconds': 26620020, 'cost': 5760000000.0}},
 'Wall Health': {1: {'seconds': 19980, 'cost': 1000000000.0},
                 2: {'seconds': 20220, 'cost': 1200000000.0},
                 3: {'seconds': 20760, 'cost': 1400000000.0},
                 4: {'seconds': 21540, 'cost': 1600000000.0},
                 5: {'seconds': 22740, 'cost': 1810000000.0},
                 6: {'seconds': 24300, 'cost': 2020000000.0},
                 7: {'seconds': 26340, 'cost': 2240000000.0},
                 8: {'seconds': 28860, 'cost': 2460000000.0},
                 9: {'seconds': 31860, 'cost': 2700000000.0},
                 10: {'seconds': 35400, 'cost': 2940000000.0},
                 11: {'seconds': 39540, 'cost': 3200000000.0},
                 12: {'seconds': 44280, 'cost': 3470000000.0},
                 13: {'seconds': 49620, 'cost': 3760000000.0},
                 14: {'seconds': 55560, 'cost': 4070000000.0},
                 15: {'seconds': 62220, 'cost': 4410000000.0},
                 16: {'seconds': 69480, 'cost': 4760000000.0},
                 17: {'seconds': 77520, 'cost': 5140000000.0},
                 18: {'seconds': 86220, 'cost': 5550000000.0},
                 19: {'seconds': 95640, 'cost': 5990000000.0},
                 20: {'seconds': 105840, 'cost': 6460000000.0},
                 21: {'seconds': 116760, 'cost': 6970000000.0},
                 22: {'seconds': 128520, 'cost': 7510000000.0},
                 23: {'seconds': 141000, 'cost': 8090000000.0},
                 24: {'seconds': 154380, 'cost': 8720000000.0},
                 25: {'seconds': 168540, 'cost': 9390000000.0},
                 26: {'seconds': 183540, 'cost': 10100000000.0},
                 27: {'seconds': 199380, 'cost': 10870000000.0},
                 28: {'seconds': 216060, 'cost': 11690000000.0},
                 29: {'seconds': 233700, 'cost': 12570000000.0},
                 30: {'seconds': 252180, 'cost': 13500000000.0},
                 31: {'seconds': 271560, 'cost': 14490000000.0},
                 32: {'seconds': 291840, 'cost': 15550000000.0},
                 33: {'seconds': 313080, 'cost': 16670000000.0},
                 34: {'seconds': 335280, 'cost': 17860000000.0},
                 35: {'seconds': 358380, 'cost': 19120000000.0},
                 36: {'seconds': 382500, 'cost': 20460000000.0},
                 37: {'seconds': 407580, 'cost': 21870000000.0},
                 38: {'seconds': 433620, 'cost': 23360000000.0},
                 39: {'seconds': 460680, 'cost': 24940000000.0},
                 40: {'seconds': 488700, 'cost': 26600000000.0},
                 41: {'seconds': 517800, 'cost': 28360000000.0},
                 42: {'seconds': 547920, 'cost': 30200000000.0},
                 43: {'seconds': 579060, 'cost': 32140000000.0},
                 44: {'seconds': 611220, 'cost': 34170000000.0},
                 45: {'seconds': 644520, 'cost': 36310000000.0},
                 46: {'seconds': 678840, 'cost': 38550000000.0},
                 47: {'seconds': 714240, 'cost': 40900000000.0},
                 48: {'seconds': 750720, 'cost': 43360000000.0},
                 49: {'seconds': 788280, 'cost': 45930000000.0},
                 50: {'seconds': 826980, 'cost': 48610000000.0}},
 'Wall Rebuild': {1: {'seconds': 49980, 'cost': 1600000000.0},
                  2: {'seconds': 90060, 'cost': 1910000000.0},
                  3: {'seconds': 130800, 'cost': 2230000000.0},
                  4: {'seconds': 173640, 'cost': 2610000000.0},
                  5: {'seconds': 220260, 'cost': 3040000000.0},
                  6: {'seconds': 272940, 'cost': 3550000000.0},
                  7: {'seconds': 334260, 'cost': 4150000000.0},
                  8: {'seconds': 407160, 'cost': 4860000000.0},
                  9: {'seconds': 494760, 'cost': 5690000000.0},
                  10: {'seconds': 600660, 'cost': 6650000000.0},
                  11: {'seconds': 728640, 'cost': 7750000000.0},
                  12: {'seconds': 882720, 'cost': 9020000000.0},
                  13: {'seconds': 1067160, 'cost': 10460000000.0},
                  14: {'seconds': 1286580, 'cost': 12080000000.0},
                  15: {'seconds': 1545720, 'cost': 13890000000.0},
                  16: {'seconds': 1849560, 'cost': 15920000000.0},
                  17: {'seconds': 2203260, 'cost': 18160000000.0},
                  18: {'seconds': 2612400, 'cost': 20640000000.0},
                  19: {'seconds': 3082440, 'cost': 23360000000.0},
                  20: {'seconds': 3619380, 'cost': 26330000000.0}},
 'Wall Regen': {1: {'seconds': 99960, 'cost': 30000000000.0},
                2: {'seconds': 120060, 'cost': 35100000000.0},
                3: {'seconds': 141360, 'cost': 40350000000.0},
                4: {'seconds': 167880, 'cost': 45720000000.0},
                5: {'seconds': 207120, 'cost': 51210000000.0},
                6: {'seconds': 270900, 'cost': 56810000000.0},
                7: {'seconds': 375240, 'cost': 62520000000.0},
                8: {'seconds': 541260, 'cost': 68320000000.0},
                9: {'seconds': 795000, 'cost': 74220000000.0},
                10: {'seconds': 1167840, 'cost': 80220000000.0},
                11: {'seconds': 1696683, 'cost': 86310000000.0},
                12: {'seconds': 1951186, 'cost': 92490000000.0},
                13: {'seconds': 2243864, 'cost': 98760000000.0},
                14: {'seconds': 2580444, 'cost': 105120000000.0},
                15: {'seconds': 2967510, 'cost': 111560000000.0},
                16: {'seconds': 3412637, 'cost': 118090000000.0},
                17: {'seconds': 3924532, 'cost': 124700000000.0},
                18: {'seconds': 4513212, 'cost': 131400000000.0},
                19: {'seconds': 5190194, 'cost': 138180000000.0},
                20: {'seconds': 5968723, 'cost': 145030000000.0},
                21: {'seconds': 6864032, 'cost': 151970000000.0},
                22: {'seconds': 7893637, 'cost': 158990000000.0},
                23: {'seconds': 9077682, 'cost': 166080000000.0},
                24: {'seconds': 10439335, 'cost': 173260000000.0},
                25: {'seconds': 12005234, 'cost': 180510000000.0},
                26: {'seconds': 13806019, 'cost': 187830000000.0},
                27: {'seconds': 15876921, 'cost': 195230000000.0},
                28: {'seconds': 18258459, 'cost': 202710000000.0},
                29: {'seconds': 20997229, 'cost': 210260000000.0},
                30: {'seconds': 24146811, 'cost': 217890000000.0}},
 'Wall Thorns': {1: {'seconds': 99960, 'cost': 30000000000.0},
                 2: {'seconds': 130020, 'cost': 38100000000.0},
                 3: {'seconds': 161340, 'cost': 46350000000.0},
                 4: {'seconds': 197880, 'cost': 54720000000.0},
                 5: {'seconds': 247140, 'cost': 63210000000.0},
                 6: {'seconds': 320880, 'cost': 71810000000.0},
                 7: {'seconds': 435240, 'cost': 80520000000.0},
                 8: {'seconds': 611280, 'cost': 89320000000.0},
                 9: {'seconds': 874980, 'cost': 98220000000.0},
                 10: {'seconds': 1257840, 'cost': 107220000000.0},
                 11: {'seconds': 1796640, 'cost': 116310000000.0},
                 12: {'seconds': 2534160, 'cost': 125490000000.0},
                 13: {'seconds': 3518940, 'cost': 134760000000.0},
                 14: {'seconds': 4805700, 'cost': 144120000000.0},
                 15: {'seconds': 6455340, 'cost': 153560000000.0},
                 16: {'seconds': 8535240, 'cost': 163090000000.0},
                 17: {'seconds': 11119320, 'cost': 172700000000.0},
                 18: {'seconds': 14288160, 'cost': 182400000000.0},
                 19: {'seconds': 18129180, 'cost': 192180000000.0},
                 20: {'seconds': 22736760, 'cost': 202030000000.0}},
 'Wall Invincibility': {1: {'seconds': 299940, 'cost': 300000000000.0},
                        2: {'seconds': 350040, 'cost': 351000000000.0},
                        3: {'seconds': 401340, 'cost': 403030000000.0},
                        4: {'seconds': 457860, 'cost': 455800000000.0},
                        5: {'seconds': 527160, 'cost': 509190000000.0},
                        6: {'seconds': 620880, 'cost': 563130000000.0},
                        7: {'seconds': 755280, 'cost': 617580000000.0},
                        8: {'seconds': 951300, 'cost': 672500000000.0},
                        9: {'seconds': 1234980, 'cost': 727860000000.0},
                        10: {'seconds': 1637820, 'cost': 783630000000.0}},
 'Wall Fortification': {1: {'seconds': 199999, 'cost': 300000000000.0},
                        2: {'seconds': 209069, 'cost': 300070000000.0},
                        3: {'seconds': 218454, 'cost': 300440000000.0},
                        4: {'seconds': 228359, 'cost': 301410000000.0},
                        5: {'seconds': 238955, 'cost': 303280000000.0},
                        6: {'seconds': 250398, 'cost': 306350000000.0},
                        7: {'seconds': 262832, 'cost': 310920000000.0},
                        8: {'seconds': 276392, 'cost': 317290000000.0},
                        9: {'seconds': 291205, 'cost': 325760000000.0},
                        10: {'seconds': 307396, 'cost': 336620000000.0},
                        11: {'seconds': 325082, 'cost': 350200000000.0},
                        12: {'seconds': 344379, 'cost': 366770000000.0},
                        13: {'seconds': 365396, 'cost': 386640000000.0},
                        14: {'seconds': 388243, 'cost': 410110000000.0},
                        15: {'seconds': 413024, 'cost': 437480000000.0},
                        16: {'seconds': 439844, 'cost': 469050000000.0},
                        17: {'seconds': 468802, 'cost': 505120000000.0},
                        18: {'seconds': 499997, 'cost': 545990000000.0},
                        19: {'seconds': 533527, 'cost': 591960000000.0},
                        20: {'seconds': 569487, 'cost': 643330000000.0},
                        21: {'seconds': 607970, 'cost': 700400000000.0},
                        22: {'seconds': 649069, 'cost': 763470000000.0},
                        23: {'seconds': 692875, 'cost': 832840000000.0},
                        24: {'seconds': 739478, 'cost': 908810000000.0},
                        25: {'seconds': 788965, 'cost': 991680000000.0},
                        26: {'seconds': 841424, 'cost': 1080000000000.0001},
                        27: {'seconds': 896941, 'cost': 1180000000000.0},
                        28: {'seconds': 955601, 'cost': 1280000000000.0},
                        29: {'seconds': 1017489, 'cost': 1400000000000.0},
                        30: {'seconds': 1082687, 'cost': 1520000000000.0},
                        31: {'seconds': 1151279, 'cost': 1650000000000.0},
                        32: {'seconds': 1223345, 'cost': 1790000000000.0},
                        33: {'seconds': 1298966, 'cost': 1940000000000.0},
                        34: {'seconds': 1378222, 'cost': 2100000000000.0},
                        35: {'seconds': 1461192, 'cost': 2270000000000.0},
                        36: {'seconds': 1547955, 'cost': 2440000000000.0},
                        37: {'seconds': 1638589, 'cost': 2630000000000.0},
                        38: {'seconds': 1733169, 'cost': 2830000000000.0},
                        39: {'seconds': 1831774, 'cost': 3040000000000.0},
                        40: {'seconds': 1934478, 'cost': 3270000000000.0},
                        41: {'seconds': 2041358, 'cost': 3500000000000.0},
                        42: {'seconds': 2152486, 'cost': 3750000000000.0},
                        43: {'seconds': 2267939, 'cost': 4010000000000.0},
                        44: {'seconds': 2387789, 'cost': 4280000000000.0005},
                        45: {'seconds': 2512109, 'cost': 4560000000000.0},
                        46: {'seconds': 2640972, 'cost': 4860000000000.0},
                        47: {'seconds': 2774450, 'cost': 5170000000000.0},
                        48: {'seconds': 2912614, 'cost': 5490000000000.0},
                        49: {'seconds': 3055535, 'cost': 5830000000000.0},
                        50: {'seconds': 3203285, 'cost': 6180000000000.0},
                        51: {'seconds': 3355933, 'cost': 6550000000000.0},
                        52: {'seconds': 3513549, 'cost': 6930000000000.0},
                        53: {'seconds': 3676202, 'cost': 7330000000000.0},
                        54: {'seconds': 3843962, 'cost': 7740000000000.0},
                        55: {'seconds': 4016897, 'cost': 8170000000000.0},
                        56: {'seconds': 4195074, 'cost': 8619999999999.999},
                        57: {'seconds': 4378563, 'cost': 9080000000000.0},
                        58: {'seconds': 4567430, 'cost': 9560000000000.0},
                        59: {'seconds': 4761742, 'cost': 10060000000000.0},
                        60: {'seconds': 4961566, 'cost': 10570000000000.0}},
 'Garlic Thorns': {1: {'seconds': 97999, 'cost': 4500.0},
                   2: {'seconds': 108300, 'cost': 7100.0},
                   3: {'seconds': 121748, 'cost': 13300.0},
                   4: {'seconds': 140393, 'cost': 26700.0},
                   5: {'seconds': 165956, 'cost': 50900.0},
                   6: {'seconds': 199990, 'cost': 89500.0},
                   7: {'seconds': 243929, 'cost': 146100.0},
                   8: {'seconds': 299125, 'cost': 224300.0},
                   9: {'seconds': 366861, 'cost': 327700.0},
                   10: {'seconds': 448368, 'cost': 459900.0}},
 'Recovery Package Amount': {1: {'seconds': 199980, 'cost': 20000000000.0},
                             2: {'seconds': 230040, 'cost': 25010000000.0},
                             3: {'seconds': 260820, 'cost': 30070000000.0},
                             4: {'seconds': 293640, 'cost': 35330000000.0},
                             5: {'seconds': 330240, 'cost': 40970000000.0},
                             6: {'seconds': 372960, 'cost': 47260000000.0},
                             7: {'seconds': 424260, 'cost': 54530000000.0},
                             8: {'seconds': 487140, 'cost': 63130000000.0},
                             9: {'seconds': 564780, 'cost': 73510000000.0},
                             10: {'seconds': 660660, 'cost': 86140000000.0},
                             11: {'seconds': 778620, 'cost': 101550000000.0},
                             12: {'seconds': 922740, 'cost': 120320000000.0},
                             13: {'seconds': 1097160, 'cost': 143080000000.0},
                             14: {'seconds': 1306620, 'cost': 170500000000.0},
                             15: {'seconds': 1555740, 'cost': 203310000000.0},
                             16: {'seconds': 1849560, 'cost': 242270000000.0},
                             17: {'seconds': 2193300, 'cost': 288200000000.0},
                             18: {'seconds': 2592360, 'cost': 341960000000.0},
                             19: {'seconds': 3052440, 'cost': 404450000000.0},
                             20: {'seconds': 3579360, 'cost': 476600000000.0}},
 'Recovery Package Max': {1: {'seconds': 199980, 'cost': 20000000000.0},
                          2: {'seconds': 230040, 'cost': 25010000000.0},
                          3: {'seconds': 260820, 'cost': 30070000000.0},
                          4: {'seconds': 293640, 'cost': 30070000000.0},
                          5: {'seconds': 330240, 'cost': 40970000000.0},
                          6: {'seconds': 372960, 'cost': 47260000000.0},
                          7: {'seconds': 424260, 'cost': 54530000000.0},
                          8: {'seconds': 487140, 'cost': 63130000000.0},
                          9: {'seconds': 564780, 'cost': 73510000000.0},
                          10: {'seconds': 660660, 'cost': 86140000000.0},
                          11: {'seconds': 778620, 'cost': 101550000000.0},
                          12: {'seconds': 922740, 'cost': 120320000000.0},
                          13: {'seconds': 1097160, 'cost': 143080000000.0},
                          14: {'seconds': 1306620, 'cost': 170500000000.0},
                          15: {'seconds': 1555740, 'cost': 203310000000.0},
                          16: {'seconds': 1849560, 'cost': 242270000000.0},
                          17: {'seconds': 2193300, 'cost': 288200000000.0},
                          18: {'seconds': 2592360, 'cost': 341960000000.0},
                          19: {'seconds': 3052440, 'cost': 404450000000.0},
                          20: {'seconds': 3579360, 'cost': 476600000000.0}},
 'Recovery Package Chance': {1: {'seconds': 199980, 'cost': 20000000000.0},
                             2: {'seconds': 230040, 'cost': 25010000000.0},
                             3: {'seconds': 260820, 'cost': 30070000000.0},
                             4: {'seconds': 293640, 'cost': 35330000000.0},
                             5: {'seconds': 330240, 'cost': 40970000000.0},
                             6: {'seconds': 372960, 'cost': 47260000000.0},
                             7: {'seconds': 424260, 'cost': 54530000000.0},
                             8: {'seconds': 487140, 'cost': 63130000000.0},
                             9: {'seconds': 564780, 'cost': 73510000000.0},
                             10: {'seconds': 660660, 'cost': 86140000000.0},
                             11: {'seconds': 778620, 'cost': 101550000000.0},
                             12: {'seconds': 922740, 'cost': 120320000000.0},
                             13: {'seconds': 1097160, 'cost': 143080000000.0},
                             14: {'seconds': 1306620, 'cost': 170500000000.0},
                             15: {'seconds': 1555740, 'cost': 203310000000.0},
                             16: {'seconds': 1849560, 'cost': 242270000000.0},
                             17: {'seconds': 2193300, 'cost': 288200000000.0},
                             18: {'seconds': 2592360, 'cost': 341960000000.0},
                             19: {'seconds': 3052440, 'cost': 404450000000.0},
                             20: {'seconds': 3579360, 'cost': 476600000000.0}},
 'Death Wave Health': {1: {'seconds': 71999, 'cost': 250000.0},
                       2: {'seconds': 102069, 'cost': 560000.0},
                       3: {'seconds': 132791, 'cost': 1100000.0},
                       4: {'seconds': 165273, 'cost': 2800000.0},
                       5: {'seconds': 200959, 'cost': 7750000.0},
                       6: {'seconds': 241565, 'cost': 19540000.0},
                       7: {'seconds': 289036, 'cost': 43580000.0},
                       8: {'seconds': 345524, 'cost': 87410000.0},
                       9: {'seconds': 413370, 'cost': 160910000.0},
                       10: {'seconds': 495089, 'cost': 276620000.0},
                       11: {'seconds': 622135, 'cost': 334620000.0},
                       12: {'seconds': 658091, 'cost': 347440000.0},
                       13: {'seconds': 694918, 'cost': 363460000.0},
                       14: {'seconds': 732653, 'cost': 383110000.0},
                       15: {'seconds': 771335, 'cost': 406870000.0},
                       16: {'seconds': 810999, 'cost': 435210000.0},
                       17: {'seconds': 851679, 'cost': 468640000.0},
                       18: {'seconds': 893410, 'cost': 507670000.0},
                       19: {'seconds': 936222, 'cost': 552830000.0},
                       20: {'seconds': 980149, 'cost': 604680000.0},
                       21: {'seconds': 1025219, 'cost': 663770000.0},
                       22: {'seconds': 1071463, 'cost': 730690000.0},
                       23: {'seconds': 1118911, 'cost': 806040000.0},
                       24: {'seconds': 1167589, 'cost': 890410000.0},
                       25: {'seconds': 1217526, 'cost': 984430000.0},
                       26: {'seconds': 1268749, 'cost': 1090000000.0},
                       27: {'seconds': 1321285, 'cost': 1200000000.0},
                       28: {'seconds': 1375159, 'cost': 1330000000.0},
                       29: {'seconds': 1430397, 'cost': 1470000000.0},
                       30: {'seconds': 1487024, 'cost': 1620000000.0}},
 'Common Enemy Attack': {1: {'seconds': 199980, 'cost': 20000000000.0},
                         2: {'seconds': 230040, 'cost': 40000000000.0},
                         3: {'seconds': 260820, 'cost': 60070000000.0},
                         4: {'seconds': 293640, 'cost': 80330000000.0},
                         5: {'seconds': 330240, 'cost': 100970000000.0},
                         6: {'seconds': 372960, 'cost': 122260000000.0},
                         7: {'seconds': 424260, 'cost': 144530000000.0},
                         8: {'seconds': 487140, 'cost': 168130000000.0},
                         9: {'seconds': 564780, 'cost': 193510000000.0},
                         10: {'seconds': 660660, 'cost': 221140000000.0},
                         11: {'seconds': 778620, 'cost': 251550000000.0},
                         12: {'seconds': 922740, 'cost': 285320000000.0},
                         13: {'seconds': 1097160, 'cost': 323080000000.0},
                         14: {'seconds': 1306620, 'cost': 365500000000.0},
                         15: {'seconds': 1555740, 'cost': 413310000000.0},
                         16: {'seconds': 1849560, 'cost': 467270000000.0},
                         17: {'seconds': 2193300, 'cost': 528200000000.0},
                         18: {'seconds': 2592360, 'cost': 596960000000.0},
                         19: {'seconds': 3052440, 'cost': 674450000000.0},
                         20: {'seconds': 3579360, 'cost': 761600000000.0},
                         21: {'seconds': 4179120, 'cost': 859420000000.0},
                         22: {'seconds': 4857960, 'cost': 968940000000.0},
                         23: {'seconds': 5622300, 'cost': 1090000000000.0},
                         24: {'seconds': 6478740, 'cost': 1230000000000.0},
                         25: {'seconds': 7434120, 'cost': 1380000000000.0},
                         26: {'seconds': 8495340, 'cost': 1550000000000.0},
                         27: {'seconds': 9669600, 'cost': 1730000000000.0},
                         28: {'seconds': 10964220, 'cost': 1930000000000.0},
                         29: {'seconds': 12386580, 'cost': 2160000000000.0},
                         30: {'seconds': 13944480, 'cost': 2400000000000.0}},
 'Fast Enemy Attack': {1: {'seconds': 199980, 'cost': 30000000000.0},
                       2: {'seconds': 230040, 'cost': 60010000000.0},
                       3: {'seconds': 260820, 'cost': 90070000000.0},
                       4: {'seconds': 293640, 'cost': 120330000000.0},
                       5: {'seconds': 330240, 'cost': 150970000000.0},
                       6: {'seconds': 372960, 'cost': 182260000000.0},
                       7: {'seconds': 424260, 'cost': 214530000000.0},
                       8: {'seconds': 487140, 'cost': 248130000000.0},
                       9: {'seconds': 564780, 'cost': 283510000000.0},
                       10: {'seconds': 660660, 'cost': 321140000000.0},
                       11: {'seconds': 778620, 'cost': 361550000000.0},
                       12: {'seconds': 922740, 'cost': 405320000000.0},
                       13: {'seconds': 1097160, 'cost': 453080000000.0},
                       14: {'seconds': 1306620, 'cost': 505500000000.0},
                       15: {'seconds': 1555740, 'cost': 563310000000.0},
                       16: {'seconds': 1849560, 'cost': 627270000000.0},
                       17: {'seconds': 2193300, 'cost': 698200000000.0},
                       18: {'seconds': 2592360, 'cost': 776960000000.0},
                       19: {'seconds': 3052440, 'cost': 864450000000.0},
                       20: {'seconds': 3579360, 'cost': 961600000000.0},
                       21: {'seconds': 4179120, 'cost': 1070000000000.0},
                       22: {'seconds': 4857960, 'cost': 1190000000000.0},
                       23: {'seconds': 5622300, 'cost': 1320000000000.0},
                       24: {'seconds': 6478740, 'cost': 1470000000000.0},
                       25: {'seconds': 7434120, 'cost': 1630000000000.0},
                       26: {'seconds': 8495340, 'cost': 1810000000000.0},
                       27: {'seconds': 9669600, 'cost': 2000000000000.0},
                       28: {'seconds': 10964220, 'cost': 2210000000000.0},
                       29: {'seconds': 12386580, 'cost': 2450000000000.0},
                       30: {'seconds': 13944480, 'cost': 2700000000000.0}},
 'Tank Enemy Attack': {1: {'seconds': 199980, 'cost': 30000000000.0},
                       2: {'seconds': 230040, 'cost': 60010000000.0},
                       3: {'seconds': 260820, 'cost': 90070000000.0},
                       4: {'seconds': 293640, 'cost': 120330000000.0},
                       5: {'seconds': 330240, 'cost': 150970000000.0},
                       6: {'seconds': 372960, 'cost': 182260000000.0},
                       7: {'seconds': 424260, 'cost': 214530000000.0},
                       8: {'seconds': 487140, 'cost': 248130000000.0},
                       9: {'seconds': 564780, 'cost': 283510000000.0},
                       10: {'seconds': 660660, 'cost': 321140000000.0},
                       11: {'seconds': 778620, 'cost': 361550000000.0},
                       12: {'seconds': 922740, 'cost': 405320000000.0},
                       13: {'seconds': 1097160, 'cost': 453080000000.0},
                       14: {'seconds': 1306620, 'cost': 505500000000.0},
                       15: {'seconds': 1555740, 'cost': 563310000000.0},
                       16: {'seconds': 1849560, 'cost': 627270000000.0},
                       17: {'seconds': 2193300, 'cost': 698200000000.0},
                       18: {'seconds': 2592360, 'cost': 776960000000.0},
                       19: {'seconds': 3052440, 'cost': 864450000000.0},
                       20: {'seconds': 3579360, 'cost': 961600000000.0},
                       21: {'seconds': 4179120, 'cost': 1070000000000.0},
                       22: {'seconds': 4857960, 'cost': 1190000000000.0},
                       23: {'seconds': 5622300, 'cost': 1320000000000.0},
                       24: {'seconds': 6478740, 'cost': 1470000000000.0},
                       25: {'seconds': 7434120, 'cost': 1630000000000.0},
                       26: {'seconds': 8495340, 'cost': 1810000000000.0},
                       27: {'seconds': 9669600, 'cost': 2000000000000.0},
                       28: {'seconds': 10964220, 'cost': 2210000000000.0},
                       29: {'seconds': 12386580, 'cost': 2450000000000.0},
                       30: {'seconds': 13944480, 'cost': 2700000000000.0}},
 'Ranged Enemy Attack': {1: {'seconds': 199980, 'cost': 30000000000.0},
                         2: {'seconds': 230040, 'cost': 60010000000.0},
                         3: {'seconds': 260820, 'cost': 90070000000.0},
                         4: {'seconds': 293640, 'cost': 120330000000.0},
                         5: {'seconds': 330240, 'cost': 150970000000.0},
                         6: {'seconds': 372960, 'cost': 182260000000.0},
                         7: {'seconds': 424260, 'cost': 214530000000.0},
                         8: {'seconds': 487140, 'cost': 248130000000.0},
                         9: {'seconds': 564780, 'cost': 283510000000.0},
                         10: {'seconds': 660660, 'cost': 321140000000.0},
                         11: {'seconds': 778620, 'cost': 361550000000.0},
                         12: {'seconds': 922740, 'cost': 405320000000.0},
                         13: {'seconds': 1097160, 'cost': 453080000000.0},
                         14: {'seconds': 1306620, 'cost': 505500000000.0},
                         15: {'seconds': 1555740, 'cost': 563310000000.0},
                         16: {'seconds': 1849560, 'cost': 627270000000.0},
                         17: {'seconds': 2193300, 'cost': 698200000000.0},
                         18: {'seconds': 2592360, 'cost': 776960000000.0},
                         19: {'seconds': 3052440, 'cost': 864450000000.0},
                         20: {'seconds': 3579360, 'cost': 961600000000.0},
                         21: {'seconds': 4179120, 'cost': 1070000000000.0},
                         22: {'seconds': 4857960, 'cost': 1190000000000.0},
                         23: {'seconds': 5622300, 'cost': 1320000000000.0},
                         24: {'seconds': 6478740, 'cost': 1470000000000.0},
                         25: {'seconds': 7434120, 'cost': 1630000000000.0},
                         26: {'seconds': 8495340, 'cost': 1810000000000.0},
                         27: {'seconds': 9669600, 'cost': 2000000000000.0},
                         28: {'seconds': 10964220, 'cost': 2210000000000.0},
                         29: {'seconds': 12386580, 'cost': 2450000000000.0},
                         30: {'seconds': 13944480, 'cost': 2700000000000.0}},
 'Boss Attack': {1: {'seconds': 199980, 'cost': 40000000000.0},
                 2: {'seconds': 230040, 'cost': 80010000000.0},
                 3: {'seconds': 260820, 'cost': 120070000000.0},
                 4: {'seconds': 293640, 'cost': 160330000000.0},
                 5: {'seconds': 330240, 'cost': 200970000000.0},
                 6: {'seconds': 372960, 'cost': 242260000000.0},
                 7: {'seconds': 424260, 'cost': 284530000000.0},
                 8: {'seconds': 487140, 'cost': 328130000000.0},
                 9: {'seconds': 564780, 'cost': 373510000000.0},
                 10: {'seconds': 660660, 'cost': 421140000000.0},
                 11: {'seconds': 778620, 'cost': 471550000000.0},
                 12: {'seconds': 922740, 'cost': 525320000000.0},
                 13: {'seconds': 1097160, 'cost': 583080000000.0},
                 14: {'seconds': 1306620, 'cost': 645500000000.0},
                 15: {'seconds': 1555740, 'cost': 713310000000.0},
                 16: {'seconds': 1849560, 'cost': 787270000000.0},
                 17: {'seconds': 2193300, 'cost': 868200000000.0},
                 18: {'seconds': 2592360, 'cost': 956960000000.0},
                 19: {'seconds': 3052440, 'cost': 1050000000000.0},
                 20: {'seconds': 3579360, 'cost': 1160000000000.0},
                 21: {'seconds': 4179120, 'cost': 1280000000000.0},
                 22: {'seconds': 4857960, 'cost': 1410000000000.0},
                 23: {'seconds': 5622300, 'cost': 1550000000000.0},
                 24: {'seconds': 6478740, 'cost': 1710000000000.0},
                 25: {'seconds': 7434120, 'cost': 1880000000000.0},
                 26: {'seconds': 8495340, 'cost': 2070000000000.0},
                 27: {'seconds': 9669600, 'cost': 2270000000000.0},
                 28: {'seconds': 10964220, 'cost': 2490000000000.0},
                 29: {'seconds': 12386580, 'cost': 2740000000000.0},
                 30: {'seconds': 13944480, 'cost': 3000000000000.0}},
 'Ray Enemy Attack': {1: {'seconds': 433880640, 'cost': 1.1625e+20}},
 'Vampire Enemy Attack': {1: {'seconds': 433880640, 'cost': 1.1625e+20}},
 'Scatter Enemy Attack': {1: {'seconds': 433880640, 'cost': 1.1625e+20}}}
NATIVE_HEALTH_LAB_MAX = {name: max(levels) if levels else 0 for name, levels in NATIVE_HEALTH_LAB_TABLES.items()}

NATIVE_HEALTH_LAB_NAMES = [
    "Death Wave Health", "Health", "Defense %", "Wall Health", "Wall Fortification",
    "Recovery Package Chance", "Recovery Package Max", "Defense Absolute",
    "Wall Thorns", "Garlic Thorns", "Common Enemy Attack", "Fast Enemy Attack",
    "Tank Enemy Attack", "Ranged Enemy Attack", "Boss Attack", "Ray Enemy Attack",
    "Vampire Enemy Attack", "Scatter Enemy Attack",
]

NATIVE_REGEN_LAB_NAMES = [
    "Health Regen", "Wall Regen", "Recovery Package Amount", "Recovery Package Chance",
    "Wall Fortification", "Wall Health", "Garlic Thorns",
]

NATIVE_HEALTH_ENHANCEMENTS = [
    "Health +", "Defense Absolute +", "Wall Health +", "Recovery Package +", "Enemy Level Skips +",
]
NATIVE_REGEN_ENHANCEMENTS = ["Health Regen +", "Wall Health +", "Recovery Package +"]

ENEMY_ATTACK_LAB_NAMES = [
    "Common Enemy Attack", "Fast Enemy Attack", "Tank Enemy Attack", "Ranged Enemy Attack",
    "Boss Attack", "Ray Enemy Attack", "Vampire Enemy Attack", "Scatter Enemy Attack",
]


def native_health_settings(profile_data: Dict[str, Any]) -> Dict[str, Any]:
    settings = profile_data.setdefault("native_health", {}).setdefault("settings", {})
    settings.setdefault("farming_defense_perk_bonus", 0.20)
    settings.setdefault("dw_health_saturation", 1.0)
    settings.setdefault("package_reliability_weight", 0.75)
    settings.setdefault("wall_weight", 0.25)
    settings.setdefault("defense_absolute_weight", 0.06)
    settings.setdefault("enemy_attack_lab_reduction", 0.004)
    settings.setdefault("vampire_bias", 1.25)
    settings.setdefault("regen_wall_weight", 1.0)
    return settings


def native_latest_run(profile_data: Dict[str, Any]) -> Dict[str, Any]:
    runs = profile_data.get("runs", [])
    return runs[-1] if isinstance(runs, list) and runs else {}


def native_latest_death(profile_data: Dict[str, Any]) -> str:
    run = native_latest_run(profile_data)
    value = str(run.get("killed_by", "") or "").strip()
    if value and value.casefold() != "unknown":
        return value
    raw = str(run.get("raw_text", "") or "")
    match = re.search(r"(?im)^Killed By\s+(.+?)\s*$", raw)
    return match.group(1).strip() if match else "Unknown"


def native_vault_active_bonus(profile_data: Dict[str, Any], name: str) -> float:
    record = profile_data.get("vault", {}).get("bonuses", {}).get(name, {})
    if isinstance(record, dict):
        return native_number(record.get("active"), 0.0)
    return native_number(record, 0.0)


def native_primary_module_record(profile_data: Dict[str, Any], slot: str, preset: str = "Farming") -> Dict[str, Any]:
    preset_data = profile_data.get("module_presets", {}).get(preset, {}).get(slot, {})
    name = preset_data.get("primary") if isinstance(preset_data, dict) else None
    if not name:
        configured = profile_data.get("modules", {}).get(slot, {})
        name = configured.get("name") if isinstance(configured, dict) else None
    if not name or name == "Any Other":
        return {}
    return profile_data.get("module_inventory", {}).get(f"{slot}::{name}", {}) or {}


def native_module_substat_bonus(profile_data: Dict[str, Any], names: list[str], preset: str = "Farming") -> float:
    aliases = {str(name).casefold().replace("/", " ").replace("%", "").strip() for name in names}
    total = 0.0
    for slot in ["Cannon", "Armor", "Generator", "Core"]:
        record = native_primary_module_record(profile_data, slot, preset)
        for sub in record.get("substats", []) if isinstance(record, dict) else []:
            key = str(sub.get("name", "")).casefold().replace("/", " ").replace("%", "").strip()
            if key in aliases:
                total += native_number(sub.get("value"), 0.0)
    return total


def native_module_unique_factor(profile_data: Dict[str, Any], module_name: str, preset: str = "Farming") -> float:
    for slot in ["Cannon", "Armor", "Generator", "Core"]:
        record = native_primary_module_record(profile_data, slot, preset)
        if record.get("name") == module_name:
            rarity = str(record.get("rarity", "")).casefold()
            if "ancestral" in rarity: return 1.0
            if "mythic" in rarity: return 0.75
            if "legendary" in rarity: return 0.50
            if "epic" in rarity: return 0.25
    return 0.0


def native_health_state(profile_data: Dict[str, Any]) -> Dict[str, Any]:
    return {
        "labs": {key: int(value or 0) for key, value in profile_data.get("labs", {}).items()},
        "enhancements": {key: int(value or 0) for key, value in profile_data.get("enhancements", {}).items()},
        "uw": json.loads(json.dumps(profile_data.get("uw", {}))),
    }


def native_enemy_attack_factor(profile_data: Dict[str, Any], labs: Dict[str, int]) -> float:
    settings = native_health_settings(profile_data)
    per_level = max(0.0, native_number(settings.get("enemy_attack_lab_reduction"), 0.004))
    latest = native_latest_run(profile_data)
    metrics = latest.get("metrics", {}) if isinstance(latest, dict) else {}
    raw = latest.get("raw_values", {}) if isinstance(latest, dict) else {}
    weights = {
        "Common Enemy Attack": native_number(raw.get("utility__basic") or raw.get("basic"), 1.0),
        "Fast Enemy Attack": native_number(raw.get("utility__fast") or raw.get("fast"), 1.0),
        "Tank Enemy Attack": native_number(raw.get("utility__tank") or raw.get("tank"), 1.0),
        "Ranged Enemy Attack": native_number(raw.get("utility__ranged") or raw.get("ranged"), 1.0),
        "Boss Attack": native_number(raw.get("utility__boss") or raw.get("boss"), 1.0),
        "Ray Enemy Attack": native_number(metrics.get("rays"), 1.0),
        "Vampire Enemy Attack": native_number(metrics.get("vampires"), 1.0),
        "Scatter Enemy Attack": native_number(metrics.get("scatters"), 1.0),
    }
    total_weight = sum(max(0.0, value) for value in weights.values()) or float(len(weights))
    weighted_reduction = 0.0
    for name, weight in weights.items():
        reduction = min(0.75, per_level * int(labs.get(name, 0) or 0))
        weighted_reduction += max(0.0, weight) / total_weight * reduction
    return 1.0 / max(0.05, 1.0 - weighted_reduction)


def native_health_components(profile_data: Dict[str, Any], state: Optional[Dict[str, Any]] = None) -> Dict[str, float]:
    state = state or native_health_state(profile_data)
    labs = state["labs"]
    enhancements = state["enhancements"]
    workshop = profile_data.get("workshop", {})
    settings = native_health_settings(profile_data)

    ws_health = max(1.0, native_number(workshop.get("Health"), 0.0) + 1.0)
    ws_regen = max(1.0, native_number(workshop.get("Health Regen"), 0.0) + 1.0)
    ws_defabs = max(1.0, native_number(workshop.get("Defense Absolute"), 0.0) + 1.0)

    armor = native_primary_module_record(profile_data, "Armor")
    armor_stat = max(1.0, native_number(armor.get("stat"), 1.0))
    health_factor = (
        ws_health
        * (1.0 + 0.03 * int(labs.get("Health", 0) or 0))
        * (1.0 + 0.01 * int(enhancements.get("Health +", 0) or 0))
        * (1.0 + native_active_relic_bonus(profile_data, "Health"))
        * (1.0 + native_vault_active_bonus(profile_data, "Health"))
        * armor_stat
    )

    dw_owned = bool(state.get("uw", {}).get("Death Wave", {}).get("owned"))
    dw_level = int(labs.get("Death Wave Health", 0) or 0)
    dw_cap = (5.0 + 0.25 * dw_level) if dw_owned and dw_level > 0 else 1.0
    saturation = min(1.0, max(0.0, native_number(settings.get("dw_health_saturation"), 1.0)))
    dw_factor = 1.0 + (dw_cap - 1.0) * saturation
    base_health = health_factor * dw_factor

    max_recovery = 1.5 + 0.03 * int(workshop.get("Max Amount", 0) or 0)
    max_recovery += 0.01 * int(labs.get("Recovery Package Max", 0) or 0)
    max_recovery *= 1.0 + native_module_substat_bonus(profile_data, ["Max Recovery"])
    max_recovery *= 1.0 + native_vault_active_bonus(profile_data, "Max Recovery")

    package_chance = 0.005 * int(workshop.get("Package Chance", 0) or 0)
    package_chance += 0.002 * int(labs.get("Recovery Package Chance", 0) or 0)
    package_chance += native_module_substat_bonus(profile_data, ["Package Chance", "Recovery Package Chance"])
    package_chance = min(1.0, max(0.0, package_chance))
    reliability_weight = min(1.0, max(0.0, native_number(settings.get("package_reliability_weight"), 0.75)))
    effective_recovery = 1.0 + (max_recovery - 1.0) * package_chance * reliability_weight
    tower_health = base_health * effective_recovery

    defense_percent = 0.005 * int(workshop.get("Defense %", 0) or 0)
    defense_percent += 0.002 * int(labs.get("Defense %", 0) or 0)
    defense_percent += native_active_relic_bonus(profile_data, "Defense %")
    defense_percent += native_vault_active_bonus(profile_data, "Defense %")
    defense_percent += native_module_substat_bonus(profile_data, ["Defense", "Defense %"])
    defense_percent += max(0.0, native_number(settings.get("farming_defense_perk_bonus"), 0.20))
    defense_percent = min(0.98, max(0.0, defense_percent))
    mitigation = 1.0 / max(0.02, 1.0 - defense_percent)

    defabs_factor = (
        ws_defabs
        * (1.0 + 0.03 * int(labs.get("Defense Absolute", 0) or 0))
        * (1.0 + 0.01 * int(enhancements.get("Defense Absolute +", 0) or 0))
        * (1.0 + native_active_relic_bonus(profile_data, "Defense Absolute"))
        * (1.0 + native_vault_active_bonus(profile_data, "Defense Absolute"))
        * (1.0 + native_module_substat_bonus(profile_data, ["Defense Absolute"]))
    )
    defabs_weight = max(0.0, native_number(settings.get("defense_absolute_weight"), 0.06))
    defabs_multiplier = 1.0 + defabs_weight * (defabs_factor / max(ws_defabs, 1.0) - 1.0)

    wall_ws = int(workshop.get("Wall Health", 0) or 0)
    wall_enabled = wall_ws > 0
    wall_base = base_health * (0.20 + 0.001 * wall_ws) if wall_enabled else 0.0
    wall_health = wall_base
    wall_health *= 1.0 + 0.02 * int(labs.get("Wall Health", 0) or 0)
    wall_health *= 1.0 + 0.20 * int(labs.get("Wall Fortification", 0) or 0)
    wall_health *= 1.0 + 0.01 * int(enhancements.get("Wall Health +", 0) or 0)
    wall_health *= 1.0 + native_module_substat_bonus(profile_data, ["Wall Health"])

    wall_thorns = int(labs.get("Wall Thorns", 0) or 0)
    wall_control_factor = 1.0 + 0.01 * wall_thorns
    death = native_latest_death(profile_data).casefold()
    garlic_level = int(labs.get("Garlic Thorns", 0) or 0)
    vampire_factor = 1.0
    if "vamp" in death:
        vampire_factor += max(0.0, native_number(settings.get("vampire_bias"), 1.25)) * 0.015 * garlic_level

    enemy_factor = native_enemy_attack_factor(profile_data, labs)
    wall_weight = max(0.0, native_number(settings.get("wall_weight"), 1.0))
    raw_pool = tower_health + wall_weight * wall_health * wall_control_factor
    ehp = raw_pool * mitigation * defabs_multiplier * enemy_factor * vampire_factor

    return {
        "Total": max(ehp, 1e-12), "Tower Health": tower_health, "Base Health": base_health,
        "Wall Health": wall_health, "Defense %": defense_percent, "Mitigation": mitigation,
        "Recovery Mult": effective_recovery, "Max Recovery": max_recovery, "Package Chance": package_chance,
        "Defense Absolute Factor": defabs_multiplier, "Enemy Attack Factor": enemy_factor,
        "Death Wave Factor": dw_factor, "Wall Control Factor": wall_control_factor,
        "Vampire Factor": vampire_factor, "Base Regen Proxy": ws_regen,
    }


def native_health_score(profile_data: Dict[str, Any], state: Optional[Dict[str, Any]] = None) -> float:
    return native_health_components(profile_data, state)["Total"]


def native_regen_components(profile_data: Dict[str, Any], state: Optional[Dict[str, Any]] = None) -> Dict[str, float]:
    state = state or native_health_state(profile_data)
    labs = state["labs"]
    enhancements = state["enhancements"]
    workshop = profile_data.get("workshop", {})
    settings = native_health_settings(profile_data)

    tower_regen = max(1.0, native_number(workshop.get("Health Regen"), 0.0) + 1.0)
    tower_regen *= 1.0 + 0.03 * int(labs.get("Health Regen", 0) or 0)
    tower_regen *= 1.0 + 0.01 * int(enhancements.get("Health Regen +", 0) or 0)
    tower_regen *= 1.0 + native_active_relic_bonus(profile_data, "Health Regen")
    tower_regen *= 1.0 + native_vault_active_bonus(profile_data, "Health Regen")
    tower_regen *= 1.0 + native_module_substat_bonus(profile_data, ["Health Regen"])

    wall_regen_percent = 0.10 * int(labs.get("Wall Regen", 0) or 0)
    wall_regen = tower_regen * wall_regen_percent
    wall_regen *= 1.0 + native_module_substat_bonus(profile_data, ["Wall Regen"])

    package_amount = 1.0 + 0.004 * int(labs.get("Recovery Package Amount", 0) or 0)
    package_amount *= 1.0 + native_module_substat_bonus(profile_data, ["Recovery Amount", "Recovery Package Amount"])
    package_amount *= 1.0 + native_active_relic_bonus(profile_data, "Recovery Amount")
    package_amount *= 1.0 + native_vault_active_bonus(profile_data, "Recovery Amount")
    package_amount *= 1.0 + 0.01 * int(enhancements.get("Recovery Package +", 0) or 0)

    package_chance = 0.005 * int(workshop.get("Package Chance", 0) or 0)
    package_chance += 0.002 * int(labs.get("Recovery Package Chance", 0) or 0)
    package_chance += native_module_substat_bonus(profile_data, ["Package Chance", "Recovery Package Chance"])
    package_chance = min(1.0, max(0.0, package_chance))
    package_sustain = tower_regen * package_amount * package_chance

    whr = native_module_unique_factor(profile_data, "Wormhole Redirector")
    recovery_regen = tower_regen * whr
    death = native_latest_death(profile_data).casefold()
    garlic = 1.0 + (0.02 * int(labs.get("Garlic Thorns", 0) or 0) if "vamp" in death else 0.0)
    wall_weight = max(0.0, native_number(settings.get("regen_wall_weight"), 1.0))
    total = (tower_regen + wall_weight * wall_regen + package_sustain + recovery_regen) * garlic

    return {
        "Total": max(total, 1e-12), "Tower Regen": tower_regen, "Wall Regen": wall_regen,
        "Wall Regen %": wall_regen_percent, "Package Sustain": package_sustain,
        "Recovery Regen": recovery_regen, "Package Chance": package_chance,
        "Package Amount": package_amount, "Wormhole Factor": whr, "Vampire Factor": garlic,
    }


def native_regen_score(profile_data: Dict[str, Any], state: Optional[Dict[str, Any]] = None) -> float:
    return native_regen_components(profile_data, state)["Total"]


def native_health_lab_allowed(profile_data: Dict[str, Any], name: str) -> tuple[bool, str]:
    maxed = profile_data.get("maxed", {}).get("labs", {}).get(name, False)
    if maxed: return False, "Gold boxed"
    if name.startswith("Wall ") and int(profile_data.get("workshop", {}).get("Wall Health", 0) or 0) <= 0:
        return False, "Wall not unlocked"
    if name == "Death Wave Health" and not profile_data.get("uw", {}).get("Death Wave", {}).get("owned"):
        return False, "Death Wave not owned"
    return True, ""


def native_survival_lab_candidates(profile_data: Dict[str, Any], state: Dict[str, Any], metric: str) -> list[Dict[str, Any]]:
    names = NATIVE_HEALTH_LAB_NAMES if metric == "health" else NATIVE_REGEN_LAB_NAMES
    score_fn = native_health_score if metric == "health" else native_regen_score
    before = score_fn(profile_data, state)
    lab_speed = native_lab_speed_multiplier(profile_data, state["labs"])
    coin_multiplier = native_lab_coin_multiplier(state["labs"])
    rows = []
    for name in names:
        allowed, reason = native_health_lab_allowed(profile_data, name)
        if not allowed: continue
        current = int(state["labs"].get(name, 0) or 0)
        next_level = current + 1
        record = NATIVE_HEALTH_LAB_TABLES.get(name, {}).get(next_level)
        if not record: continue
        next_state = json.loads(json.dumps(state))
        next_state["labs"][name] = next_level
        after = score_fn(profile_data, next_state)
        relative_gain = after / before - 1.0
        if relative_gain <= 0: continue
        adjusted_seconds = native_number(record.get("seconds"), 0.0) / max(lab_speed, 1e-9)
        adjusted_cost = native_number(record.get("cost"), 0.0) * coin_multiplier
        roi = relative_gain / max(adjusted_seconds / 86400.0, 1e-12)
        rows.append({
            "Upgrade": name, "Level": next_level, "Cost Numeric": adjusted_cost,
            "Cost": format_large_number(adjusted_cost), "Duration Seconds": adjusted_seconds,
            "Duration": native_format_duration(adjusted_seconds), "Relative Gain": relative_gain,
            "Gain %": relative_gain * 100.0, "ROI Numeric": roi, "ROI": roi,
            "Result Numeric": after, "Result": f"x{after:.6f}",
            "Confidence": "High tables / Medium model",
            "Why": f"Estimated {relative_gain * 100:.3f}% {('eHP' if metric == 'health' else 'sustain')} gain over {adjusted_seconds/86400.0:.2f} adjusted lab days.",
        })
    return sorted(rows, key=lambda row: row["ROI Numeric"], reverse=True)


def build_native_health_lab_path(profile_data: Dict[str, Any], steps: int = 50) -> list[Dict[str, Any]]:
    state = native_health_state(profile_data)
    cumulative = 0.0
    rows = []
    for rank in range(1, steps + 1):
        candidates = native_survival_lab_candidates(profile_data, state, "health")
        if not candidates: break
        chosen = dict(candidates[0])
        state["labs"][chosen["Upgrade"]] = int(chosen["Level"])
        cumulative += native_number(chosen.get("Duration Seconds"), 0.0)
        chosen.update({"Rank": rank, "Path": "native_health_lab", "Resource": "Time", "Cumulative": native_format_duration(cumulative)})
        rows.append(chosen)
    return rows


def build_native_regen_lab_path(profile_data: Dict[str, Any], steps: int = 50) -> list[Dict[str, Any]]:
    state = native_health_state(profile_data)
    cumulative = 0.0
    rows = []
    for rank in range(1, steps + 1):
        candidates = native_survival_lab_candidates(profile_data, state, "regen")
        if not candidates: break
        chosen = dict(candidates[0])
        state["labs"][chosen["Upgrade"]] = int(chosen["Level"])
        cumulative += native_number(chosen.get("Duration Seconds"), 0.0)
        chosen.update({"Rank": rank, "Path": "native_regen_lab", "Resource": "Time", "Cumulative": native_format_duration(cumulative)})
        rows.append(chosen)
    return rows


def native_health_stone_candidates(profile_data: Dict[str, Any], state: Dict[str, Any]) -> list[Dict[str, Any]]:
    # Stone spending only affects eHP here when Death Wave health is not reliably saturated.
    settings = native_health_settings(profile_data)
    saturation = min(1.0, max(0.0, native_number(settings.get("dw_health_saturation"), 1.0)))
    if saturation >= 0.999 or not state.get("uw", {}).get("Death Wave", {}).get("owned"):
        return []
    before = native_health_score(profile_data, state)
    rows = []
    mapping = {"DW | Quantity": "Quantity", "DW | Cooldown": "Cooldown"}
    for upgrade, attr in mapping.items():
        table = NATIVE_DAMAGE_UW_TABLES.get(upgrade, [])
        current_value = native_number(state["uw"].get("Death Wave", {}).get("attributes", {}).get(attr), 1 if attr == "Quantity" else 300)
        index = min(range(len(table)), key=lambda i: abs(native_number(table[i].get("value")) - current_value)) if table else -1
        next_index = index + 1
        if next_index < 0 or next_index >= len(table): continue
        record = table[next_index]
        cost = native_number(record.get("cost"), 0.0)
        if cost <= 0: continue
        next_state = json.loads(json.dumps(state))
        next_state["uw"].setdefault("Death Wave", {}).setdefault("attributes", {})[attr] = record.get("value")
        # Convert the stone step into a modest saturation improvement.
        old_sat = saturation
        if attr == "Quantity": new_sat = min(1.0, old_sat + 0.12)
        else: new_sat = min(1.0, old_sat + max(0.0, current_value - native_number(record.get("value"))) / 600.0)
        old_setting = settings.get("dw_health_saturation", 1.0)
        settings["dw_health_saturation"] = new_sat
        after = native_health_score(profile_data, next_state)
        settings["dw_health_saturation"] = old_setting
        relative_gain = after / before - 1.0
        if relative_gain <= 0: continue
        rows.append({
            "Upgrade": upgrade, "Level": f"lvl {next_index}", "Value": record.get("value"),
            "Cost Numeric": cost, "Cost": cost, "Relative Gain": relative_gain, "Gain %": relative_gain * 100.0,
            "ROI Numeric": relative_gain / cost, "ROI": relative_gain / cost, "Result Numeric": after,
            "Result": f"x{after:.6f}", "Confidence": "Low-Medium",
            "Why": f"Improves estimated Death Wave health saturation from {old_sat*100:.1f}% to {new_sat*100:.1f}%.",
            "New Saturation": new_sat,
        })
    return sorted(rows, key=lambda row: row["ROI Numeric"], reverse=True)


def build_native_health_stone_path(profile_data: Dict[str, Any], steps: int = 25) -> list[Dict[str, Any]]:
    state = native_health_state(profile_data)
    cumulative = 0.0
    rows = []
    original = native_health_settings(profile_data).get("dw_health_saturation", 1.0)
    saturation = native_number(original, 1.0)
    try:
        for rank in range(1, steps + 1):
            native_health_settings(profile_data)["dw_health_saturation"] = saturation
            candidates = native_health_stone_candidates(profile_data, state)
            if not candidates: break
            chosen = dict(candidates[0])
            _, attr = chosen["Upgrade"].split(" | ", 1)
            state["uw"]["Death Wave"]["attributes"][attr] = chosen["Value"]
            saturation = native_number(chosen.pop("New Saturation"), saturation)
            cumulative += chosen["Cost Numeric"]
            chosen.update({"Rank": rank, "Path": "native_health_stone", "Resource": "Stones", "Cumulative": f"{cumulative:,.0f}"})
            rows.append(chosen)
    finally:
        native_health_settings(profile_data)["dw_health_saturation"] = original
    return rows


def build_native_survival_coin_path(profile_data: Dict[str, Any], metric: str, steps: int = 50) -> list[Dict[str, Any]]:
    state = native_health_state(profile_data)
    names = NATIVE_HEALTH_ENHANCEMENTS if metric == "health" else NATIVE_REGEN_ENHANCEMENTS
    score_fn = native_health_score if metric == "health" else native_regen_score
    discount_lab = "Enhancement Defense - Coin Discount"
    discount_level = int(profile_data.get("labs", {}).get(discount_lab, 0) or 0)
    discount = max(0.05, 1.0 - 0.01 * discount_level)
    costs = NATIVE_ECON_ENHANCEMENT_COSTS.get("Coin Bonus +", [])
    cumulative = 0.0
    rows = []
    for rank in range(1, steps + 1):
        before = score_fn(profile_data, state)
        candidates = []
        for name in names:
            current = int(state["enhancements"].get(name, 0) or 0)
            next_level = current + 1
            if next_level >= len(costs) or next_level > ENHANCEMENT_MAX_LEVELS.get(name, 200): continue
            cost = native_number(costs[next_level], 0.0) * discount
            if cost <= 0: continue
            next_state = json.loads(json.dumps(state))
            next_state["enhancements"][name] = next_level
            after = score_fn(profile_data, next_state)
            relative_gain = after / before - 1.0
            if relative_gain <= 0: continue
            candidates.append({
                "Upgrade": name, "Level": next_level, "Cost Numeric": cost, "Cost": format_large_number(cost),
                "Relative Gain": relative_gain, "Gain %": relative_gain * 100.0,
                "ROI Numeric": relative_gain / max(cost / 1e9, 1e-12), "ROI": relative_gain / max(cost / 1e9, 1e-12),
                "Result Numeric": after, "Result": f"x{after:.6f}", "Confidence": "Medium",
                "Why": f"Estimated {relative_gain*100:.3f}% {('eHP' if metric == 'health' else 'sustain')} gain per enhancement step.",
            })
        if not candidates: break
        chosen = max(candidates, key=lambda row: row["ROI Numeric"])
        state["enhancements"][chosen["Upgrade"]] = int(chosen["Level"])
        cumulative += chosen["Cost Numeric"]
        chosen.update({"Rank": rank, "Path": f"native_{metric}_coin", "Resource": "Coins", "Cumulative": format_large_number(cumulative)})
        rows.append(chosen)
    return rows


def build_native_health_paths(profile_data: Dict[str, Any], steps: int = 50) -> Dict[str, list[Dict[str, Any]]]:
    result = {
        "health_lab": build_native_health_lab_path(profile_data, steps),
        "health_stone": build_native_health_stone_path(profile_data, min(steps, 25)),
        "health_coin": build_native_survival_coin_path(profile_data, "health", steps),
        "regen_lab": build_native_regen_lab_path(profile_data, steps),
        "regen_coin": build_native_survival_coin_path(profile_data, "regen", steps),
    }
    profile_data.setdefault("native_health", {})["last_run"] = {
        "at": datetime.now(timezone.utc).isoformat(), "steps": steps,
        "coverage": {
            "health_lab": "Exact lab cost/time tables; relative eHP model.",
            "health_stone": "DW saturation support only; empty when saturation is already 100%.",
            "health_coin": "Shared enhancement cost curve; relative health effects.",
            "regen_lab": "Exact lab cost/time tables; wall/tower/package sustain model.",
            "regen_coin": "Shared enhancement cost curve; relative sustain effects.",
        },
    }
    return result


# -----------------------------------------------------------------------------
# BUILD AUDIT, BATTLE REPORT, AND ANALYZER HELPERS
# -----------------------------------------------------------------------------


# -----------------------------------------------------------------------------
# NATIVE ECONOMY ENGINE
# -----------------------------------------------------------------------------

NATIVE_ECON_LAB_NAMES = [
    "Coins / Kill Bonus", "Golden Tower Bonus", "Golden Tower Duration",
    "Black Hole Coin Bonus", "Spotlight Coin Bonus", "Death Wave Coin Bonus",
    "Extra Black Hole",
]


def native_number(value: Any, default: float = 0.0) -> float:
    try:
        return float(value)
    except (TypeError, ValueError):
        return float(default)


def native_active_relic_bonus(profile_data: Dict[str, Any], bonus_name: str) -> float:
    record = profile_data.get("relics", {}).get("bonuses", {}).get(bonus_name, {})
    if isinstance(record, dict):
        return native_number(record.get("active"), 0.0)
    return native_number(record, 0.0)


def native_lab_speed_multiplier(profile_data: Dict[str, Any], labs: Optional[Dict[str, int]] = None) -> float:
    labs = labs or profile_data.get("labs", {})
    settings = profile_data.get("native_econ", {}).get("settings", {})
    override = native_number(settings.get("lab_speed_multiplier_override"), 0.0)
    if override > 0:
        return override
    lab_level = int(labs.get("Labs Speed", 0) or 0)
    relic_bonus = native_active_relic_bonus(profile_data, "Lab Speed")
    return max(0.01, (1.0 + 0.02 * lab_level) * (1.0 + relic_bonus))


def native_lab_coin_multiplier(labs: Dict[str, int]) -> float:
    # Labs Coin Discount grants 0.3% per completed level.
    return max(0.05, 1.0 - 0.003 * int(labs.get("Labs Coin Discount", 0) or 0))


def native_uw_value(profile_data: Dict[str, Any], uw_name: str, attribute: str, fallback: float) -> float:
    uw = profile_data.get("uw", {}).get(uw_name, {})
    attrs = uw.get("attributes", {}) if isinstance(uw, dict) else {}
    return native_number(attrs.get(attribute), fallback)


def native_econ_context(profile_data: Dict[str, Any]) -> Dict[str, float]:
    labs = profile_data.get("labs", {})
    gt = profile_data.get("uw", {}).get("Golden Tower", {})
    gt_stone_bonus = native_uw_value(profile_data, "Golden Tower", "Multiplier", 5.0)
    gt_stone_duration = native_uw_value(profile_data, "Golden Tower", "Duration", 15.0)
    gt_total_bonus_reported = max(
        native_number(gt.get("bonus_or_damage"), 0.0),
        native_number(gt.get("gt_bonus"), 0.0),
        gt_stone_bonus + 0.15 * int(labs.get("Golden Tower Bonus", 0) or 0),
    )
    gt_total_duration_reported = max(
        native_number(gt.get("duration"), 0.0),
        gt_stone_duration + int(labs.get("Golden Tower Duration", 0) or 0),
    )
    return {
        "gt_bonus_residual": max(0.0, gt_total_bonus_reported - gt_stone_bonus - 0.15 * int(labs.get("Golden Tower Bonus", 0) or 0)),
        "gt_duration_residual": max(0.0, gt_total_duration_reported - gt_stone_duration - int(labs.get("Golden Tower Duration", 0) or 0)),
    }


def native_initial_stone_state(profile_data: Dict[str, Any]) -> Dict[str, float]:
    return {
        "GT | Bonus": native_uw_value(profile_data, "Golden Tower", "Multiplier", 5.0),
        "GT | Duration": native_uw_value(profile_data, "Golden Tower", "Duration", 15.0),
        "GT | Cooldown": native_uw_value(profile_data, "Golden Tower", "Cooldown", 300.0),
        "BH | Size": native_uw_value(profile_data, "Black Hole", "Size", 30.0),
        "BH | Duration": native_uw_value(profile_data, "Black Hole", "Duration", 15.0),
        "BH | Cooldown": native_uw_value(profile_data, "Black Hole", "Cooldown", 200.0),
        "SL | Angle": native_uw_value(profile_data, "Spotlight", "Angle", 0.0),
        "SL | Quantity": native_uw_value(profile_data, "Spotlight", "Quantity", 0.0),
        "DW | Quantity": native_uw_value(profile_data, "Death Wave", "Quantity", 0.0),
        "DW | Cooldown": native_uw_value(profile_data, "Death Wave", "Cooldown", 300.0),
    }


def native_econ_score(
    profile_data: Dict[str, Any],
    labs: Optional[Dict[str, int]] = None,
    stone_state: Optional[Dict[str, float]] = None,
) -> float:
    """Return a relative coin-production index.

    The index models synced GT/BH windows, Spotlight coverage, Death Wave tags,
    Coins/Kill, and current UW timing. Constant account multipliers cancel when
    comparing one upgrade to another, so the absolute value is intentionally a
    relative index rather than projected coins/hour.
    """
    labs = dict(labs or profile_data.get("labs", {}))
    stones = dict(stone_state or native_initial_stone_state(profile_data))
    settings = profile_data.get("native_econ", {}).get("settings", {})
    context = native_econ_context(profile_data)

    cpk_factor = 1.0 + 0.02 * int(labs.get("Coins / Kill Bonus", 0) or 0)

    gt_owned = bool(profile_data.get("uw", {}).get("Golden Tower", {}).get("owned"))
    bh_owned = bool(profile_data.get("uw", {}).get("Black Hole", {}).get("owned"))
    sl_owned = bool(profile_data.get("uw", {}).get("Spotlight", {}).get("owned"))
    dw_owned = bool(profile_data.get("uw", {}).get("Death Wave", {}).get("owned"))

    gt_bonus = (
        stones["GT | Bonus"]
        + 0.15 * int(labs.get("Golden Tower Bonus", 0) or 0)
        + context["gt_bonus_residual"]
    ) if gt_owned else 1.0
    gt_duration = (
        stones["GT | Duration"]
        + int(labs.get("Golden Tower Duration", 0) or 0)
        + context["gt_duration_residual"]
    ) if gt_owned else 1.0
    gt_cooldown = max(1.0, stones["GT | Cooldown"] if gt_owned else 1.0)

    if gt_owned:
        activation_value = gt_bonus * gt_duration / gt_cooldown
    else:
        activation_value = 1.0

    if gt_owned and bh_owned:
        bh_quantity = 1 + int(labs.get("Extra Black Hole", 0) or 0)
        bh_coin = 1.0 + 0.5 * int(labs.get("Black Hole Coin Bonus", 0) or 0)
        bh_perk_bonus = native_number(settings.get("bh_perk_duration_bonus"), 12.0)
        bh_duration = stones["BH | Duration"] + bh_perk_bonus
        bh_cooldown = max(1.0, stones["BH | Cooldown"])
        divisor = max(1.0, native_number(settings.get("bh_coverage_divisor"), 70.0))
        capture = min(1.0, bh_quantity * stones["BH | Size"] / divisor)
        alignment = 1.0 if abs(gt_cooldown - bh_cooldown) < 1e-9 else min(gt_cooldown, bh_cooldown) / max(gt_cooldown, bh_cooldown)
        overlap = min(gt_duration, bh_duration) * capture * alignment
        activation_value = gt_bonus * (gt_duration + overlap * (bh_coin - 1.0)) / gt_cooldown

    spotlight_factor = 1.0
    if sl_owned:
        sl_coin = 1.0 + 0.1 * int(labs.get("Spotlight Coin Bonus", 0) or 0)
        coverage = min(1.0, stones["SL | Angle"] * stones["SL | Quantity"] / 360.0)
        spotlight_factor = 1.0 + coverage * (sl_coin - 1.0)

    death_wave_factor = 1.0
    if dw_owned:
        dw_coin = 1.5 + 0.05 * int(labs.get("Death Wave Coin Bonus", 0) or 0)
        tag_per_quantity = native_number(settings.get("dw_tag_share_per_quantity"), 0.0382747832266563)
        # Quantity is verified against Effective Paths. Cooldown is deliberately
        # omitted from the v1.0 tag model until its tagging behavior is verified.
        tag_share = min(1.0, tag_per_quantity * stones["DW | Quantity"])
        death_wave_factor = 1.0 + tag_share * (dw_coin - 1.0)

    return max(1e-12, cpk_factor * activation_value * spotlight_factor * death_wave_factor)


def native_format_duration(seconds: float) -> str:
    seconds = max(0, int(round(seconds)))
    days, seconds = divmod(seconds, 86400)
    hours, seconds = divmod(seconds, 3600)
    minutes, seconds = divmod(seconds, 60)
    parts = []
    if days: parts.append(f"{days}d")
    if hours or days: parts.append(f"{hours}h")
    if minutes or hours or days: parts.append(f"{minutes}m")
    if not parts: parts.append(f"{seconds}s")
    return " ".join(parts)


def native_format_resource(value: float, resource: str) -> str:
    if resource == "Time":
        return native_format_duration(value)
    if resource in {"Stones", "Keys"}:
        return f"{value:,.0f}"
    return format_large_number(value)


def native_econ_lab_candidates(profile_data: Dict[str, Any], labs: Dict[str, int]) -> list[Dict[str, Any]]:
    before = native_econ_score(profile_data, labs=labs)
    speed = native_lab_speed_multiplier(profile_data, labs)
    coin_multiplier = native_lab_coin_multiplier(labs)
    rows = []
    for name in NATIVE_ECON_LAB_NAMES:
        current = int(labs.get(name, 0) or 0)
        max_level = NATIVE_ECON_LAB_MAX.get(name, 0)
        if current >= max_level:
            continue
        next_level = current + 1
        base = NATIVE_ECON_LAB_TABLES.get(name, {}).get(next_level)
        if not base:
            continue
        next_labs = dict(labs)
        next_labs[name] = next_level
        after = native_econ_score(profile_data, labs=next_labs)
        relative_gain = after / before - 1.0
        adjusted_seconds = base["seconds"] / speed
        adjusted_cost = base["cost"] * coin_multiplier
        roi = relative_gain / max(adjusted_seconds / 86400.0, 1e-12)
        rows.append({
            "Upgrade": name,
            "Level": next_level,
            "Cost Numeric": adjusted_cost,
            "Cost": format_large_number(adjusted_cost),
            "Duration Seconds": adjusted_seconds,
            "Duration": native_format_duration(adjusted_seconds),
            "Relative Gain": relative_gain,
            "Gain %": relative_gain * 100.0,
            "ROI Numeric": roi,
            "ROI": roi,
            "Result Numeric": after,
            "Result": f"x{after:.6f}",
            "Why": f"Estimated {relative_gain * 100:.3f}% coin-index gain over {native_format_duration(adjusted_seconds)}.",
        })
    return sorted(rows, key=lambda row: row["ROI Numeric"], reverse=True)


def build_native_econ_lab_path(profile_data: Dict[str, Any], steps: int = 50) -> list[Dict[str, Any]]:
    labs = {name: int(value or 0) for name, value in profile_data.get("labs", {}).items()}
    result = []
    cumulative_seconds = 0.0
    for rank in range(1, steps + 1):
        candidates = native_econ_lab_candidates(profile_data, labs)
        if not candidates:
            break
        chosen = dict(candidates[0])
        labs[chosen["Upgrade"]] = int(chosen["Level"])
        cumulative_seconds += chosen["Duration Seconds"]
        chosen.update({
            "Rank": rank,
            "Path": "native_econ_lab",
            "Resource": "Time",
            "Cumulative": native_format_duration(cumulative_seconds),
        })
        result.append(chosen)
    return result


def build_native_discount_path(profile_data: Dict[str, Any], steps: int = 50) -> list[Dict[str, Any]]:
    labs = {name: int(value or 0) for name, value in profile_data.get("labs", {}).items()}
    current = int(labs.get("Labs Coin Discount", 0) or 0)
    max_level = NATIVE_ECON_LAB_MAX.get("Labs Coin Discount", 0)
    cumulative_seconds = 0.0
    rows = []
    for rank in range(1, steps + 1):
        next_level = current + 1
        if next_level > max_level:
            break
        base = NATIVE_ECON_LAB_TABLES["Labs Coin Discount"].get(next_level)
        if not base:
            break
        speed = native_lab_speed_multiplier(profile_data, labs)
        duration = base["seconds"] / speed
        current_value = 1.0 / native_lab_coin_multiplier(labs)
        next_labs = dict(labs)
        next_labs["Labs Coin Discount"] = next_level
        next_value = 1.0 / native_lab_coin_multiplier(next_labs)
        relative_gain = next_value / current_value - 1.0
        roi = relative_gain / max(duration / 86400.0, 1e-12)
        cumulative_seconds += duration
        rows.append({
            "Rank": rank,
            "Upgrade": "Labs Coin Discount",
            "Level": next_level,
            "Duration Seconds": duration,
            "Duration": native_format_duration(duration),
            "ROI Numeric": roi,
            "ROI": roi,
            "Result Numeric": next_value,
            "Result": next_value,
            "Relative Gain": relative_gain,
            "Gain %": relative_gain * 100.0,
            "Cumulative": native_format_duration(cumulative_seconds),
            "Why": f"Improves effective lab-coin value by {relative_gain * 100:.3f}%.",
        })
        labs = next_labs
        current = next_level
    return rows


def native_table_index(table: list[Dict[str, float]], value: float) -> int:
    if not table:
        return -1
    return min(range(len(table)), key=lambda index: abs(table[index]["value"] - value))


def native_stone_candidate_is_allowed(profile_data: Dict[str, Any], state: Dict[str, float], upgrade: str) -> tuple[bool, str]:
    settings = profile_data.get("native_econ", {}).get("settings", {})
    if upgrade in {"GT | Cooldown", "BH | Cooldown"}:
        synced = abs(state["GT | Cooldown"] - state["BH | Cooldown"]) < 1e-9
        if synced and not bool(settings.get("allow_desync_cooldowns", False)):
            return False, "Excluded because buying only one cooldown would break the current GT/BH sync."
    if upgrade == "DW | Cooldown":
        return False, "Death Wave cooldown economy impact is not yet verified in the v1.0 tagging model."
    return True, ""


def native_econ_stone_candidates(profile_data: Dict[str, Any], labs: Dict[str, int], state: Dict[str, float]) -> list[Dict[str, Any]]:
    before = native_econ_score(profile_data, labs=labs, stone_state=state)
    rows = []
    for upgrade, table in NATIVE_ECON_UW_TABLES.items():
        allowed, exclusion = native_stone_candidate_is_allowed(profile_data, state, upgrade)
        if not allowed:
            continue
        current_value = state.get(upgrade)
        if current_value is None:
            continue
        current_index = native_table_index(table, current_value)
        next_index = current_index + 1
        if current_index < 0 or next_index >= len(table):
            continue
        next_record = table[next_index]
        cost = next_record["cost"]
        if cost <= 0:
            continue
        next_state = dict(state)
        next_state[upgrade] = next_record["value"]
        after = native_econ_score(profile_data, labs=labs, stone_state=next_state)
        relative_gain = after / before - 1.0
        roi = relative_gain / cost
        rows.append({
            "Upgrade": upgrade,
            "Level": f"lvl {next_index}",
            "Value": next_record["value"],
            "Cost Numeric": cost,
            "Cost": cost,
            "Relative Gain": relative_gain,
            "Gain %": relative_gain * 100.0,
            "ROI Numeric": roi,
            "ROI": roi,
            "Result Numeric": after,
            "Result": f"x{after:.6f}",
            "Why": f"Estimated {relative_gain * 100:.3f}% coin-index gain for {cost:,.0f} stones.",
        })
    return sorted(rows, key=lambda row: row["ROI Numeric"], reverse=True)


def build_native_econ_stone_path(profile_data: Dict[str, Any], steps: int = 50) -> list[Dict[str, Any]]:
    labs = {name: int(value or 0) for name, value in profile_data.get("labs", {}).items()}
    state = native_initial_stone_state(profile_data)
    cumulative = 0.0
    rows = []
    for rank in range(1, steps + 1):
        candidates = native_econ_stone_candidates(profile_data, labs, state)
        if not candidates:
            break
        chosen = dict(candidates[0])
        state[chosen["Upgrade"]] = native_number(chosen["Value"])
        cumulative += chosen["Cost Numeric"]
        chosen.update({
            "Rank": rank,
            "Path": "native_econ_stone",
            "Resource": "Stones",
            "Cumulative": f"{cumulative:,.0f}",
        })
        rows.append(chosen)
    return rows


def native_enhancement_discount(profile_data: Dict[str, Any]) -> float:
    level = int(profile_data.get("labs", {}).get("Enhancement Utility - Coin Discount", 0) or 0)
    # The exact late-game discount curve is retained as an explicit confidence
    # limitation. At level 0 this is exact, which covers the supplied reference.
    return max(0.05, 1.0 - 0.01 * level)


def build_native_econ_coin_path(profile_data: Dict[str, Any], steps: int = 50) -> list[Dict[str, Any]]:
    level = int(profile_data.get("enhancements", {}).get("Coin Bonus +", 0) or 0)
    costs = NATIVE_ECON_ENHANCEMENT_COSTS.get("Coin Bonus +", [])
    discount = native_enhancement_discount(profile_data)
    cumulative = 0.0
    rows = []
    current_multiplier = 1.0 + 0.01 * level
    for rank in range(1, steps + 1):
        next_level = level + 1
        if next_level >= len(costs) or next_level > ENHANCEMENT_MAX_LEVELS.get("Coin Bonus +", 200):
            break
        cost = costs[next_level] * discount
        next_multiplier = 1.0 + 0.01 * next_level
        relative_gain = next_multiplier / current_multiplier - 1.0
        roi = relative_gain / max(cost / 1e9, 1e-12)
        cumulative += cost
        rows.append({
            "Rank": rank,
            "Upgrade": "Coin Bonus +",
            "Level": next_level,
            "Cost Numeric": cost,
            "Cost": format_large_number(cost),
            "Relative Gain": relative_gain,
            "Gain %": relative_gain * 100.0,
            "ROI Numeric": roi,
            "ROI": roi,
            "Result Numeric": next_multiplier,
            "Result": f"x{next_multiplier:.2f}",
            "Cumulative": format_large_number(cumulative),
            "Why": f"Verified +1% enhancement step; {relative_gain * 100:.3f}% relative gain.",
        })
        level = next_level
        current_multiplier = next_multiplier
    return rows


def build_native_econ_paths(profile_data: Dict[str, Any], steps: int = 50) -> Dict[str, list[Dict[str, Any]]]:
    result = {
        "econ_lab": build_native_econ_lab_path(profile_data, steps),
        "econ_stone": build_native_econ_stone_path(profile_data, steps),
        "econ_coin": build_native_econ_coin_path(profile_data, steps),
        "econ_discount": build_native_discount_path(profile_data, steps),
    }
    profile_data.setdefault("native_econ", {})["last_run"] = {
        "at": datetime.now(timezone.utc).isoformat(),
        "steps": steps,
        "coverage": {
            "econ_lab": "Verified formulas and source tables",
            "econ_discount": "Verified formulas and source tables",
            "econ_stone": "Verified source tables; BH/DW coverage assumptions configurable",
            "econ_coin": "Verified Coin Bonus+ path only",
        },
    }
    return result


def normalize_econ_upgrade_name(value: Any) -> str:
    text = str(value or "").strip().lower()
    replacements = {
        "golden tower": "gt", "black hole": "bh", "spotlight": "sl", "death wave": "dw",
        "multiplier": "bonus", "coin bonus": "coin bonus",
    }
    for old, new in replacements.items():
        text = text.replace(old, new)
    return re.sub(r"[^a-z0-9]+", " ", text).strip()


def compare_native_path(native_rows: list[Dict[str, Any]], reference_rows: list[Dict[str, Any]], limit: int = 50) -> list[Dict[str, Any]]:
    rows = []
    for index in range(min(limit, max(len(native_rows), len(reference_rows)))):
        native = native_rows[index] if index < len(native_rows) else {}
        reference = reference_rows[index] if index < len(reference_rows) else {}
        native_roi = native_number(native.get("ROI Numeric", native.get("ROI")), 0.0)
        reference_roi = native_number(reference.get("ROI Numeric", reference.get("ROI")), 0.0)
        roi_difference = ((native_roi / reference_roi) - 1.0) * 100.0 if reference_roi else None
        native_name = normalize_econ_upgrade_name(native.get("Upgrade"))
        reference_name = normalize_econ_upgrade_name(reference.get("Upgrade"))
        rows.append({
            "Rank": index + 1,
            "Native Upgrade": native.get("Upgrade", "—"),
            "Native Level": native.get("Level", "—"),
            "Reference Upgrade": reference.get("Upgrade", "—"),
            "Reference Level": reference.get("Level", "—"),
            "Upgrade Match": bool(native_name and native_name == reference_name),
            "Native ROI": native_roi or None,
            "Reference ROI": reference_roi or None,
            "ROI Difference %": roi_difference,
        })
    return rows


def native_path_match_summary(comparison: list[Dict[str, Any]]) -> Dict[str, Any]:
    comparable = [row for row in comparison if row.get("Reference Upgrade") not in (None, "—", "")]
    if not comparable:
        return {"rows": 0, "matches": 0, "match_rate": None, "median_roi_difference": None}
    matches = sum(bool(row.get("Upgrade Match")) for row in comparable)
    differences = sorted(abs(row["ROI Difference %"]) for row in comparable if row.get("ROI Difference %") is not None)
    median = differences[len(differences) // 2] if differences else None
    return {"rows": len(comparable), "matches": matches, "match_rate": matches / len(comparable), "median_roi_difference": median}

NUMBER_SUFFIXES = {
    "": 1.0,
    "K": 1e3,
    "M": 1e6,
    "B": 1e9,
    "T": 1e12,
    "q": 1e15,
    "Q": 1e18,
    "s": 1e21,
    "S": 1e24,
    "O": 1e27,
    "N": 1e30,
    "D": 1e33,
}


def parse_tower_number(value: Any) -> Optional[float]:
    """Parse Tower-style values such as 207.40M, $1.95B, 10.54Q, or x11.4."""
    if value is None:
        return None
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return float(value)
    text = str(value).strip().replace(",", "")
    if not text or text in {"-", "N/A"}:
        return None
    text = text.replace("$", "").replace("%", "").replace("×", "x")
    if text.lower().startswith("x"):
        text = text[1:]
    match = re.fullmatch(r"([+-]?\d+(?:\.\d+)?)\s*([KMBTqQsSOND]?)", text)
    if not match:
        try:
            return float(text)
        except ValueError:
            return None
    number = float(match.group(1))
    return number * NUMBER_SUFFIXES.get(match.group(2), 1.0)


def parse_duration_seconds(value: Any) -> Optional[int]:
    """Parse values such as '1d 8h 5m 27s' or '7h 40m 36s'."""
    if value is None:
        return None
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return int(value)
    text = str(value).strip().lower()
    if not text:
        return None
    units = {"d": 86400, "h": 3600, "m": 60, "s": 1}
    total = 0
    found = False
    for amount, unit in re.findall(r"(\d+(?:\.\d+)?)\s*([dhms])", text):
        total += int(float(amount) * units[unit])
        found = True
    return total if found else None


def canonical_report_key(key: str) -> str:
    cleaned = re.sub(r"[^a-z0-9]+", "_", key.strip().lower()).strip("_")
    aliases = {
        "battle_date": "battle_date",
        "game_time": "game_time",
        "real_time": "real_time",
        "tier": "tier",
        "wave": "wave",
        "killed_by": "killed_by",
        "coins_earned": "coins_earned",
        "coins_per_hour": "coins_per_hour",
        "cash_earned": "cash_earned",
        "cells_earned": "cells_earned",
        "reroll_shards_earned": "reroll_shards_earned",
        "damage_dealt": "damage_dealt",
        "damage_taken": "damage_taken",
        "damage_taken_wall": "damage_taken_wall",
        "lifesteal": "lifesteal",
        "projectiles_damage": "projectiles_damage",
        "thorn_damage": "thorn_damage",
        "orb_damage": "orb_damage",
        "black_hole_damage": "black_hole_damage",
        "death_ray_damage": "death_ray_damage",
        "waves_skipped": "waves_skipped",
        "recovery_packages": "recovery_packages",
        "free_attack_upgrade": "free_attack_upgrade",
        "free_defense_upgrade": "free_defense_upgrade",
        "free_utility_upgrade": "free_utility_upgrade",
        "coins_from_golden_tower": "coins_from_golden_tower",
        "coins_from_black_hole": "coins_from_black_hole",
        "coins_from_spotlight": "coins_from_spotlight",
        "coins_from_death_wave": "coins_from_death_wave",
        "total_enemies": "total_enemies",
        "vampires": "vampires",
        "rays": "rays",
        "scatters": "scatters",
        "protector": "protectors",
        "protectors": "protectors",
        "destroyed_by_orbs": "destroyed_by_orbs",
        "destroyed_by_thorns": "destroyed_by_thorns",
        "destroyed_by_death_ray": "destroyed_by_death_ray",
    }
    return aliases.get(cleaned, cleaned)


def parse_battle_report(report_text: str) -> Dict[str, Any]:
    """Parse a copied in-game Battle Report using the standalone robust parser."""
    return parse_battle_report_text(report_text)

def format_large_number(value: Any) -> str:
    number = parse_tower_number(value)
    if number is None:
        return "—"
    if number == 0:
        return "0"
    for suffix, divisor in [("D", 1e33), ("N", 1e30), ("O", 1e27), ("S", 1e24), ("s", 1e21), ("Q", 1e18), ("q", 1e15), ("T", 1e12), ("B", 1e9), ("M", 1e6), ("K", 1e3)]:
        if abs(number) >= divisor:
            return f"{number / divisor:.2f}{suffix}"
    return f"{number:,.0f}"


def run_history_rows(profile: Dict[str, Any]) -> list[Dict[str, Any]]:
    rows = []
    for run in profile.get("runs", []):
        rows.append({
            "ID": run.get("id"),
            "Battle Date": run.get("battle_date"),
            "Tier": run.get("tier", 0),
            "Wave": run.get("wave", 0),
            "Killed By": run.get("killed_by", "Unknown"),
            "Real Time": run.get("real_time", ""),
            "Coins": run.get("coins_earned", 0),
            "Coins / Hour": run.get("coins_per_hour", 0),
            "Cells": run.get("cells_earned", 0),
            "Cells / Hour": run.get("cells_per_hour", 0),
        })
    return rows


def farming_tier_rows(profile: Dict[str, Any]) -> list[Dict[str, Any]]:
    grouped: Dict[int, list[Dict[str, Any]]] = {}
    for run in profile.get("runs", []):
        tier = int(run.get("tier", 0) or 0)
        if tier > 0:
            grouped.setdefault(tier, []).append(run)
    rows = []
    for tier, runs in sorted(grouped.items()):
        cph_values = [float(r.get("coins_per_hour", 0) or 0) for r in runs if float(r.get("coins_per_hour", 0) or 0) > 0]
        cell_values = [float(r.get("cells_per_hour", 0) or 0) for r in runs if float(r.get("cells_per_hour", 0) or 0) > 0]
        rows.append({
            "Tier": tier,
            "Runs": len(runs),
            "Best Wave": max(int(r.get("wave", 0) or 0) for r in runs),
            "Average CPH": sum(cph_values) / len(cph_values) if cph_values else 0,
            "Best CPH": max(cph_values) if cph_values else 0,
            "Average Cells/H": sum(cell_values) / len(cell_values) if cell_values else 0,
        })
    return rows


def profile_audit(profile: Dict[str, Any]) -> list[Dict[str, str]]:
    issues: list[Dict[str, str]] = []

    def add(severity: str, category: str, item: str, detail: str, action: str = "") -> None:
        issues.append({"Severity": severity, "Category": category, "Item": item, "Details": detail, "Suggested Action": action})

    # Level and Gold Box consistency.
    for section, metadata in [("workshop", WORKSHOP_MAX_LEVELS), ("labs", LAB_MAX_LEVELS), ("enhancements", ENHANCEMENT_MAX_LEVELS)]:
        values = profile.get(section, {})
        gold = profile.get("maxed", {}).get(section, {})
        for name, cap in metadata.items():
            raw = values.get(name)
            if raw is None:
                continue
            try:
                value = int(raw)
            except (TypeError, ValueError):
                add("Error", section.title(), name, f"Non-numeric level: {raw!r}", "Enter a whole-number level.")
                continue
            if value < 0:
                add("Error", section.title(), name, f"Negative level ({value}).", "Set the level to 0 or higher.")
            if value > cap:
                add("Error", section.title(), name, f"Level {value:,} exceeds the known maximum of {cap:,}.", "Clamp the value to the known maximum.")
            if gold.get(name, False) and value < cap:
                add("Warning", section.title(), name, f"Marked Gold at level {value:,}, below maximum {cap:,}.", "Use the safe-fix button or uncheck Gold.")
            if value >= cap and not gold.get(name, False):
                add("Info", section.title(), name, "At maximum but not marked Gold.", "Enable Auto Gold or mark it Gold manually.")

    # UW validation and sync checks.
    for uw_name in UW_NAMES:
        uw = profile.get("uw", {}).get(uw_name, {})
        if not uw.get("owned"):
            continue
        attrs = uw.get("attributes", {})
        missing = [name for name in UW_ATTRIBUTE_META[uw_name] if name not in attrs]
        if missing:
            add("Warning", "Ultimate Weapons", uw_name, "Missing attributes: " + ", ".join(missing), "Re-import the UW workbook or enter them manually.")
        for attr, meta in UW_ATTRIBUTE_META[uw_name].items():
            if attr not in attrs:
                continue
            value = float(attrs[attr])
            if meta.get("lower_is_better"):
                start = float(meta.get("start", value))
                if value < float(meta["max"]) or value > start:
                    add("Error", "Ultimate Weapons", f"{uw_name} — {attr}", f"Value {value:g} is outside the expected range {meta['max']:g}–{start:g}.")
            elif value < 0 or value > float(meta["max"]):
                add("Error", "Ultimate Weapons", f"{uw_name} — {attr}", f"Value {value:g} exceeds the expected maximum {meta['max']:g}.")

    gt = profile.get("uw", {}).get("Golden Tower", {})
    bh = profile.get("uw", {}).get("Black Hole", {})
    if gt.get("owned") and bh.get("owned"):
        gt_cd = to_number(gt.get("attributes", {}).get("Cooldown"))
        bh_cd = to_number(bh.get("attributes", {}).get("Cooldown"))
        if gt_cd and bh_cd and abs(gt_cd - bh_cd) > 0.01:
            add("Warning", "Ultimate Weapons", "GT/BH Sync", f"GT cooldown is {gt_cd:g}s and BH cooldown is {bh_cd:g}s.", "Confirm that this is an intentional partial sync.")

    # Modules.
    for slot, mod in profile.get("modules", {}).items():
        rarity = mod.get("rarity", "")
        level = int(mod.get("level", 0) or 0)
        if rarity and rarity in MODULE_RARITY_MAX_LEVELS and level > MODULE_RARITY_MAX_LEVELS[rarity]:
            add("Error", "Modules", slot, f"Level {level} exceeds {rarity}'s cap of {MODULE_RARITY_MAX_LEVELS[rarity]}.")
        if not mod.get("name"):
            add("Info", "Modules", slot, "No equipped module selected.", "Choose the active module or import Modules.")

    # Cards.
    for card, data in profile.get("cards", {}).get("items", {}).items():
        level = int(data.get("level", 0) or 0)
        mastery = int(data.get("mastery", 0) or 0)
        if level > 7 or level < 0:
            add("Error", "Cards", card, f"Card level {level} is outside 0–7.")
        if mastery > 9 or mastery < 0:
            add("Error", "Cards", card, f"Mastery {mastery} is outside 0–9.")

    coverage = completeness_rows(profile)
    for row in coverage:
        if row["Coverage %"] == 0:
            add("Warning", "Profile Coverage", row["Section"], "No data is available for this section.", "Import its companion workbook or enter it manually.")
        elif row["Coverage %"] < 50:
            add("Info", "Profile Coverage", row["Section"], f"Only {row['Coverage %']:.1f}% populated.")

    if not profile.get("runs"):
        add("Info", "Battle History", "Run data", "No Battle Reports have been saved.", "Paste a report to improve farming and bottleneck analysis.")
    return issues


def apply_safe_audit_fixes(profile: Dict[str, Any]) -> int:
    changed = 0
    for section, metadata in [("workshop", WORKSHOP_MAX_LEVELS), ("labs", LAB_MAX_LEVELS), ("enhancements", ENHANCEMENT_MAX_LEVELS)]:
        values = profile.setdefault(section, {})
        maxed = profile.setdefault("maxed", {}).setdefault(section, {})
        for name, cap in metadata.items():
            if name not in values:
                continue
            try:
                value = int(values[name])
            except (TypeError, ValueError):
                continue
            clamped = max(0, min(value, cap))
            if clamped != value:
                values[name] = clamped
                changed += 1
            should_gold = clamped >= cap
            if bool(maxed.get(name, False)) != should_gold:
                maxed[name] = should_gold
                changed += 1
    return changed


def ratio(value: Any, maximum: Any) -> float:
    try:
        maximum_f = float(maximum)
        value_f = float(value or 0)
    except (TypeError, ValueError):
        return 0.0
    if maximum_f <= 0:
        return 0.0
    return max(0.0, min(1.0, value_f / maximum_f))


def profile_confidence(profile: Dict[str, Any]) -> float:
    rows = completeness_rows(profile)
    important = {"Workshop", "Laboratory", "Ultimate Weapons", "Module Inventory", "Cards", "Relics"}
    values = [float(row["Coverage %"]) for row in rows if row["Section"] in important]
    return sum(values) / len(values) if values else 0.0


def build_analysis(profile: Dict[str, Any]) -> Dict[str, Any]:
    ws = profile.get("workshop", {})
    labs = profile.get("labs", {})
    uw = profile.get("uw", {})

    gt = uw.get("Golden Tower", {})
    bh = uw.get("Black Hole", {})
    sl = uw.get("Spotlight", {})
    cl = uw.get("Chain Lightning", {})
    bh_qty = 1 + int(labs.get("Extra Black Hole", 0) or 0) if bh.get("owned") else 0
    gt_cd = to_number(gt.get("attributes", {}).get("Cooldown"))
    bh_cd = to_number(bh.get("attributes", {}).get("Cooldown"))
    synced = bool(gt_cd and bh_cd and abs(gt_cd - bh_cd) <= 0.01)

    economy = 100 * (
        0.18 * ratio(labs.get("Coins / Kill Bonus"), LAB_MAX_LEVELS["Coins / Kill Bonus"]) +
        0.12 * ratio(ws.get("Coin / Kill Bonus"), WORKSHOP_MAX_LEVELS["Coin / Kill Bonus"]) +
        0.12 * ratio(labs.get("Black Hole Coin Bonus"), LAB_MAX_LEVELS["Black Hole Coin Bonus"]) +
        0.10 * ratio(labs.get("Golden Tower Bonus"), LAB_MAX_LEVELS["Golden Tower Bonus"]) +
        0.08 * ratio(labs.get("Golden Tower Duration"), LAB_MAX_LEVELS["Golden Tower Duration"]) +
        0.12 * (1.0 if gt.get("owned") else 0.0) +
        0.12 * (1.0 if bh.get("owned") else 0.0) +
        0.08 * (1.0 if bh_qty >= 2 else 0.0) +
        0.08 * (1.0 if synced else 0.0)
    )

    damage = 100 * (
        0.12 * ratio(ws.get("Damage"), WORKSHOP_MAX_LEVELS["Damage"]) +
        0.10 * ratio(ws.get("Attack Speed"), WORKSHOP_MAX_LEVELS["Attack Speed"]) +
        0.09 * ratio(ws.get("Critical Factor"), WORKSHOP_MAX_LEVELS["Critical Factor"]) +
        0.08 * ratio(ws.get("Damage / Meter"), WORKSHOP_MAX_LEVELS["Damage / Meter"]) +
        0.06 * ratio(ws.get("Super Critical Chance"), WORKSHOP_MAX_LEVELS["Super Critical Chance"]) +
        0.06 * ratio(ws.get("Super Critical Mult"), WORKSHOP_MAX_LEVELS["Super Critical Mult"]) +
        0.13 * ratio(labs.get("Damage"), LAB_MAX_LEVELS["Damage"]) +
        0.11 * ratio(labs.get("Attack Speed"), LAB_MAX_LEVELS["Attack Speed"]) +
        0.09 * ratio(labs.get("Critical Factor"), LAB_MAX_LEVELS["Critical Factor"]) +
        0.06 * (1.0 if sl.get("owned") else 0.0) +
        0.05 * (1.0 if cl.get("owned") else 0.0) +
        0.05 * ratio(labs.get("Shock Multiplier"), LAB_MAX_LEVELS["Shock Multiplier"])
    )

    survival = 100 * (
        0.18 * ratio(ws.get("Health"), WORKSHOP_MAX_LEVELS["Health"]) +
        0.13 * ratio(ws.get("Defense %"), WORKSHOP_MAX_LEVELS["Defense %"]) +
        0.08 * ratio(ws.get("Thorn Damage"), WORKSHOP_MAX_LEVELS["Thorn Damage"]) +
        0.08 * ratio(ws.get("Enemy Attack Level Skip"), WORKSHOP_MAX_LEVELS["Enemy Attack Level Skip"]) +
        0.10 * ratio(ws.get("Recovery Amount"), WORKSHOP_MAX_LEVELS["Recovery Amount"]) +
        0.10 * ratio(ws.get("Max Amount"), WORKSHOP_MAX_LEVELS["Max Amount"]) +
        0.08 * ratio(ws.get("Package Chance"), WORKSHOP_MAX_LEVELS["Package Chance"]) +
        0.10 * ratio(labs.get("Health"), LAB_MAX_LEVELS["Health"]) +
        0.07 * ratio(labs.get("Defense %"), LAB_MAX_LEVELS["Defense %"]) +
        0.08 * ratio(labs.get("Recovery Package Chance"), LAB_MAX_LEVELS["Recovery Package Chance"])
    )

    wall_present = int(ws.get("Wall Health", 0) or 0) > 0
    regen = 100 * (
        0.20 * ratio(ws.get("Health Regen"), WORKSHOP_MAX_LEVELS["Health Regen"]) +
        0.14 * ratio(labs.get("Health Regen"), LAB_MAX_LEVELS["Health Regen"]) +
        0.14 * ratio(ws.get("Recovery Amount"), WORKSHOP_MAX_LEVELS["Recovery Amount"]) +
        0.14 * ratio(ws.get("Package Chance"), WORKSHOP_MAX_LEVELS["Package Chance"]) +
        0.10 * ratio(labs.get("Recovery Package Amount"), LAB_MAX_LEVELS["Recovery Package Amount"]) +
        0.10 * ratio(labs.get("Recovery Package Max"), LAB_MAX_LEVELS["Recovery Package Max"]) +
        0.10 * ratio(labs.get("Recovery Package Chance"), LAB_MAX_LEVELS["Recovery Package Chance"]) +
        0.08 * (ratio(labs.get("Wall Regen"), LAB_MAX_LEVELS["Wall Regen"]) if wall_present else 0.0)
    )

    scores = {
        "Economy": round(economy, 1),
        "Damage": round(damage, 1),
        "Survivability": round(survival, 1),
        "Regen / Recovery": round(regen, 1),
    }
    weakest = min(scores, key=scores.get)
    latest = profile.get("runs", [])[-1] if profile.get("runs") else None
    killed_by = str(latest.get("killed_by", "")) if latest else ""

    recommendations: list[str] = []
    if weakest == "Economy":
        recommendations.extend(["Compare Coins/Kill, Black Hole Coin Bonus, Golden Tower Bonus, and GT duration in the future ROI engine.", "Verify GT and BH cooldown synchronization."])
    elif weakest == "Damage":
        recommendations.extend(["Review Damage, Critical Factor, Damage/Meter, Super Crit, and damage-oriented UW development.", "Use the latest run to distinguish boss/elite damage from general wave clear."])
    elif weakest == "Survivability":
        recommendations.extend(["Review Health, Defense %, recovery-package scaling, and Enemy Attack Level Skip.", "Check whether the run ends from a single large hit or sustained pressure."])
    else:
        recommendations.extend(["Review Health Regen, recovery packages, and Wall Regen if the Wall is unlocked.", "Compare recovery throughput against the cause of death in recent reports."])

    lower_kill = killed_by.lower()
    if "vampire" in lower_kill:
        recommendations.append("Recent Vampire death: inspect Garlic Thorns, elite damage/control, and recovery-package coverage.")
    elif "ray" in lower_kill:
        recommendations.append("Recent Ray death: inspect burst survivability, target priority, and elite control.")
    elif "scatter" in lower_kill:
        recommendations.append("Recent Scatter death: inspect multi-target damage, control, and package consistency.")
    elif "boss" in lower_kill:
        recommendations.append("Recent Boss death: inspect single-target damage, Plasma Cannon/Thorns interaction, and effective HP.")

    confidence = profile_confidence(profile)
    return {
        "scores": scores,
        "weakest": weakest,
        "confidence": round(confidence, 1),
        "bh_quantity": bh_qty,
        "gt_bh_synced": synced,
        "latest_killed_by": killed_by or "No report saved",
        "recommendations": recommendations,
        "method": "Heuristic account-development scoring; not the Effective Paths ROI engine.",
    }


# -----------------------------------------------------------------------------
# COMBINED RECOMMENDATION ENGINE (v1.3)
# -----------------------------------------------------------------------------
# The native engines use different ROI units: gain/day for labs, gain/stone for
# UWs, and gain/billion coins for enhancements.  This layer therefore does not
# pretend those raw ROI values are directly interchangeable.  It normalizes
# each path internally, then applies transparent bottleneck, recent-death,
# confidence, affordability, and rank weights to create an everyday shortlist.

COMBINED_PATH_META: Dict[str, Dict[str, str]] = {
    "econ_lab": {"domain": "Economy", "resource": "Lab", "label": "Economy Lab"},
    "econ_discount": {"domain": "Economy", "resource": "Lab", "label": "Lab Discount"},
    "econ_stone": {"domain": "Economy", "resource": "Stones", "label": "Economy Stones"},
    "econ_coin": {"domain": "Economy", "resource": "Coins", "label": "Economy Coins"},
    "damage_lab": {"domain": "Damage", "resource": "Lab", "label": "Damage Lab"},
    "damage_stone": {"domain": "Damage", "resource": "Stones", "label": "Damage Stones"},
    "damage_coin": {"domain": "Damage", "resource": "Coins", "label": "Damage Coins"},
    "damage_key": {"domain": "Damage", "resource": "Keys", "label": "Power Vault Keys"},
    "health_lab": {"domain": "Survivability", "resource": "Lab", "label": "eHP Lab"},
    "health_stone": {"domain": "Survivability", "resource": "Stones", "label": "eHP Stones"},
    "health_coin": {"domain": "Survivability", "resource": "Coins", "label": "eHP Coins"},
    "regen_lab": {"domain": "Regen / Recovery", "resource": "Lab", "label": "Regen Lab"},
    "regen_coin": {"domain": "Regen / Recovery", "resource": "Coins", "label": "Regen Coins"},
}


def combined_reference_rank(profile_data: Dict[str, Any], path_key: str, upgrade: Any) -> Optional[int]:
    reference = profile_data.get("roi_reference", {})
    paths = reference.get("paths", {}) if isinstance(reference, dict) else {}
    path = paths.get(path_key, {}) if isinstance(paths, dict) else {}
    target = normalize_econ_upgrade_name(upgrade)
    if not target:
        return None
    for index, row in enumerate(path.get("rows", []) or [], start=1):
        if normalize_econ_upgrade_name(row.get("Upgrade")) == target:
            return int(row.get("Rank") or index)
    return None


def combined_confidence_factor(value: Any) -> float:
    text = str(value or "Medium").lower()
    if "high" in text and "medium" not in text:
        return 1.00
    if "high" in text and "medium" in text:
        return 0.95
    if "medium" in text and "low" not in text:
        return 0.88
    if "low-medium" in text or "medium-low" in text:
        return 0.78
    if "low" in text:
        return 0.68
    return 0.85


def combined_resource_balance(profile_data: Dict[str, Any], resource: str) -> Optional[float]:
    key = {"Coins": "coins", "Stones": "stones", "Keys": "keys"}.get(resource)
    if key is None:
        return None
    value = native_number(profile_data.get("resources", {}).get(key), 0.0)
    return value if value > 0 else None


def combined_affordability(profile_data: Dict[str, Any], resource: str, row: Dict[str, Any]) -> tuple[str, Optional[bool]]:
    if resource == "Lab":
        coin_cost = native_number(row.get("Cost Numeric"), 0.0)
        coin_balance = native_number(profile_data.get("resources", {}).get("coins"), 0.0)
        if coin_cost > 0 and coin_balance > 0:
            affordable = coin_cost <= coin_balance
            return ("Affordable" if affordable else "Needs more coins", affordable)
        return "Queueable", True
    balance = combined_resource_balance(profile_data, resource)
    cost = native_number(row.get("Cost Numeric", row.get("Cost")), 0.0)
    if balance is None:
        return "Balance not entered", None
    affordable = cost <= balance + 1e-9
    return ("Affordable" if affordable else "Unaffordable", affordable)


def combined_domain_weight(
    domain: str,
    analysis: Dict[str, Any],
    death: str,
    upgrade: str,
    apply_death_weighting: bool,
    focus: str,
) -> tuple[float, list[str]]:
    weight = 1.0
    reasons: list[str] = []
    weakest = str(analysis.get("weakest", ""))
    if domain == weakest:
        weight *= 1.28
        reasons.append(f"{domain} is your lowest development score")

    if focus != "Balanced":
        focus_domain = {
            "Economy": "Economy",
            "Damage": "Damage",
            "Survival": "Survivability",
            "Recovery": "Regen / Recovery",
        }.get(focus)
        if focus_domain == domain:
            weight *= 1.30
            reasons.append(f"manual {focus.lower()} focus")

    if apply_death_weighting:
        lower = death.lower()
        name = upgrade.lower()
        if "vampire" in lower:
            if domain == "Regen / Recovery":
                weight *= 1.35
                reasons.append("recent Vampire death favors sustain")
            elif domain == "Survivability":
                weight *= 1.16
            elif domain == "Damage":
                weight *= 1.08
            if any(term in name for term in ["garlic", "thorn", "regen", "recovery", "package"]):
                weight *= 1.18
                reasons.append("direct Vampire-response upgrade")
        elif "boss" in lower:
            if domain == "Damage":
                weight *= 1.28
                reasons.append("recent Boss death favors single-target damage")
            elif domain == "Survivability":
                weight *= 1.14
        elif "ray" in lower:
            if domain == "Survivability":
                weight *= 1.30
                reasons.append("recent Ray death favors burst eHP")
            elif domain == "Regen / Recovery":
                weight *= 1.08
        elif "scatter" in lower:
            if domain == "Damage":
                weight *= 1.20
                reasons.append("recent Scatter death favors clear/control damage")
            elif domain == "Regen / Recovery":
                weight *= 1.15
        elif "fast" in lower:
            if domain in {"Damage", "Survivability"}:
                weight *= 1.12
                reasons.append("recent Fast death favors kill speed or eHP")

    if domain == "Economy" and not bool(analysis.get("gt_bh_synced", False)):
        if any(term in upgrade.lower() for term in ["golden tower", "black hole", "gt |", "bh |"]) and "cooldown" not in upgrade.lower():
            weight *= 1.08
            reasons.append("GT/BH economy core needs refinement")
    return weight, reasons


def build_combined_recommendations(
    profile_data: Dict[str, Any],
    steps: int = 12,
    candidates_per_path: int = 3,
    apply_death_weighting: bool = True,
    focus: str = "Balanced",
) -> Dict[str, Any]:
    econ = build_native_econ_paths(profile_data, steps)
    damage = build_native_damage_paths(profile_data, steps)
    health = build_native_health_paths(profile_data, steps)
    paths: Dict[str, list[Dict[str, Any]]] = {}
    paths.update({key: value for key, value in econ.items() if isinstance(value, list)})
    paths.update({key: value for key, value in damage.items() if isinstance(value, list)})
    paths.update({key: value for key, value in health.items() if isinstance(value, list)})

    analysis = build_analysis(profile_data)
    death = native_latest_death(profile_data) or str(analysis.get("latest_killed_by", ""))
    rows: list[Dict[str, Any]] = []

    for path_key, meta in COMBINED_PATH_META.items():
        path_rows = paths.get(path_key, []) or []
        if not path_rows:
            continue
        usable = [row for row in path_rows if not recommendation_is_gold_boxed(profile_data, path_key, str(row.get("Upgrade", "")))]
        if not usable:
            continue
        positive_rois = [max(0.0, native_number(row.get("ROI Numeric", row.get("ROI")), 0.0)) for row in usable]
        top_roi = max(positive_rois) if positive_rois else 0.0
        positive_gains = [max(0.0, native_number(row.get("Gain %"), 0.0)) for row in usable]
        top_gain = max(positive_gains) if positive_gains else 0.0

        for index, source_row in enumerate(usable[:max(1, candidates_per_path)], start=1):
            row = dict(source_row)
            upgrade = str(row.get("Upgrade", "Unknown"))
            rank = int(row.get("Rank") or index)
            roi = max(0.0, native_number(row.get("ROI Numeric", row.get("ROI")), 0.0))
            gain = max(0.0, native_number(row.get("Gain %"), 0.0))
            roi_norm = roi / top_roi if top_roi > 0 else 0.0
            gain_norm = gain / top_gain if top_gain > 0 else 0.0
            rank_norm = 1.0 / (rank ** 0.55)
            confidence_factor = combined_confidence_factor(row.get("Confidence"))
            domain_weight, urgency_reasons = combined_domain_weight(
                meta["domain"], analysis, death, upgrade, apply_death_weighting, focus
            )
            affordability_text, affordable = combined_affordability(profile_data, meta["resource"], row)
            affordability_factor = 1.04 if affordable is True else (0.82 if affordable is False else 0.94)
            reference_rank = combined_reference_rank(profile_data, path_key, upgrade)
            reference_factor = 1.04 if reference_rank is not None and reference_rank <= 5 else 1.0
            priority = 100.0 * (0.60 * roi_norm + 0.24 * rank_norm + 0.16 * gain_norm)
            priority *= domain_weight * confidence_factor * affordability_factor * reference_factor

            reason_parts = urgency_reasons[:]
            if reference_rank is not None:
                reason_parts.append(f"Effective Paths reference rank {reference_rank}")
            if affordable is False:
                reason_parts.append("save-up target")
            elif affordable is True and meta["resource"] != "Lab":
                reason_parts.append("currently affordable")
            model_why = str(row.get("Why", "")).strip()
            if model_why:
                reason_parts.append(model_why)

            rows.append({
                "Priority Index": priority,
                "Domain": meta["domain"],
                "Resource": meta["resource"],
                "Path": meta["label"],
                "Path Key": path_key,
                "Path Rank": rank,
                "Upgrade": upgrade,
                "Next Level": row.get("Level", "—"),
                "Cost / Time": row.get("Cost") or row.get("Duration") or "—",
                "Cost Numeric": native_number(row.get("Cost Numeric"), 0.0),
                "Estimated Gain %": gain,
                "Native ROI": row.get("ROI"),
                "Affordability": affordability_text,
                "Affordable Bool": affordable,
                "Reference Rank": reference_rank,
                "Confidence": row.get("Confidence", "Medium"),
                "Why": "; ".join(reason_parts),
            })

    # Keep the strongest occurrence when the same upgrade appears more than once
    # in the same resource class while retaining cross-resource alternatives.
    deduped: Dict[tuple[str, str], Dict[str, Any]] = {}
    for row in rows:
        key = (str(row["Resource"]), normalize_econ_upgrade_name(row["Upgrade"]))
        if key not in deduped or row["Priority Index"] > deduped[key]["Priority Index"]:
            deduped[key] = row
    ranked = sorted(deduped.values(), key=lambda item: item["Priority Index"], reverse=True)

    weakest_domain = str(analysis.get("weakest", ""))
    bottleneck_rows = [row for row in ranked if row["Domain"] == weakest_domain]
    if "vampire" in death.lower():
        extra = [row for row in ranked if row["Domain"] in {"Regen / Recovery", "Survivability"}]
        seen = {(row["Resource"], row["Upgrade"]) for row in bottleneck_rows}
        bottleneck_rows.extend(row for row in extra if (row["Resource"], row["Upgrade"]) not in seen)
        bottleneck_rows.sort(key=lambda item: item["Priority Index"], reverse=True)

    return {
        "rows": ranked,
        "by_resource": {
            resource: [row for row in ranked if row["Resource"] == resource]
            for resource in ["Lab", "Coins", "Stones", "Keys"]
        },
        "affordable": [row for row in ranked if row.get("Affordable Bool") is True],
        "long_term": [row for row in ranked if row.get("Affordable Bool") is False],
        "bottleneck": bottleneck_rows,
        "analysis": analysis,
        "latest_death": death or "No report saved",
        "method": (
            "Cross-category priority index. Raw ROI is normalized only within its own path because lab-time, stone, coin, and key ROI units are not directly comparable."
        ),
    }


def combined_display_frame(rows: list[Dict[str, Any]], limit: int = 12) -> pd.DataFrame:
    columns = [
        "Priority Index", "System", "Domain", "Resource", "Upgrade", "Next Level",
        "Cost / Time", "Estimated Gain %", "Affordability", "Reference Rank",
        "Confidence", "Why",
    ]
    frame = pd.DataFrame(rows[:limit])
    if frame.empty:
        return frame
    return frame[[column for column in columns if column in frame.columns]]

# -----------------------------------------------------------------------------
# STYLING
# -----------------------------------------------------------------------------

st.markdown(
    """
    <style>
        .gold-badge {display:inline-block;padding:.18rem .55rem;border-radius:.45rem;
            background:linear-gradient(135deg,#8d6e00,#f6d365,#fff1a8);
            color:#171000;font-weight:800;border:1px solid #ffe28a;margin-left:.35rem;}
        .known-cap {display:inline-block;color:#8fc7ff;font-size:.82rem;margin-left:.35rem;}
        .section-note {color:#aeb7c2;font-size:.9rem;}
    </style>
    """,
    unsafe_allow_html=True,
)


def level_input_with_gold(section: str, item_name: str) -> None:
    profile = st.session_state.profile
    revision = st.session_state.profile_revision
    profile[section].setdefault(item_name, 0)
    profile["maxed"][section].setdefault(item_name, False)
    profile["custom_max"][section].setdefault(item_name, 0)

    cap = effective_max(section, item_name)
    value = int(profile[section].get(item_name, 0) or 0)
    if cap is not None:
        value = min(value, cap)
        if profile["settings"].get("auto_gold_at_max", True) and value >= cap:
            profile["maxed"][section][item_name] = True

    row = st.container(border=True)
    with row:
        label_col, gold_col = st.columns([5, 1.3])
        gold_key = f"gold_{section}_{item_name}_{revision}"
        level_key = f"level_{section}_{item_name}_{revision}"
        with gold_col:
            gold = st.checkbox(
                "Gold Box", value=bool(profile["maxed"][section].get(item_name, False)),
                key=gold_key, help="Set this entry to its Effective Paths maximum."
            )
            profile["maxed"][section][item_name] = gold
        if gold and cap is not None:
            value = cap
            profile[section][item_name] = cap
        with label_col:
            badge = '<span class="gold-badge">GOLD</span>' if gold else ""
            st.markdown(
                f"**{item_name}** {badge} <span class='known-cap'>{format_cap_text(section, item_name)}</span>",
                unsafe_allow_html=True,
            )
            kwargs: Dict[str, Any] = {
                "label": f"{item_name} level", "min_value": 0, "value": int(value),
                "step": 1, "key": level_key, "label_visibility": "collapsed",
                "disabled": gold,
            }
            if cap is not None:
                kwargs["max_value"] = cap
            profile[section][item_name] = int(st.number_input(**kwargs))


def uw_attribute_input(uw_name: str, attribute: str) -> None:
    profile = st.session_state.profile
    revision = st.session_state.profile_revision
    uw = profile["uw"].setdefault(uw_name, {"owned": False, "attributes": {}})
    attrs = uw.setdefault("attributes", {})
    gold_map = profile["maxed"]["uw"].setdefault(uw_name, {})
    meta = UW_ATTRIBUTE_META[uw_name][attribute]
    default = meta.get("start", 0) if meta.get("lower_is_better") else 0
    value = attrs.get(attribute, default)
    gold = bool(gold_map.get(attribute, False))

    if profile["settings"].get("auto_gold_at_max", True):
        if meta.get("lower_is_better") and float(value) <= float(meta["max"]):
            gold = True
        elif not meta.get("lower_is_better") and float(value) >= float(meta["max"]):
            gold = True
    if gold:
        value = meta["max"]
        attrs[attribute] = value
    gold_map[attribute] = gold

    with st.container(border=True):
        c1, c2 = st.columns([5, 1.3])
        with c2:
            gold = st.checkbox(
                "Gold Box", value=gold,
                key=f"uw_gold_{uw_name}_{attribute}_{revision}",
            )
            gold_map[attribute] = gold
        if gold:
            value = meta["max"]
            attrs[attribute] = value
        with c1:
            badge = '<span class="gold-badge">GOLD</span>' if gold else ""
            if meta.get("display") == "percent":
                cap_text = f"Max {float(meta['max']) * 100:g}%"
            else:
                cap_text = f"Max {meta['max']:g}"
            st.markdown(
                f"**{attribute}** {badge} <span class='known-cap'>{cap_text}</span>",
                unsafe_allow_html=True,
            )
            is_int = isinstance(meta["max"], int) and not isinstance(meta["max"], bool)
            if is_int:
                attrs[attribute] = int(st.number_input(
                    attribute, min_value=0, max_value=int(meta.get("start", meta["max"]) if meta.get("lower_is_better") else meta["max"]),
                    value=int(value), step=max(1, int(meta.get("step", 1))),
                    key=f"uw_attr_{uw_name}_{attribute}_{revision}", label_visibility="collapsed", disabled=gold,
                ))
            else:
                max_input = float(meta.get("start", meta["max"]) if meta.get("lower_is_better") else meta["max"])
                attrs[attribute] = float(st.number_input(
                    attribute, min_value=0.0, max_value=max_input,
                    value=float(value), step=float(meta.get("step", 0.1)),
                    format="%.3f" if meta.get("display") == "percent" else "%.2f",
                    key=f"uw_attr_{uw_name}_{attribute}_{revision}", label_visibility="collapsed", disabled=gold,
                ))


# -----------------------------------------------------------------------------
# SESSION INITIALIZATION
# -----------------------------------------------------------------------------

if "profile" not in st.session_state:
    st.session_state.profile = default_profile()
else:
    st.session_state.profile = ensure_profile_shape(st.session_state.profile)
if "profile_revision" not in st.session_state:
    st.session_state.profile_revision = 0


# v1.6 modular engine APIs. The original implementations remain above as a
# safe rollback reference, while all UI calls below use the standalone modules.
from .engines.economy import native_econ_context, native_econ_score, build_native_econ_paths
from .engines.damage import native_damage_settings, native_damage_components, build_native_damage_paths
from .engines.health import native_health_settings, native_health_components, build_native_health_paths
from .engines.regen import native_regen_components
from .engines.combined import build_combined_recommendations, build_progression_recommendations
from .regression import run_engine_health, bundled_data_status
from .calibration import build_calibration_report, calibration_snapshot, compare_snapshots, PATH_LABELS
from .quality import profile_quality_report, apply_safe_fixes as apply_quality_safe_fixes
from .explanations import recommendation_explanation
from .battle_ui import render_battle_learning_page
from .planner import (
    GOAL_OPTIONS, GOAL_FOCUS, ensure_planner_state, effective_income_rates,
    build_progression_plan, build_weekly_report, queue_add, queue_set_status,
    queue_toggle_lock, queue_move, queue_remove, queue_undo, queue_rows,
)
from .visual_ui import (
    apply_visual_theme, render_app_header, render_card_deck_page, render_icon_studio_page,
    render_module_forge_page, render_overview_page, render_relic_gallery_page,
    render_sync_center_page, render_visual_sidebar, render_grouped_navigation,
)
from .visual_models import build_card_report, build_module_forge_report, build_relic_report, build_sync_report
from .save_import_ui import render_player_save_import

profile = st.session_state.profile
ensure_planner_state(profile)
ensure_battle_learning_state(profile)
revision = st.session_state.profile_revision

render_visual_sidebar(profile, APP_VERSION)
apply_visual_theme(profile)
render_app_header(profile, APP_VERSION)

page = render_grouped_navigation()

st.sidebar.divider()
st.sidebar.subheader("Display")
profile["settings"]["auto_gold_at_max"] = st.sidebar.checkbox(
    "Auto Gold at maximum", value=bool(profile["settings"].get("auto_gold_at_max", True))
)
profile["settings"]["show_only_incomplete"] = st.sidebar.checkbox(
    "Show only incomplete entries", value=bool(profile["settings"].get("show_only_incomplete", False))
)
st.sidebar.divider()
if st.sidebar.button("Save current profile", use_container_width=True):
    save_profile(profile["name"], profile)
    st.sidebar.success("Profile saved.")


# -----------------------------------------------------------------------------
# PAGES
# -----------------------------------------------------------------------------

if page == "Overview 2.0":
    render_overview_page(profile)

elif page == "Sync Center":
    render_sync_center_page(profile)

elif page == "Card Deck":
    render_card_deck_page(profile)

elif page == "Module Forge":
    render_module_forge_page(profile)

elif page == "Relic Gallery":
    render_relic_gallery_page(profile)

elif page == "Icon Studio":
    render_icon_studio_page(profile)

elif page == "Setup Wizard":
    st.header("Profile Setup Wizard")
    st.caption("A guided path from a blank install to trustworthy standalone recommendations. Effective Paths is optional.")

    coverage = completeness_rows(profile)
    total_filled = sum(row["Filled"] for row in coverage)
    total_expected = sum(row["Expected"] for row in coverage)
    coverage_percent = 100.0 * total_filled / total_expected if total_expected else 0.0
    quality_preview = profile_quality_report(profile)
    c1, c2, c3, c4 = st.columns(4)
    c1.metric("Profile", profile.get("name", "default"))
    c2.metric("Coverage", f"{coverage_percent:.1f}%")
    c3.metric("Quality", f"{quality_preview.get('score', 0)}/100")
    c4.metric("Battle reports", len(profile.get("runs", [])))
    st.progress(min(max(coverage_percent / 100.0, 0.0), 1.0))

    wizard_tabs = st.tabs(["1. Profile", "2. Import Data", "3. Resources & Goal", "4. Review & Finish"])

    with wizard_tabs[0]:
        st.subheader("Choose how to begin")
        mode = st.radio(
            "Starting point",
            ["Keep current profile", "Create a blank profile", "Load a saved profile", "Restore a profile JSON"],
            horizontal=True,
            key="wizard_start_mode",
        )
        if mode == "Keep current profile":
            st.success(f"Continuing with **{profile.get('name', 'default')}**.")
        elif mode == "Create a blank profile":
            new_name = st.text_input("New profile name", value="new_profile", key="wizard_new_profile_name")
            if st.button("Create blank profile", type="primary", key="wizard_create_blank"):
                new_profile = default_profile()
                new_profile["name"] = safe_profile_filename(new_name)
                new_profile["setup_wizard"]["last_step"] = "Import Data"
                st.session_state.profile = new_profile
                save_profile(new_profile["name"], new_profile)
                bump_revision(); st.rerun()
        elif mode == "Load a saved profile":
            saved_profiles = list_profiles()
            selected_profile = st.selectbox(
                "Saved profile", saved_profiles if saved_profiles else ["No saved profiles"],
                disabled=not saved_profiles, key="wizard_saved_profile",
            )
            if st.button("Load selected profile", type="primary", disabled=not saved_profiles, key="wizard_load_profile"):
                st.session_state.profile = load_profile(selected_profile)
                bump_revision(); st.rerun()
        else:
            profile_upload = st.file_uploader("Profile JSON", type=["json"], key="wizard_profile_json")
            if profile_upload is not None:
                try:
                    restored_profile = ensure_profile_shape(json.loads(profile_upload.getvalue().decode("utf-8")))
                    st.success(f"Ready to restore profile **{restored_profile.get('name', 'uploaded')}**.")
                    if st.button("Restore uploaded profile", type="primary", key="wizard_restore_json"):
                        st.session_state.profile = restored_profile
                        save_profile(restored_profile.get("name", "uploaded"), restored_profile)
                        bump_revision(); st.rerun()
                except Exception as exc:
                    st.error(f"Could not read profile JSON: {exc}")

    with wizard_tabs[1]:
        st.subheader("Import player data")
        st.write("The IDS companion bundle is the fastest complete setup. You can also skip imports and enter values manually later.")
        ids_uploads = st.file_uploader(
            "IDS companion workbooks", type=["xlsx"], accept_multiple_files=True, key="wizard_ids_bundle"
        )
        wizard_results = []
        if ids_uploads:
            for upload in ids_uploads:
                try:
                    wizard_results.append(parse_companion_upload(upload))
                except Exception as exc:
                    wizard_results.append({"filename": upload.name, "kind": "unknown", "version": "", "patch": {}, "recognized": 0, "error": str(exc)})
            st.dataframe(
                [{
                    "File": row.get("filename"), "Type": row.get("kind"), "Version": row.get("version"),
                    "Recognized": row.get("recognized", 0), "Status": row.get("error", "Ready"),
                } for row in wizard_results],
                use_container_width=True, hide_index=True,
            )
            successful = [row for row in wizard_results if not row.get("error")]
            replace_imported = st.checkbox("Replace imported sections", value=False, key="wizard_replace_ids")
            if st.button("Apply IDS companion bundle", type="primary", disabled=not successful, key="wizard_apply_ids"):
                apply_companion_bundle(successful, replace=replace_imported)
                profile["setup_wizard"]["last_step"] = "Resources & Goal"
                save_profile(profile["name"], profile)
                bump_revision(); st.rerun()

        st.divider()
        render_player_save_import(
            profile,
            save_profile=save_profile,
            bump_revision=bump_revision,
            key_prefix="wizard_save",
        )

        st.divider()
        st.markdown("**Optional Effective Paths import**")
        st.caption("Use this to fill Master Sheet inputs and/or save a regression reference. The standalone engines do not require it.")
        wizard_ep = st.file_uploader("Filled Effective Paths workbook", type=["xlsx", "csv"], key="wizard_ep_upload")
        if wizard_ep is not None:
            try:
                ep_inputs = parse_uploaded_effective_paths(wizard_ep)
                st.success("Effective Paths inputs parsed.")
                metrics = st.columns(4)
                metrics[0].metric("Workshop", len(ep_inputs.get("workshop", {})))
                metrics[1].metric("Labs", len(ep_inputs.get("labs", {})))
                metrics[2].metric("Enhancements", len(ep_inputs.get("enhancements", {})))
                metrics[3].metric("UWs", len(ep_inputs.get("uw", {})))
                import_ep_inputs = st.checkbox("Apply profile inputs", value=True, key="wizard_apply_ep_inputs")
                import_ep_reference = st.checkbox(
                    "Also import ROI calibration reference", value=Path(wizard_ep.name).suffix.lower() == ".xlsx",
                    disabled=Path(wizard_ep.name).suffix.lower() != ".xlsx", key="wizard_apply_ep_reference",
                )
                if st.button("Apply selected Effective Paths data", type="primary", key="wizard_apply_ep"):
                    if import_ep_inputs:
                        apply_import(ep_inputs, replace=False)
                    if import_ep_reference:
                        reference = parse_effective_paths_roi_reference(wizard_ep)
                        apply_roi_reference(reference)
                    profile["setup_wizard"]["last_step"] = "Resources & Goal"
                    save_profile(profile["name"], profile)
                    bump_revision(); st.rerun()
            except Exception as exc:
                st.error(f"Could not parse Effective Paths: {exc}")

    with wizard_tabs[2]:
        st.subheader("Enter current balances and goal")
        resource_names = ["coins", "stones", "gems", "medals", "keys", "bits", "reroll_shards", "module_shards"]
        for start in range(0, len(resource_names), 4):
            resource_columns = st.columns(4)
            for column, currency in zip(resource_columns, resource_names[start:start + 4]):
                profile["resources"][currency] = int(column.number_input(
                    currency.replace("_", " ").title(), min_value=0, value=int(profile["resources"].get(currency, 0) or 0),
                    step=1, key=f"wizard_resource_{currency}_{revision}",
                ))
        goal_options = ["Balanced", "Economy", "Damage", "Survival", "Recovery"]
        combined_settings = profile.setdefault("combined_recommendations", {}).setdefault("settings", {})
        current_goal = str(combined_settings.get("focus", "Balanced"))
        if current_goal not in goal_options:
            current_goal = "Balanced"
        goal = st.selectbox("Primary goal", goal_options, index=goal_options.index(current_goal), key="wizard_goal")
        combined_settings["focus"] = goal
        player_cols = st.columns(2)
        profile["player"]["farming_tier"] = player_cols[0].text_input(
            "Typical farming tier", value=str(profile["player"].get("farming_tier", "")), key=f"wizard_farming_tier_{revision}"
        )
        profile["player"]["tourney_league"] = player_cols[1].text_input(
            "Tournament league", value=str(profile["player"].get("tourney_league", "")), key=f"wizard_tourney_{revision}"
        )
        use_death = st.checkbox(
            "Use recent Battle Report deaths in recommendations",
            value=bool(combined_settings.get("death_weighting", True)), key="wizard_death_weighting",
        )
        combined_settings["death_weighting"] = use_death
        if st.button("Save resources and goal", type="primary", key="wizard_save_goal"):
            profile["setup_wizard"]["last_step"] = "Review & Finish"
            save_profile(profile["name"], profile)
            st.success("Resources and recommendation goal saved.")

    with wizard_tabs[3]:
        st.subheader("Review setup")
        st.dataframe(coverage, use_container_width=True, hide_index=True)
        quality = profile_quality_report(profile)
        q1, q2, q3, q4 = st.columns(4)
        q1.metric("Quality score", f"{quality.get('score', 0)}/100")
        q2.metric("Errors", quality.get("counts", {}).get("Error", 0))
        q3.metric("Warnings", quality.get("counts", {}).get("Warning", 0))
        q4.metric("Information", quality.get("counts", {}).get("Info", 0))
        blocking = [row for row in quality.get("issues", []) if row.get("Severity") == "Error"]
        if blocking:
            st.error("Resolve the error-level findings before treating recommendations as reliable.")
            st.dataframe(blocking, use_container_width=True, hide_index=True)
        elif quality.get("counts", {}).get("Warning", 0):
            st.warning("Setup is usable, but Data Quality contains warnings worth reviewing.")
        else:
            st.success("No error or warning-level data-quality findings were detected.")
        if st.button("Finish setup and save profile", type="primary", key="wizard_finish"):
            profile["setup_wizard"]["completed"] = True
            profile["setup_wizard"]["last_step"] = "Complete"
            save_profile(profile["name"], profile)
            st.success("Setup complete. Open Recommendation Dashboard for the practical shortlist.")

elif page == "Profile Setup":
    st.header("Profile Setup")
    profile_name = st.text_input("Profile name", value=profile.get("name", "default"), key=f"profile_name_{revision}")
    col1, col2, col3 = st.columns(3)
    with col1:
        if st.button("Create / Save Profile", use_container_width=True):
            clean_name = safe_profile_filename(profile_name)
            profile["name"] = clean_name
            save_profile(clean_name, profile)
            st.success(f"Saved profile: {clean_name}")
    profiles = list_profiles()
    with col2:
        selected = st.selectbox("Load profile", profiles if profiles else ["No saved profiles"], disabled=not profiles, key=f"profile_load_{revision}")
        if st.button("Load", disabled=not profiles, use_container_width=True):
            st.session_state.profile = load_profile(selected)
            bump_revision(); st.rerun()
    with col3:
        st.write("")
        if st.button("New Blank Profile", use_container_width=True):
            st.session_state.profile = default_profile()
            bump_revision(); st.rerun()

elif page == "Import / Export":
    st.header("Import / Export")
    bundle_tab, save_tab, ep_tab, roi_tab, profile_tab = st.tabs(
        ["IDS Companion Bundle", "Game Save", "Effective Paths Inputs", "ROI Reference", "Profile JSON"]
    )

    with bundle_tab:
        st.write("Upload any combination of the IDS companion `.xlsx` workbooks. The app identifies each file from its EXPORT signature and merges it into the canonical profile.")
        uploads = st.file_uploader("IDS companion workbooks", type=["xlsx"], accept_multiple_files=True, key="ids_bundle")
        if uploads:
            results = []
            for upload in uploads:
                try:
                    results.append(parse_companion_upload(upload))
                except Exception as exc:
                    results.append({"filename": upload.name, "kind": "unknown", "version": "", "patch": {}, "recognized": 0, "error": str(exc)})
            st.session_state.bundle_preview = results
        results = st.session_state.get("bundle_preview", [])
        if results:
            audit_rows = [{"File": r["filename"], "Type": r.get("kind"), "Version": r.get("version"), "Recognized groups": r.get("recognized", 0), "Status": r.get("error", "Ready")} for r in results]
            st.dataframe(audit_rows, use_container_width=True, hide_index=True)
            successful = [r for r in results if not r.get("error")]
            replace_bundle = st.checkbox("Replace imported sections instead of merging", value=False, key="replace_bundle")
            if st.button("Apply IDS bundle", type="primary", disabled=not successful):
                apply_companion_bundle(successful, replace=replace_bundle)
                save_profile(profile["name"], profile)
                st.session_state.pop("bundle_preview", None)
                bump_revision(); st.rerun()

    with save_tab:
        render_player_save_import(
            profile,
            save_profile=save_profile,
            bump_revision=bump_revision,
            key_prefix="import_save",
        )

    with ep_tab:
        st.write("Upload a filled Effective Paths workbook or a CSV export of its **Master Sheet**. This fills the optimizer-focused fields only.")
        uploaded = st.file_uploader("Effective Paths file", type=["xlsx", "csv"], key="ep_upload")
        if uploaded is not None:
            try:
                imported = parse_uploaded_effective_paths(uploaded)
                st.session_state.import_preview = imported
                st.session_state.import_filename = uploaded.name
                st.success("File parsed successfully. Review the preview before applying it.")
            except Exception as exc:
                st.session_state.pop("import_preview", None)
                st.error(str(exc))
        imported = st.session_state.get("import_preview")
        if imported:
            m1, m2, m3, m4 = st.columns(4)
            m1.metric("Workshop", len(imported["workshop"])); m2.metric("Labs", len(imported["labs"])); m3.metric("Enhancements", len(imported["enhancements"])); m4.metric("Ultimate Weapons", len(imported["uw"]))
            replace = st.checkbox("Replace existing Workshop, Lab, Enhancement and UW values", value=False, key="replace_ep")
            if st.button("Apply Effective Paths import", type="primary"):
                apply_import(imported, replace); save_profile(profile["name"], profile)
                st.session_state.pop("import_preview", None); bump_revision(); st.rerun()

    with roi_tab:
        st.write(
            "Upload the filled and recalculated Effective Paths `.xlsx` workbook. "
            "This imports the cached eEcon, eDamage, eHP, and wall-regen recommendation paths as a verification reference."
        )
        st.info(
            "v0.9 reads the spreadsheet's calculated outputs; it does not yet claim that every formula has been independently recreated in Python. "
            "These saved paths are the regression target for the native engine."
        )
        roi_upload = st.file_uploader("Filled Effective Paths workbook", type=["xlsx"], key="roi_reference_upload")
        if roi_upload is not None:
            try:
                reference = parse_effective_paths_roi_reference(roi_upload)
                st.session_state.roi_reference_preview = reference
                st.success("ROI paths parsed successfully. Review the coverage below before applying.")
            except Exception as exc:
                st.session_state.pop("roi_reference_preview", None)
                st.error(str(exc))
        reference = st.session_state.get("roi_reference_preview")
        if reference:
            preview_rows = []
            for key, path in reference.get("paths", {}).items():
                preview_rows.append({
                    "Path": path.get("title", key),
                    "Rows": len(path.get("rows", [])),
                    "First recommendation": path.get("rows", [{}])[0].get("Upgrade") if path.get("rows") else "—",
                })
            st.dataframe(preview_rows, use_container_width=True, hide_index=True)
            if reference.get("warnings"):
                for warning in reference["warnings"]:
                    st.warning(warning)
            if st.button("Apply ROI reference", type="primary", key="apply_roi_reference"):
                apply_roi_reference(reference)
                save_profile(profile["name"], profile)
                st.session_state.pop("roi_reference_preview", None)
                bump_revision(); st.rerun()
        current_reference = profile.get("roi_reference", {})
        if current_reference.get("imported_at"):
            st.subheader("Current saved reference")
            source = current_reference.get("source", {})
            st.json(source, expanded=False)
            st.download_button(
                "Download ROI reference JSON",
                data=json.dumps(current_reference, indent=2),
                file_name=f"{safe_profile_filename(profile['name'])}_roi_reference.json",
                mime="application/json",
                key="download_roi_reference",
            )

    with profile_tab:
        st.download_button("Download current profile JSON", data=json.dumps(profile, indent=2), file_name=f"{safe_profile_filename(profile['name'])}.json", mime="application/json")
        json_upload = st.file_uploader("Restore a profile JSON", type=["json"], key="json_restore")
        if json_upload is not None and st.button("Load uploaded JSON"):
            try:
                st.session_state.profile = ensure_profile_shape(json.loads(json_upload.getvalue().decode("utf-8")))
                bump_revision(); st.rerun()
            except Exception as exc:
                st.error(f"Could not load JSON: {exc}")

elif page == "Dashboard":
    st.header("Dashboard")
    st.subheader(profile.get("name", "default"))
    resource_names = ["coins", "stones", "gems", "medals", "keys", "bits", "reroll_shards", "module_shards"]
    for start in range(0, len(resource_names), 4):
        cols = st.columns(4)
        for col, currency in zip(cols, resource_names[start:start + 4]):
            profile["resources"][currency] = int(col.number_input(currency.replace("_", " ").title(), min_value=0, value=int(profile["resources"].get(currency, 0) or 0), step=1, key=f"resource_{currency}_{revision}"))
    coverage = completeness_rows(profile)
    total_filled = sum(r["Filled"] for r in coverage); total_expected = sum(r["Expected"] for r in coverage)
    overall = 100 * total_filled / total_expected if total_expected else 0
    st.metric("Profile coverage", f"{overall:.1f}%")
    st.dataframe(coverage, use_container_width=True, hide_index=True)
    if profile["metadata"].get("last_import"):
        st.info(f"Last import: {profile['metadata']['last_import']['source']} at {profile['metadata']['last_import']['at']}")

elif page == "Profile Completeness":
    st.header("Profile Completeness")
    coverage = completeness_rows(profile)
    st.dataframe(coverage, use_container_width=True, hide_index=True)
    missing_sources = [r["Section"] for r in coverage if r["Coverage %"] < 100]
    if missing_sources:
        st.warning("Incomplete sections: " + ", ".join(missing_sources))
    else:
        st.success("All represented sections are complete.")
    if profile.get("import_audit"):
        st.subheader("Latest Import Audit")
        st.dataframe(profile["import_audit"], use_container_width=True, hide_index=True)

elif page == "Data Quality":
    st.header("Data Quality")
    st.caption("Checks the profile before recommendations are trusted. Safe fixes only rename known aliases and correct Gold flags; they do not delete data or clamp levels.")

    quality = profile_quality_report(profile)
    q1, q2, q3, q4 = st.columns(4)
    q1.metric("Quality score", f"{quality.get('score', 0)}/100")
    q2.metric("Errors", quality.get("counts", {}).get("Error", 0))
    q3.metric("Warnings", quality.get("counts", {}).get("Warning", 0))
    q4.metric("Info", quality.get("counts", {}).get("Info", 0))

    if quality.get("overall") == "PASS":
        st.success("No error or warning-level data-quality findings were detected.")
    elif quality.get("overall") == "WARN":
        st.warning("The profile is usable, but warning-level findings may lower recommendation confidence.")
    else:
        st.error("Error-level profile findings should be corrected before relying on recommendations.")

    quality_tabs = st.tabs(["Findings", "Engine Readiness", "Safe Fixes", "Export"])
    with quality_tabs[0]:
        issues = quality.get("issues", [])
        if issues:
            severity_options = ["Error", "Warning", "Info"]
            selected_severity = st.multiselect(
                "Severity", severity_options, default=severity_options, key="quality_severity_filter"
            )
            categories = sorted({row.get("Category", "Other") for row in issues})
            selected_categories = st.multiselect(
                "Category", categories, default=categories, key="quality_category_filter"
            )
            filtered = [
                row for row in issues
                if row.get("Severity") in selected_severity and row.get("Category") in selected_categories
            ]
            st.dataframe(pd.DataFrame(filtered), use_container_width=True, hide_index=True)
        else:
            st.success("No findings.")

    with quality_tabs[1]:
        st.dataframe(pd.DataFrame(quality.get("readiness", [])), use_container_width=True, hide_index=True)
        st.info("Optional Effective Paths calibration and Battle Reports improve confidence, but are not required for standalone engine execution.")

    with quality_tabs[2]:
        safe_issues = [row for row in quality.get("issues", []) if row.get("Safe Fix")]
        if safe_issues:
            st.write("The following findings can be corrected without changing levels or deleting entries:")
            st.dataframe(pd.DataFrame(safe_issues), use_container_width=True, hide_index=True)
            confirm_safe = st.checkbox("Apply alias migrations and Gold-flag corrections", key="confirm_quality_safe_fixes")
            if st.button("Apply safe fixes", type="primary", disabled=not confirm_safe, key="apply_quality_safe_fixes"):
                result = apply_quality_safe_fixes(profile)
                save_profile(profile["name"], profile)
                st.success(f"Applied {result.get('changed', 0)} safe changes.")
                if result.get("changes"):
                    with st.expander("Changes applied"):
                        for change in result["changes"]:
                            st.write(f"- {change}")
                bump_revision(); st.rerun()
        else:
            st.info("No safe automatic fixes are currently available.")

    with quality_tabs[3]:
        st.download_button(
            "Download data-quality report",
            data=json.dumps(quality, indent=2),
            file_name=f"{safe_profile_filename(profile['name'])}_data_quality.json",
            mime="application/json",
            use_container_width=True,
        )

elif page == "Build Audit":
    st.header("Build Audit")
    st.caption("Checks imported and manually entered data for missing fields, impossible values, cap mismatches, and configuration conflicts.")
    issues = profile_audit(profile)
    counts = {severity: sum(1 for issue in issues if issue["Severity"] == severity) for severity in ["Error", "Warning", "Info"]}
    c1, c2, c3, c4 = st.columns(4)
    c1.metric("Errors", counts["Error"])
    c2.metric("Warnings", counts["Warning"])
    c3.metric("Info", counts["Info"])
    c4.metric("Total Findings", len(issues))

    if issues:
        severity_filter = st.multiselect("Severity", ["Error", "Warning", "Info"], default=["Error", "Warning", "Info"])
        category_options = sorted({issue["Category"] for issue in issues})
        category_filter = st.multiselect("Category", category_options, default=category_options)
        filtered = [issue for issue in issues if issue["Severity"] in severity_filter and issue["Category"] in category_filter]
        st.dataframe(pd.DataFrame(filtered), use_container_width=True, hide_index=True)
    else:
        st.success("No audit findings.")

    col1, col2 = st.columns(2)
    with col1:
        if st.button("Apply safe cap and Gold fixes", use_container_width=True):
            changed = apply_safe_audit_fixes(profile)
            save_profile(profile["name"], profile)
            st.success(f"Applied {changed} safe changes.")
            bump_revision(); st.rerun()
    with col2:
        st.download_button(
            "Download audit JSON",
            data=json.dumps(issues, indent=2),
            file_name=f"{safe_profile_filename(profile['name'])}_audit.json",
            mime="application/json",
            use_container_width=True,
        )

    if profile.get("import_audit"):
        st.subheader("Latest Import Audit")
        st.dataframe(profile["import_audit"], use_container_width=True, hide_index=True)

elif page == "Battle Reports":
    st.header("Battle Reports")
    st.caption(
        "Paste one or several complete in-game Battle Reports. Each report normally begins with the "
        "'Battle Report' heading. Imported reports are saved to the current profile and immediately "
        "become available to Battle Learning and farming-tier comparisons."
    )
    st.info(
        "Use **Import report(s)** for the one-click workflow. **Preview only** lets you review the batch "
        "before saving it. Duplicate runs are skipped automatically."
    )
    report_text = st.text_area(
        "Paste Battle Report(s)", height=360, key="battle_report_text",
        placeholder="Battle Report\nBattle Date ...\nTier ...\nWave ...\n\nBattle Report\nBattle Date ...",
    )

    flash = st.session_state.pop("battle_import_flash", None)
    if flash:
        added = int(flash.get("added", 0))
        duplicates = int(flash.get("duplicates", 0))
        parse_errors = flash.get("parse_errors", [])
        invalid = int(flash.get("invalid", 0))
        if added:
            st.success(
                f"Imported and saved {added} report(s). "
                f"Skipped {duplicates} duplicate(s) and {invalid} invalid parsed row(s)."
            )
        elif duplicates and not parse_errors and not invalid:
            st.warning(f"No new reports were added; all {duplicates} parsed report(s) were already saved.")
        else:
            st.warning("No reports were imported.")
        for item in parse_errors:
            st.error(f"Report {item.get('report', '?')}: {item.get('error', 'Could not parse report')}")

    c1, c2, c3 = st.columns([1.3, 1, 1])
    with c1:
        if st.button("Import report(s)", type="primary", use_container_width=True, disabled=not report_text.strip()):
            batch = parse_battle_report_batch(report_text)
            result = import_runs(
                profile,
                batch.get("parsed", []),
                allow_duplicates=False,
                batch_label="Battle Reports page paste import",
            )
            if result.get("added"):
                save_profile(profile["name"], profile)
                bump_revision()
            st.session_state["battle_import_flash"] = {
                "added": len(result.get("added", [])),
                "duplicates": len(result.get("duplicates", [])),
                "invalid": len(result.get("invalid", [])),
                "parse_errors": batch.get("errors", []),
            }
            st.session_state.pop("battle_batch_preview", None)
            st.rerun()
    with c2:
        if st.button("Preview only", use_container_width=True, disabled=not report_text.strip()):
            batch = parse_battle_report_batch(report_text)
            prepared = prepare_import_batch(profile.get("runs", []), batch.get("parsed", []))
            st.session_state["battle_batch_preview"] = {
                "parsed": batch.get("parsed", []),
                "errors": batch.get("errors", []),
                "unique": prepared.get("unique", []),
                "duplicates": prepared.get("duplicates", []),
                "invalid": prepared.get("invalid", []),
            }
            st.rerun()
    with c3:
        if st.button("Clear preview", use_container_width=True):
            st.session_state.pop("battle_batch_preview", None)
            st.rerun()

    preview = st.session_state.get("battle_batch_preview")
    if preview:
        st.subheader("Batch Preview")
        p1, p2, p3, p4 = st.columns(4)
        p1.metric("New reports", len(preview.get("unique", [])))
        p2.metric("Duplicates", len(preview.get("duplicates", [])))
        p3.metric("Parse errors", len(preview.get("errors", [])))
        p4.metric("Invalid rows", len(preview.get("invalid", [])))

        preview_rows = []
        for status, collection in (("New", preview.get("unique", [])), ("Duplicate", preview.get("duplicates", []))):
            for run in collection:
                preview_rows.append({
                    "Status": status,
                    "Battle Date": run.get("battle_date", ""),
                    "Tier": run.get("tier", 0),
                    "Wave": run.get("wave", 0),
                    "Killed By": run.get("killed_by", "Unknown"),
                    "Coins": run.get("coins_earned", 0),
                    "Coins / Hour": run.get("coins_per_hour", 0),
                    "Cells": run.get("cells_earned", 0),
                    "Cells / Hour": run.get("cells_per_hour", 0),
                })
        if preview_rows:
            st.dataframe(
                pd.DataFrame(preview_rows), use_container_width=True, hide_index=True,
                column_config={
                    "Coins": st.column_config.NumberColumn(format="%.3e"),
                    "Coins / Hour": st.column_config.NumberColumn(format="%.3e"),
                    "Cells / Hour": st.column_config.NumberColumn(format="%.1f"),
                },
            )
        for item in preview.get("errors", []):
            st.error(f"Report {item.get('report', '?')}: {item.get('error', 'Could not parse report')}")
        for item in preview.get("invalid", []):
            st.warning(f"Invalid parsed row: {item.get('reason', 'Unknown reason')}")

        if st.button(
            "Import new reports from preview", type="primary",
            disabled=not preview.get("unique"), key="battle_import_preview_batch",
        ):
            result = import_runs(
                profile,
                preview.get("parsed", []),
                allow_duplicates=False,
                batch_label="Battle Reports page reviewed import",
            )
            if result.get("added"):
                save_profile(profile["name"], profile)
                bump_revision()
            st.session_state["battle_import_flash"] = {
                "added": len(result.get("added", [])),
                "duplicates": len(result.get("duplicates", [])),
                "invalid": len(result.get("invalid", [])),
                "parse_errors": preview.get("errors", []),
            }
            st.session_state.pop("battle_batch_preview", None)
            st.rerun()

    rows = run_history_rows(profile)
    if rows:
        st.subheader("Saved Runs")
        display = pd.DataFrame(rows)
        st.dataframe(
            display.drop(columns=["ID"]),
            use_container_width=True,
            hide_index=True,
            column_config={
                "Coins": st.column_config.NumberColumn(format="%.3e"),
                "Coins / Hour": st.column_config.NumberColumn(format="%.3e"),
                "Cells / Hour": st.column_config.NumberColumn(format="%.1f"),
            },
        )
        run_options = {f"T{row['Tier']} W{row['Wave']} — {row['Battle Date'] or row['ID']}": row["ID"] for row in rows}
        selected_label = st.selectbox("Select a run to inspect or delete", list(run_options))
        selected_id = run_options[selected_label]
        selected_run = next(run for run in profile["runs"] if run.get("id") == selected_id)
        with st.expander("Selected run details"):
            st.json({k: v for k, v in selected_run.items() if k != "raw_text"}, expanded=False)
        if st.button("Delete selected run"):
            profile["runs"] = [run for run in profile["runs"] if run.get("id") != selected_id]
            save_profile(profile["name"], profile)
            bump_revision(); st.rerun()

        tier_rows = farming_tier_rows(profile)
        st.subheader("Farming Tier Comparison")
        st.dataframe(
            pd.DataFrame(tier_rows), use_container_width=True, hide_index=True,
            column_config={
                "Average CPH": st.column_config.NumberColumn(format="%.3e"),
                "Best CPH": st.column_config.NumberColumn(format="%.3e"),
                "Average Cells/H": st.column_config.NumberColumn(format="%.1f"),
            },
        )
        eligible = [row for row in tier_rows if row["Average CPH"] > 0]
        if eligible:
            best = max(eligible, key=lambda row: row["Average CPH"])
            st.success(f"Current best measured coin-farming tier: Tier {best['Tier']} at {format_large_number(best['Average CPH'])} coins/hour average.")
    else:
        st.info("No reports saved yet.")

elif page == "Battle Learning":
    render_battle_learning_page(
        profile,
        save_profile=save_profile,
        safe_profile_filename=safe_profile_filename,
        parse_battle_report=parse_battle_report,
        parse_tower_number=parse_tower_number,
        bump_revision=bump_revision,
        app_version=APP_VERSION,
    )

elif page == "Build Analyzer":
    st.header("Build Analyzer")
    analysis = build_analysis(profile)
    scores = analysis["scores"]
    c1, c2, c3, c4 = st.columns(4)
    c1.metric("Economy", f"{scores['Economy']:.1f}/100")
    c2.metric("Damage", f"{scores['Damage']:.1f}/100")
    c3.metric("Survivability", f"{scores['Survivability']:.1f}/100")
    c4.metric("Regen / Recovery", f"{scores['Regen / Recovery']:.1f}/100")

    c5, c6, c7 = st.columns(3)
    c5.metric("Data Confidence", f"{analysis['confidence']:.1f}%")
    c6.metric("Black Holes", analysis["bh_quantity"])
    c7.metric("GT/BH Sync", "Yes" if analysis["gt_bh_synced"] else "No / Unknown")

    st.subheader("Current Development Bottleneck")
    st.warning(f"**{analysis['weakest']}** is the lowest relative development score.")
    st.write(f"Latest recorded cause of death: **{analysis['latest_killed_by']}**")

    st.subheader("Recommended Review")
    for recommendation in analysis["recommendations"]:
        st.write(f"- {recommendation}")

    st.subheader("Score Details")
    st.dataframe(
        pd.DataFrame([{"Area": name, "Score": score} for name, score in scores.items()]),
        use_container_width=True,
        hide_index=True,
        column_config={"Score": st.column_config.ProgressColumn(min_value=0, max_value=100, format="%.1f")},
    )
    st.info(analysis["method"] + " Scores compare progression within each category; they do not estimate waves or calculate true ROI yet.")

elif page == "Recommendation Dashboard":
    st.header("Recommendation Dashboard")
    st.caption(
        "This is the everyday decision screen. It combines native economy, damage, eHP, and regen paths without pretending their raw ROI units are directly interchangeable."
    )

    settings = profile.setdefault("combined_recommendations", {}).setdefault("settings", {})
    with st.expander("Priority controls", expanded=False):
        c1, c2, c3 = st.columns(3)
        with c1:
            candidates_per_path = st.slider(
                "Candidates from each native path", 1, 5,
                int(settings.get("candidates_per_path", 3)), key="combined_candidates_per_path"
            )
            settings["candidates_per_path"] = candidates_per_path
        with c2:
            focus_options = ["Balanced", "Economy", "Damage", "Survival", "Recovery"]
            focus_default = str(settings.get("focus", "Balanced"))
            if focus_default not in focus_options:
                focus_default = "Balanced"
            focus = st.selectbox("Manual focus", focus_options, index=focus_options.index(focus_default), key="combined_focus")
            settings["focus"] = focus
        with c3:
            death_weighting = st.checkbox(
                "Use recent cause of death", value=bool(settings.get("death_weighting", True)), key="combined_death_weighting"
            )
            settings["death_weighting"] = death_weighting

    with st.spinner("Building combined native paths..."):
        combined = build_combined_recommendations(
            profile, steps=15, candidates_per_path=candidates_per_path,
            apply_death_weighting=death_weighting, focus=focus,
        )

    analysis = combined["analysis"]
    c1, c2, c3, c4 = st.columns(4)
    c1.metric("Current bottleneck", analysis.get("weakest", "Unknown"))
    c2.metric("Latest death", combined.get("latest_death", "No report saved"))
    c3.metric("Profile confidence", f"{analysis.get('confidence', 0):.1f}%")
    c4.metric("GT/BH sync", "Yes" if analysis.get("gt_bh_synced") else "No / Unknown")

    top_rows = combined.get("rows", [])
    if top_rows:
        top = top_rows[0]
        st.success(
            f"Top current priority: **{top['Upgrade']}** ({top['Domain']} · {top['Resource']}) — "
            f"{top['Affordability']}."
        )
        st.caption(top.get("Why", ""))
    else:
        st.warning("No eligible native recommendations were generated. Check profile imports, unlocks, and Gold Boxes.")

    if top_rows:
        st.subheader("Why this recommendation?")
        explanation_options = {
            f"#{index + 1} · {row.get('Upgrade')} · {row.get('Resource')}": row
            for index, row in enumerate(top_rows[:15])
        }
        selected_explanation = st.selectbox(
            "Recommendation to explain", list(explanation_options), key="combined_explanation_select"
        )
        explanation = recommendation_explanation(
            explanation_options[selected_explanation], profile, analysis, combined.get("latest_death", "")
        )
        st.write(explanation["Summary"])
        e1, e2 = st.columns(2)
        with e1:
            st.markdown("**Inputs used**")
            st.dataframe(
                [{"Input": key, "Value": value} for key, value in explanation["Inputs"].items()],
                use_container_width=True, hide_index=True,
            )
        with e2:
            st.markdown("**Caveats**")
            for caveat in explanation["Caveats"]:
                st.write(f"- {caveat}")
        if explanation.get("Source explanation"):
            st.caption(explanation["Source explanation"])

    st.subheader("Whole-account opportunity cost")
    account_resources = ["Gems", "Medals", "Bits", "Keys", "Reroll Shards", "Module Shards", "Milestone / Event", "Action"]
    selected_account_resource = st.selectbox(
        "Compare another resource", account_resources, key="combined_account_resource"
    )
    account_rows = combined.get("by_resource", {}).get(selected_account_resource, [])
    if account_rows:
        account_top = account_rows[0]
        st.info(
            f"Best {selected_account_resource} use: **{account_top.get('Upgrade')}** "
            f"({account_top.get('Domain')} · {account_top.get('Affordability')})."
        )
        st.dataframe(combined_display_frame(account_rows, 12), use_container_width=True, hide_index=True)
    else:
        st.caption(f"No eligible {selected_account_resource} recommendations were generated from the current profile.")

    tabs = st.tabs([
        "Best Overall", "Start This Lab", "Spend Coins", "Spend Stones",
        "Fix Bottleneck", "Long-Term", "Method",
    ])

    with tabs[0]:
        st.subheader("Best overall shortlist")
        st.dataframe(
            combined_display_frame(top_rows, 15), use_container_width=True, hide_index=True,
            column_config={
                "Priority Index": st.column_config.NumberColumn(format="%.1f"),
                "Estimated Gain %": st.column_config.NumberColumn(format="%.3f%%"),
            },
        )
        st.info("Use this as a shortlist, then open the dedicated native page for the full path and assumptions.")

    with tabs[1]:
        rows = combined.get("by_resource", {}).get("Lab", [])
        if rows:
            first = rows[0]
            st.success(f"Start: **{first['Upgrade']}** → {first['Next Level']} ({first['Cost / Time']})")
            st.dataframe(combined_display_frame(rows, 15), use_container_width=True, hide_index=True)
        else:
            st.info("No eligible lab recommendations are available.")

    with tabs[2]:
        rows = combined.get("by_resource", {}).get("Coins", [])
        affordable_rows = [row for row in rows if row.get("Affordable Bool") is True]
        if affordable_rows:
            first = affordable_rows[0]
            st.success(f"Affordable now: **{first['Upgrade']}** → {first['Next Level']} ({first['Cost / Time']})")
        elif rows:
            st.warning(f"No modeled coin upgrade is currently affordable. Best save-up target: **{rows[0]['Upgrade']}**.")
        st.dataframe(combined_display_frame(rows, 15), use_container_width=True, hide_index=True)

    with tabs[3]:
        rows = combined.get("by_resource", {}).get("Stones", [])
        affordable_rows = [row for row in rows if row.get("Affordable Bool") is True]
        if affordable_rows:
            first = affordable_rows[0]
            st.success(f"Affordable now: **{first['Upgrade']}** → {first['Next Level']} ({first['Cost / Time']})")
        elif rows:
            st.warning(f"No modeled stone upgrade is currently affordable. Best save-up target: **{rows[0]['Upgrade']}**.")
        st.dataframe(combined_display_frame(rows, 15), use_container_width=True, hide_index=True)

    with tabs[4]:
        rows = combined.get("bottleneck", [])
        st.warning(f"Current bottleneck weighting: **{analysis.get('weakest', 'Unknown')}**")
        st.write(f"Recent death signal: **{combined.get('latest_death', 'No report saved')}**")
        st.dataframe(combined_display_frame(rows, 15), use_container_width=True, hide_index=True)

    with tabs[5]:
        rows = combined.get("long_term", [])
        if rows:
            st.caption("High-priority upgrades that are currently above the entered resource balance.")
            st.dataframe(combined_display_frame(rows, 15), use_container_width=True, hide_index=True)
        else:
            st.info("No explicitly unaffordable recommendations were found, or resource balances have not been entered.")

    with tabs[6]:
        st.write(combined.get("method"))
        st.markdown(
            """
            **Priority Index inputs**
            - Native ROI normalized inside each individual path
            - Position within that path
            - Estimated relative gain
            - Current weakest development area
            - Recent cause of death, when enabled
            - Model confidence
            - Affordability
            - Agreement with an imported Effective Paths reference

            A lab ROI, stone ROI, and coin ROI use different units. The dashboard deliberately avoids treating their raw numbers as directly comparable.
            """
        )
        st.subheader("Development scores")
        score_rows = [{"Area": key, "Score": value} for key, value in analysis.get("scores", {}).items()]
        st.dataframe(score_rows, use_container_width=True, hide_index=True)

elif page == "Progression Planner":
    st.header("Progression Planner")
    st.caption(
        "Turns whole-account recommendations into a daily plan, five-slot lab schedule, milestone route, "
        "resource forecast, and persistent upgrade queue. Effective Paths remains optional."
    )

    planner_state = ensure_planner_state(profile)
    planner_settings = planner_state["settings"]

    with st.expander("Planning target", expanded=True):
        c1, c2, c3, c4 = st.columns([2.2, 1, 1, 1])
        with c1:
            current_goal = str(planner_settings.get("goal", "Balanced progression"))
            if current_goal not in GOAL_OPTIONS:
                current_goal = "Balanced progression"
            planner_settings["goal"] = st.selectbox(
                "Primary goal", GOAL_OPTIONS, index=GOAL_OPTIONS.index(current_goal), key="planner_goal"
            )
            planner_settings["focus"] = GOAL_FOCUS.get(planner_settings["goal"], "Balanced")
        with c2:
            planner_settings["target_tier"] = int(st.number_input(
                "Target tier", min_value=1, max_value=25,
                value=int(planner_settings.get("target_tier", 10) or 10), step=1, key="planner_target_tier"
            ))
        with c3:
            planner_settings["target_wave"] = int(st.number_input(
                "Target wave", min_value=1, max_value=100000,
                value=int(planner_settings.get("target_wave", 4500) or 4500), step=100, key="planner_target_wave"
            ))
        with c4:
            planner_settings["lab_slots"] = int(st.number_input(
                "Lab slots", min_value=1, max_value=8,
                value=int(planner_settings.get("lab_slots", 5) or 5), step=1, key="planner_lab_slots"
            ))
        c5, c6, c7 = st.columns(3)
        with c5:
            planner_settings["planning_horizon_days"] = int(st.number_input(
                "Planning horizon (days)", min_value=1, max_value=90,
                value=int(planner_settings.get("planning_horizon_days", 7) or 7), step=1,
                key="planner_horizon"
            ))
        with c6:
            planner_settings["hours_played_per_day"] = float(st.number_input(
                "Expected farming hours/day", min_value=0.0, max_value=24.0,
                value=float(planner_settings.get("hours_played_per_day", 4.0) or 0.0), step=0.5,
                key="planner_hours_day"
            ))
        with c7:
            planner_settings["use_death_weighting"] = st.checkbox(
                "Use latest cause of death", value=bool(planner_settings.get("use_death_weighting", True)),
                key="planner_use_death"
            )
        if st.button("Save planning target", key="planner_save_target"):
            save_profile(profile["name"], profile)
            st.success("Planning target saved.")

    with st.spinner("Building the progression plan..."):
        progression_plan = build_progression_plan(profile)

    queue_active = queue_rows(profile, include_finished=False)
    c1, c2, c3, c4, c5 = st.columns(5)
    c1.metric("Goal focus", progression_plan.get("focus", "Balanced"))
    c2.metric("Active queue", len(queue_active))
    c3.metric("Lab slots planned", len(progression_plan.get("lab_plan", [])))
    c4.metric("Current bottleneck", progression_plan.get("analysis", {}).get("weakest", "Unknown"))
    c5.metric("Latest death", progression_plan.get("latest_death", "No report saved"))

    planner_tabs = st.tabs([
        "Today", "Lab Slots", "Milestone Route", "Resource Forecast",
        "Upgrade Queue", "Weekly Report", "Rates & Export",
    ])

    with planner_tabs[0]:
        st.subheader("What to do next")
        daily_actions = progression_plan.get("daily_actions", [])
        if daily_actions:
            top = daily_actions[0]
            st.success(f"Top action: **{top.get('Action')}** — {top.get('When')}.")
            if top.get("Why"):
                st.caption(top.get("Why"))
            daily_frame = pd.DataFrame([
                {key: row.get(key) for key in [
                    "Order", "Action", "System", "Resource", "Next Level", "When",
                    "Planner Score", "Confidence", "Why"
                ]}
                for row in daily_actions
            ])
            st.dataframe(
                daily_frame, use_container_width=True, hide_index=True,
                column_config={"Planner Score": st.column_config.NumberColumn(format="%.1f")},
            )

            action_options = {
                f"{row.get('Order')}. {row.get('Action')} · {row.get('Resource')}": row.get("Source Row", {})
                for row in daily_actions
            }
            selected_action = st.selectbox("Add an action to the upgrade queue", list(action_options), key="planner_daily_add")
            if st.button("Add selected action", type="primary", key="planner_add_daily"):
                queue_add(profile, action_options[selected_action], source="Daily plan")
                save_profile(profile["name"], profile)
                st.success("Action added to the queue.")
                st.rerun()
        else:
            st.warning("No actions were generated. Import or enter more profile data and resource balances.")

        st.info(progression_plan.get("method", ""))

    with planner_tabs[1]:
        st.subheader("Recommended lab-slot schedule")
        lab_plan = progression_plan.get("lab_plan", [])
        if lab_plan:
            lab_frame = pd.DataFrame([
                {key: row.get(key) for key in [
                    "Slot", "Research", "Next Level", "Domain", "Estimated Duration",
                    "Designation", "Replacement", "Planner Score", "Confidence", "Why"
                ]}
                for row in lab_plan
            ])
            st.dataframe(
                lab_frame, use_container_width=True, hide_index=True,
                column_config={"Planner Score": st.column_config.NumberColumn(format="%.1f")},
            )
            st.caption(
                "Durations are extracted from native-engine results when available. A replacement is the next ranked lab, "
                "not a promise that it will remain optimal after the profile changes."
            )
            c1, c2 = st.columns(2)
            with c1:
                selected_lab_label = st.selectbox(
                    "Add one lab", [f"Slot {row['Slot']}: {row['Research']}" for row in lab_plan], key="planner_lab_add_select"
                )
                selected_lab_index = int(selected_lab_label.split(":", 1)[0].replace("Slot", "").strip()) - 1
                if st.button("Add selected lab to queue", key="planner_add_lab"):
                    queue_add(profile, lab_plan[selected_lab_index].get("Source Row", {}), source="Lab plan")
                    save_profile(profile["name"], profile)
                    st.success("Lab added to the queue.")
                    st.rerun()
            with c2:
                if st.button("Add all planned labs to queue", key="planner_add_all_labs"):
                    for row in lab_plan:
                        queue_add(profile, row.get("Source Row", {}), source="Lab plan")
                    save_profile(profile["name"], profile)
                    st.success("All planned labs were added to the queue.")
                    st.rerun()
        else:
            st.info("No eligible lab recommendations were generated.")

    with planner_tabs[2]:
        st.subheader("Staged progression route")
        st.warning(
            f"Target: **{progression_plan.get('goal')}**, Tier {planner_settings.get('target_tier')}, "
            f"wave {int(planner_settings.get('target_wave', 0)):,}. The app plans upgrades but does not predict an exact wave outcome."
        )
        for stage_name, stage_rows in progression_plan.get("stages", {}).items():
            with st.expander(stage_name, expanded=stage_name in {"Immediate", "Next 7 days"}):
                if stage_rows:
                    st.dataframe(pd.DataFrame(stage_rows), use_container_width=True, hide_index=True)
                else:
                    st.info("No actions currently fall into this stage.")

    with planner_tabs[3]:
        st.subheader("Resource affordability forecast")
        forecast_rows = progression_plan.get("forecast", [])
        if forecast_rows:
            st.dataframe(
                pd.DataFrame(forecast_rows), use_container_width=True, hide_index=True,
                column_config={
                    "Current Balance": st.column_config.NumberColumn(format="%.2f"),
                    "Estimated Gain / Day": st.column_config.NumberColumn(format="%.2f"),
                    "Target Cost": st.column_config.NumberColumn(format="%.2f"),
                },
            )
            st.caption(
                "Coins/hour and cells/hour fall back to the median of recent battle reports when the manual rate is zero. "
                "Other currencies require manual rates. Unpriced strategic recommendations cannot receive a date."
            )
        else:
            st.info("No priced resource targets were found.")

    with planner_tabs[4]:
        st.subheader("Persistent upgrade queue")
        queue_filter = st.radio(
            "Show", ["Active only", "All items"], horizontal=True, key="planner_queue_filter"
        )
        queue_items = queue_rows(profile, include_finished=queue_filter == "All items")
        c1, c2, c3 = st.columns([1, 1, 4])
        with c1:
            if st.button("Undo last queue action", disabled=not planner_state.get("queue_history"), key="planner_queue_undo"):
                description = queue_undo(profile)
                save_profile(profile["name"], profile)
                st.success(f"Undid: {description}")
                st.rerun()
        with c2:
            queue_csv = pd.DataFrame(queue_items).to_csv(index=False).encode("utf-8") if queue_items else b""
            st.download_button(
                "Download queue CSV", data=queue_csv, file_name=f"{safe_profile_filename(profile['name'])}_upgrade_queue.csv",
                mime="text/csv", disabled=not queue_items, key="planner_queue_download"
            )

        if not queue_items:
            st.info("The queue is empty. Add actions from Today, Lab Slots, or the selection below.")
        else:
            for item in queue_items:
                item_id = str(item.get("id"))
                locked = bool(item.get("locked"))
                status = str(item.get("status", "Planned"))
                title = f"{int(item.get('order', 0))}. {item.get('upgrade')} · {item.get('resource')} · {status}"
                with st.container(border=True):
                    st.markdown(f"**{title}** {'🔒' if locked else ''}")
                    st.caption(
                        f"{item.get('system')} · Next: {item.get('next_level')} · {item.get('cost_text')} · "
                        f"Priority {float(item.get('priority', 0)):.1f} · {item.get('confidence')}"
                    )
                    if item.get("why"):
                        st.write(item.get("why"))
                    b1, b2, b3, b4, b5, b6, b7 = st.columns(7)
                    if b1.button("↑", key=f"queue_up_{item_id}"):
                        queue_move(profile, item_id, -1); save_profile(profile["name"], profile); st.rerun()
                    if b2.button("↓", key=f"queue_down_{item_id}"):
                        queue_move(profile, item_id, 1); save_profile(profile["name"], profile); st.rerun()
                    if b3.button("Start", key=f"queue_start_{item_id}", disabled=status == "In Progress"):
                        queue_set_status(profile, item_id, "In Progress"); save_profile(profile["name"], profile); st.rerun()
                    if b4.button("Complete", key=f"queue_complete_{item_id}", disabled=status == "Completed"):
                        queue_set_status(profile, item_id, "Completed"); save_profile(profile["name"], profile); st.rerun()
                    if b5.button("Skip", key=f"queue_skip_{item_id}", disabled=status == "Skipped"):
                        queue_set_status(profile, item_id, "Skipped"); save_profile(profile["name"], profile); st.rerun()
                    if b6.button("Unlock" if locked else "Lock", key=f"queue_lock_{item_id}"):
                        queue_toggle_lock(profile, item_id); save_profile(profile["name"], profile); st.rerun()
                    if b7.button("Remove", key=f"queue_remove_{item_id}", disabled=locked):
                        queue_remove(profile, item_id); save_profile(profile["name"], profile); st.rerun()

        st.divider()
        ranked_rows = progression_plan.get("ranked_rows", [])[:40]
        queue_options = {
            f"{row.get('Upgrade')} · {row.get('Resource')} · {float(row.get('Planner Score', 0)):.1f}": row
            for row in ranked_rows
        }
        if queue_options:
            selected_queue_action = st.selectbox("Add another recommendation", list(queue_options), key="planner_queue_add_select")
            if st.button("Add recommendation", key="planner_queue_add_button"):
                queue_add(profile, queue_options[selected_queue_action], source="Planner recommendation")
                save_profile(profile["name"], profile)
                st.success("Recommendation added.")
                st.rerun()

    with planner_tabs[5]:
        st.subheader("Weekly progression report")
        weekly = build_weekly_report(profile, progression_plan)
        current = weekly.get("current", {})
        trends = weekly.get("trends", {})
        c1, c2, c3, c4 = st.columns(4)
        coin_delta = trends.get("coins_per_hour_percent")
        cell_delta = trends.get("cells_per_hour_percent")
        wave_delta = trends.get("best_wave_percent")
        c1.metric("Runs", current.get("runs", 0))
        c2.metric(
            "Avg coins/hour", f"{float(current.get('avg_coins_per_hour', 0)):,.0f}",
            delta=f"{coin_delta:+.1f}%" if isinstance(coin_delta, (int, float)) else None,
        )
        c3.metric(
            "Avg cells/hour", f"{float(current.get('avg_cells_per_hour', 0)):,.0f}",
            delta=f"{cell_delta:+.1f}%" if isinstance(cell_delta, (int, float)) else None,
        )
        c4.metric(
            "Best wave", f"{int(current.get('best_wave', 0)):,}",
            delta=f"{wave_delta:+.1f}%" if isinstance(wave_delta, (int, float)) else None,
        )
        st.write(f"Current bottleneck: **{weekly.get('current_bottleneck')}** · Latest death: **{weekly.get('latest_death')}**")
        completed_rows = weekly.get("completed_upgrades", [])
        if completed_rows:
            st.subheader("Completed this week")
            st.dataframe(pd.DataFrame(completed_rows), use_container_width=True, hide_index=True)
        else:
            st.info("No queue items were marked completed during this reporting period.")
        st.subheader("Next-week priorities")
        st.dataframe(pd.DataFrame(weekly.get("next_week_priorities", [])), use_container_width=True, hide_index=True)
        st.caption(weekly.get("note", ""))
        st.download_button(
            "Download weekly report JSON", data=json.dumps(weekly, indent=2, default=str),
            file_name=f"{safe_profile_filename(profile['name'])}_weekly_report.json", mime="application/json",
            key="planner_weekly_download",
        )

    with planner_tabs[6]:
        st.subheader("Income rates")
        income_rates = planner_settings.setdefault("income_rates", {})
        derived_rates = effective_income_rates(profile)
        st.caption(
            "Leave coins/hour or cells/hour at zero to use the median of recent battle reports. "
            "Manual values override battle-report-derived rates."
        )
        rate_columns = st.columns(3)
        rate_specs = [
            ("coins_per_hour", "Coins per farming hour"),
            ("cells_per_hour", "Cells per farming hour"),
            ("stones_per_week", "Stones per week"),
            ("gems_per_day", "Gems per day"),
            ("medals_per_week", "Medals per week"),
            ("keys_per_week", "Keys per week"),
            ("bits_per_day", "Bits per day"),
            ("reroll_shards_per_day", "Reroll shards per day"),
            ("module_shards_per_day", "Module shards per day"),
        ]
        for index, (field, label) in enumerate(rate_specs):
            with rate_columns[index % 3]:
                income_rates[field] = float(st.number_input(
                    label, min_value=0.0, value=float(income_rates.get(field, 0.0) or 0.0),
                    step=1.0, format="%.2f", key=f"planner_rate_{field}"
                ))
                if field in {"coins_per_hour", "cells_per_hour"} and income_rates[field] <= 0:
                    st.caption(f"Derived: {float(derived_rates.get(field, 0.0)):,.2f}")
        if st.button("Save income rates", type="primary", key="planner_save_rates"):
            save_profile(profile["name"], profile)
            st.success("Income rates saved.")
            st.rerun()

        export_payload = {
            "app_version": APP_VERSION,
            "profile": profile.get("name"),
            "generated_at": progression_plan.get("generated_at"),
            "settings": planner_settings,
            "daily_actions": [
                {key: row.get(key) for key in ["Order", "Action", "System", "Resource", "Next Level", "When", "Planner Score", "Confidence", "Why"]}
                for row in progression_plan.get("daily_actions", [])
            ],
            "lab_plan": [
                {key: row.get(key) for key in ["Slot", "Research", "Next Level", "Domain", "Estimated Duration", "Designation", "Replacement", "Planner Score", "Confidence", "Why"]}
                for row in progression_plan.get("lab_plan", [])
            ],
            "stages": progression_plan.get("stages", {}),
            "forecast": progression_plan.get("forecast", []),
            "queue": queue_rows(profile, include_finished=True),
            "method": progression_plan.get("method"),
        }
        st.download_button(
            "Download complete progression plan", data=json.dumps(export_payload, indent=2, default=str),
            file_name=f"{safe_profile_filename(profile['name'])}_progression_plan_v{APP_VERSION}.json",
            mime="application/json", key="planner_full_export",
        )


elif page == "Whole Account":
    st.header("Whole Account Optimizer")
    st.caption(
        "Compares cards, modules, relics, themes, bots, guardians, and Vault choices. "
        "Exact cost curves are used only where verified; other rows are clearly labeled strategic recommendations."
    )
    account_settings = profile.setdefault("combined_recommendations", {}).setdefault("settings", {})
    focus_options = ["Balanced", "Economy", "Damage", "Survival", "Recovery", "Modules"]
    current_focus = str(account_settings.get("focus", "Balanced"))
    if current_focus not in focus_options:
        current_focus = "Balanced"
    c1, c2 = st.columns(2)
    with c1:
        account_focus = st.selectbox(
            "Whole-account focus", focus_options, index=focus_options.index(current_focus), key="whole_account_focus"
        )
        account_settings["focus"] = account_focus
    with c2:
        account_death = st.checkbox(
            "Use latest cause of death", value=bool(account_settings.get("death_weighting", True)), key="whole_account_death"
        )
        account_settings["death_weighting"] = account_death

    progression = build_combined_recommendations(
        profile, steps=15, candidates_per_path=3,
        focus=account_focus, apply_death_weighting=account_death
    )
    rows = progression.get("rows", [])
    systems = progression.get("by_system", {})
    resource_groups = progression.get("by_resource", {})
    c1, c2, c3, c4 = st.columns(4)
    c1.metric("Eligible actions", len(rows))
    c2.metric("Systems represented", len([name for name, values in systems.items() if values]))
    c3.metric("Current bottleneck", progression.get("analysis", {}).get("weakest", "Unknown"))
    c4.metric("Latest death", progression.get("latest_death", "No report saved"))

    if rows:
        top = rows[0]
        st.success(
            f"Top non-native action: **{top.get('Upgrade')}** — "
            f"{top.get('Resource')} · {top.get('Affordability')}."
        )
        st.caption(top.get("Why", ""))
    else:
        st.warning("No whole-account recommendations were generated. Import or enter the relevant systems first.")

    view_mode = st.radio("Group recommendations by", ["System", "Resource"], horizontal=True, key="whole_account_group")
    groups = systems if view_mode == "System" else {k: v for k, v in resource_groups.items() if v}
    if groups:
        selected_group = st.selectbox(view_mode, list(groups), key="whole_account_selected_group")
        selected_rows = groups.get(selected_group, [])
        st.dataframe(
            combined_display_frame(selected_rows, 25), use_container_width=True, hide_index=True,
            column_config={"Priority Index": st.column_config.NumberColumn(format="%.1f")},
        )
    else:
        st.info("No populated recommendation groups are available.")

    st.subheader("System coverage")
    coverage_rows = []
    expected_systems = [
        "Laboratory", "Workshop / Enhancements", "Ultimate Weapons",
        "Cards", "Modules", "Bots", "Guardians", "Vault", "Relics", "Themes & Songs"
    ]
    for system in expected_systems:
        system_rows = systems.get(system, [])
        coverage_rows.append({
            "System": system,
            "Recommendations": len(system_rows),
            "Top action": system_rows[0].get("Upgrade") if system_rows else "No eligible action / insufficient data",
            "Confidence": system_rows[0].get("Confidence") if system_rows else "—",
        })
    st.dataframe(coverage_rows, use_container_width=True, hide_index=True)
    st.info(progression.get("method", ""))

elif page == "Calibration Center":
    st.header("Calibration Center")
    st.caption("Compares standalone native-engine rankings with an imported Effective Paths ROI reference. Name aliases and close rank matches are handled explicitly.")

    calibration_steps = st.slider("Ranks to compare per path", 5, 30, 15, key="calibration_steps")
    with st.spinner("Running native engines and comparing paths..."):
        calibration_report = build_calibration_report(profile, steps=calibration_steps)

    counts = calibration_report.get("counts", {})
    c1, c2, c3, c4, c5 = st.columns(5)
    c1.metric("Overall", calibration_report.get("overall", "Unknown"))
    agreement = calibration_report.get("overall_agreement_percent")
    c2.metric("Agreement", f"{agreement:.1f}%" if isinstance(agreement, (int, float)) else "No reference")
    c3.metric("Exact", counts.get("Exact", 0))
    c4.metric("Close", counts.get("Close", 0))
    c5.metric("Different", counts.get("Different", 0))

    if not calibration_report.get("reference_loaded"):
        st.info("No Effective Paths ROI reference is loaded. The standalone engines still work; import a filled workbook under Import / Export → ROI Reference only when you want calibration.")
    elif calibration_report.get("overall") == "WARN":
        st.warning("At least one path materially differs from the saved reference. Review the first differing ranks below before changing formulas.")
    else:
        st.success("All comparable paths are exact or close matches to the saved reference.")

    calibration_tabs = st.tabs(["Path Summary", "Rank Detail", "Version History", "Export"])
    with calibration_tabs[0]:
        summary_rows = calibration_report.get("summary", [])
        status_filter = st.multiselect(
            "Status", ["Exact", "Close", "Different", "No reference"],
            default=["Exact", "Close", "Different", "No reference"], key="calibration_status_filter",
        )
        filtered_summary = [row for row in summary_rows if row.get("Status") in status_filter]
        st.dataframe(
            pd.DataFrame(filtered_summary), use_container_width=True, hide_index=True,
            column_config={"Weighted Agreement %": st.column_config.ProgressColumn(min_value=0, max_value=100, format="%.1f%%")},
        )

    with calibration_tabs[1]:
        summary_rows = calibration_report.get("summary", [])
        path_options = {f"{row.get('Path')} · {row.get('Status')}": row.get("Path Key") for row in summary_rows}
        if path_options:
            selected_path_label = st.selectbox("Path", list(path_options), key="calibration_path_select")
            selected_path = path_options[selected_path_label]
            selected_summary = next((row for row in summary_rows if row.get("Path Key") == selected_path), {})
            s1, s2, s3 = st.columns(3)
            s1.metric("Native top", selected_summary.get("Native Top", "—"))
            s2.metric("Reference top", selected_summary.get("Reference Top", "—"))
            s3.metric("Agreement", f"{selected_summary.get('Weighted Agreement %', 0):.1f}%")
            st.caption(selected_summary.get("Note", ""))
            detail_rows = calibration_report.get("details", {}).get(selected_path, [])
            st.dataframe(pd.DataFrame(detail_rows), use_container_width=True, hide_index=True)
        else:
            st.info("No paths were generated.")

    with calibration_tabs[2]:
        history = profile.setdefault("calibration", {}).setdefault("history", [])
        previous = history[-1] if history else None
        if previous:
            st.write(f"Previous saved snapshot: **v{previous.get('app_version', 'unknown')}** · {previous.get('saved_at', '')}")
            delta_rows = compare_snapshots(calibration_report, previous)
            st.dataframe(
                pd.DataFrame(delta_rows), use_container_width=True, hide_index=True,
                column_config={"Agreement Change": st.column_config.NumberColumn(format="%+.1f")},
            )
        else:
            st.info("No calibration snapshots have been saved yet.")
        if st.button("Save calibration snapshot", type="primary", key="save_calibration_snapshot"):
            snapshot = calibration_snapshot(calibration_report, APP_VERSION)
            history.append(snapshot)
            profile["calibration"]["history"] = history[-20:]
            profile["calibration"]["last_report"] = {
                "generated_at": calibration_report.get("generated_at"),
                "overall": calibration_report.get("overall"),
                "overall_agreement_percent": calibration_report.get("overall_agreement_percent"),
                "counts": calibration_report.get("counts", {}),
            }
            save_profile(profile["name"], profile)
            st.success("Calibration snapshot saved to the profile.")
            bump_revision(); st.rerun()

    with calibration_tabs[3]:
        st.download_button(
            "Download calibration report",
            data=json.dumps(calibration_report, indent=2),
            file_name=f"{safe_profile_filename(profile['name'])}_calibration_v{APP_VERSION}.json",
            mime="application/json",
            use_container_width=True,
        )

elif page == "Native eEcon":
    st.header("Native eEcon")
    st.caption(
        "This page calculates economy paths in Python from the canonical profile. "
        "It does not require Effective Paths at runtime. The imported ROI reference is used only for regression comparison."
    )

    settings = profile.setdefault("native_econ", {}).setdefault("settings", {})
    with st.expander("Model assumptions and controls", expanded=False):
        c1, c2 = st.columns(2)
        with c1:
            settings["bh_perk_duration_bonus"] = st.number_input(
                "Black Hole farming-perk duration bonus (seconds)", min_value=0.0, max_value=30.0,
                value=float(settings.get("bh_perk_duration_bonus", 12.0)), step=1.0,
                help="Default 12 seconds. This keeps current farming GT/BH windows modeled correctly."
            )
            settings["bh_coverage_divisor"] = st.number_input(
                "Black Hole coverage divisor", min_value=10.0, max_value=200.0,
                value=float(settings.get("bh_coverage_divisor", 70.0)), step=1.0,
                help="Lower values assume each Black Hole captures a larger share of enemies."
            )
        with c2:
            settings["dw_tag_share_per_quantity"] = st.number_input(
                "Death Wave tag share per quantity", min_value=0.0, max_value=1.0,
                value=float(settings.get("dw_tag_share_per_quantity", 0.0382747832266563)), step=0.001,
                format="%.6f",
                help="Default is regression-calibrated to the supplied Effective Paths reference."
            )
            settings["lab_speed_multiplier_override"] = st.number_input(
                "Effective lab-speed multiplier override (0 = automatic)", min_value=0.0, max_value=100.0,
                value=float(settings.get("lab_speed_multiplier_override", 0.0)), step=0.01,
                help="Automatic uses the Labs Speed level and active Lab Speed relic bonus."
            )
        settings["allow_desync_cooldowns"] = st.checkbox(
            "Allow single cooldown purchases that break a current GT/BH sync",
            value=bool(settings.get("allow_desync_cooldowns", False)),
        )
        st.write(f"Automatic effective lab-speed multiplier: **x{native_lab_speed_multiplier(profile):.4f}**")
        st.write(f"Current lab coin-cost multiplier: **x{native_lab_coin_multiplier(profile.get('labs', {})):.4f}**")

    steps = st.slider("Path length", 5, 100, 50, key="native_econ_steps")
    native_paths = build_native_econ_paths(profile, steps)
    reference_paths = profile.get("roi_reference", {}).get("paths", {})

    score_now = native_econ_score(profile)
    gt_cd = native_initial_stone_state(profile)["GT | Cooldown"]
    bh_cd = native_initial_stone_state(profile)["BH | Cooldown"]
    c1, c2, c3, c4 = st.columns(4)
    c1.metric("Native coin index", f"x{score_now:.4f}")
    c2.metric("Effective lab speed", f"x{native_lab_speed_multiplier(profile):.3f}")
    c3.metric("GT/BH cooldown", f"{gt_cd:.0f}s / {bh_cd:.0f}s")
    c4.metric("Sync", "Synced" if abs(gt_cd - bh_cd) < 1e-9 else "Not synced")

    tabs = st.tabs(["Lab path", "Stone path", "Coin path", "Discount path", "Regression", "Coverage"])
    path_display = {
        "econ_lab": ["Rank", "Upgrade", "Level", "Cost", "Duration", "Gain %", "ROI", "Cumulative", "Why"],
        "econ_stone": ["Rank", "Upgrade", "Level", "Value", "Cost", "Gain %", "ROI", "Cumulative", "Why"],
        "econ_coin": ["Rank", "Upgrade", "Level", "Cost", "Gain %", "ROI", "Cumulative", "Why"],
        "econ_discount": ["Rank", "Upgrade", "Level", "Duration", "Gain %", "ROI", "Cumulative", "Why"],
    }
    for tab, key in zip(tabs[:4], ["econ_lab", "econ_stone", "econ_coin", "econ_discount"]):
        with tab:
            rows = native_paths[key]
            if rows:
                frame = pd.DataFrame(rows)
                columns = [column for column in path_display[key] if column in frame.columns]
                st.dataframe(frame[columns], use_container_width=True, hide_index=True)
                current_resource = {
                    "econ_lab": None,
                    "econ_stone": native_number(profile.get("resources", {}).get("stones"), 0),
                    "econ_coin": native_number(profile.get("resources", {}).get("coins"), 0),
                    "econ_discount": None,
                }[key]
                first = rows[0]
                if current_resource is not None:
                    affordable = current_resource >= native_number(first.get("Cost Numeric"), 0)
                    st.info(
                        f"Immediate top choice: **{first['Upgrade']} → {first['Level']}**. "
                        f"Current balance is {'sufficient' if affordable else 'not sufficient'} for its listed cost."
                    )
                else:
                    st.info(f"Immediate top choice: **{first['Upgrade']} → {first['Level']}**.")
            else:
                st.info("No eligible upgrades remain in this path.")

    with tabs[4]:
        if not reference_paths:
            st.warning("No Effective Paths ROI reference is saved. Native paths still work, but regression comparison is unavailable.")
        else:
            regression_tabs = st.tabs(["Lab", "Stone", "Coin", "Discount"])
            for reg_tab, key in zip(regression_tabs, ["econ_lab", "econ_stone", "econ_coin", "econ_discount"]):
                with reg_tab:
                    comparison = compare_native_path(native_paths[key], reference_paths.get(key, {}).get("rows", []), limit=steps)
                    summary = native_path_match_summary(comparison)
                    if summary["rows"]:
                        a, b, c = st.columns(3)
                        a.metric("Comparable rows", summary["rows"])
                        b.metric("Rank matches", summary["matches"])
                        c.metric("Match rate", f"{summary['match_rate'] * 100:.1f}%")
                        st.dataframe(pd.DataFrame(comparison), use_container_width=True, hide_index=True)
                    else:
                        st.info("The saved reference has no cached rows for this path.")

    with tabs[5]:
        coverage_rows = [
            {"Path": "Economy lab", "Status": "Native", "Confidence": "High", "Notes": "Source cost/time curves; marginal formulas regression-verified."},
            {"Path": "Lab discount", "Status": "Native", "Confidence": "High", "Notes": "Matches the 0.3%/level discount curve and adjusted lab time."},
            {"Path": "Economy stones", "Status": "Native", "Confidence": "Medium-High", "Notes": "Exact UW values/costs. DW cooldown excluded; BH/DW coverage assumptions shown above."},
            {"Path": "Economy coins", "Status": "Partial native", "Confidence": "High for Coin Bonus+", "Notes": "Coin Bonus+ is exact. ELS+, Free Upgrades+, and survival-driven coin gains are not yet ranked."},
            {"Path": "Damage", "Status": "Native", "Confidence": "Mixed", "Notes": "Native lab/stone/coin/key paths are available under Native eDamage."},
            {"Path": "Health / Regen", "Status": "Native", "Confidence": "Mixed", "Notes": "Native eHP and eRegen paths are available under their dedicated pages."},
        ]
        st.dataframe(coverage_rows, use_container_width=True, hide_index=True)
        st.warning(
            "The native coin index is a relative upgrade-comparison model, not a coins/hour prediction. "
            "It deliberately avoids claiming wave gains or exact income without run-distribution data."
        )


elif page == "Native eDamage":
    st.header("Native eDamage")
    st.caption(
        "Standalone damage-upgrade paths using exact source cost tables and a transparent relative-damage model. "
        "Effective Paths remains the regression reference while coverage is expanded."
    )

    settings = native_damage_settings(profile)
    with st.expander("Damage model settings", expanded=False):
        settings["cl_weight"] = st.number_input("Chain Lightning contribution weight", min_value=0.0, max_value=1.0, value=float(settings.get("cl_weight", 0.12)), step=0.001, format="%.6f")
        settings["dw_weight"] = st.number_input("Death Wave contribution weight", min_value=0.0, max_value=1.0, value=float(settings.get("dw_weight", 0.0025)), step=0.0005, format="%.5f")
        settings["sm_weight"] = st.number_input("Smart Missiles contribution weight", min_value=0.0, max_value=1.0, value=float(settings.get("sm_weight", 0.0040)), step=0.0005, format="%.5f")
        settings["shock_events_per_cycle"] = st.number_input("Shock event-density estimate", min_value=0.1, max_value=100.0, value=float(settings.get("shock_events_per_cycle", 50.0)), step=0.5)
        st.info("The default Chain Lightning and shock values are calibrated against the bundled Effective Paths regression reference. Change them only for testing.")

    steps = st.slider("Path length", 5, 100, 50, key="native_damage_steps")
    native_paths = build_native_damage_paths(profile, steps)
    reference_paths = profile.get("roi_reference", {}).get("paths", {})
    components = native_damage_components(profile)

    c1, c2, c3, c4 = st.columns(4)
    c1.metric("Native damage index", f"x{components['Total']:.4f}")
    c2.metric("Bullet component", f"{components['Bullet']:.3f}")
    c3.metric("CL component", f"{components['Chain Lightning']:.3f}")
    c4.metric("Spotlight coverage", f"{components['Spotlight Coverage'] * 100:.1f}%")

    with st.expander("Current damage composition", expanded=False):
        composition_rows = [
            {"Component": "Bullet", "Value": components["Bullet"]},
            {"Component": "Chain Lightning", "Value": components["Chain Lightning"]},
            {"Component": "Death Wave", "Value": components["Death Wave"]},
            {"Component": "Smart Missiles", "Value": components["Smart Missiles"]},
            {"Component": "Black Hole", "Value": components["Black Hole"]},
            {"Component": "Spotlight factor", "Value": components["Spotlight Factor"]},
            {"Component": "Shock factor", "Value": components["Shock Factor"]},
            {"Component": "Shock uptime estimate", "Value": components["Shock Uptime"]},
        ]
        st.dataframe(composition_rows, use_container_width=True, hide_index=True)

    tabs = st.tabs(["Lab path", "Stone path", "Coin path", "Key path", "Regression", "Coverage"])
    path_display = {
        "damage_lab": ["Rank", "Upgrade", "Level", "Cost", "Duration", "Gain %", "ROI", "Confidence", "Cumulative", "Why"],
        "damage_stone": ["Rank", "Upgrade", "Level", "Value", "Cost", "Gain %", "ROI", "Confidence", "Cumulative", "Why"],
        "damage_coin": ["Rank", "Upgrade", "Level", "Cost", "Gain %", "ROI", "Confidence", "Cumulative", "Why"],
        "damage_key": ["Rank", "Upgrade", "Level", "Cost", "Gain %", "ROI", "Confidence", "Cumulative", "Why"],
    }
    for tab, key in zip(tabs[:4], ["damage_lab", "damage_stone", "damage_coin", "damage_key"]):
        with tab:
            rows = native_paths[key]
            if rows:
                frame = pd.DataFrame(rows)
                columns = [column for column in path_display[key] if column in frame.columns]
                st.dataframe(frame[columns], use_container_width=True, hide_index=True)
                first = rows[0]
                balance = None
                if key == "damage_stone": balance = native_number(profile.get("resources", {}).get("stones"), 0)
                elif key == "damage_coin": balance = native_number(profile.get("resources", {}).get("coins"), 0)
                if balance is not None:
                    affordable = balance >= native_number(first.get("Cost Numeric"), 0)
                    st.info(f"Immediate top choice: **{first['Upgrade']} → {first['Level']}**. Listed balance is {'sufficient' if affordable else 'not sufficient'} for this cost.")
                else:
                    st.info(f"Immediate top choice: **{first['Upgrade']} → {first['Level']}**.")
            else:
                st.info("No eligible upgrades remain in this path.")

    with tabs[4]:
        if not reference_paths:
            st.warning("No Effective Paths ROI reference is saved. Native paths still work, but regression comparison is unavailable.")
        else:
            regression_tabs = st.tabs(["Lab", "Stone", "Coin", "Key"])
            for reg_tab, key in zip(regression_tabs, ["damage_lab", "damage_stone", "damage_coin", "damage_key"]):
                with reg_tab:
                    comparison = compare_native_path(native_paths[key], reference_paths.get(key, {}).get("rows", []), limit=steps)
                    summary = native_path_match_summary(comparison)
                    if summary["rows"]:
                        a, b, c = st.columns(3)
                        a.metric("Comparable rows", summary["rows"])
                        b.metric("Rank matches", summary["matches"])
                        c.metric("Match rate", f"{summary['match_rate'] * 100:.1f}%")
                        st.dataframe(pd.DataFrame(comparison), use_container_width=True, hide_index=True)
                    else:
                        st.info("The saved reference has no cached rows for this path.")

    with tabs[5]:
        coverage_rows = [
            {"Path": "Damage labs", "Status": "Native", "Confidence": "High tables / mixed effects", "Notes": "Exact lab costs and durations. Core and shock effects are transparent and regression-calibrated."},
            {"Path": "Damage stones", "Status": "Native", "Confidence": "High CL tables / medium model", "Notes": "Exact UW value/cost curves for CL, SM, DW, BH and SL."},
            {"Path": "Damage coins", "Status": "Native", "Confidence": "Medium", "Notes": "Exact shared enhancement cost curve; relative enhancement effects are modeled."},
            {"Path": "Damage keys", "Status": "Partial native", "Confidence": "Medium-Low", "Notes": "First-tier Power vault effects only. Prerequisite graph and higher tiers are not yet enforced."},
            {"Path": "Modules / cards / masteries", "Status": "Partial", "Confidence": "Pending", "Notes": "Imported data is retained, but several unique-effect interactions are not yet in eDamage."},
        ]
        st.dataframe(coverage_rows, use_container_width=True, hide_index=True)
        st.warning("The native damage index is a relative comparison model, not a prediction of displayed in-game damage or maximum wave.")


elif page == "Native eHP":
    st.header("Native eHP")
    st.caption(
        "Standalone effective-health upgrade paths using the canonical profile. Exact lab cost/time tables are embedded; "
        "the survivability score is relative and is not a guaranteed wave prediction."
    )
    settings = native_health_settings(profile)
    with st.expander("Model assumptions and controls", expanded=False):
        c1, c2 = st.columns(2)
        with c1:
            settings["farming_defense_perk_bonus"] = st.number_input(
                "Farming defense-perk bonus", min_value=0.0, max_value=0.50,
                value=float(settings.get("farming_defense_perk_bonus", 0.20)), step=0.01, format="%.3f",
                help="Use 0 for tournament-style comparison or your expected farming perk contribution."
            )
            settings["dw_health_saturation"] = st.slider(
                "Death Wave health saturation", 0.0, 1.0,
                float(settings.get("dw_health_saturation", 1.0)), 0.01,
                help="1.00 means the run reaches the current DW health cap. Lower values allow DW quantity/cooldown to appear in the stone path."
            )
            settings["package_reliability_weight"] = st.slider(
                "Recovery-package reliability weight", 0.0, 1.0,
                float(settings.get("package_reliability_weight", 0.75)), 0.05,
            )
        with c2:
            settings["wall_weight"] = st.number_input(
                "Wall contribution weight", min_value=0.0, max_value=5.0,
                value=float(settings.get("wall_weight", 0.25)), step=0.1,
            )
            settings["defense_absolute_weight"] = st.number_input(
                "Defense Absolute contribution weight", min_value=0.0, max_value=1.0,
                value=float(settings.get("defense_absolute_weight", 0.06)), step=0.01, format="%.3f",
            )
            settings["enemy_attack_lab_reduction"] = st.number_input(
                "Enemy-attack reduction per lab level", min_value=0.0, max_value=0.02,
                value=float(settings.get("enemy_attack_lab_reduction", 0.004)), step=0.001, format="%.4f",
            )
        st.write(f"Latest recorded cause of death: **{native_latest_death(profile)}**")

    steps = st.slider("Path length", 5, 100, 50, key="native_health_steps")
    native_paths = build_native_health_paths(profile, steps)
    reference_paths = profile.get("roi_reference", {}).get("paths", {})
    components = native_health_components(profile)

    c1, c2, c3, c4 = st.columns(4)
    c1.metric("Native eHP index", f"x{components['Total']:.4f}")
    c2.metric("Defense % modeled", f"{components['Defense %']*100:.2f}%")
    c3.metric("Recovery multiplier", f"x{components['Recovery Mult']:.3f}")
    c4.metric("DW health factor", f"x{components['Death Wave Factor']:.2f}")

    with st.expander("Current eHP composition", expanded=False):
        rows = [
            {"Component": "Tower health pool", "Value": components["Tower Health"]},
            {"Component": "Wall health pool", "Value": components["Wall Health"]},
            {"Component": "Defense mitigation", "Value": components["Mitigation"]},
            {"Component": "Defense Absolute factor", "Value": components["Defense Absolute Factor"]},
            {"Component": "Enemy attack-lab factor", "Value": components["Enemy Attack Factor"]},
            {"Component": "Package chance", "Value": components["Package Chance"]},
            {"Component": "Max recovery", "Value": components["Max Recovery"]},
            {"Component": "Wall control factor", "Value": components["Wall Control Factor"]},
            {"Component": "Vampire-specific factor", "Value": components["Vampire Factor"]},
        ]
        st.dataframe(rows, use_container_width=True, hide_index=True)

    tabs = st.tabs(["Lab path", "Stone path", "Coin path", "Regression", "Coverage"])
    displays = {
        "health_lab": ["Rank", "Upgrade", "Level", "Cost", "Duration", "Gain %", "ROI", "Confidence", "Cumulative", "Why"],
        "health_stone": ["Rank", "Upgrade", "Level", "Value", "Cost", "Gain %", "ROI", "Confidence", "Cumulative", "Why"],
        "health_coin": ["Rank", "Upgrade", "Level", "Cost", "Gain %", "ROI", "Confidence", "Cumulative", "Why"],
    }
    for tab, key in zip(tabs[:3], ["health_lab", "health_stone", "health_coin"]):
        with tab:
            rows = native_paths[key]
            if rows:
                frame = pd.DataFrame(rows)
                st.dataframe(frame[[c for c in displays[key] if c in frame.columns]], use_container_width=True, hide_index=True)
                first = rows[0]
                balance = native_number(profile.get("resources", {}).get("stones" if key == "health_stone" else "coins"), 0) if key != "health_lab" else None
                if balance is None:
                    st.info(f"Immediate top choice: **{first['Upgrade']} → {first['Level']}**.")
                else:
                    st.info(f"Immediate top choice: **{first['Upgrade']} → {first['Level']}**. Current balance is {'sufficient' if balance >= native_number(first.get('Cost Numeric'),0) else 'not sufficient'}.")
            else:
                if key == "health_stone" and float(settings.get("dw_health_saturation", 1.0)) >= 0.999:
                    st.info("No direct health-stone step is modeled because Death Wave health saturation is set to 100%.")
                else:
                    st.info("No eligible upgrades remain in this path.")

    with tabs[3]:
        comparison_tabs = st.tabs(["Lab", "Stone", "Coin"])
        for subtab, key in zip(comparison_tabs, ["health_lab", "health_stone", "health_coin"]):
            with subtab:
                comparison = compare_native_path(native_paths[key], reference_paths.get(key, {}).get("rows", []), limit=steps)
                summary = native_path_match_summary(comparison)
                if summary["rows"]:
                    a,b,c = st.columns(3)
                    a.metric("Comparable rows", summary["rows"])
                    b.metric("Rank matches", summary["matches"])
                    c.metric("Match rate", f"{summary['match_rate']*100:.1f}%")
                    st.dataframe(pd.DataFrame(comparison), use_container_width=True, hide_index=True)
                else:
                    st.info("The saved Effective Paths reference has no cached rows for this path.")

    with tabs[4]:
        st.dataframe([
            {"Path": "Health labs", "Status": "Native", "Confidence": "High tables / medium model", "Notes": "Exact time/cost tables and transparent eHP effects."},
            {"Path": "Health stones", "Status": "Conditional native", "Confidence": "Low-Medium", "Notes": "DW quantity/cooldown only when DW health saturation is below 100%."},
            {"Path": "Health coins", "Status": "Native", "Confidence": "Medium", "Notes": "Shared enhancement cost curve with health, wall, recovery and defense effects."},
            {"Path": "Modules", "Status": "Partial", "Confidence": "Medium", "Notes": "Equipped base stat and matching sub-effects are included; not every unique effect is modeled."},
        ], use_container_width=True, hide_index=True)
        st.warning("The eHP index is a relative ranking tool. It does not predict an exact maximum wave or displayed in-game HP.")

elif page == "Native eRegen":
    st.header("Native eRegen")
    st.caption(
        "Standalone tower, wall, and recovery-package sustain paths. Exact lab costs and durations are embedded; "
        "the sustain index is relative."
    )
    settings = native_health_settings(profile)
    with st.expander("Model assumptions and controls", expanded=False):
        settings["regen_wall_weight"] = st.number_input(
            "Wall-regen contribution weight", min_value=0.0, max_value=5.0,
            value=float(settings.get("regen_wall_weight", 1.0)), step=0.1,
        )
        st.write(f"Farming armor module: **{native_primary_module_record(profile, 'Armor').get('name','None')}**")
        st.write(f"Wormhole Redirector modeled factor: **{native_module_unique_factor(profile, 'Wormhole Redirector'):.2f}**")
        st.write(f"Latest recorded cause of death: **{native_latest_death(profile)}**")

    steps = st.slider("Path length", 5, 100, 50, key="native_regen_steps")
    native_paths = build_native_health_paths(profile, steps)
    reference_paths = profile.get("roi_reference", {}).get("paths", {})
    components = native_regen_components(profile)

    c1,c2,c3,c4 = st.columns(4)
    c1.metric("Native sustain index", f"x{components['Total']:.4f}")
    c2.metric("Tower regen proxy", f"{components['Tower Regen']:.3f}")
    c3.metric("Wall regen proxy", f"{components['Wall Regen']:.3f}")
    c4.metric("Wall regen lab", f"{components['Wall Regen %']*100:.0f}%")

    with st.expander("Current sustain composition", expanded=False):
        st.dataframe([
            {"Component": "Tower regen", "Value": components["Tower Regen"]},
            {"Component": "Wall regen", "Value": components["Wall Regen"]},
            {"Component": "Package sustain", "Value": components["Package Sustain"]},
            {"Component": "Recovery-health regen", "Value": components["Recovery Regen"]},
            {"Component": "Package chance", "Value": components["Package Chance"]},
            {"Component": "Package amount", "Value": components["Package Amount"]},
            {"Component": "Vampire-specific factor", "Value": components["Vampire Factor"]},
        ], use_container_width=True, hide_index=True)

    tabs = st.tabs(["Lab path", "Coin path", "Regression", "Coverage"])
    for tab, key in zip(tabs[:2], ["regen_lab", "regen_coin"]):
        with tab:
            rows = native_paths[key]
            if rows:
                frame = pd.DataFrame(rows)
                columns = [c for c in ["Rank","Upgrade","Level","Cost","Duration","Gain %","ROI","Confidence","Cumulative","Why"] if c in frame.columns]
                st.dataframe(frame[columns], use_container_width=True, hide_index=True)
                first = rows[0]
                st.info(f"Immediate top choice: **{first['Upgrade']} → {first['Level']}**.")
            else:
                st.info("No eligible upgrades remain in this path.")

    with tabs[2]:
        comparison = compare_native_path(native_paths["regen_lab"], reference_paths.get("regen_lab", {}).get("rows", []), limit=steps)
        summary = native_path_match_summary(comparison)
        if summary["rows"]:
            a,b,c = st.columns(3)
            a.metric("Comparable rows", summary["rows"])
            b.metric("Rank matches", summary["matches"])
            c.metric("Match rate", f"{summary['match_rate']*100:.1f}%")
            st.dataframe(pd.DataFrame(comparison), use_container_width=True, hide_index=True)
        else:
            st.info("The saved Effective Paths workbook had no cached Wall Regen path, so native results cannot yet be regression-compared.")

    with tabs[3]:
        st.dataframe([
            {"Path": "Regen labs", "Status": "Native", "Confidence": "High tables / medium model", "Notes": "Health Regen, Wall Regen, package amount/chance and supporting wall labs."},
            {"Path": "Regen coins", "Status": "Native", "Confidence": "Medium", "Notes": "Health Regen+, Wall Health+ and Recovery Package+ enhancement comparison."},
            {"Path": "Wormhole Redirector", "Status": "Partial", "Confidence": "Medium-Low", "Notes": "Rarity-based sustain factor; exact unique-effect caps remain updateable metadata."},
        ], use_container_width=True, hide_index=True)
        st.warning("The sustain index compares marginal upgrades; it is not an exact health-per-second display prediction.")

elif page == "ROI Paths":
    st.header("Effective Paths ROI Reference")
    reference = profile.get("roi_reference", {})
    paths = reference.get("paths", {}) if isinstance(reference, dict) else {}
    if not paths:
        st.warning("No ROI reference is saved. Import a recalculated Effective Paths workbook under Import / Export → ROI Reference.")
    else:
        source = reference.get("source", {})
        c1, c2, c3 = st.columns(3)
        c1.metric("Effective Paths", source.get("effective_paths_version", "Unknown"))
        c2.metric("Populated paths", sum(1 for p in paths.values() if p.get("rows")))
        c3.metric("Recommendation rows", sum(len(p.get("rows", [])) for p in paths.values()))
        st.caption(
            f"Source: {source.get('filename', 'Unknown')} · imported {source.get('imported_at', 'Unknown')} · mode: {source.get('mode', 'reference')}"
        )
        if reference.get("warnings"):
            with st.expander("Import warnings"):
                for warning in reference.get("warnings", []):
                    st.warning(warning)

        category = st.radio("Category", ["Economy", "Damage", "Health", "Wall Regen"], horizontal=True)
        category_keys = {
            "Economy": ["econ_lab", "econ_stone", "econ_coin", "econ_discount"],
            "Damage": ["damage_lab", "damage_stone", "damage_coin", "damage_key"],
            "Health": ["health_lab", "health_stone", "health_coin"],
            "Wall Regen": ["regen_lab"],
        }
        available = [key for key in category_keys[category] if key in paths]
        if not available:
            st.info("No path tables are available for this category.")
        else:
            labels = {key: paths[key].get("title", key) for key in available}
            selected_key = st.selectbox("Path", available, format_func=lambda key: labels[key])
            rows = roi_path_rows(profile, selected_key)
            show_gold = st.checkbox("Include recommendations already Gold Boxed", value=False, key=f"roi_gold_{selected_key}")
            filtered = [row for row in rows if show_gold or not recommendation_is_gold_boxed(profile, selected_key, row.get("Upgrade", ""))]
            search_roi = st.text_input("Filter upgrades", key=f"roi_search_{selected_key}").strip().casefold()
            if search_roi:
                filtered = [row for row in filtered if search_roi in str(row.get("Upgrade", "")).casefold()]

            row_count = len(filtered)
            if row_count == 0:
                top_n = 0
            elif row_count <= 5:
                top_n = row_count
                st.caption(f"Showing all {row_count} matching recommendation row{'s' if row_count != 1 else ''}.")
            else:
                slider_max = min(100, row_count)
                slider_default = min(25, slider_max)
                top_n = st.slider(
                    "Rows to show",
                    min_value=5,
                    max_value=slider_max,
                    value=slider_default,
                    key=f"roi_rows_{selected_key}",
                )

            display_rows = filtered[:top_n]
            if display_rows:
                display_columns = ["Rank", "Upgrade", "Level", "Cost", "Duration", "ROI", "Result", "Cumulative"]
                frame = pd.DataFrame(display_rows)
                frame = frame[[column for column in display_columns if column in frame.columns]]
                st.dataframe(frame, use_container_width=True, hide_index=True)
                first = display_rows[0]
                st.success(
                    f"Next reference recommendation: {first.get('Upgrade')} · level {first.get('Level')} · "
                    f"cost {first.get('Cost') or first.get('Duration') or '—'} · ROI {first.get('ROI', '—')}"
                )
            else:
                st.info("No recommendations match the current filter.")

        st.divider()
        st.subheader("Cross-path shortlist")
        include_gold = st.checkbox("Include Gold Boxed entries in shortlist", value=False, key="roi_shortlist_gold")
        per_path = st.slider("Recommendations per path", 1, 10, 3, key="roi_shortlist_count")
        summary = roi_summary_rows(profile, include_gold=include_gold, top_per_path=per_path)
        if summary:
            st.dataframe(pd.DataFrame(summary), use_container_width=True, hide_index=True)
        else:
            st.info("No reference recommendations are available.")

elif page == "Player":
    st.header("Player & Progression")
    player_data = profile["player"]
    c1, c2, c3 = st.columns(3)
    player_data["player_id"] = c1.text_input("Player ID", value=str(player_data.get("player_id", "")), key=f"player_id_{revision}")
    player_data["farming_tier"] = c2.text_input("Farming Tier", value=str(player_data.get("farming_tier", "")), key=f"farming_tier_{revision}")
    player_data["tourney_league"] = c3.text_input("Tournament League", value=str(player_data.get("tourney_league", "")), key=f"league_{revision}")
    c4, c5 = st.columns(2)
    player_data["lifetime_coins"] = int(c4.number_input("Lifetime Coins", min_value=0, value=int(player_data.get("lifetime_coins", 0) or 0), step=1, key=f"lifetime_{revision}"))
    player_data["coin_multiplier"] = c5.text_input("Coin Multiplier", value=str(player_data.get("coin_multiplier", "")), key=f"coin_mult_{revision}")
    st.subheader("Tier Progress")
    tier_rows = []
    for tier_num in range(1, 22):
        tier = f"Tier {tier_num}"
        data = player_data.setdefault("tiers", {}).setdefault(tier, {"wave": 0, "attack": 0, "defense": 0, "utility": 0, "ultimate_weapon": 0, "premium": 0})
        tier_rows.append({"Tier": tier, **data})
    edited = st.data_editor(pd.DataFrame(tier_rows), use_container_width=True, hide_index=True, disabled=["Tier"], key=f"tiers_editor_{revision}")
    for row in edited.to_dict("records"):
        tier = row.pop("Tier"); player_data["tiers"][tier] = {k: int(v or 0) for k, v in row.items()}

elif page == "Workshop":
    st.header("Workshop")
    query = st.text_input("Filter Workshop entries", key=f"workshop_filter_{revision}").strip().lower()
    tabs = st.tabs(list(WORKSHOP_GROUPS.keys()))
    for tab, (_, stats) in zip(tabs, WORKSHOP_GROUPS.items()):
        with tab:
            visible = 0
            for stat in stats:
                if query and query not in stat.lower(): continue
                if profile["settings"].get("show_only_incomplete") and is_maxed("workshop", stat): continue
                visible += 1; level_input_with_gold("workshop", stat)
            if visible == 0: st.caption("No matching entries.")

elif page == "Labs":
    st.header("Labs")
    query = st.text_input("Filter Lab entries", key=f"lab_filter_{revision}").strip().lower()
    tabs = st.tabs(list(LAB_GROUPS.keys()))
    for tab, (_, labs) in zip(tabs, LAB_GROUPS.items()):
        with tab:
            visible = 0
            for lab in labs:
                if query and query not in lab.lower(): continue
                if profile["settings"].get("show_only_incomplete") and is_maxed("labs", lab): continue
                visible += 1; level_input_with_gold("labs", lab)
            if visible == 0: st.caption("No matching entries.")

elif page == "Enhancements":
    st.header("Workshop Enhancements")
    query = st.text_input("Filter Enhancements", key=f"enh_filter_{revision}").strip().lower()
    tabs = st.tabs(list(ENHANCEMENT_GROUPS.keys()))
    for tab, (_, entries) in zip(tabs, ENHANCEMENT_GROUPS.items()):
        with tab:
            for entry in entries:
                if query and query not in entry.lower(): continue
                if profile["settings"].get("show_only_incomplete") and is_maxed("enhancements", entry): continue
                level_input_with_gold("enhancements", entry)

elif page == "Ultimate Weapons":
    st.header("Ultimate Weapons")
    selected_uw = st.selectbox("Ultimate Weapon", UW_NAMES, key=f"uw_select_{revision}")
    uw = profile["uw"].setdefault(selected_uw, {"owned": False, "attributes": {}})
    uw["owned"] = st.checkbox("Owned", value=bool(uw.get("owned", False)), key=f"uw_owned_{selected_uw}_{revision}")
    if selected_uw == "Black Hole" and uw["owned"]:
        quantity = 1 + int(profile["labs"].get("Extra Black Hole", 0) or 0)
        st.metric("Black Hole quantity", quantity, help="Base Black Hole plus the Extra Black Hole lab.")
    for attribute in UW_ATTRIBUTE_META[selected_uw]:
        uw_attribute_input(selected_uw, attribute)

elif page == "Modules":
    st.header("Modules")
    equipped_tab, inventory_tab, presets_tab = st.tabs(["Equipped / Manual", "Inventory", "Presets"])
    with equipped_tab:
        for slot in ["Cannon", "Armor", "Generator", "Core"]:
            with st.container(border=True):
                st.subheader(slot)
                mod = profile["modules"].setdefault(slot, {})
                profile["maxed"]["modules"].setdefault(slot, False)
                options = MODULE_OPTIONS_BY_SLOT[slot]
                current_name = mod.get("name", "") if mod.get("name", "") in options else ""
                mod["name"] = st.selectbox(f"{slot} Module", options, index=options.index(current_name), key=f"module_name_{slot}_{revision}")
                current_rarity = mod.get("rarity", "Epic") if mod.get("rarity", "Epic") in RARITY_OPTIONS else "Epic"
                gold = st.checkbox("Gold Box", value=bool(profile["maxed"]["modules"].get(slot, False)), key=f"module_gold_{slot}_{revision}")
                profile["maxed"]["modules"][slot] = gold
                if gold: current_rarity = "Ancestral 5*"; mod["rarity"] = current_rarity; mod["level"] = 300
                rarity = st.selectbox(f"{slot} Rarity", RARITY_OPTIONS, index=RARITY_OPTIONS.index(current_rarity), key=f"module_rarity_{slot}_{revision}", disabled=gold)
                mod["rarity"] = rarity; cap = MODULE_RARITY_MAX_LEVELS[rarity]
                mod["level"] = int(st.number_input(f"{slot} Level (max {cap})", min_value=0, max_value=cap, value=min(int(mod.get("level", 0) or 0), cap), step=1, key=f"module_level_{slot}_{revision}", disabled=gold))
                mod["substats"] = st.text_area(f"{slot} sub-effects (one per line)", value="\n".join(mod.get("substats", [])) if isinstance(mod.get("substats", []), list) else str(mod.get("substats", "")), key=f"module_substats_{slot}_{revision}", height=100).splitlines()
    with inventory_tab:
        inventory = profile.get("module_inventory", {})
        rows = []
        for key, item in inventory.items():
            rows.append({"Key": key, "Slot": item.get("slot"), "Name": item.get("name"), "Rarity": item.get("rarity"), "Level": item.get("level"), "Stat": item.get("stat"), "Sub-effects": len(item.get("substats", []))})
        if rows: st.dataframe(rows, use_container_width=True, hide_index=True)
        else: st.info("Import the Modules companion workbook to populate the full inventory.")
    with presets_tab:
        presets = profile.get("module_presets", {})
        if presets:
            for preset, slots in presets.items():
                st.subheader(preset)
                st.json(slots, expanded=False)
        else: st.info("No module presets imported yet.")

elif page == "Cards":
    st.header("Cards")
    profile["cards"]["slots"] = int(st.number_input("Card slots", min_value=0, value=int(profile["cards"].get("slots", 0) or 0), step=1, key=f"card_slots_{revision}"))
    card_rows = []
    for name in CARD_NAMES:
        item = profile["cards"]["items"].setdefault(name, {"level": 0, "mastery": 0})
        card_rows.append({"Card": name, "Level": int(item.get("level", 0)), "Mastery": int(item.get("mastery", 0))})
    edited = st.data_editor(pd.DataFrame(card_rows), use_container_width=True, hide_index=True, disabled=["Card"], column_config={"Level": st.column_config.NumberColumn(min_value=0, max_value=7, step=1), "Mastery": st.column_config.NumberColumn(min_value=0, max_value=9, step=1)}, key=f"cards_editor_{revision}")
    for row in edited.to_dict("records"): profile["cards"]["items"][row["Card"]] = {"level": int(row["Level"] or 0), "mastery": int(row["Mastery"] or 0)}

elif page == "Relics":
    st.header("Relics")
    items = profile["relics"].get("items", {})
    st.metric("Owned relics", sum(bool(v.get("owned")) for v in items.values()), delta=f"{len(items)} catalogued")
    search = st.text_input("Filter relics", key=f"relic_search_{revision}").lower()
    rows = []
    for name, item in items.items():
        if search and search not in name.lower() and search not in str(item.get("bonus_type", "")).lower(): continue
        rows.append({"Relic": name, "Owned": bool(item.get("owned")), "Rarity": item.get("rarity"), "Bonus Type": item.get("bonus_type"), "Value": item.get("value"), "Type": item.get("type")})
    if rows:
        edited = st.data_editor(pd.DataFrame(rows), use_container_width=True, hide_index=True, disabled=["Relic", "Rarity", "Bonus Type", "Value", "Type"], key=f"relic_editor_{revision}")
        for row in edited.to_dict("records"): profile["relics"]["items"].setdefault(row["Relic"], {})["owned"] = bool(row["Owned"])
    else: st.info("Import the Relics companion workbook to populate individual relics.")

elif page == "Themes & Songs":
    st.header("Themes & Songs")
    items = profile["themes"].get("items", {})
    rows = [{"Key": key, "Type": item.get("type"), "Name": item.get("name"), "Owned": bool(item.get("owned")), "Source": item.get("event") or item.get("tier") or item.get("source")} for key, item in items.items()]
    if rows:
        edited = st.data_editor(pd.DataFrame(rows), use_container_width=True, hide_index=True, disabled=["Key", "Type", "Name", "Source"], key=f"themes_editor_{revision}")
        for row in edited.to_dict("records"): profile["themes"]["items"].setdefault(row["Key"], {})["owned"] = bool(row["Owned"])
    else: st.info("Import the Themes & Songs companion workbook to populate ownership.")

elif page == "Bots":
    st.header("Bots")
    selected = st.selectbox("Bot", BOT_NAMES, key=f"bot_select_{revision}")
    bot = profile["bots"].setdefault(selected, {"unlocked": False, "attributes": {}, "plus": {}})
    bot["unlocked"] = st.checkbox("Unlocked", value=bool(bot.get("unlocked")), key=f"bot_unlock_{selected}_{revision}")
    for attr in BOT_ATTRIBUTES[selected]:
        val = bot.setdefault("attributes", {}).get(attr, 0)
        bot["attributes"][attr] = st.number_input(attr, value=float(val or 0), step=0.01, key=f"bot_{selected}_{attr}_{revision}")
    if bot.get("plus"): st.json(bot["plus"], expanded=False)

elif page == "Guardians":
    st.header("Guardians")
    selected = st.selectbox("Guardian", GUARDIAN_NAMES, key=f"guardian_select_{revision}")
    guardian = profile["guardians"].setdefault(selected, {"unlocked": False, "attributes": {}, "bits_spent": 0})
    guardian["unlocked"] = st.checkbox("Unlocked", value=bool(guardian.get("unlocked")), key=f"guardian_unlock_{selected}_{revision}")
    guardian["bits_spent"] = int(st.number_input("Bits Spent", min_value=0, value=int(guardian.get("bits_spent", 0) or 0), step=1, key=f"guardian_bits_{selected}_{revision}"))
    for attr in GUARDIAN_ATTRIBUTES[selected]:
        val = guardian.setdefault("attributes", {}).get(attr, 0)
        guardian["attributes"][attr] = st.number_input(attr, value=float(val or 0), step=0.01, key=f"guardian_{selected}_{attr}_{revision}")

elif page == "Vault":
    st.header("Vault")
    vault = profile["vault"]
    vault["keys_spent"] = int(st.number_input("Keys Spent", min_value=0, value=int(vault.get("keys_spent", 0) or 0), step=1, key=f"vault_keys_{revision}"))
    bonus_rows = [{"Bonus": k, "Active": v.get("active"), "Total": v.get("total")} for k, v in vault.get("bonuses", {}).items()]
    unlock_rows = [{"Unlock": k, "Owned": bool(v)} for k, v in vault.get("unlocks", {}).items()]
    if bonus_rows:
        edited = st.data_editor(pd.DataFrame(bonus_rows), use_container_width=True, hide_index=True, disabled=["Bonus", "Total"], key=f"vault_bonus_{revision}")
        for row in edited.to_dict("records"): vault["bonuses"].setdefault(row["Bonus"], {})["active"] = row["Active"]
    if unlock_rows:
        edited = st.data_editor(pd.DataFrame(unlock_rows), use_container_width=True, hide_index=True, disabled=["Unlock"], key=f"vault_unlock_{revision}")
        for row in edited.to_dict("records"): vault["unlocks"][row["Unlock"]] = bool(row["Owned"])
    if not bonus_rows and not unlock_rows: st.info("Import the Vault companion workbook to populate bonuses and unlocks.")

elif page == "Optimizer":
    st.header("Optimizer")
    reference = profile.get("roi_reference", {})
    paths = reference.get("paths", {}) if isinstance(reference, dict) else {}
    native_paths = build_native_econ_paths(profile, 50)
    native_health_paths = build_native_health_paths(profile, 50)

    cpk = int(profile["labs"].get("Coins / Kill Bonus", 0) or 0)
    gt = profile["uw"].get("Golden Tower", {})
    bh = profile["uw"].get("Black Hole", {})
    bh_qty = 1 + int(profile["labs"].get("Extra Black Hole", 0) or 0) if bh.get("owned") else 0
    c1, c2, c3, c4 = st.columns(4)
    c1.metric("Coins/Kill Lab", cpk)
    c2.metric("Golden Tower", "Owned" if gt.get("owned") else "Missing")
    c3.metric("Black Hole", "Owned" if bh.get("owned") else "Missing")
    c4.metric("Black Hole Quantity", bh_qty)

    st.subheader("Standalone economy recommendations")
    immediate_rows = []
    for native_key, label in [("econ_lab", "Lab"), ("econ_stone", "Stones"), ("econ_coin", "Coins"), ("econ_discount", "Discount")]:
        if native_paths.get(native_key):
            first = native_paths[native_key][0]
            immediate_rows.append({
                "Resource": label,
                "Upgrade": first.get("Upgrade"),
                "Next level": first.get("Level"),
                "Cost / Time": first.get("Cost") or first.get("Duration"),
                "Estimated gain %": first.get("Gain %"),
                "Native ROI": first.get("ROI"),
            })
    if immediate_rows:
        st.dataframe(immediate_rows, use_container_width=True, hide_index=True)
        st.caption("These economy rows are calculated natively. Open Native eEcon for full paths, assumptions, and regression details.")

    st.subheader("Standalone survival recommendations")
    survival_rows = []
    for native_key, label in [("health_lab", "eHP Lab"), ("health_coin", "eHP Coins"), ("regen_lab", "Regen Lab"), ("regen_coin", "Regen Coins")]:
        if native_health_paths.get(native_key):
            first = native_health_paths[native_key][0]
            survival_rows.append({
                "Path": label, "Upgrade": first.get("Upgrade"), "Next level": first.get("Level"),
                "Cost / Time": first.get("Cost") or first.get("Duration"),
                "Estimated gain %": first.get("Gain %"), "Native ROI": first.get("ROI"),
            })
    if survival_rows:
        st.dataframe(survival_rows, use_container_width=True, hide_index=True)
        st.caption("These rows are calculated natively. Open Native eHP or Native eRegen for assumptions and regression details.")

    if paths:
        source = reference.get("source", {})
        st.success(
            f"ROI reference connected: {source.get('effective_paths_version', 'Effective Paths')} from {source.get('filename', 'workbook')}."
        )
        st.caption(
            "Recommendations below are imported calculated outputs from Effective Paths and filtered against exact Gold Box matches where possible. "
            "They are the reference layer while the standalone Python formulas are implemented and regression-tested."
        )
        tabs = st.tabs(["Next upgrades", "Economy", "Damage", "Health / Regen", "Reference status"])
        with tabs[0]:
            top_each = st.slider("Top entries from each path", 1, 8, 3, key="optimizer_top_each")
            shortlist = roi_summary_rows(profile, include_gold=False, top_per_path=top_each)
            if shortlist:
                st.dataframe(pd.DataFrame(shortlist), use_container_width=True, hide_index=True)
            else:
                st.info("No non-maxed recommendations were found.")
        with tabs[1]:
            for key in ["econ_lab", "econ_stone", "econ_coin", "econ_discount"]:
                path = paths.get(key, {})
                rows = [r for r in path.get("rows", []) if not recommendation_is_gold_boxed(profile, key, r.get("Upgrade", ""))]
                st.subheader(path.get("title", key))
                if rows:
                    st.dataframe(pd.DataFrame(rows[:10])[[c for c in ["Rank", "Upgrade", "Level", "Cost", "Duration", "ROI", "Result"] if c in pd.DataFrame(rows[:10]).columns]], use_container_width=True, hide_index=True)
                else:
                    st.caption("No cached rows available.")
        with tabs[2]:
            for key in ["damage_lab", "damage_stone", "damage_coin", "damage_key"]:
                path = paths.get(key, {})
                rows = [r for r in path.get("rows", []) if not recommendation_is_gold_boxed(profile, key, r.get("Upgrade", ""))]
                st.subheader(path.get("title", key))
                if rows:
                    frame = pd.DataFrame(rows[:10])
                    st.dataframe(frame[[c for c in ["Rank", "Upgrade", "Level", "Cost", "Duration", "ROI", "Result"] if c in frame.columns]], use_container_width=True, hide_index=True)
                else:
                    st.caption("No cached rows available.")
        with tabs[3]:
            for key in ["health_lab", "health_stone", "health_coin", "regen_lab"]:
                path = paths.get(key, {})
                rows = [r for r in path.get("rows", []) if not recommendation_is_gold_boxed(profile, key, r.get("Upgrade", ""))]
                st.subheader(path.get("title", key))
                if rows:
                    frame = pd.DataFrame(rows[:10])
                    st.dataframe(frame[[c for c in ["Rank", "Upgrade", "Level", "Cost", "Duration", "ROI", "Result"] if c in frame.columns]], use_container_width=True, hide_index=True)
                else:
                    st.caption("No cached rows available.")
        with tabs[4]:
            status_rows = []
            for key, spec in ROI_PATH_SPECS.items():
                path = paths.get(key, {})
                status_rows.append({
                    "Path": spec["title"],
                    "Rows": len(path.get("rows", [])),
                    "Status": "Ready" if path.get("rows") else "Missing cached output",
                    "Native Python formula": "Ready" if key.startswith(("econ_", "damage_", "health_", "regen_")) else "Pending",
                })
            st.dataframe(status_rows, use_container_width=True, hide_index=True)
            st.info(
                "Economy, damage, health, and regeneration now have native paths; remaining work is calibration, combined prioritization, and update automation."
            )
    else:
        recommendations = []
        if gt.get("owned") and bh.get("owned") and bh_qty >= 2:
            recommendations.append("The GT + 2BH economy core is online.")
            for lab in ["Black Hole Coin Bonus", "Golden Tower Bonus"]:
                if not profile["maxed"]["labs"].get(lab, False):
                    recommendations.append(f"Keep comparing {lab} against your other economy labs.")
        elif gt.get("owned") and bh.get("owned"):
            recommendations.append("GT + BH is online; Extra Black Hole remains a major milestone.")
        else:
            recommendations.append("The GT/BH economy core is incomplete.")
        if not profile["maxed"]["labs"].get("Coins / Kill Bonus", False):
            recommendations.append("Coins / Kill Bonus remains eligible for ROI comparison.")
        st.warning("No Effective Paths ROI reference has been imported yet.")
        for recommendation in recommendations:
            st.write(f"- {recommendation}")
        st.info("Import a recalculated Effective Paths workbook under Import / Export → ROI Reference to activate ranked paths.")

elif page == "Game Data":
    st.header("Game Data Coverage")
    tabs = st.tabs(["Workshop", "Labs", "Enhancements", "UWs", "Modules", "Cards", "Bots", "Guardians"])
    with tabs[0]: st.dataframe([{"Entry": k, "Maximum": v} for k, v in WORKSHOP_MAX_LEVELS.items()], use_container_width=True, hide_index=True)
    with tabs[1]: st.dataframe([{"Entry": k, "Maximum": v} for k, v in LAB_MAX_LEVELS.items()], use_container_width=True, hide_index=True)
    with tabs[2]: st.dataframe([{"Entry": k, "Maximum": v} for k, v in ENHANCEMENT_MAX_LEVELS.items()], use_container_width=True, hide_index=True)
    with tabs[3]:
        rows = []
        for uw, attrs in UW_ATTRIBUTE_META.items():
            for attr, meta in attrs.items(): rows.append({"Ultimate Weapon": uw, "Attribute": attr, "Maximum": meta["max"]})
        st.dataframe(rows, use_container_width=True, hide_index=True)
    with tabs[4]:
        st.dataframe([{"Rarity": k, "Max Level": v} for k, v in MODULE_RARITY_MAX_LEVELS.items()], use_container_width=True, hide_index=True)
        for slot, names in MODULE_OPTIONS_BY_SLOT.items(): st.write(f"**{slot}:** {', '.join(n for n in names if n)}")
    with tabs[5]: st.dataframe([{"Card": name, "Max Level": 7, "Max Mastery": 9} for name in CARD_NAMES], use_container_width=True, hide_index=True)
    with tabs[6]:
        for name in BOT_NAMES: st.write(f"**{name}:** {', '.join(BOT_ATTRIBUTES[name])}")
    with tabs[7]:
        for name in GUARDIAN_NAMES: st.write(f"**{name}:** {', '.join(GUARDIAN_ATTRIBUTES[name])}")

elif page == "System & Updates":
    st.header("System & Updates")
    st.caption("Backups, workbook compatibility, reviewed game-data updates, visual-preview health, data-change detection, and diagnostics.")

    backup_tab, compatibility_tab, data_update_tab, engine_health_tab, diagnostics_tab, about_tab = st.tabs([
        "Backups", "Workbook Compatibility", "Data Update Center", "Engine Health", "Diagnostics", "About"
    ])

    with backup_tab:
        st.subheader("Profile backups")
        current_path = PROFILE_DIR / f"{safe_profile_filename(profile.get('name', 'default'))}.json"
        col1, col2 = st.columns(2)
        with col1:
            if st.button("Create backup now", use_container_width=True):
                if not current_path.exists():
                    save_profile(profile.get("name", "default"), profile)
                backup = create_profile_backup(current_path, PROFILE_DIR, reason="manual")
                if backup:
                    st.success(f"Backup created: {backup.name}")
                else:
                    st.warning("No saved profile existed to back up.")
        with col2:
            st.info("Every profile save now creates an automatic backup before replacing the existing file.")

        backups = list_profile_backups(PROFILE_DIR, profile.get("name"))
        if backups:
            st.dataframe(pd.DataFrame(backups)[["file", "created", "size_kb", "reason"]], use_container_width=True, hide_index=True)
            selected_backup = st.selectbox("Backup to restore", [row["file"] for row in backups], key="restore_backup_select")
            selected_row = next(row for row in backups if row["file"] == selected_backup)
            confirm_restore = st.checkbox("I understand this will replace the currently saved profile", key="confirm_restore")
            if st.button("Restore selected backup", type="primary", disabled=not confirm_restore):
                restore_profile_backup(Path(selected_row["path"]), current_path, PROFILE_DIR)
                st.session_state.profile = load_profile(profile.get("name", "default"))
                bump_revision()
                st.success("Backup restored.")
                st.rerun()
        else:
            st.info("No backups exist for this profile yet.")

    with compatibility_tab:
        st.subheader("Workbook version compatibility")
        st.write("Upload one or more IDS companion or Effective Paths workbooks. Nothing is imported on this page.")
        update_uploads = st.file_uploader(
            "Workbooks to inspect", type=["xlsx"], accept_multiple_files=True, key="update_check_uploads"
        )
        if update_uploads:
            compatibility_rows = []
            effective_paths_upload = None
            for upload in update_uploads:
                payload = upload.getvalue()
                info = workbook_compatibility(upload.name, payload)
                compatibility_rows.append({
                    "File": info.get("filename"),
                    "Type": info.get("kind"),
                    "Found version": info.get("version") or "Unknown",
                    "Supported version": info.get("supported_version") or "Unknown",
                    "Status": info.get("status"),
                    "Sheets": info.get("sheets", "—"),
                    "Error": info.get("error", ""),
                })
                if info.get("kind") == "Effective Paths":
                    effective_paths_upload = payload
            st.dataframe(pd.DataFrame(compatibility_rows), use_container_width=True, hide_index=True)
            if any(row["Status"] == "Newer than supported" for row in compatibility_rows):
                st.warning("A newer workbook version was detected. Review data changes before importing it into a production profile.")

            if effective_paths_upload is not None:
                st.subheader("Effective Paths maximum-level comparison")
                try:
                    incoming_caps = extract_effective_paths_caps(effective_paths_upload)
                    workshop_changes = compare_cap_maps(WORKSHOP_MAX_LEVELS, incoming_caps.get("workshop", {}))
                    lab_changes = compare_cap_maps(LAB_MAX_LEVELS, incoming_caps.get("labs", {}))
                    c1, c2 = st.columns(2)
                    c1.metric("Workshop changes", len(workshop_changes))
                    c2.metric("Lab changes", len(lab_changes))
                    if workshop_changes:
                        st.write("**Workshop changes**")
                        st.dataframe(pd.DataFrame(workshop_changes), use_container_width=True, hide_index=True)
                    if lab_changes:
                        st.write("**Lab changes**")
                        st.dataframe(pd.DataFrame(lab_changes), use_container_width=True, hide_index=True)
                    if not workshop_changes and not lab_changes:
                        st.success("No Workshop or Lab maximum-level changes were detected.")
                except Exception as exc:
                    st.error(f"Could not compare Effective Paths metadata: {exc}")

        st.subheader("Supported workbook versions")
        st.dataframe(
            pd.DataFrame([{"Workbook": name, "Supported version": version} for name, version in SUPPORTED_WORKBOOK_VERSIONS.items()]),
            use_container_width=True, hide_index=True,
        )

    with data_update_tab:
        st.subheader("Game-data update workflow")
        st.caption(
            "Review newer Effective Paths and IDS companion workbooks, export a change manifest, "
            "and activate a reversible local metadata overlay. Uploaded workbooks are never copied into the update package."
        )

        active_update = get_active_update()
        active_health = active_update_health()
        active_cols = st.columns(4)
        active_cols[0].metric("Overlay status", active_health.get("Status", "INFO"))
        active_cols[1].metric("Update ID", active_update.get("update_id", "Bundled data"))
        active_cols[2].metric("Effective Paths", active_update.get("effective_paths_version", "Bundled"))
        active_cols[3].metric("Applied", str(active_update.get("applied_at", "—"))[:19])
        st.caption(active_health.get("Detail", ""))

        if active_update:
            rollback_confirm = st.checkbox(
                "I understand rollback restores the metadata bundled with this app release",
                key="game_update_rollback_confirm",
            )
            if st.button(
                "Roll back active metadata update",
                disabled=not rollback_confirm,
                key="game_update_rollback_button",
                use_container_width=True,
            ):
                archived = rollback_active_update()
                if archived:
                    st.success(f"The active overlay was rolled back and archived as {archived.name}.")
                    st.warning("Stop and restart Streamlit before using recommendations so every engine reloads the bundled metadata.")
                else:
                    st.info("No active update was found.")

        st.divider()
        update_files = st.file_uploader(
            "Effective Paths and/or IDS companion workbooks",
            type=["xlsx"],
            accept_multiple_files=True,
            key="game_data_update_uploads",
            help="A full bundle gives the clearest version audit. Effective Paths is required only when maximum-level metadata needs to change.",
        )
        if st.button(
            "Analyze uploaded update bundle",
            type="primary",
            disabled=not bool(update_files),
            key="analyze_game_data_update",
            use_container_width=True,
        ):
            try:
                st.session_state.game_data_update_report = analyze_update_bundle(
                    [(upload.name, upload.getvalue()) for upload in update_files or []],
                    WORKSHOP_MAX_LEVELS,
                    LAB_MAX_LEVELS,
                    SUPPORTED_WORKBOOK_VERSIONS,
                    app_version=APP_VERSION,
                )
            except Exception as exc:
                st.session_state.pop("game_data_update_report", None)
                st.error(f"The update bundle could not be analyzed: {exc}")

        update_report = st.session_state.get("game_data_update_report")
        if update_report:
            risk = update_report.get("risk_level", "UNKNOWN")
            validation = update_report.get("validation", "UNKNOWN")
            coverage = update_report.get("coverage", {})
            result_cols = st.columns(4)
            result_cols[0].metric("Validation", validation)
            result_cols[1].metric("Risk", risk)
            result_cols[2].metric("Workshop coverage", f"{coverage.get('Workshop %', 0)}%")
            result_cols[3].metric("Lab coverage", f"{coverage.get('Labs %', 0)}%")

            if risk == "BLOCKED":
                st.error("This candidate is blocked and cannot be staged or applied.")
            elif risk == "REVIEW":
                st.warning("This candidate changes versions or maximum-level metadata and requires explicit review.")
            else:
                st.success("No blocking or review-level differences were found.")

            if update_report.get("blocked_reasons"):
                st.write("**Blocking findings**")
                for reason in update_report["blocked_reasons"]:
                    st.write(f"- {reason}")
            if update_report.get("review_reasons"):
                st.write("**Review findings**")
                for reason in update_report["review_reasons"]:
                    st.write(f"- {reason}")
            for note in update_report.get("notes", []):
                st.caption(note)

            st.write("**Workbook audit**")
            files_frame = pd.DataFrame(update_report.get("files", []))
            if not files_frame.empty:
                display_columns = [
                    column for column in [
                        "File", "Type", "Version", "Supported version", "Compatibility",
                        "Sheets", "Size KB", "Has EXPORT", "Has _IDS", "Error",
                    ] if column in files_frame.columns
                ]
                st.dataframe(files_frame[display_columns], use_container_width=True, hide_index=True)

            change_tabs = st.tabs(["Workshop changes", "Lab changes", "Candidate JSON"])
            with change_tabs[0]:
                workshop_change_rows = update_report.get("workshop_changes", [])
                if workshop_change_rows:
                    st.dataframe(pd.DataFrame(workshop_change_rows), use_container_width=True, hide_index=True)
                else:
                    st.success("No Workshop maximum-level changes detected.")
            with change_tabs[1]:
                lab_change_rows = update_report.get("lab_changes", [])
                if lab_change_rows:
                    st.dataframe(pd.DataFrame(lab_change_rows), use_container_width=True, hide_index=True)
                else:
                    st.success("No Lab maximum-level changes detected.")
            with change_tabs[2]:
                st.json(update_report.get("candidate", {}), expanded=False)

            candidate = update_report.get("candidate", {})
            action_cols = st.columns(3)
            with action_cols[0]:
                st.download_button(
                    "Download review package",
                    data=export_update_bundle(update_report),
                    file_name=f"tower_data_update_{candidate.get('update_id', 'candidate')}.zip",
                    mime="application/zip",
                    use_container_width=True,
                )
            with action_cols[1]:
                if st.button(
                    "Stage candidate",
                    disabled=risk == "BLOCKED",
                    key="stage_game_data_candidate",
                    use_container_width=True,
                ):
                    try:
                        staged_path = stage_update_candidate(candidate)
                        st.success(f"Staged as {staged_path.name}.")
                    except Exception as exc:
                        st.error(str(exc))
            with action_cols[2]:
                review_confirm = st.checkbox(
                    "Reviewed",
                    value=False,
                    disabled=risk != "REVIEW",
                    key="confirm_game_data_review",
                    help="Required when a workbook is newer/older or maximum-level metadata changes.",
                )
                allow_apply = risk == "SAFE" or (risk == "REVIEW" and review_confirm)
                if st.button(
                    "Apply metadata overlay",
                    disabled=not allow_apply or risk == "BLOCKED",
                    key="apply_game_data_candidate",
                    use_container_width=True,
                ):
                    try:
                        active_path = apply_update_candidate(candidate, allow_review=review_confirm)
                        st.success(f"Activated {active_path.name}.")
                        st.warning("Stop Streamlit with Ctrl+C and restart it before using recommendations. This reloads the same metadata in every engine.")
                    except Exception as exc:
                        st.error(str(exc))

        st.divider()
        st.write("**Update history**")
        update_history = list_update_history()
        if update_history:
            st.dataframe(pd.DataFrame(update_history), use_container_width=True, hide_index=True)
        else:
            st.info("No metadata updates have been applied or rolled back yet.")
        st.info(
            "Automatic updates are deliberately limited to recognized workbook versions and complete Workshop/Lab maximum maps. "
            "Formula, cost, value, UW, module, card, relic, bot, guardian, and Vault model changes still require a tested app release."
        )

    with engine_health_tab:
        st.subheader("Standalone engine health")
        st.caption("Runs engine execution, profile quality, bundled data, and optional Effective Paths calibration checks against the current profile.")
        health_report = run_engine_health(profile, steps=10)
        overall = health_report.get("overall", "UNKNOWN")
        quality_health = health_report.get("quality", {})
        h1, h2, h3, h4 = st.columns(4)
        h1.metric("Overall", overall)
        h2.metric("Profile quality", f"{quality_health.get('score', 0)}/100")
        h3.metric("Quality errors", quality_health.get("counts", {}).get("Error", 0))
        h4.metric("Calibration", health_report.get("calibration", {}).get("overall", "INFO"))
        if overall == "PASS":
            st.success("All required standalone engine and profile checks passed.")
        elif overall == "WARN":
            st.warning("The engines run, but warning-level quality or calibration differences need review.")
        else:
            st.error("An engine, bundled-data, or profile-quality check failed.")

        health_tabs = st.tabs(["Engine Execution", "Calibration", "Profile Quality", "Bundled Data", "Actions"])
        with health_tabs[0]:
            st.dataframe(pd.DataFrame(health_report.get("engines", [])), use_container_width=True, hide_index=True)
        with health_tabs[1]:
            st.dataframe(pd.DataFrame(health_report.get("comparisons", [])), use_container_width=True, hide_index=True)
            st.caption("No imported reference is an INFO result, not an engine failure.")
        with health_tabs[2]:
            quality_issues = quality_health.get("issues", [])
            if quality_issues:
                st.dataframe(pd.DataFrame(quality_issues), use_container_width=True, hide_index=True)
            else:
                st.success("No data-quality findings.")
        with health_tabs[3]:
            st.dataframe(pd.DataFrame(health_report.get("game_data", [])), use_container_width=True, hide_index=True)
        with health_tabs[4]:
            actions = health_report.get("actions", [])
            if actions:
                for action in actions:
                    st.write(f"- {action}")
            else:
                st.success("No corrective actions are currently suggested.")

        st.download_button(
            "Download engine health report",
            data=json.dumps(health_report, indent=2),
            file_name=f"{safe_profile_filename(profile.get('name', 'profile'))}_engine_health.json",
            mime="application/json",
            use_container_width=True,
        )

    with diagnostics_tab:
        st.subheader("Application self-check")
        issues = profile_audit(profile)
        self_check = {
            "profile_name": profile.get("name"),
            "profile_shape": "OK" if isinstance(profile, dict) and all(key in profile for key in ["workshop", "labs", "uw", "modules"]) else "FAIL",
            "audit_findings": len(issues),
            "econ_paths": {key: len(value) for key, value in build_native_econ_paths(profile, 10).items()},
            "damage_paths": {key: len(value) for key, value in build_native_damage_paths(profile, 10).items()},
            "health_paths": {key: len(value) for key, value in build_native_health_paths(profile, 10).items()},
            "roi_reference_loaded": bool(profile.get("roi_reference", {}).get("imported_at")),
            "data_quality": profile_quality_report(profile).get("overall"),
            "calibration": build_calibration_report(profile, 5).get("overall"),
            "planner_actions": len(build_progression_plan(profile).get("daily_actions", [])),
            "battle_learning_runs": len(build_battle_learning_report(profile).get("runs", [])),
            "battle_learning_bottlenecks": len(build_battle_learning_report(profile).get("bottlenecks", [])),
            "visual_sync_status": build_sync_report(profile).get("status"),
            "visual_card_slots": build_card_report(profile).get("slots"),
            "visual_module_records": build_module_forge_report(profile).get("module_names"),
            "visual_relic_records": build_relic_report(profile).get("total"),
            "custom_icon_files": custom_icon_count(),
            "fixed_icon_slots_ready": sum(1 for row in fixed_icon_status() if row.get("exists")),
            "saved_profile_exists": (PROFILE_DIR / f"{safe_profile_filename(profile.get('name', 'default'))}.json").exists(),
        }
        st.json(self_check, expanded=True)
        if self_check["profile_shape"] == "OK":
            st.success("Core profile and native engines completed the self-check.")
        else:
            st.error("The profile structure failed the self-check.")

        diagnostic_bytes = build_diagnostic_zip(profile, APP_VERSION, issues, self_check)
        st.download_button(
            "Download diagnostic package",
            data=diagnostic_bytes,
            file_name=f"{safe_profile_filename(profile.get('name', 'profile'))}_tower_optimizer_diagnostics.zip",
            mime="application/zip",
            use_container_width=True,
        )
        st.caption("The diagnostic ZIP contains the current profile, audit findings, self-check results, and basic system information. It does not include source workbooks.")

    with about_tab:
        st.subheader("Tower Optimizer")
        st.write(f"Application version: **{APP_VERSION}**")
        st.write(f"Embedded Effective Paths reference: **{DATA_SOURCE_VERSION}**")
        st.write("v2.0 Preview 6 adds one-click saving and multi-report paste importing while retaining the robust current-format parser, GitHub-ready foundation, expandable navigation, Icon Studio, and every prior preview feature.")
        st.write("v1.10 added a staged, reversible game-data update workflow with workbook hashing, version audits, coverage checks, review manifests, and local maximum-level overlays.")
        st.write("v1.9 added battle-history analytics, farming-tier recommendations, bulk report import, upgrade impact comparisons, and conservative feedback into recommendation priorities.")
        st.write("v1.8 added daily planning, lab-slot scheduling, resource forecasting, a persistent upgrade queue, and weekly progression reports.")
        st.write("v1.7 added whole-account opportunity-cost recommendations for cards, modules, relics, themes, bots, guardians, and the Vault.")
        st.info("Effective Paths remains optional for normal operation. Local updates never overwrite formula or cost tables; those still arrive through tested app releases.")

elif page == "Raw Profile":
    st.header("Raw Profile JSON")
    st.json(profile)
    st.download_button("Download profile JSON", data=json.dumps(profile, indent=2), file_name=f"{safe_profile_filename(profile['name'])}.json", mime="application/json")
