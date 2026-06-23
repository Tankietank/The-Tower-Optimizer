"""Fill Effective Paths Master Sheet cells from a Tower Optimizer profile."""
from __future__ import annotations

from typing import Any, Dict, Mapping, MutableMapping, Optional, Tuple

from openpyxl.worksheet.worksheet import Worksheet

from .engines.core import (
    ENHANCEMENT_MAX_LEVELS,
    LAB_ALIASES,
    LAB_MAX_LEVELS,
    UW_ATTRIBUTE_META,
    UW_NAMES,
    WORKSHOP_MAX_LEVELS,
)
from .quality import WORKSHOP_ALIASES


def _clean(value: Any) -> str:
    return "" if value is None else str(value).strip()


def _header_matches(text: str, needle: str) -> bool:
    folded = text.casefold()
    target = needle.casefold()
    if folded == target:
        return True
    # Writable workbooks expose HYPERLINK(...) formulas instead of display text.
    return target in folded


def _detect_columns(sheet: Worksheet, *, max_row: int = 12) -> Dict[str, int]:
    needles = {
        "lab_header": "Go to my Laboratory Sheet",
        "workshop_header": "Go to my Workshop Sheet",
        "uw_header": "Go to my Ultimate Weapon Sheet",
    }
    found: Dict[str, list[tuple[int, int]]] = {key: [] for key in needles}
    for row_idx in range(1, max_row + 1):
        for col_idx in range(1, 40):
            text = _clean(sheet.cell(row_idx, col_idx).value)
            for key, needle in needles.items():
                if _header_matches(text, needle):
                    found[key].append((row_idx, col_idx))
    if not found["lab_header"] or len(found["workshop_header"]) < 2 or not found["uw_header"]:
        # EP v5.06.04.00 Master Sheet layout (fallback when headers are stripped).
        return {
            "header_row": 1,
            "lab_col": 5,
            "workshop_col": 10,
            "enhancement_col": 15,
            "uw_header_col": 20,
        }
    header_row = min(
        found["lab_header"][0][0],
        found["workshop_header"][0][0],
        found["uw_header"][0][0],
    )
    return {
        "header_row": header_row,
        "lab_col": found["lab_header"][0][1],
        "workshop_col": found["workshop_header"][0][1],
        "enhancement_col": found["workshop_header"][1][1],
        "uw_header_col": found["uw_header"][0][1],
    }


def fill_master_sheet_from_profile(
    sheet: Worksheet,
    profile: Mapping[str, Any],
    *,
    max_row: int = 120,
) -> Dict[str, int]:
    """Write workshop, labs, enhancements, and UW levels into Master Sheet rows."""
    cols = _detect_columns(sheet)
    header_row = cols["header_row"]
    lab_col = cols["lab_col"]
    workshop_col = cols["workshop_col"]
    enhancement_col = cols["enhancement_col"]
    uw_header_col = cols["uw_header_col"]
    uw_name_col = uw_header_col + 1
    uw_attr_col = uw_header_col + 2
    uw_value_col = uw_header_col + 3

    labs = profile.get("labs") or {}
    workshop = profile.get("workshop") or {}
    enhancements = profile.get("enhancements") or {}
    uw_data = profile.get("uw") or {}

    stats = {"labs": 0, "workshop": 0, "enhancements": 0, "uw_attributes": 0, "uw_owned": 0}

    current_uw: Optional[str] = None
    for row_idx in range(header_row + 1, max_row + 1):
        lab_name = _clean(sheet.cell(row_idx, lab_col).value)
        if lab_name:
            canonical_lab = LAB_ALIASES.get(lab_name, lab_name)
            if canonical_lab in LAB_MAX_LEVELS and canonical_lab in labs:
                sheet.cell(row_idx, lab_col + 1).value = int(labs[canonical_lab])
                stats["labs"] += 1

        workshop_name = _clean(sheet.cell(row_idx, workshop_col).value)
        if workshop_name:
            canonical_ws = WORKSHOP_ALIASES.get(workshop_name, workshop_name)
            if canonical_ws in WORKSHOP_MAX_LEVELS and canonical_ws in workshop:
                level = int(workshop[canonical_ws])
                sheet.cell(row_idx, workshop_col + 1).value = level
                sheet.cell(row_idx, workshop_col + 2).value = level
                stats["workshop"] += 1

        enhancement_name = _clean(sheet.cell(row_idx, enhancement_col).value)
        if enhancement_name in ENHANCEMENT_MAX_LEVELS and enhancement_name in enhancements:
            sheet.cell(row_idx, enhancement_col + 2).value = int(enhancements[enhancement_name])
            stats["enhancements"] += 1

        name_or_status = _clean(sheet.cell(row_idx, uw_name_col).value)
        attribute = _clean(sheet.cell(row_idx, uw_attr_col).value)

        if name_or_status in UW_NAMES:
            current_uw = name_or_status
        elif name_or_status == "UW Unlocked" and current_uw and current_uw in uw_data:
            owned = bool((uw_data[current_uw] or {}).get("owned"))
            sheet.cell(row_idx, uw_name_col).value = "UW Unlocked" if owned else "UW Locked"
            stats["uw_owned"] += 1
        elif current_uw and current_uw in uw_data and attribute in UW_ATTRIBUTE_META.get(current_uw, {}):
            attrs = (uw_data[current_uw] or {}).get("attributes") or {}
            if attribute in attrs:
                value = attrs[attribute]
                meta = UW_ATTRIBUTE_META[current_uw][attribute]
                sheet.cell(row_idx, uw_value_col).value = int(round(value)) if isinstance(meta.get("max"), int) else float(value)
                stats["uw_attributes"] += 1

    return stats


def read_master_sheet_profile(sheet: Worksheet) -> Dict[str, Dict[str, Any]]:
    """Best-effort read-back using the same Master Sheet layout."""
    cols = _detect_columns(sheet)
    header_row = cols["header_row"]
    lab_col = cols["lab_col"]
    workshop_col = cols["workshop_col"]
    enhancement_col = cols["enhancement_col"]
    uw_header_col = cols["uw_header_col"]
    uw_name_col = uw_header_col + 1
    uw_attr_col = uw_header_col + 2
    uw_value_col = uw_header_col + 3

    result: Dict[str, Dict[str, Any]] = {
        "labs": {},
        "workshop": {},
        "enhancements": {},
        "uw": {},
    }
    current_uw: Optional[str] = None
    for row_idx in range(header_row + 1, 121):
        lab_name = _clean(sheet.cell(row_idx, lab_col).value)
        lab_level = sheet.cell(row_idx, lab_col + 1).value
        if lab_name:
            canonical = LAB_ALIASES.get(lab_name, lab_name)
            if canonical in LAB_MAX_LEVELS and isinstance(lab_level, (int, float)):
                result["labs"][canonical] = int(round(lab_level))

        workshop_name = _clean(sheet.cell(row_idx, workshop_col).value)
        workshop_level = sheet.cell(row_idx, workshop_col + 1).value
        if workshop_name:
            canonical = WORKSHOP_ALIASES.get(workshop_name, workshop_name)
            if canonical in WORKSHOP_MAX_LEVELS and isinstance(workshop_level, (int, float)):
                result["workshop"][canonical] = int(round(workshop_level))

        enhancement_name = _clean(sheet.cell(row_idx, enhancement_col).value)
        enhancement_level = sheet.cell(row_idx, enhancement_col + 2).value
        if enhancement_name in ENHANCEMENT_MAX_LEVELS and isinstance(enhancement_level, (int, float)):
            result["enhancements"][enhancement_name] = int(round(enhancement_level))

        name_or_status = _clean(sheet.cell(row_idx, uw_name_col).value)
        attribute = _clean(sheet.cell(row_idx, uw_attr_col).value)
        value = sheet.cell(row_idx, uw_value_col).value
        if name_or_status in UW_NAMES:
            current_uw = name_or_status
            result["uw"].setdefault(current_uw, {"owned": None, "attributes": {}})
        elif name_or_status in {"UW Unlocked", "UW Locked"} and current_uw:
            result["uw"].setdefault(current_uw, {"owned": None, "attributes": {}})["owned"] = name_or_status == "UW Unlocked"
        elif current_uw and attribute in UW_ATTRIBUTE_META.get(current_uw, {}) and isinstance(value, (int, float)):
            result["uw"].setdefault(current_uw, {"owned": None, "attributes": {}})["attributes"][attribute] = value

    return result
