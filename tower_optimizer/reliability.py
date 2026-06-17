from __future__ import annotations

import io
import json
import os
import platform
import re
import shutil
import sys
import tempfile
import zipfile
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Dict, Iterable, List, Mapping, Optional

from openpyxl import load_workbook

SUPPORTED_WORKBOOK_VERSIONS: Dict[str, str] = {
    "Effective Paths": "5.06.04.00",
    "IDS Master": "3.1.2",
    "Player & Stuff": "4.1.6",
    "Laboratory": "3.0.5",
    "Workshop": "3.0.1",
    "Ultimate Weapon": "3.1.2",
    "Modules": "6.1.2",
    "Cards": "3.0.4",
    "Relics": "3.1.3",
    "Themes & Songs": "3.0.2",
    "Bots": "3.1",
    "Guardians": "3.0.2",
    "Vault": "3.1.1",
}

WORKBOOK_FILENAME_PATTERNS = [
    "Effective Paths",
    "IDS Master",
    "Player & Stuff",
    "Laboratory",
    "Workshop",
    "Ultimate Weapon",
    "Modules",
    "Cards",
    "Relics",
    "Themes & Songs",
    "Bots",
    "Guardians",
    "Vault",
]


def utc_now() -> str:
    return datetime.now(timezone.utc).isoformat()


def safe_name(value: str) -> str:
    cleaned = re.sub(r"[^A-Za-z0-9_.-]+", "_", str(value)).strip("._")
    return cleaned or "profile"


def _backup_root(profile_dir: Path) -> Path:
    root = profile_dir.parent / "backups" / "profiles"
    root.mkdir(parents=True, exist_ok=True)
    return root


def create_profile_backup(path: Path, profile_dir: Optional[Path] = None, reason: str = "manual") -> Optional[Path]:
    path = Path(path)
    if not path.exists():
        return None
    profile_dir = Path(profile_dir or path.parent)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S_%f")
    destination = _backup_root(profile_dir) / f"{path.stem}__{timestamp}__{safe_name(reason)}.json"
    shutil.copy2(path, destination)
    return destination


def atomic_save_json(path: Path, data: Mapping[str, Any], profile_dir: Optional[Path] = None) -> Path:
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    profile_dir = Path(profile_dir or path.parent)
    if path.exists():
        create_profile_backup(path, profile_dir, reason="autosave")

    fd, temp_name = tempfile.mkstemp(prefix=f".{path.stem}_", suffix=".tmp", dir=str(path.parent))
    try:
        with os.fdopen(fd, "w", encoding="utf-8") as handle:
            json.dump(data, handle, indent=2, ensure_ascii=False)
            handle.flush()
            os.fsync(handle.fileno())
        os.replace(temp_name, path)
    except Exception:
        try:
            os.unlink(temp_name)
        except OSError:
            pass
        raise
    return path


def list_profile_backups(profile_dir: Path, profile_name: Optional[str] = None) -> List[Dict[str, Any]]:
    root = _backup_root(Path(profile_dir))
    prefix = f"{safe_name(profile_name)}__" if profile_name else ""
    rows: List[Dict[str, Any]] = []
    for path in sorted(root.glob(f"{prefix}*.json"), key=lambda p: p.stat().st_mtime, reverse=True):
        stat = path.stat()
        parts = path.stem.split("__")
        rows.append({
            "file": path.name,
            "path": str(path),
            "profile": parts[0] if parts else path.stem,
            "created": datetime.fromtimestamp(stat.st_mtime).isoformat(timespec="seconds"),
            "size_kb": round(stat.st_size / 1024, 1),
            "reason": parts[-1] if len(parts) >= 3 else "unknown",
        })
    return rows


def restore_profile_backup(backup_path: Path, target_path: Path, profile_dir: Optional[Path] = None) -> Path:
    backup_path = Path(backup_path)
    target_path = Path(target_path)
    if not backup_path.exists():
        raise FileNotFoundError(backup_path)
    # Validate JSON before replacing a working profile.
    with open(backup_path, "r", encoding="utf-8") as handle:
        json.load(handle)
    if target_path.exists():
        create_profile_backup(target_path, Path(profile_dir or target_path.parent), reason="before_restore")
    target_path.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(backup_path, target_path)
    return target_path


def parse_version(value: str) -> tuple[int, ...]:
    text = str(value or "")
    match = re.search(r"(?:^|[^0-9])v?\s*(\d+(?:\.\d+)+)", text, flags=re.I)
    if not match:
        return ()
    return tuple(int(part) for part in match.group(1).split("."))


def compare_versions(found: str, supported: str) -> str:
    left = parse_version(found)
    right = parse_version(supported)
    if not left or not right:
        return "Unknown"
    length = max(len(left), len(right))
    left += (0,) * (length - len(left))
    right += (0,) * (length - len(right))
    if left > right:
        return "Newer than supported"
    if left < right:
        return "Older than supported"
    return "Supported"


def identify_workbook(filename: str) -> Dict[str, str]:
    name = Path(filename).stem
    kind = "Unknown"
    for candidate in WORKBOOK_FILENAME_PATTERNS:
        if candidate.lower() in name.lower():
            kind = candidate
            break
    match = re.search(r"v\s*(\d+(?:\.\d+)+)", name, flags=re.I)
    version = match.group(1) if match else ""
    supported = SUPPORTED_WORKBOOK_VERSIONS.get(kind, "")
    return {
        "kind": kind,
        "version": version,
        "supported_version": supported,
        "status": compare_versions(version, supported) if supported else "Unknown",
    }


def workbook_compatibility(filename: str, payload: bytes) -> Dict[str, Any]:
    info: Dict[str, Any] = {"filename": filename, **identify_workbook(filename)}
    try:
        wb = load_workbook(io.BytesIO(payload), read_only=True, data_only=False, keep_links=False)
        info["sheets"] = len(wb.sheetnames)
        info["has_export"] = "EXPORT" in wb.sheetnames
        info["has_ids"] = "IDS" in wb.sheetnames
        info["has_private_ids"] = "_IDS" in wb.sheetnames
        if info["kind"] == "Unknown":
            names = set(wb.sheetnames)
            if {"eEcon", "eDamage", "eHP"}.issubset(names):
                info["kind"] = "Effective Paths"
                info["supported_version"] = SUPPORTED_WORKBOOK_VERSIONS["Effective Paths"]
                info["status"] = compare_versions(info.get("version", ""), info["supported_version"])
        wb.close()
    except Exception as exc:
        info["error"] = str(exc)
    return info


def _fallback_formula_value(value: Any) -> Any:
    if not isinstance(value, str):
        return value
    # Effective Paths exported by Excel frequently wraps cached values in
    # IFERROR(__xludf.DUMMYFUNCTION(...), <fallback>).  Pull the fallback.
    if "DUMMYFUNCTION" not in value:
        return value
    match = re.search(r',\s*("(?:[^"]|"")*"|TRUE|FALSE|-?\d+(?:\.\d+)?)\s*\)\s*$', value, flags=re.I)
    if not match:
        return value
    token = match.group(1)
    if token.upper() == "TRUE":
        return True
    if token.upper() == "FALSE":
        return False
    if token.startswith('"') and token.endswith('"'):
        return token[1:-1].replace('""', '"')
    try:
        number = float(token)
        return int(number) if number.is_integer() else number
    except ValueError:
        return token


def extract_effective_paths_caps(payload: bytes) -> Dict[str, Dict[str, int]]:
    wb = load_workbook(io.BytesIO(payload), read_only=True, data_only=False, keep_links=False)
    if "_IDS" not in wb.sheetnames:
        wb.close()
        return {"workshop": {}, "labs": {}}
    ws = wb["_IDS"]
    workshop: Dict[str, int] = {}
    labs: Dict[str, int] = {}

    # Read the relevant cells in one pass. Repeated ws.cell access is very slow
    # on read-only worksheets because it can restart XML iteration.
    for row_number, row in enumerate(
        ws.iter_rows(min_row=1, max_row=260, max_col=17, values_only=True), start=1
    ):
        if row_number >= 2:
            name = _fallback_formula_value(row[0])
            maximum = _fallback_formula_value(row[3])
            if isinstance(name, str) and name.strip() and isinstance(maximum, (int, float)):
                labs[name.strip()] = int(maximum)
        if row_number >= 4:
            name = _fallback_formula_value(row[5])
            maximum = _fallback_formula_value(row[16])
            if isinstance(name, str) and name.strip() and isinstance(maximum, (int, float)):
                workshop[name.strip()] = int(maximum)
    wb.close()
    return {"workshop": workshop, "labs": labs}


def compare_cap_maps(current: Mapping[str, int], incoming: Mapping[str, int]) -> List[Dict[str, Any]]:
    rows: List[Dict[str, Any]] = []
    names = sorted(set(current) | set(incoming))
    for name in names:
        old = current.get(name)
        new = incoming.get(name)
        if old is None:
            status = "Added"
        elif new is None:
            status = "Missing from upload"
        elif int(old) != int(new):
            status = "Maximum changed"
        else:
            continue
        rows.append({"Entry": name, "Embedded maximum": old, "Uploaded maximum": new, "Change": status})
    return rows


def build_diagnostic_zip(
    profile: Mapping[str, Any],
    app_version: str,
    audit_issues: Iterable[Mapping[str, Any]],
    self_check: Mapping[str, Any],
) -> bytes:
    buffer = io.BytesIO()
    system = {
        "generated_at": utc_now(),
        "app_version": app_version,
        "python": sys.version,
        "platform": platform.platform(),
        "profile_name": profile.get("name", "unknown"),
        "supported_workbooks": SUPPORTED_WORKBOOK_VERSIONS,
    }
    # Diagnostics intentionally omit raw imported workbook files. The profile is
    # included because this package is generated explicitly by the user.
    with zipfile.ZipFile(buffer, "w", compression=zipfile.ZIP_DEFLATED) as archive:
        archive.writestr("system.json", json.dumps(system, indent=2, default=str))
        archive.writestr("profile.json", json.dumps(profile, indent=2, default=str))
        archive.writestr("audit.json", json.dumps(list(audit_issues), indent=2, default=str))
        archive.writestr("self_check.json", json.dumps(dict(self_check), indent=2, default=str))
    return buffer.getvalue()
