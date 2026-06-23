"""Parse Effective Paths ROI reference workbooks without Streamlit."""
from __future__ import annotations

import io
import re
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Dict

from openpyxl import load_workbook

from .engines.core import parse_tower_number

ROI_PATH_SPECS: Dict[str, Dict[str, Any]] = {
    "econ_lab": {
        "sheet": "eEcon", "start_row": 6, "end_row": 100,
        "columns": {"Upgrade": "F", "Level": "G", "Cost": "H", "Duration": "I", "ROI": "J", "Result": "K", "Cumulative": "L"},
        "title": "Economy · Lab time path", "resource": "Time", "metric": "CPK",
    },
    "econ_stone": {
        "sheet": "eEcon", "start_row": 6, "end_row": 100,
        "columns": {"Upgrade": "N", "Level": "O", "Cost": "P", "ROI": "Q", "Result": "R", "Cumulative": "S"},
        "title": "Economy · Stone path", "resource": "Stones", "metric": "CPK",
    },
    "econ_coin": {
        "sheet": "eEcon", "start_row": 6, "end_row": 100,
        "columns": {"Upgrade": "U", "Level": "V", "Cost": "W", "ROI": "X", "Result": "Y", "Cumulative": "Z"},
        "title": "Economy · Coin path", "resource": "Coins", "metric": "CPK",
    },
    "econ_discount": {
        "sheet": "eEcon", "start_row": 6, "end_row": 100,
        "columns": {"Upgrade": "AB", "Level": "AC", "Duration": "AD", "ROI": "AE", "Result": "AF", "Cumulative": "AG"},
        "title": "Economy · Lab discount path", "resource": "Time", "metric": "Effective coin value",
    },
    "damage_lab": {
        "sheet": "eDamage", "start_row": 6, "end_row": 150,
        "columns": {"Upgrade": "F", "Level": "G", "Cost": "H", "Duration": "I", "ROI": "J", "Result": "K", "Cumulative": "L"},
        "title": "Damage · Lab time path", "resource": "Time", "metric": "eDMG",
    },
    "damage_stone": {
        "sheet": "eDamage", "start_row": 6, "end_row": 150,
        "columns": {"Upgrade": "N", "Level": "O", "Cost": "P", "ROI": "Q", "Result": "R", "Cumulative": "S"},
        "title": "Damage · Stone path", "resource": "Stones", "metric": "eDMG",
    },
    "damage_coin": {
        "sheet": "eDamage", "start_row": 6, "end_row": 150,
        "columns": {"Upgrade": "U", "Level": "V", "Cost": "W", "ROI": "X", "Result": "Y", "Cumulative": "Z"},
        "title": "Damage · Coin path", "resource": "Coins", "metric": "eDMG",
    },
    "damage_key": {
        "sheet": "eDamage", "start_row": 6, "end_row": 150,
        "columns": {"Upgrade": "AB", "Level": "AC", "Cost": "AD", "ROI": "AE", "Result": "AF", "Cumulative": "AG"},
        "title": "Damage · Key path", "resource": "Keys", "metric": "eDMG",
    },
    "health_lab": {
        "sheet": "eHP", "start_row": 6, "end_row": 150,
        "columns": {"Upgrade": "F", "Level": "G", "Cost": "H", "Duration": "I", "ROI": "J", "Result": "K", "Cumulative": "L"},
        "title": "Health · Lab time path", "resource": "Time", "metric": "eHP",
    },
    "health_stone": {
        "sheet": "eHP", "start_row": 6, "end_row": 150,
        "columns": {"Upgrade": "N", "Level": "O", "Cost": "P", "ROI": "Q", "Result": "R", "Cumulative": "S"},
        "title": "Health · Stone path", "resource": "Stones", "metric": "eHP",
    },
    "health_coin": {
        "sheet": "eHP", "start_row": 6, "end_row": 150,
        "columns": {"Upgrade": "U", "Level": "V", "Cost": "W", "ROI": "X", "Result": "Y", "Cumulative": "Z"},
        "title": "Health · Coin path", "resource": "Coins", "metric": "eHP",
    },
    "regen_lab": {
        "sheet": "eHP", "start_row": 6, "end_row": 150,
        "columns": {"Upgrade": "AB", "Level": "AC", "Cost": "AD", "Duration": "AE", "ROI": "AF", "Result": "AG", "Cumulative": "AH"},
        "title": "Wall regen · Lab time path", "resource": "Time", "metric": "Wall regen",
    },
}


def _serialize_sheet_value(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.isoformat()
    if hasattr(value, "total_seconds") and not isinstance(value, (str, bytes)):
        try:
            seconds = int(value.total_seconds())
            days, seconds = divmod(seconds, 86400)
            hours, seconds = divmod(seconds, 3600)
            minutes, seconds = divmod(seconds, 60)
            parts = []
            if days:
                parts.append(f"{days}d")
            if hours or days:
                parts.append(f"{hours}h")
            if minutes or hours or days:
                parts.append(f"{minutes}m")
            parts.append(f"{seconds}s")
            return " ".join(parts)
        except Exception:
            return str(value)
    if isinstance(value, (int, float, bool, str)):
        return value
    return str(value)


def _clean_cell(value: Any) -> str:
    return str(value or "").strip()


def parse_roi_path_table(workbook: Any, path_key: str, spec: Dict[str, Any]) -> list[Dict[str, Any]]:
    sheet_name = spec["sheet"]
    if sheet_name not in workbook.sheetnames:
        return []
    sheet = workbook[sheet_name]
    rows: list[Dict[str, Any]] = []
    blank_streak = 0
    for row_number in range(int(spec["start_row"]), int(spec["end_row"]) + 1):
        row: Dict[str, Any] = {}
        for field, column in spec["columns"].items():
            row[field] = _serialize_sheet_value(sheet[f"{column}{row_number}"].value)
        upgrade = row.get("Upgrade")
        roi = row.get("ROI")
        if upgrade in (None, ""):
            blank_streak += 1
            if blank_streak >= 12 and rows:
                break
            continue
        blank_streak = 0
        upgrade_text = str(upgrade).strip()
        if not upgrade_text or upgrade_text.lower().startswith("not seeing path"):
            continue
        if all(row.get(field) in (None, "") for field in ["Level", "Cost", "Duration", "ROI", "Result"]):
            continue
        row["Rank"] = len(rows) + 1
        row["Path"] = path_key
        row["Resource"] = spec.get("resource")
        row["Metric"] = spec.get("metric")
        try:
            row["ROI Numeric"] = float(roi) if roi not in (None, "") else None
        except (TypeError, ValueError):
            row["ROI Numeric"] = None
        row["Cost Numeric"] = parse_tower_number(row.get("Cost"))
        rows.append(row)
    return rows


def detect_effective_paths_version(workbook: Any, filename: str) -> str:
    for sheet_name, cell in [("eEcon", "F3"), ("eDamage", "F3"), ("eHP", "F3")]:
        if sheet_name in workbook.sheetnames:
            text = _clean_cell(workbook[sheet_name][cell].value)
            match = re.search(r"v(\d+(?:\.\d+)*)", text, re.IGNORECASE)
            if match:
                return f"v{match.group(1)}"
    match = re.search(r"v\s*(\d+(?:\.\d+)+)", filename, re.IGNORECASE)
    return f"v{match.group(1)}" if match else "Unknown"


def parse_effective_paths_roi_reference_bytes(payload: bytes, *, filename: str) -> Dict[str, Any]:
    if Path(filename).suffix.lower() != ".xlsx":
        raise ValueError("ROI reference import requires the filled Effective Paths .xlsx workbook.")
    workbook = load_workbook(io.BytesIO(payload), read_only=False, data_only=True)
    missing = sorted({spec["sheet"] for spec in ROI_PATH_SPECS.values()} - set(workbook.sheetnames))
    if missing:
        raise ValueError("Missing Effective Paths result sheets: " + ", ".join(missing))

    paths: Dict[str, Any] = {}
    warnings_list: list[str] = []
    for path_key, spec in ROI_PATH_SPECS.items():
        rows = parse_roi_path_table(workbook, path_key, spec)
        paths[path_key] = {
            "title": spec["title"],
            "resource": spec["resource"],
            "metric": spec["metric"],
            "rows": rows,
        }
        if not rows:
            warnings_list.append(
                f"{spec['title']} had no cached recommendation rows. Recalculate the workbook and import again."
            )

    nonempty = sum(1 for path in paths.values() if path.get("rows"))
    if nonempty == 0:
        raise ValueError(
            "The workbook contains the result sheets, but no cached ROI paths were readable. "
            "Recalculate and save the workbook before importing it."
        )

    return {
        "source": {
            "filename": filename,
            "effective_paths_version": detect_effective_paths_version(workbook, filename),
            "imported_at": datetime.now(timezone.utc).isoformat(),
            "mode": "cached spreadsheet outputs",
        },
        "paths": paths,
        "warnings": warnings_list,
    }
